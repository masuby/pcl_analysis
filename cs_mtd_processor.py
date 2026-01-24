import pandas as pd
import numpy as np
import os
import re
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import os
import glob

# Configuration
# Define the base paths
base_dir = r"C:\Users\Daniel\Desktop\code\pcl\MTD"
input_folder = os.path.join(base_dir, "ROW_FILES")

# Output directory will be inside CS_MTD folder
output_base = os.path.join(base_dir, "CS_MTD")
output_dir = os.path.join(output_base, "split_supervisions_cs_mtd")

# Automatically find the CS MTD file
cs_files = glob.glob(os.path.join(input_folder, "*CS*MTD*.xlsx"))

if not cs_files:
    # Try alternative patterns
    cs_files = glob.glob(os.path.join(input_folder, "*CS*.xlsx"))
    
if not cs_files:
    raise FileNotFoundError(f"No CS MTD file found in {input_folder}")

# Use the first matching file
input_file = cs_files[0]
print(f"Found input file: {input_file}")

# Create output directories if they don't exist
if not os.path.exists(output_base):
    os.makedirs(output_base)
    print(f"Created base output directory: {output_base}")

if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    print(f"Created split files output directory: {output_dir}")
else:
    print(f"Output directory already exists: {output_dir}")

# Create output directory
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Function to clean and normalize branch names for matching
def normalize_branch_name(name):
    if pd.isna(name):
        return ""
    name_str = str(name).upper()
    name_str = re.sub(r'\s+', ' ', name_str)
    name_str = re.sub(r'CS\s*', '', name_str)
    name_str = re.sub(r'\([^)]*\)', '', name_str)
    name_str = name_str.replace('REGION', '').strip()
    name_str = name_str.replace('  ', ' ').strip()
    return name_str

# FIXED: Keep numbers as numeric
def format_numbers(df):
    df_formatted = df.copy()
    for col in df_formatted.columns:
        df_formatted[col] = pd.to_numeric(df_formatted[col], errors='ignore')
    return df_formatted

# Step 1: Read Excel file
print("Reading Excel file...")
xls = pd.ExcelFile(input_file)
sheet_names = xls.sheet_names

# Find sheets
mtd_sheet_name = None
listing_sheet_name = None

for sheet in sheet_names:
    sheet_upper = sheet.upper()
    if 'MTD' in sheet_upper and 'CS' in sheet_upper:
        mtd_sheet_name = sheet
    elif 'LISTING' in sheet_upper:
        listing_sheet_name = sheet

print(f"Found sheets: MTD='{mtd_sheet_name}', LISTING='{listing_sheet_name}'")

# Read MTD sheet
mtd_raw = pd.read_excel(input_file, sheet_name=mtd_sheet_name, header=None)

# Find header row
mtd_header_row = None
for i in range(min(10, len(mtd_raw))):
    if any('BRANCH' in str(val).upper() for val in mtd_raw.iloc[i].astype(str).tolist()):
        mtd_header_row = i
        break

print(f"\nMTD header row found at: {mtd_header_row}")

# Read with header
mtd_df = pd.read_excel(input_file, sheet_name=mtd_sheet_name, header=mtd_header_row)

# Read Sales Listing sheet
listing_raw = pd.read_excel(input_file, sheet_name=listing_sheet_name, header=None)
listing_header_row = None
for i in range(min(10, len(listing_raw))):
    if any('SUPERVISION' in str(val).upper() for val in listing_raw.iloc[i].astype(str).tolist()):
        listing_header_row = i
        break

sales_listing_df = pd.read_excel(input_file, sheet_name=listing_sheet_name, header=listing_header_row)

# Identify supervision column in listing
supervision_column = None
for col in sales_listing_df.columns:
    if 'SUPERVISION' in str(col).upper():
        supervision_column = col
        break

unique_supervisions = sales_listing_df[supervision_column].dropna().unique()
print(f"\nFound {len(unique_supervisions)} supervisions: {list(unique_supervisions)}")

# Extract date
date_info = ""
for i in range(min(3, len(mtd_raw))):
    for val in mtd_raw.iloc[i].astype(str).tolist():
        if 'MTD' in val.upper() and '2025' in val:
            date_info = val
            break
    if date_info:
        break

print(f"\nReport date: {date_info}")

# Find branch column in MTD
branch_column_mtd = None
for col in mtd_df.columns:
    if 'BRANCH' in str(col).upper():
        branch_column_mtd = col
        break

print(f"\nBranch column in MTD: {branch_column_mtd}")

# ============================================
# DETECT PERCENTAGE COLUMNS FROM HEADERS
# ============================================
# Find columns that have "%" in their header names (e.g., "NEW %", "REFINANCE %")
percentage_columns = []
for col in mtd_df.columns:
    col_str = str(col).upper().strip()
    if "%" in col_str:
        percentage_columns.append(col)
        print(f"  Found percentage column: '{col}'")

if percentage_columns:
    print(f"\n  Identified {len(percentage_columns)} percentage columns (will format as percentage without conversion)")

# ============================================
# NEW LOGIC: Find ALL supervision rows in MTD sheet
# ============================================

# First, let's identify what rows contain supervision names
# Supervision names are in the branch column but are not actual branches
print("\nScanning MTD sheet for supervision rows...")

# Get all unique values in the branch column
if branch_column_mtd:
    all_branch_values = mtd_df[branch_column_mtd].dropna().unique()
    print(f"Found {len(all_branch_values)} unique values in branch column")
    
    # Identify which ones might be supervisions (based on sales listing supervisions)
    possible_supervisions = []
    for value in all_branch_values:
        value_str = str(value).strip()
        # Check if this matches any supervision from sales listing
        for supervision in unique_supervisions:
            if value_str == supervision or supervision in value_str:
                possible_supervisions.append(value_str)
                break
    
    print(f"Possible supervision rows in MTD: {possible_supervisions}")

# Process each supervision
for supervision in unique_supervisions:
    print(f"\n{'='*60}")
    print(f"Processing supervision: {supervision}")
    print(f"{'='*60}")
    
    try:
        # ============================================
        # STRATEGY: Find supervision row and collect data below it
        # ============================================
        supervision_mtd_indices = []
        
        if branch_column_mtd:
            # Look for EXACT match of supervision in branch column
            # DO NOT normalize supervision names!
            exact_supervision_rows = mtd_df[mtd_df[branch_column_mtd] == supervision]
            
            if len(exact_supervision_rows) > 0:
                print(f"  Found exact supervision row(s) for '{supervision}'")
                
                for sup_idx in exact_supervision_rows.index:
                    print(f"  Supervision at row {sup_idx}")
                    
                    # INCLUDE the supervision row itself
                    supervision_mtd_indices.append(sup_idx)
                    
                    # Start collecting from the next row
                    current_idx = sup_idx + 1
                    
                    while current_idx < len(mtd_df):
                        if current_idx in mtd_df.index:
                            current_row = mtd_df.loc[current_idx]
                            branch_val = current_row.get(branch_column_mtd, None)
                            
                            if pd.notna(branch_val):
                                branch_str = str(branch_val).strip()
                                
                                # Check if we hit another supervision
                                # Compare with all other supervisions (EXACT match, no normalization)
                                is_another_supervision = False
                                for other_sup in unique_supervisions:
                                    if other_sup != supervision and branch_str == other_sup:
                                        is_another_supervision = True
                                        print(f"    Stopping at another supervision: '{branch_str}' at row {current_idx}")
                                        break
                                
                                # Check for "Grand Total"
                                if 'GRAND TOTAL' in branch_str.upper():
                                    print(f"    Stopping at 'Grand Total' at row {current_idx}")
                                    break
                                elif is_another_supervision:
                                    break
                                else:
                                    # This is a branch under this supervision
                                    supervision_mtd_indices.append(current_idx)
                                    print(f"    Added branch: '{branch_str}' at row {current_idx}")
                            
                            # Also check if we hit an empty row (end of data section)
                            elif current_idx > sup_idx + 50:  # Limit search to 50 rows
                                print(f"    Stopping after searching 50 rows")
                                break
                        
                        current_idx += 1
            else:
                print(f"  WARNING: No exact match found for supervision '{supervision}'")
                print(f"  Trying to find partial match...")
                
                # Try partial matching (but still not normalizing)
                for idx, row in mtd_df.iterrows():
                    branch_val = row.get(branch_column_mtd, None)
                    if pd.notna(branch_val):
                        branch_str = str(branch_val).strip()
                        # Check if supervision is in branch string or vice versa
                        if supervision in branch_str or branch_str in supervision:
                            print(f"  Found partial match at row {idx}: '{branch_str}'")
                            
                            # Include this row
                            supervision_mtd_indices.append(idx)
                            
                            # Collect branches below
                            current_idx = idx + 1
                            
                            while current_idx < len(mtd_df):
                                if current_idx in mtd_df.index:
                                    current_row = mtd_df.loc[current_idx]
                                    next_branch_val = current_row.get(branch_column_mtd, None)
                                    
                                    if pd.notna(next_branch_val):
                                        next_branch_str = str(next_branch_val).strip()
                                        
                                        # Check if we hit another supervision
                                        is_another_supervision = False
                                        for other_sup in unique_supervisions:
                                            if other_sup != supervision and next_branch_str == other_sup:
                                                is_another_supervision = True
                                                print(f"    Stopping at another supervision: '{next_branch_str}' at row {current_idx}")
                                                break
                                        
                                        if 'GRAND TOTAL' in next_branch_str.upper():
                                            print(f"    Stopping at 'Grand Total' at row {current_idx}")
                                            break
                                        elif is_another_supervision:
                                            break
                                        else:
                                            supervision_mtd_indices.append(current_idx)
                                            print(f"    Added branch: '{next_branch_str}' at row {current_idx}")
                                    
                                    elif current_idx > idx + 50:
                                        print(f"    Stopping after searching 50 rows")
                                        break
                                
                                current_idx += 1
                            break
        
        # Sort and get unique indices
        supervision_mtd_indices = sorted(set(supervision_mtd_indices))
        
        # Get the actual branch names
        matched_mtd_branches = []
        if branch_column_mtd and supervision_mtd_indices:
            for idx in supervision_mtd_indices:
                if idx in mtd_df.index:
                    branch_val = mtd_df.loc[idx, branch_column_mtd]
                    if pd.notna(branch_val):
                        matched_mtd_branches.append(branch_val)
        
        print(f"\n  Final MTD rows for '{supervision}':")
        print(f"  - Row indices ({len(supervision_mtd_indices)}): {supervision_mtd_indices}")
        print(f"  - Branches/Supervisions: {matched_mtd_branches}")
        
        if not supervision_mtd_indices:
            print(f"  ERROR: No MTD rows found for {supervision}")
            continue
        
        # Extract MTD data for this supervision (INCLUDING supervision row)
        supervision_mtd_data = mtd_df.iloc[supervision_mtd_indices].copy()
        
        # Get sales listing data for this supervision
        supervision_sales_data = sales_listing_df[sales_listing_df[supervision_column] == supervision].copy()
        
        supervision_mtd_data_formatted = format_numbers(supervision_mtd_data)
        supervision_sales_data_formatted = format_numbers(supervision_sales_data)

        # Save file
        supervision_clean = supervision.replace(" ", "-").replace("/", "-").replace("(", "").replace(")", "")
        filename = f"CS_{supervision_clean}_{date_info.replace(' ', '-')}.xlsx"
        file_path = os.path.join(output_dir, filename)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            workbook = writer.book
            
            bold_font = Font(bold=True)
            orange_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            # Write MTD sheet
            mtd_sheet = workbook.create_sheet(mtd_sheet_name, 0)
            current_row = 1

            # Write headers
            for idx, row in mtd_raw.iloc[:mtd_header_row+1].iterrows():
                for col_idx, val in enumerate(row.values, 1):
                    cell = mtd_sheet.cell(row=current_row, column=col_idx, value=val)
                    cell.font = bold_font
                    cell.fill = white_fill
                current_row += 1

            # Write MTD rows (INCLUDING supervision row)
            # Get column names to check if they are percentage columns
            supervision_column_names = list(supervision_mtd_data_formatted.columns)
            
            for data_row_idx, row in enumerate(dataframe_to_rows(supervision_mtd_data_formatted, index=False, header=False)):
                excel_row = current_row + data_row_idx
                
                for col_idx, val in enumerate(row, 1):
                    cell = mtd_sheet.cell(row=excel_row, column=col_idx, value=val)
                    cell.font = bold_font
                    cell.fill = orange_fill

                    # Apply formatting based on percentage column detection
                    if isinstance(val, (int, float)):
                        # Check if this column is a percentage column (based on header name)
                        # col_idx is 1-based Excel column, but we need 0-based index for list
                        col_name_idx = col_idx - 1
                        if col_name_idx < len(supervision_column_names):
                            col_name = supervision_column_names[col_name_idx]
                            is_percentage_col = col_name in percentage_columns
                            
                            if is_percentage_col:
                                # Format as percentage - keep original value, no conversion needed
                                # Values like 79.37 will display as 79.37%, values like 150 will display as 150.00%
                                cell.number_format = "0.00%"
                            else:
                                # Original logic: percentage if between 0 and 1, otherwise accounting
                                if 0 < val < 1:
                                    cell.number_format = "0.00%"
                                else:
                                    cell.number_format = "#,##0.00"  # Accounting style
                        else:
                            # Fallback: original logic if column name not found
                            if 0 < val < 1:
                                cell.number_format = "0.00%"
                            else:
                                cell.number_format = "#,##0.00"

            # Write Sales Listing sheet
            listing_sheet = workbook.create_sheet(listing_sheet_name, 1)
            current_row = 1

            for idx, row in listing_raw.iloc[:listing_header_row+1].iterrows():
                for col_idx, val in enumerate(row.values, 1):
                    cell = listing_sheet.cell(row=current_row, column=col_idx, value=val)
                    cell.font = bold_font
                    cell.fill = white_fill
                current_row += 1

            for r_idx, row in enumerate(dataframe_to_rows(supervision_sales_data_formatted, index=False, header=False), current_row):
                for col_idx, val in enumerate(row, 1):
                    cell = listing_sheet.cell(row=r_idx, column=col_idx, value=val)
                    cell.font = bold_font
                    cell.fill = orange_fill

                    # FIXED: Apply accounting & percent formats
                    if isinstance(val, (int, float)):
                        if 0 < val < 1:
                            cell.number_format = "0.00%"
                        else:
                            cell.number_format = "#,##0.00"

            # Auto-adjust column widths
            for sheet in workbook.sheetnames:
                ws = workbook[sheet]
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
                    ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

        print(f"\n  ✓ Saved: {filename}")

    except Exception as e:
        print(f"\n  ✗ Error on {supervision}: {str(e)}")
        import traceback
        traceback.print_exc()

print("\nProcessing Complete!")