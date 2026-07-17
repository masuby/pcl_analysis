import os
import shutil
import subprocess
import sys
import glob
import re as _re_date
import openpyxl
import xlwings as xw
import time
from dotenv import load_dotenv

# ============================================================================
# Ask for the report date up front (used to date-stamp the final report file).
# Enter as DD/MM, e.g. 15/06. The day is appended to the generated report name
# at the very end:  "Management Report2026-06.xlsx" -> "Management Report2026-06-15.xlsx"
# ============================================================================
def _parse_ddmm(raw):
    m = _re_date.match(r'^(\d{1,2})\s*/\s*(\d{1,2})$', (raw or "").strip())
    if m:
        d, mo = int(m.group(1)), int(m.group(2))
        if 1 <= d <= 31 and 1 <= mo <= 12:
            return d, mo
    return None


def _send_choice() -> str:
    """--send yes|no from the web orchestrator (default: no).
    'no' = build the report but DO NOT run send_email_tz.exe (test run)."""
    import argparse as _ap
    _p = _ap.ArgumentParser(add_help=False)
    _p.add_argument("--send", default="no")
    _known, _ = _p.parse_known_args()
    return "yes" if str(_known.send).strip().lower() in ("yes", "actual", "y") else "no"


SEND_EMAILS = _send_choice()


def _prompt_report_day():
    # Accept --date DD/MM (web orchestrator) or PCL_REPORT_DATE env; else prompt.
    import argparse as _ap
    _p = _ap.ArgumentParser(add_help=False)
    _p.add_argument("--date", default=os.environ.get("PCL_REPORT_DATE"))
    _known, _ = _p.parse_known_args()
    parsed = _parse_ddmm(_known.date) if _known.date else None
    if parsed:
        return parsed
    while True:
        try:
            raw = input("Enter report date as DD/MM (e.g. 15/06): ").strip()
        except EOFError:
            raise SystemExit("No --date provided and no interactive input available.")
        parsed = _parse_ddmm(raw)
        if parsed:
            return parsed
        print("  ⚠️  Invalid format. Please enter the date as DD/MM, e.g. 15/06.")

REPORT_DAY, REPORT_MONTH = _prompt_report_day()
print(f"📅 Report will be date-stamped as day {REPORT_DAY:02d} (month {REPORT_MONTH:02d}).")

# ============================================================================
# CONFIGURATION
# ============================================================================

# Source folder (dynamic — orchestrator sets PCL_AUTOMATION_ROOT; fallback = this
# script's folder, i.e. .../Automation/Management).
_MGMT_DIR = os.environ.get("PCL_MANAGEMENT_DIR") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "Management")
    if os.environ.get("PCL_AUTOMATION_ROOT") else os.path.dirname(os.path.abspath(__file__)))
source_folder = os.path.join(_MGMT_DIR, "ROW_FILES")

# Destination folder (where files are copied to and cleaned)
destination_folder = os.environ.get("PCL_MANAGEMENT_OUT", os.path.join(_MGMT_DIR, "OUTPUT"))

# Path to the executable. The exe reads its input .xlsx from its working dir
# (destination_folder, where STEP 3 copies them), so only the BINARY has to be
# located. It's kept in the Email folder; look there first, then a few sensible
# places, with a PCL_MGMT_EXE env override for the server.
_EXE_NAME = "plattz_generateReportv4.exe"


def _find_report_exe():
    override = os.environ.get("PCL_MGMT_EXE")
    if override and os.path.exists(override):
        return override
    for _base in (os.path.join(_MGMT_DIR, "Email"), _MGMT_DIR,
                  destination_folder, source_folder):
        _cand = os.path.join(_base, _EXE_NAME)
        if os.path.exists(_cand):
            return _cand
    return os.path.join(_MGMT_DIR, "Email", _EXE_NAME)  # best-guess default


exe_path = _find_report_exe()

# Files to copy
files_to_copy = ["Loan.xlsx", "Clientstz.xlsx", "Settlements.xlsx", "Users.xlsx"]

# ============================================================================
# STEP 1: Delete all xlsx files in destination folder
# ============================================================================

print("="*80)
print("STEP 1: Cleaning destination folder")
print("="*80)

if not os.path.exists(destination_folder):
    print(f"Creating destination folder: {destination_folder}")
    os.makedirs(destination_folder)
else:
    print(f"Destination folder exists: {destination_folder}")

# Find all xlsx files in destination folder
xlsx_files = glob.glob(os.path.join(destination_folder, "*.xlsx"))
print(f"Found {len(xlsx_files)} xlsx file(s) in destination folder")

# Delete all xlsx files
deleted_count = 0
for file_path in xlsx_files:
    try:
        os.remove(file_path)
        deleted_count += 1
        print(f"  Deleted: {os.path.basename(file_path)}")
    except Exception as e:
        print(f"  ⚠️ Error deleting {os.path.basename(file_path)}: {e}")

print(f"Deleted {deleted_count} xlsx file(s)")

# ============================================================================
# STEP 2: Modify Loan.xlsx worksheet name before copying
# ============================================================================

print("\n" + "="*80)
print("STEP 2: Modifying Loan.xlsx worksheet name")
print("="*80)

loan_file_path = os.path.join(source_folder, "Loan.xlsx")
if os.path.exists(loan_file_path):
    try:
        wb = openpyxl.load_workbook(loan_file_path)
        ws = wb.active
        ws.title = "Loan Accounts"
        wb.save(loan_file_path)
        print("  ✅ Renamed worksheet to 'Loan Accounts' in Loan.xlsx")
    except Exception as e:
        print(f"  ⚠️ Error modifying Loan.xlsx: {e}")
else:
    print(f"  ⚠️ Loan.xlsx not found at: {loan_file_path}")

# ============================================================================
# STEP 3: Copy required files to destination folder
# ============================================================================

print("\n" + "="*80)
print("STEP 3: Copying files to destination folder")
print("="*80)

copied_count = 0
for file_name in files_to_copy:
    source_path = os.path.join(source_folder, file_name)
    destination_path = os.path.join(destination_folder, file_name)
    
    if os.path.exists(source_path):
        try:
            shutil.copy2(source_path, destination_path)
            copied_count += 1
            print(f"  ✅ Copied: {file_name}")
        except Exception as e:
            print(f"  ⚠️ Error copying {file_name}: {e}")
    else:
        print(f"  ⚠️ Source file not found: {source_path}")

print(f"Copied {copied_count} file(s) to destination folder")

# ============================================================================
# STEP 4: Run the executable with interactive inputs
# ============================================================================

print("\n" + "="*80)
print("STEP 4: Running plattz_generateReportv4.exe")
print("="*80)

# Change to destination folder for relative paths
os.chdir(destination_folder)

# Prepare inputs for the executable
# The executable asks for:
# 1. Users.xlsx
# 2. Clientstz.xlsx
# 3. Loan.xlsx
# 4. Settlements.xlsx

inputs = [
    "Users.xlsx\n",
    "Clientstz.xlsx\n",
    "Loan.xlsx\n",
    "Settlements.xlsx\n"
]

# Combine all inputs into a single string
input_string = "".join(inputs)

print(f"Executable path: {exe_path}")
print("Inputs to be sent:")
print("  - Users.xlsx")
print("  - Clientstz.xlsx")
print("  - Loan.xlsx")
print("  - Settlements.xlsx")
print("\nStarting executable...")

try:
    # Run the executable
    # Using Popen with stdin to send inputs
    process = subprocess.Popen(
        exe_path,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        cwd=destination_folder
    )
    
    # Send inputs and get output
    stdout, stderr = process.communicate(input=input_string, timeout=None)
    
    # Print output
    if stdout:
        print("\nExecutable output:")
        print(stdout)
    
    if stderr:
        print("\nExecutable errors/warnings:")
        print(stderr)
    
    # Check return code
    if process.returncode == 0:
        print("\n✅ Executable completed successfully!")
    else:
        print(f"\n⚠️ Executable exited with code: {process.returncode}")
        
except FileNotFoundError:
    print(f"\n❌ Error: Executable not found at: {exe_path}")
    print("Please update the 'exe_path' variable in the script with the correct path.")
except subprocess.TimeoutExpired:
    print("\n⚠️ Executable timed out")
except Exception as e:
    print(f"\n❌ Error running executable: {e}")

# ============================================================================
# STEP 4b: Run create_product_sales_report.py to generate product totals
# ============================================================================

print("\n" + "="*80)
print("STEP 4b: Running create_product_sales_report.py")
print("="*80)

script_dir = os.path.dirname(os.path.abspath(__file__))
product_sales_script = os.path.join(script_dir, "create_product_sales_report.py")
product_sales_report_path = os.path.join(script_dir, "Product_Sales_Report.xlsx")

if os.path.exists(product_sales_script):
    try:
        result = subprocess.run(
            [sys.executable, product_sales_script],
            cwd=script_dir,
            capture_output=False,
            text=True,
            timeout=180
        )
        if result.returncode == 0:
            print("✅ create_product_sales_report.py completed successfully")
        else:
            print(f"⚠️ create_product_sales_report.py exited with code {result.returncode}")
    except subprocess.TimeoutExpired:
        print("⚠️ create_product_sales_report.py timed out")
    except Exception as e:
        print(f"⚠️ Error running create_product_sales_report.py: {e}")
else:
    print(f"⚠️ Script not found: {product_sales_script}")

# ============================================================================
# STEP 5: Process Management file - Update Country row calculations
# ============================================================================

print("\n" + "="*80)
print("STEP 5: Processing Management file - Updating Country row")
print("="*80)

# Wait a moment for the file to be fully created
time.sleep(2)

# ----------------------------------------------------------------------------
# STEP 5a: Add loan-count columns (# of New Loan / Repeat Loan / Reactivated /
# Refinanced clients) to the Country sheet first, via add_loan_counts_to_country.py.
# This inserts the count columns (matched from Product_Sales_Report.xlsx) and
# colours them, before the amount/Country calculations below.
# ----------------------------------------------------------------------------
add_counts_script = os.path.join(script_dir, "add_loan_counts_to_country.py")
if os.path.exists(add_counts_script):
    try:
        result = subprocess.run(
            [sys.executable, add_counts_script],
            cwd=script_dir,
            capture_output=False,
            text=True,
            timeout=180
        )
        if result.returncode == 0:
            print("✅ add_loan_counts_to_country.py completed successfully")
        else:
            print(f"⚠️ add_loan_counts_to_country.py exited with code {result.returncode}")
    except subprocess.TimeoutExpired:
        print("⚠️ add_loan_counts_to_country.py timed out")
    except Exception as e:
        print(f"⚠️ Error running add_loan_counts_to_country.py: {e}")
else:
    print(f"⚠️ Script not found: {add_counts_script}")

# Load product totals from Product_Sales_Report.xlsx: per-product rows + Grand Total (for Country)
# Product Sales has Summary by Product block: header "Product","New Business",... then one row per product, then "Grand Total (Mapped)".
product_totals_from_report = None  # Grand Total (Mapped) -> for Country row
product_totals_by_name = {}         # product_name -> {New Business, Repeat Business, Reactivation, Refinance, Total}
if os.path.exists(product_sales_report_path):
    try:
        wb_prod = openpyxl.load_workbook(product_sales_report_path, data_only=True)
        if "Product Sales" in wb_prod.sheetnames:
            ws_prod = wb_prod["Product Sales"]
            grand_total_row = None
            summary_header_row = None
            for row in range(2, ws_prod.max_row + 1):
                a1 = ws_prod.cell(row, 1).value
                if a1 is None:
                    continue
                name = str(a1).strip()
                if name == "Grand Total (Mapped)":
                    grand_total_row = row
                    product_totals_from_report = {
                        "New Business": ws_prod.cell(row, 2).value or 0,
                        "Repeat Business": ws_prod.cell(row, 3).value or 0,
                        "Reactivation": ws_prod.cell(row, 4).value or 0,
                        "Refinance": ws_prod.cell(row, 5).value or 0,
                        "Total": ws_prod.cell(row, 6).value or 0,
                    }
                    break
            for row in range(2, ws_prod.max_row + 1):
                a1 = ws_prod.cell(row, 1).value
                if a1 is not None and str(a1).strip() == "Product":
                    summary_header_row = row
                    break
            if summary_header_row is not None and grand_total_row is not None:
                for row in range(summary_header_row + 1, grand_total_row):
                    name = ws_prod.cell(row, 1).value
                    if name is None or not str(name).strip():
                        continue
                    name = str(name).strip()
                    try:
                        b = ws_prod.cell(row, 2).value or 0
                        c = ws_prod.cell(row, 3).value or 0
                        d = ws_prod.cell(row, 4).value or 0
                        e = ws_prod.cell(row, 5).value or 0
                        f = ws_prod.cell(row, 6).value or 0
                        for v in (b, c, d, e, f):
                            if not isinstance(v, (int, float)):
                                float(v)
                        product_totals_by_name[name] = {
                            "New Business": b,
                            "Repeat Business": c,
                            "Reactivation": d,
                            "Refinance": e,
                            "Total": f,
                        }
                    except (TypeError, ValueError):
                        pass
            if product_totals_from_report:
                print(f"Loaded Grand Total (Country): New Business={product_totals_from_report['New Business']}, "
                      f"Repeat={product_totals_from_report['Repeat Business']}, "
                      f"Reactivation={product_totals_from_report['Reactivation']}, "
                      f"Refinance={product_totals_from_report['Refinance']}, "
                      f"Disbursement Total={product_totals_from_report['Total']}")
            print(f"Loaded {len(product_totals_by_name)} product rows for branch alignment.")
        wb_prod.close()
    except Exception as e:
        print(f"⚠️ Could not load product totals: {e}")
else:
    print("⚠️ Product_Sales_Report.xlsx not found; Country/branch rows will use summed values only")

# Find the Management file
management_files = glob.glob(os.path.join(destination_folder, "Management*.xlsx"))
if not management_files:
    # Try with .xlsm extension as well
    management_files = glob.glob(os.path.join(destination_folder, "Management*.xlsm"))

management_file = None  # Initialize variable
if not management_files:
    print("⚠️ No Management file found in destination folder.")
    print("⚠️ Skipping Country row update.")
else:
    management_file = management_files[0]  # Take the first one found
    print(f"Found Management file: {os.path.basename(management_file)}")
    
    try:
        # Open the workbook with xlwings
        app = xw.App(visible=False, add_book=False)
        wb = app.books.open(management_file)
        
        # Get the Country sheet
        try:
            ws = wb.sheets["Country"]
            print("✅ Opened 'Country' sheet")
        except:
            print("⚠️ 'Country' sheet not found. Available sheets:")
            for sheet in wb.sheets:
                print(f"  - {sheet.name}")
            wb.close()
            app.quit()
            raise Exception("Country sheet not found")
        
        # Find the Branch column
        branch_col = None
        header_row = 1  # Assuming headers are in row 1
        
        # Search for "Branch" in the header row
        max_col = 100  # Reasonable limit
        for col in range(1, max_col + 1):
            cell_value = ws.range(header_row, col).value
            if cell_value and str(cell_value).strip().lower() == "branch":
                branch_col = col
                break
        
        if branch_col is None:
            wb.close()
            app.quit()
            raise Exception("Branch column not found in Country sheet")
        
        print(f"Found Branch column at column {branch_col}")
        
        # Find the Country row
        country_row = None
        
        # Find the last row with data in Branch column
        # Start from a reasonable max row and go up
        max_row = 10000  # Start with a large number
        last_row = None
        for row in range(max_row, header_row, -1):
            if ws.range(row, branch_col).value is not None:
                last_row = row
                break
        
        if last_row is None:
            last_row = header_row + 100  # Fallback
        
        for row in range(header_row + 1, last_row + 1):
            branch_value = ws.range(row, branch_col).value
            if branch_value and str(branch_value).strip().lower() == "country":
                country_row = row
                print(f"Found Country row at row {row}")
                break
        
        if country_row is None:
            wb.close()
            app.quit()
            raise Exception("Country row not found in Branch column")
        
        # Resolve column numbers for product-report columns (used for branch and Country updates)
        def _find_col(name):
            for col in range(1, max_col + 1):
                v = ws.range(header_row, col).value
                if v and str(v).strip().lower() == name.lower():
                    return col
            return None
        col_new_business = _find_col("New Business")
        col_repeat_business = _find_col("Repeat Business")
        col_reactivation = _find_col("Reactivation")
        col_refinance = _find_col("Refinance")
        col_disbursements_this_month = _find_col("Disbursements This Month")
        col_number_of_loans = _find_col("Number of loans")
        col_average_loan_size = _find_col("Average loan size")
        product_cols_ok = all((
            col_new_business, col_repeat_business, col_reactivation,
            col_refinance, col_disbursements_this_month
        ))

        # Update each Branch row from Product Sales (only where Product name matches Branch name)
        if product_totals_by_name and product_cols_ok:
            updated_branch_count = 0
            for row in range(header_row + 1, last_row + 1):
                if row == country_row:
                    continue
                branch_value = ws.range(row, branch_col).value
                if branch_value is None:
                    continue
                key = str(branch_value).strip()
                totals = product_totals_by_name.get(key)
                if totals is None:
                    for pk, pv in product_totals_by_name.items():
                        if str(pk).strip().lower() == key.lower():
                            totals = pv
                            break
                if totals is None:
                    continue
                def _num(x):
                    try:
                        return float(x) if x is not None else 0
                    except (TypeError, ValueError):
                        return 0
                ws.range(row, col_new_business).value = _num(totals.get("New Business"))
                ws.range(row, col_repeat_business).value = _num(totals.get("Repeat Business"))
                ws.range(row, col_reactivation).value = _num(totals.get("Reactivation"))
                ws.range(row, col_refinance).value = _num(totals.get("Refinance"))
                ws.range(row, col_disbursements_this_month).value = _num(totals.get("Total"))
                print(f"  ✅ Updated branch '{key}' from product report")
                updated_branch_count += 1
            print(f"  Updated {updated_branch_count} branch row(s) from product report.")
        elif product_totals_by_name and not product_cols_ok:
            print("  ⚠️ Could not find all product columns; skipping branch row updates.")

        # Fixed row range for Country totals
        start_sum_row = 3
        end_sum_row = 16
        print(f"\nSumming rows {start_sum_row} to {end_sum_row} for Country totals (other columns)")
        
        # Columns to process (use product Grand Total for these on Country row when available)
        columns_from_product_report = ("New Business", "Repeat Business", "Reactivation", "Refinance")
        disbursement_column_names = ("Disbursements This Month",)  # single canonical name (with 's')

        # Average loan size is NOT summed; it is set after this loop as Disbursements This Month / Number of loans
        columns_to_process = {
            "Target": lambda sum_val: sum_val - 5,  # Sum then subtract 5
            "New Business": lambda sum_val: sum_val,
            "Repeat Business": lambda sum_val: sum_val,
            "Reactivation": lambda sum_val: sum_val,
            "Disbursements This Month": lambda sum_val: sum_val,
            "Daily Target": lambda sum_val: sum_val,
            "Number of loans": lambda sum_val: sum_val,
            "In arrears": lambda sum_val: sum_val,
            "Value in arrears": lambda sum_val: sum_val,
            "Portfolio": lambda sum_val: sum_val,
            "Active Reps": lambda sum_val: sum_val
        }

        # Find each column and update the Country row
        for col_name, calculation_func in columns_to_process.items():
            # Find the column (exact match first, then case-insensitive)
            col_num = None
            for col in range(1, max_col + 1):
                cell_value = ws.range(header_row, col).value
                if cell_value:
                    cv_strip = str(cell_value).strip()
                    if cv_strip == col_name or cv_strip.lower() == col_name.lower():
                        col_num = col
                        break

            if col_num is None:
                print(f"⚠️ Column '{col_name}' not found, skipping...")
                continue

            # Use product report totals for these columns when available
            if product_totals_from_report:
                if col_name in columns_from_product_report and col_name in product_totals_from_report:
                    final_value = product_totals_from_report[col_name]
                    try:
                        final_value = float(final_value)
                    except (TypeError, ValueError):
                        final_value = 0
                    ws.range(country_row, col_num).value = final_value
                    print(f"  ✅ Updated Country row (from product report): {col_name} = {final_value}")
                    continue
                if col_name in disbursement_column_names and "Total" in product_totals_from_report:
                    final_value = product_totals_from_report["Total"]
                    try:
                        final_value = float(final_value)
                    except (TypeError, ValueError):
                        final_value = 0
                    ws.range(country_row, col_num).value = final_value
                    print(f"  ✅ Updated Country row (from product report): {col_name} = {final_value}")
                    continue

            print(f"Processing column '{col_name}' at column {col_num}...")

            # Sum fixed rows for Country totals
            total_sum = 0
            for row in range(start_sum_row, end_sum_row + 1):
                cell_value = ws.range(row, col_num).value
                if cell_value is not None:
                    try:
                        if isinstance(cell_value, (int, float)):
                            total_sum += cell_value
                        elif isinstance(cell_value, str):
                            cleaned = cell_value.replace(',', '').strip()
                            if cleaned:
                                total_sum += float(cleaned)
                    except (ValueError, TypeError):
                        pass

            final_value = calculation_func(total_sum)
            ws.range(country_row, col_num).value = final_value
            print(f"  ✅ Updated Country row: {final_value}")

        # Average loan size (Country) = Disbursements This Month / Number of loans (not sum of averages)
        if col_average_loan_size and col_disbursements_this_month and col_number_of_loans:
            disbursements_val = ws.range(country_row, col_disbursements_this_month).value
            num_loans_val = ws.range(country_row, col_number_of_loans).value
            try:
                disbursements_val = float(disbursements_val) if disbursements_val is not None else 0
                num_loans_val = float(num_loans_val) if num_loans_val is not None else 0
            except (TypeError, ValueError):
                disbursements_val = 0
                num_loans_val = 0
            avg_loan_size = (disbursements_val / num_loans_val) if num_loans_val else 0
            ws.range(country_row, col_average_loan_size).value = avg_loan_size
            print(f"  ✅ Updated Country row: Average loan size = Disbursements This Month / Number of loans = {avg_loan_size}")
        
        # Save and close
        wb.save()
        wb.close()
        app.quit()
        
        print("\n✅ Successfully updated Country row in Management file")
        
        # Run correct_management_report.py to fix Active/Inactive client numbers
        print("\n" + "-"*60)
        print("Running correct_management_report.py to correct client counts...")
        print("-"*60)
        script_dir = os.path.dirname(os.path.abspath(__file__))
        correct_script = os.path.join(script_dir, "correct _management_report.py")
        if os.path.exists(correct_script):
            try:
                result = subprocess.run(
                    [sys.executable, correct_script],
                    cwd=script_dir,
                    capture_output=False,
                    text=True,
                    timeout=120
                )
                if result.returncode == 0:
                    print("✅ correct_management_report.py completed successfully")
                else:
                    print(f"⚠️ correct_management_report.py exited with code {result.returncode}")
            except subprocess.TimeoutExpired:
                print("⚠️ correct_management_report.py timed out")
            except Exception as e:
                print(f"⚠️ Error running correct_management_report.py: {e}")
        else:
            print(f"⚠️ Script not found: {correct_script}")
        
    except Exception as e:
        print(f"\n❌ Error processing Management file: {e}")
        try:
            if 'wb' in locals():
                wb.close()
            if 'app' in locals():
                app.quit()
        except:
            pass

# ============================================================================
# STEP 6: Copy Management file to Email folder and process
# ============================================================================

print("\n" + "="*80)
print("STEP 6: Processing Email folder")
print("="*80)

# Email tooling folder (dynamic). NOTE: these are Windows .exe helpers — this
# email step is Windows-only and won't run on a Linux server as-is.
email_folder = os.environ.get("PCL_MANAGEMENT_EMAIL_DIR", os.path.join(_MGMT_DIR, "Email"))
split_workbook_exe = os.path.join(email_folder, "splitWorkBook.exe")
send_email_exe = os.path.join(email_folder, "send_email_tz (1).exe")

# Check if management file exists (find it again or use the one from STEP 5)
if management_file is None:
    # Try to find it again
    management_files = glob.glob(os.path.join(destination_folder, "Management*.xlsx"))
    if not management_files:
        management_files = glob.glob(os.path.join(destination_folder, "Management*.xlsm"))
    if management_files:
        management_file = management_files[0]

if management_file is None or not os.path.exists(management_file):
    print("⚠️ No Management file found. Skipping email processing.")
else:
    management_filename = os.path.basename(management_file)
    email_management_file = os.path.join(email_folder, management_filename)
    
    try:
        # Create email folder if it doesn't exist
        if not os.path.exists(email_folder):
            os.makedirs(email_folder)
            print(f"Created email folder: {email_folder}")
        
        # Step 6.1: Delete all Excel files in Email folder except Email.xlsx
        print("\n--- Step 6.1: Cleaning Email folder ---")
        all_files = os.listdir(email_folder)
        deleted_count = 0
        excel_extensions = ['.xlsx', '.xlsm', '.xls']
        
        for filename in all_files:
            file_path = os.path.join(email_folder, filename)
            # Only delete Excel files, skip directories, .exe files, and Email.xlsx
            if os.path.isfile(file_path):
                file_ext = os.path.splitext(filename)[1].lower()
                if file_ext in excel_extensions and filename.lower() != "email.xlsx":
                    try:
                        os.remove(file_path)
                        deleted_count += 1
                        print(f"  🗑️  Deleted: {filename}")
                    except Exception as e:
                        print(f"  ⚠️  Error deleting {filename}: {e}")
        
        print(f"Deleted {deleted_count} Excel file(s) from Email folder")
        
        # Step 6.2: Copy Management file to Email folder
        print("\n--- Step 6.2: Copying Management file to Email folder ---")
        try:
            shutil.copy2(management_file, email_management_file)
            print(f"  ✅ Copied: {management_filename}")
        except Exception as e:
            print(f"  ❌ Error copying Management file: {e}")
            raise
        
        # Step 6.3: Run splitWorkBook.exe
        print("\n--- Step 6.3: Running splitWorkBook.exe ---")
        if not os.path.exists(split_workbook_exe):
            print(f"  ⚠️  splitWorkBook.exe not found at: {split_workbook_exe}")
        else:
            try:
                print(f"  Running: {os.path.basename(split_workbook_exe)}")
                process_split = subprocess.Popen(
                    split_workbook_exe,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True,
                    cwd=email_folder
                )
                
                stdout, stderr = process_split.communicate(timeout=60)
                
                if stdout:
                    print("  Output:", stdout)
                if stderr:
                    print("  Warnings/Errors:", stderr)
                
                if process_split.returncode == 0:
                    print("  ✅ splitWorkBook.exe completed successfully")
                else:
                    print(f"  ⚠️  splitWorkBook.exe exited with code: {process_split.returncode}")
                    
            except subprocess.TimeoutExpired:
                print("  ⚠️  splitWorkBook.exe timed out")
            except Exception as e:
                print(f"  ❌ Error running splitWorkBook.exe: {e}")
        
        # Step 6.4: Send emails automatically (no confirmation prompt)
        print("\n--- Step 6.4: Sending emails automatically ---")

        # Load .env file
        load_dotenv()
        
        email_address = os.getenv('MANAGEMENT_EMAIL')
        email_password = os.getenv('MANAGEMENT_EMAIL_PASSWORD')
        
        if not email_address or not email_password:
            print("  ❌ Error: MANAGEMENT_EMAIL or MANAGEMENT_EMAIL_PASSWORD not found in .env file")
            print("  Please ensure your .env file contains both variables")
        else:
            print(f"  ✅ Found email credentials in .env file")
            print(f"  Email: {email_address}")
            
            # Step 6.5: Run send_email_tz (1).exe
            print("\n--- Step 6.5: Running send_email_tz (1).exe ---")
            if SEND_EMAILS != "yes":
                print("  ⏭  Test run (--send no) — SKIPPING the email send.")
                print("     The report was still built and split; re-run with")
                print("     'Yes (all recipients)' to actually send.")
            elif not os.path.exists(send_email_exe):
                print(f"  ❌ send_email_tz (1).exe not found at: {send_email_exe}")
            else:
                try:
                    print(f"  Running: {os.path.basename(send_email_exe)}")
                    
                    # Prepare inputs
                    email_inputs = f"{email_address}\n{email_password}\n"
                    
                    process_email = subprocess.Popen(
                        send_email_exe,
                        stdin=subprocess.PIPE,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        cwd=email_folder
                    )
                    
                    stdout, stderr = process_email.communicate(input=email_inputs, timeout=300)
                    
                    if stdout:
                        print("\n  Executable output:")
                        print(stdout)
                    if stderr:
                        print("\n  Executable errors/warnings:")
                        print(stderr)
                    
                    if process_email.returncode == 0:
                        print("\n  ✅ Email sending completed successfully")
                        
                        # Step 6.6: Clean up Email folder after email sending
                        print("\n--- Step 6.6: Cleaning Email folder after email sending ---")
                        all_files_after = os.listdir(email_folder)
                        deleted_after_count = 0
                        excel_extensions = ['.xlsx', '.xlsm', '.xls']
                        
                        for filename in all_files_after:
                            file_path = os.path.join(email_folder, filename)
                            # Only delete Excel files, skip directories, .exe files, and Email.xlsx
                            if os.path.isfile(file_path):
                                file_ext = os.path.splitext(filename)[1].lower()
                                if file_ext in excel_extensions and filename.lower() != "email.xlsx":
                                    try:
                                        os.remove(file_path)
                                        deleted_after_count += 1
                                        print(f"  🗑️  Deleted: {filename}")
                                    except Exception as e:
                                        print(f"  ⚠️  Error deleting {filename}: {e}")
                        
                        print(f"Deleted {deleted_after_count} Excel file(s) from Email folder")
                        print("  ✅ Email folder cleaned - only Email.xlsx and .exe files remain")
                    else:
                        print(f"\n  ⚠️  Email sending exited with code: {process_email.returncode}")
                        
                except subprocess.TimeoutExpired:
                    print("  ⚠️  send_email_tz (1).exe timed out")
                except Exception as e:
                    print(f"  ❌ Error running send_email_tz (1).exe: {e}")
        
    except Exception as e:
        print(f"\n❌ Error in email processing: {e}")

# ============================================================================
# FINAL STEP: Append the entered day to the generated report file name.
#   "Management Report2026-06.xlsx" -> "Management Report2026-06-15.xlsx"
# ============================================================================
print("\n" + "="*80)
print("FINAL STEP: Date-stamping the report file")
print("="*80)
try:
    candidates = glob.glob(os.path.join(destination_folder, "Management Report*.xlsx"))
    renamed_any = False
    for rf in candidates:
        base = os.path.basename(rf)
        # Match the undated monthly report, e.g. "Management Report2026-06.xlsx"
        m = _re_date.match(r'^(Management Report\d{4}-\d{2})\.xlsx$', base, _re_date.IGNORECASE)
        if not m:
            continue  # already dated or a different file — leave it alone
        new_name = f"{m.group(1)}-{REPORT_DAY:02d}.xlsx"
        new_path = os.path.join(destination_folder, new_name)
        if os.path.abspath(new_path) != os.path.abspath(rf):
            if os.path.exists(new_path):
                os.remove(new_path)
            os.rename(rf, new_path)
            renamed_any = True
            print(f"  ✅ Renamed: {base} -> {new_name}")
    if not renamed_any:
        print("  ⚠️  No undated 'Management Report2026-MM.xlsx' found to date-stamp.")
except Exception as e:
    print(f"  ⚠️  Error date-stamping the report: {e}")

# ============================================================================
# END PROCESS: upload the freshly generated report to the live PCL system.
# Runs direct_upload_files_to_db.py --commit, which logs in over SSH, dedups by
# (department, type, date) and uploads only files not already in the DB.
# ============================================================================
_UPLOAD_SCRIPT = os.path.join(
    os.environ.get("PCL_AUTOMATION_ROOT")
    or os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "CRM", "direct_upload_files_to_db.py")
print("\n" + "="*80)
print("END PROCESS: uploading new report(s) to the live PCL system ...")
print("="*80)
if SEND_EMAILS != "yes":
    print("  ⏭  Test run (--send no) — SKIPPING the live DB upload.")
    print(f'  To upload manually:  python "{_UPLOAD_SCRIPT}" --commit')
else:
    try:
        subprocess.run([sys.executable, _UPLOAD_SCRIPT, "--commit"], check=False)
    except Exception as _up_err:
        print(f"  Upload step could not run: {_up_err}")
        print(f'  You can upload manually:  python "{_UPLOAD_SCRIPT}" --commit')

print("\n" + "="*80)
print("PROCESS MANAGEMENT COMPLETED")
print("="*80)