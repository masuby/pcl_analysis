"""
Checker - Compare rows between Raphael and Masubi sheets
=========================================================
Finds rows that exist in one sheet but not the other (based on common columns).
Colors non-matching rows in Debug.xlsx and exports missing rows to a new Excel file.
Aim: Understand which rows cause Net Disbursement difference.
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill

# Paths
DEBUG_FILE = os.path.join(os.path.dirname(__file__), "Debug.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "Debug_Missing_Rows.xlsx")

# Common columns for comparison (both sheets have these)
COMMON_COLS = ["Loan Name", "Loan Amount", "Branch", "Activation Date (Loan)", "Net disbursement", "Type of business"]

# Highlight colors
FILL_MISSING = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Light red


def _normalize(val):
    """Normalize value for comparison."""
    if val is None:
        return ""
    if isinstance(val, (datetime,)):
        return val.strftime("%Y-%m-%d %H:%M:%S") if hasattr(val, "strftime") else str(val)
    if isinstance(val, float):
        return round(val, 2) if val == val else ""  # handle NaN
    return str(val).strip()


def _row_signature(row_dict, col_names):
    """Create a tuple signature for a row (common columns only)."""
    return tuple(_normalize(row_dict.get(c)) for c in col_names)


def _get_col_indices(ws, col_names):
    """Get column indices for given header names. Returns dict {col_name: 1-based index}."""
    headers = [cell.value for cell in ws[1]]
    result = {}
    for cn in col_names:
        for i, h in enumerate(headers, 1):
            if h and str(h).strip() == cn:
                result[cn] = i
                break
        if cn not in result:
            raise KeyError(f"Column '{cn}' not found. Headers: {headers}")
    return result


def _read_sheet_data(wb, sheet_name, common_cols):
    """Read all data rows as list of (signature, row_dict, row_idx)."""
    ws = wb[sheet_name]
    col_idx = _get_col_indices(ws, common_cols)
    data = []
    for row_idx in range(2, ws.max_row + 1):
        row_dict = {c: ws.cell(row_idx, col_idx[c]).value for c in common_cols}
        sig = _row_signature(row_dict, common_cols)
        data.append((sig, row_dict, row_idx))
    return data, col_idx


def main():
    print("=" * 60)
    print("Checker - Compare Raphael vs Masubi")
    print("=" * 60)

    if not os.path.exists(DEBUG_FILE):
        print(f"[ERROR] File not found: {DEBUG_FILE}")
        return

    print(f"\nLoading {DEBUG_FILE}...")
    wb = openpyxl.load_workbook(DEBUG_FILE, data_only=False)

    if "Raphael" not in wb.sheetnames or "Masubi" not in wb.sheetnames:
        print("[ERROR] Required sheets 'Raphael' and 'Masubi' not found")
        return

    # Determine which common cols exist in each sheet
    raph_headers = [cell.value for cell in wb["Raphael"][1]]
    masubi_headers = [cell.value for cell in wb["Masubi"][1]]
    raph_common = [c for c in COMMON_COLS if c in raph_headers]
    masubi_common = [c for c in COMMON_COLS if c in masubi_headers]
    common = [c for c in COMMON_COLS if c in raph_headers and c in masubi_headers]

    if len(common) < len(COMMON_COLS):
        print(f"[WARN] Using common columns only: {common}")

    # Read data
    raph_data, raph_col_idx = _read_sheet_data(wb, "Raphael", common)
    masubi_data, masubi_col_idx = _read_sheet_data(wb, "Masubi", common)

    raph_sigs = {s for s, _, _ in raph_data}
    masubi_sigs = {s for s, _, _ in masubi_data}

    # Rows in Raphael NOT in Masubi
    in_raph_not_masubi = [(s, rd, ri) for s, rd, ri in raph_data if s not in masubi_sigs]
    # Rows in Masubi NOT in Raphael
    in_masubi_not_raph = [(s, rd, ri) for s, rd, ri in masubi_data if s not in raph_sigs]

    print(f"\nRaphael: {len(raph_data)} rows | Masubi: {len(masubi_data)} rows")
    print(f"Rows in Raphael but NOT in Masubi: {len(in_raph_not_masubi)}")
    print(f"Rows in Masubi but NOT in Raphael: {len(in_masubi_not_raph)}")

    # Color non-matching rows in Debug.xlsx
    ws_raph = wb["Raphael"]
    ws_masubi = wb["Masubi"]

    for _, _, row_idx in in_raph_not_masubi:
        for col_idx in range(1, ws_raph.max_column + 1):
            ws_raph.cell(row_idx, col_idx).fill = FILL_MISSING

    for _, _, row_idx in in_masubi_not_raph:
        for col_idx in range(1, ws_masubi.max_column + 1):
            ws_masubi.cell(row_idx, col_idx).fill = FILL_MISSING

    wb.save(DEBUG_FILE)
    print(f"\n[OK] Colored non-matching rows in {DEBUG_FILE}")

    # Create output Excel with missing rows (copy full rows from source)
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    # Sheet: In Raphael but NOT in Masubi
    ws1 = out_wb.create_sheet("In_Raphael_Not_Masubi")
    for c in range(1, ws_raph.max_column + 1):
        ws1.cell(1, c).value = ws_raph.cell(1, c).value
    for r, (_, _, row_idx) in enumerate(in_raph_not_masubi, 2):
        for c in range(1, ws_raph.max_column + 1):
            ws1.cell(r, c).value = ws_raph.cell(row_idx, c).value

    # Sheet: In Masubi but NOT in Raphael
    ws2 = out_wb.create_sheet("In_Masubi_Not_Raphael")
    for c in range(1, ws_masubi.max_column + 1):
        ws2.cell(1, c).value = ws_masubi.cell(1, c).value
    for r, (_, _, row_idx) in enumerate(in_masubi_not_raph, 2):
        for c in range(1, ws_masubi.max_column + 1):
            ws2.cell(r, c).value = ws_masubi.cell(row_idx, c).value

    out_wb.save(OUTPUT_FILE)
    print(f"[OK] Exported missing rows to {OUTPUT_FILE}")

    # Summary
    print("\n" + "=" * 60)
    print("Summary")
    print("=" * 60)
    if in_raph_not_masubi:
        net_raph = sum(
            float(r.get("Net disbursement") or 0)
            for _, r, _ in in_raph_not_masubi
        )
        print(f"Net disbursement in Raphael-only rows: {net_raph:,.2f}")
    if in_masubi_not_raph:
        net_masubi = sum(
            float(r.get("Net disbursement") or 0)
            for _, r, _ in in_masubi_not_raph
        )
        print(f"Net disbursement in Masubi-only rows: {net_masubi:,.2f}")


if __name__ == "__main__":
    main()
