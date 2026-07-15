"""
Add loan-count columns to the Management Report's Country sheet
===============================================================
Beside each disbursement-amount column we add the matching COUNT column, filled
by a VLOOKUP-style match against Product_Sales_Report.xlsx (raw rows):

    New Business      -> # of New Loan            (CLIENT TYPE == NEW)
    Repeat Business   -> # of Repeat Loan         (CLIENT TYPE == REPEAT)
    Reactivation      -> # of Reactivated clients (Type of business == REACTIVATION)
    Refinance         -> # of Refinanced clients  (Type of business == REFINANCE)

Each Country-sheet row is matched by its Branch value:
    "Country"        -> grand totals (all rows)
    a product name   -> counts for that product   (e.g. CS, LBF, SME …)
    a branch name    -> counts for that branch     (e.g. Musoma, Bukoba …)
Rows that match neither (e.g. zone roll-ups) are left blank.

Re-running is safe: if the count columns already exist they are refreshed in place
(no duplicate columns).  openpyxl only (formatting preserved on the unchanged cells).
"""

from __future__ import annotations

import copy
import glob
import os
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCT_SALES_REPORT_PATH = os.path.join(SCRIPT_DIR, "Product_Sales_Report.xlsx")
ZONE_CLUSTER_PATH = os.path.join(SCRIPT_DIR, "Zone and cluster.xlsx")
# Management report output — must match process_management.py's destination_folder.
_MGMT_DIR = os.environ.get("PCL_MANAGEMENT_DIR") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "Management")
    if os.environ.get("PCL_AUTOMATION_ROOT") else SCRIPT_DIR)
DESTINATION_FOLDER = os.environ.get("PCL_MANAGEMENT_OUT", os.path.join(_MGMT_DIR, "OUTPUT"))
HEADER_ROW = 1

# amount column -> (count column header, count key)
COUNT_SPECS = [
    ("New Business", "# of New Loan", "new"),
    ("Repeat Business", "# of Repeat Loan", "repeat"),
    ("Reactivation", "# of Reactivated clients", "reactivation"),
    ("Refinance", "# of Refinanced clients", "refinance"),
]


def _blank_counts() -> dict:
    return {"new": 0, "repeat": 0, "reactivation": 0, "refinance": 0}


def load_counts():
    """Return (by_product, by_branch, by_zone, by_cluster, grand) loan counts
    from the raw Product Sales rows (uses the Zone / Cluster columns added by
    create_product_sales_report.py)."""
    by_product: dict[str, dict] = defaultdict(_blank_counts)
    by_branch: dict[str, dict] = defaultdict(_blank_counts)
    by_zone: dict[str, dict] = defaultdict(_blank_counts)
    by_cluster: dict[str, dict] = defaultdict(_blank_counts)
    grand = _blank_counts()

    wb = openpyxl.load_workbook(PRODUCT_SALES_REPORT_PATH, data_only=True)
    ws = wb["Product Sales"]
    hdr = {str(ws.cell(1, c).value).strip().lower(): c for c in range(1, ws.max_column + 1)
           if ws.cell(1, c).value is not None}
    c_prod = hdr.get("products")
    c_branch = hdr.get("branch")
    c_ct = hdr.get("client type")
    c_tob = hdr.get("type of business")
    c_zone = hdr.get("zone")
    c_cluster = hdr.get("cluster")

    for row in range(2, ws.max_row + 1):
        a1 = ws.cell(row, 1).value
        if a1 is not None and str(a1).strip() == "Summary by Product":
            break  # raw data ends here; the summary block follows
        ct = str(ws.cell(row, c_ct).value or "").strip().upper() if c_ct else ""
        tob = str(ws.cell(row, c_tob).value or "").strip().upper() if c_tob else ""
        if ct not in ("NEW", "REPEAT") and tob not in ("REACTIVATION", "REFINANCE"):
            continue  # not a real data row
        prod = str(ws.cell(row, c_prod).value or "").strip() if c_prod else ""
        branch = str(ws.cell(row, c_branch).value or "").strip() if c_branch else ""
        zone = str(ws.cell(row, c_zone).value or "").strip() if c_zone else ""
        cluster = str(ws.cell(row, c_cluster).value or "").strip() if c_cluster else ""

        def bump(d):
            if ct == "NEW":
                d["new"] += 1
            elif ct == "REPEAT":
                d["repeat"] += 1
            if tob == "REACTIVATION":
                d["reactivation"] += 1
            elif tob == "REFINANCE":
                d["refinance"] += 1

        if prod:
            bump(by_product[prod])
        if branch:
            bump(by_branch[branch])
        if zone:
            bump(by_zone[zone])
        if cluster:
            bump(by_cluster[cluster])
        bump(grand)
    wb.close()
    return dict(by_product), dict(by_branch), dict(by_zone), dict(by_cluster), grand


def load_zone_cluster_names():
    """Distinct Zone and Cluster names (master list of which Country rows are
    zones / clusters) from 'Zone and cluster.xlsx'."""
    zones, clusters = set(), set()
    if not os.path.exists(ZONE_CLUSTER_PATH):
        print(f"[WARN] Zone and cluster file not found: {ZONE_CLUSTER_PATH}")
        return zones, clusters
    wb = openpyxl.load_workbook(ZONE_CLUSTER_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = {str(ws.cell(1, c).value).strip().lower(): c
           for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None}
    zi, ci = hdr.get("zone"), hdr.get("cluster")
    for r in range(2, ws.max_row + 1):
        if zi and ws.cell(r, zi).value and str(ws.cell(r, zi).value).strip():
            zones.add(str(ws.cell(r, zi).value).strip())
        if ci and ws.cell(r, ci).value and str(ws.cell(r, ci).value).strip():
            clusters.add(str(ws.cell(r, ci).value).strip())
    wb.close()
    return zones, clusters


def find_col(ws, name, max_col=200):
    for col in range(1, max_col + 1):
        v = ws.cell(HEADER_ROW, col).value
        if v and str(v).strip().lower() == name.lower():
            return col
    return None


def main():
    print("=" * 60)
    print("Add loan-count columns to Country sheet")
    print("=" * 60)
    if not os.path.exists(PRODUCT_SALES_REPORT_PATH):
        print(f"Missing: {PRODUCT_SALES_REPORT_PATH}")
        return

    by_product, by_branch, by_zone, by_cluster, grand = load_counts()
    zone_names, cluster_names = load_zone_cluster_names()
    # case-insensitive lookup helpers
    prod_ci = {k.lower(): v for k, v in by_product.items()}
    branch_ci = {k.lower(): v for k, v in by_branch.items()}
    zone_ci = {k.lower(): v for k, v in by_zone.items()}
    cluster_ci = {k.lower(): v for k, v in by_cluster.items()}
    zone_names_ci = {z.lower() for z in zone_names}
    cluster_names_ci = {c.lower() for c in cluster_names}
    print(f"Loaded counts: {len(by_product)} products, {len(by_branch)} branches, "
          f"{len(by_zone)} zones, {len(by_cluster)} clusters.")
    print(f"Grand totals: {grand}")

    files = glob.glob(os.path.join(DESTINATION_FOLDER, "Management*.xlsx")) or \
        glob.glob(os.path.join(DESTINATION_FOLDER, "Management*.xlsm"))
    if not files:
        print(f"No Management file in {DESTINATION_FOLDER}")
        return
    mgmt = files[0]
    print(f"Management file: {os.path.basename(mgmt)}")

    is_xlsm = mgmt.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(mgmt, data_only=True, keep_vba=is_xlsm)
    if "Country" not in wb.sheetnames:
        print("'Country' sheet not found.")
        wb.close()
        return
    ws = wb["Country"]

    # ---- insert (or locate) the 4 count columns ----
    already = find_col(ws, "# of New Loan") is not None
    if already:
        print("Count columns already present — refreshing values in place.")
        count_cols = {key: find_col(ws, header) for _amt, header, key in COUNT_SPECS}
    else:
        # base amount columns (locate BEFORE inserting)
        base = {amt: find_col(ws, amt) for amt, _h, _k in COUNT_SPECS}
        missing = [a for a, c in base.items() if c is None]
        if missing:
            print(f"Amount columns not found: {missing}")
            wb.close()
            return
        # insert right-to-left so indices stay valid; insert AFTER each amount col
        for amt, header, key in sorted(COUNT_SPECS, key=lambda s: -base[s[0]]):
            ws.insert_cols(base[amt] + 1)
        # recompute positions and write headers
        count_cols = {}
        for amt, header, key in COUNT_SPECS:
            amt_col = find_col(ws, amt)
            new_col = amt_col + 1
            hc = ws.cell(HEADER_ROW, new_col)
            hc.value = header
            # copy header look from the amount header for a consistent style
            src = ws.cell(HEADER_ROW, amt_col)
            hc.font = copy.copy(src.font)
            hc.fill = copy.copy(src.fill)
            hc.border = copy.copy(src.border)
            hc.alignment = copy.copy(src.alignment)
            count_cols[key] = new_col
        print("Inserted 4 count columns beside their amount columns.")

    # ---- fill values + match each row's colour to the neighbouring amount cell ----
    branch_col = find_col(ws, "Branch")
    num_loans_col = find_col(ws, "Number of loans")  # = # New Loan + # Repeat Loan
    last_row = ws.max_row
    filled = 0
    for row in range(HEADER_ROW + 1, last_row + 1):
        bval = ws.cell(row, branch_col).value
        key = str(bval).strip().lower() if bval is not None else ""
        counts = None
        if key == "country":
            counts = grand
        elif key in prod_ci:
            counts = prod_ci[key]
        elif key in zone_names_ci:                      # zone row -> zone counts (0 if none)
            counts = zone_ci.get(key, _blank_counts())
        elif key in cluster_names_ci:                   # cluster row -> cluster counts (0 if none)
            counts = cluster_ci.get(key, _blank_counts())
        elif key in branch_ci:
            counts = branch_ci[key]
        # branch present but no match in product sales -> count 0 (not blank)
        if counts is None and key:
            counts = _blank_counts()

        for _amt, _header, ckey in COUNT_SPECS:
            cc = count_cols[ckey]
            dst = ws.cell(row, cc)
            src = ws.cell(row, cc - 1)  # the amount column to its left
            # inherit the row's look (zebra / highlight) from the neighbour
            dst.fill = copy.copy(src.fill)
            dst.font = copy.copy(src.font)
            dst.border = copy.copy(src.border)
            dst.alignment = copy.copy(src.alignment)
            dst.number_format = "#,##0"
            dst.value = int(counts.get(ckey, 0)) if counts is not None else None
        # Number of loans = total loan count = # New Loan + # Repeat Loan
        if counts is not None and num_loans_col:
            nl = ws.cell(row, num_loans_col)
            nl.value = int(counts.get("new", 0)) + int(counts.get("repeat", 0))
            nl.number_format = "#,##0"
        if counts is not None:
            filled += 1

    # ---- highlight the Disbursements This Month column in red ----
    disb_col = find_col(ws, "Disbursements This Month")
    if disb_col:
        red_header = PatternFill(start_color="FFC00000", end_color="FFC00000", fill_type="solid")
        red_data = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
        hc = ws.cell(HEADER_ROW, disb_col)
        hc.fill = red_header
        hc.font = Font(bold=True, color="FFFFFFFF")
        for row in range(HEADER_ROW + 1, last_row + 1):
            bval = ws.cell(row, branch_col).value
            if bval is None or not str(bval).strip():
                continue
            ws.cell(row, disb_col).fill = red_data
        print("Highlighted 'Disbursements This Month' column in red.")

    # ---- colour the '# of Reactivated clients' column cyan ----
    react_count_col = count_cols.get("reactivation")
    if react_count_col:
        cyan_header = PatternFill(start_color="FF00BCD4", end_color="FF00BCD4", fill_type="solid")
        cyan_data = PatternFill(start_color="FFE0F7FA", end_color="FFE0F7FA", fill_type="solid")
        hc = ws.cell(HEADER_ROW, react_count_col)
        hc.fill = cyan_header
        hc.font = Font(bold=True, color="FFFFFFFF")
        for row in range(HEADER_ROW + 1, last_row + 1):
            bval = ws.cell(row, branch_col).value
            if bval is None or not str(bval).strip():
                continue
            ws.cell(row, react_count_col).fill = cyan_data
        print("Coloured '# of Reactivated clients' column cyan.")

    wb.save(mgmt)
    wb.close()
    print(f"Filled count columns for {filled} row(s).")
    print(f"Saved: {mgmt}")


if __name__ == "__main__":
    main()
