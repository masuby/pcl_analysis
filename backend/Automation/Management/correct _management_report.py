#!/usr/bin/env python3
"""
Process Management Report - Uses Clients file as source of truth (no Created date filter)

Input locations:
- Client file, Zone and cluster file: same folder as this script (Management folder in pcl)
- Management Report: C:\\Users\\Daniel\\Desktop\\Management

Files:
- Management Report* (e.g. Management Report2026-02.xlsx) - from Desktop/Management
- Client* (e.g. Clientstz.xlsx) - from script location
- Zone and cluster* (e.g. Zone and cluster.xlsx) - from script location

Output (Desktop/Management folder):
- branch_summary.xlsx
- Management report updated with corrected client counts (in-place)

No date filtering: uses ALL Active/Inactive clients (no Created column required).

Usage:
  python correct_management_report.py
"""

import os
import re
import sys
from pathlib import Path

import warnings
warnings.filterwarnings("ignore", message="Workbook contains no default style")

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Install: pip install openpyxl")
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
# Client and Zone/cluster files: same location as this script
# Management report output — must match process_management.py's destination_folder.
_MGMT_DIR = Path(os.environ.get("PCL_MANAGEMENT_DIR") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "Management")
    if os.environ.get("PCL_AUTOMATION_ROOT") else str(SCRIPT_DIR)))
MANAGEMENT_FOLDER = Path(os.environ.get("PCL_MANAGEMENT_OUT", str(_MGMT_DIR / "OUTPUT")))


def find_file(folder: Path, *prefixes: str) -> Path | None:
    """Find first .xlsx file whose name (lowercase) starts with any of the prefixes."""
    folder = Path(folder)
    if not folder.is_dir():
        return None
    lower_prefixes = [p.lower() for p in prefixes]
    for f in folder.glob("*.xlsx"):
        name = f.stem.lower()
        for pre in lower_prefixes:
            if name.startswith(pre):
                return f
    return None


def load_zone_cluster_lookup(path: Path) -> list:
    """Load Zone and cluster.xlsx as lookup (branch -> zone, cluster, product)."""
    if not path or not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    zi = next((i for i, h in enumerate(headers) if h and "zone" in str(h).lower() and "cluster" not in str(h).lower()), 0)
    bi = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), 1)
    ci = next((i for i, h in enumerate(headers) if h and "cluster" in str(h).lower()), 2)
    pi = next((i for i, h in enumerate(headers) if h and "product" in str(h).lower()), 3)
    out = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(zi, bi, ci, pi):
            continue
        branch = str(row[bi] or "").strip()
        if not branch or branch in seen:
            continue
        seen.add(branch)
        out.append({
            "branch": branch,
            "zone": str(row[zi] or "").strip(),
            "cluster": str(row[ci] or "").strip(),
            "product": str(row[pi] or "").strip(),
        })
    wb.close()
    return out


def vlookup_branch(branch: str, lookup: list) -> dict:
    x = str(branch or "").strip()
    matches = [r for r in lookup if r.get("branch") == x]
    if matches:
        return matches[0]
    return {"zone": "ERR", "cluster": "ERR", "product": "ERR"}


def inspect_clients(client_path: Path):
    """Print column names and sample data from a Client file."""
    wb = load_workbook(client_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    print("Columns:", headers)
    rows = list(ws.iter_rows(min_row=2, max_row=6, values_only=True))
    print("Sample rows (first 5):")
    for i, r in enumerate(rows):
        print(f"  {i+2}:", (r[:10] if r and len(r) > 10 else r))
    wb.close()


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if "--inspect" in sys.argv:
        # Inspect Client file: python script.py --inspect [folder_or_file]
        path = Path(args[0]) if args else SCRIPT_DIR
        if path.is_file():
            inspect_clients(path)
        else:
            client_file = find_file(path, "Client")
            if client_file:
                print(f"File: {client_file}")
                inspect_clients(client_file)
            else:
                print("No Client*.xlsx found in", path)
        return

    # Client and Zone/cluster: script location; Management file: Desktop/Management
    client_file = find_file(SCRIPT_DIR, "Client")
    zone_file = find_file(SCRIPT_DIR, "Zone and cluster")
    mgmt_file = find_file(MANAGEMENT_FOLDER, "Management Report")
    work_dir = MANAGEMENT_FOLDER  # for output (branch_summary, mgmt update)

    if not mgmt_file:
        print(f"Error: No Management Report file found in {MANAGEMENT_FOLDER}")
        sys.exit(1)
    if not client_file:
        print(f"Error: No Client file found in {SCRIPT_DIR}")
        sys.exit(1)
    if not zone_file:
        print(f"Error: No Zone and cluster file found in {SCRIPT_DIR}")
        sys.exit(1)

    print("=" * 60)
    print("  Process Management Report")
    print("=" * 60)
    print(f"  Client file:      {client_file}")
    print(f"  Zone/cluster:     {zone_file}")
    print(f"  Management file:  {mgmt_file}")
    print(f"  Output folder:    {work_dir}")
    print()

    # --- Load clients: Active/Inactive only (no date filter) ---
    wb_cl = load_workbook(client_file, read_only=True, data_only=True)
    ws_cl = wb_cl.active
    headers = [c.value for c in ws_cl[1]]
    branch_col = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), -1)
    state_col = next((i for i, h in enumerate(headers) if h and "client" in str(h).lower() and "state" in str(h).lower()), -1)

    if branch_col < 0 or state_col < 0:
        wb_cl.close()
        print("Error: Clients file missing Branch or Client State column")
        print("Columns:", headers)
        sys.exit(1)

    by_branch = {}
    for row in ws_cl.iter_rows(min_row=2, values_only=True):
        row = list(row) if row else []
        if len(row) <= max(branch_col, state_col):
            continue
        state = str(row[state_col] or "").strip().lower()
        if state not in ("active", "inactive"):
            continue
        branch = str(row[branch_col] or "").strip()
        if not branch:
            continue
        if branch not in by_branch:
            by_branch[branch] = {"active": 0, "inactive": 0}
        if state == "inactive":
            by_branch[branch]["inactive"] += 1
        else:
            by_branch[branch]["active"] += 1
    wb_cl.close()

    zone_cluster = load_zone_cluster_lookup(zone_file)
    zone_sums = {}
    cluster_sums = {}
    product_sums = {}
    for branch, counts in by_branch.items():
        lookup = vlookup_branch(branch, zone_cluster)
        z, c, p = lookup["zone"], lookup["cluster"], lookup["product"]
        if z == "ERR" or c == "ERR" or p == "ERR":
            continue
        for d, k in [(zone_sums, z), (cluster_sums, c), (product_sums, p)]:
            if k and k not in d:
                d[k] = {"active": 0, "inactive": 0}
            if k:
                d[k]["active"] += counts["active"]
                d[k]["inactive"] += counts["inactive"]

    # --- Save branch_summary.xlsx ---
    out_summary = work_dir / "branch_summary.xlsx"
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Branch Summary"
    ws3["A1"] = "Branch"
    ws3["B1"] = "Zone"
    ws3["C1"] = "Cluster"
    ws3["D1"] = "Product"
    ws3["E1"] = "Active"
    ws3["F1"] = "Inactive"
    ws3["G1"] = "Total"
    row_num = 2
    for branch, counts in sorted(by_branch.items()):
        lookup = vlookup_branch(branch, zone_cluster)
        z = lookup.get("zone", "ERR") if lookup.get("zone") != "ERR" else "ERR"
        c = lookup.get("cluster", "ERR") if lookup.get("cluster") != "ERR" else "ERR"
        p = lookup.get("product", "ERR") if lookup.get("product") != "ERR" else "ERR"
        ws3.cell(row=row_num, column=1, value=branch)
        ws3.cell(row=row_num, column=2, value=z)
        ws3.cell(row=row_num, column=3, value=c)
        ws3.cell(row=row_num, column=4, value=p)
        ws3.cell(row=row_num, column=5, value=counts["active"])
        ws3.cell(row=row_num, column=6, value=counts["inactive"])
        ws3.cell(row=row_num, column=7, value=counts["active"] + counts["inactive"])
        row_num += 1
    if zone_sums:
        ws3.cell(row=row_num, column=1, value="--- ZONE TOTALS ---")
        row_num += 1
        for z, v in sorted(zone_sums.items()):
            ws3.cell(row=row_num, column=1, value=z)
            ws3.cell(row=row_num, column=2, value=z)
            ws3.cell(row=row_num, column=5, value=v["active"])
            ws3.cell(row=row_num, column=6, value=v["inactive"])
            ws3.cell(row=row_num, column=7, value=v["active"] + v["inactive"])
            row_num += 1
    if cluster_sums:
        ws3.cell(row=row_num, column=1, value="--- CLUSTER TOTALS ---")
        row_num += 1
        for c, v in sorted(cluster_sums.items()):
            ws3.cell(row=row_num, column=1, value=c)
            ws3.cell(row=row_num, column=3, value=c)
            ws3.cell(row=row_num, column=5, value=v["active"])
            ws3.cell(row=row_num, column=6, value=v["inactive"])
            ws3.cell(row=row_num, column=7, value=v["active"] + v["inactive"])
            row_num += 1
    if product_sums:
        ws3.cell(row=row_num, column=1, value="--- PRODUCT TOTALS ---")
        row_num += 1
        for p, v in sorted(product_sums.items()):
            ws3.cell(row=row_num, column=1, value=p)
            ws3.cell(row=row_num, column=4, value=p)
            ws3.cell(row=row_num, column=5, value=v["active"])
            ws3.cell(row=row_num, column=6, value=v["inactive"])
            ws3.cell(row=row_num, column=7, value=v["active"] + v["inactive"])
            row_num += 1
    for col, w in [("A", 35), ("B", 22), ("C", 18), ("D", 12), ("E", 10), ("F", 10), ("G", 10)]:
        ws3.column_dimensions[col].width = w
    wb3.save(out_summary)
    wb3.close()
    print(f"Saved: {out_summary.name} ({len(by_branch)} branches)")

    # --- Update management report ---
    def is_leaf(val):
        if not val or not isinstance(val, str):
            return False
        v = val.lower()
        return not re.search(r"cluster|zone|^cs$|^lbf$|zanzibar|call center|^sme$|maziwa|agrifinance|lbf zone|smes", v)

    def match_branch(report_br, client_br):
        r = str(report_br or "").strip().lower()
        c = str(client_br or "").strip().lower()
        return r == c or r in c

    def get_counts(report_br):
        for cb, counts in by_branch.items():
            if match_branch(report_br, cb):
                return counts
        return None

    wb_mgmt = load_workbook(mgmt_file, data_only=False, rich_text=True, keep_links=True)
    country_sheet = next((s for s in wb_mgmt.sheetnames if "country" in s.lower()), wb_mgmt.sheetnames[0])
    ws = wb_mgmt[country_sheet]

    headers_m = [c.value for c in ws[1]]
    idx = {}
    for i, h in enumerate(headers_m):
        if not h:
            continue
        h = str(h).strip().lower()
        if "branch" in h:
            idx["branch"] = i
        elif "number" in h and "clients" in h:
            idx["num_clients"] = i
        elif "inactive" in h and "clients" in h:
            idx["inactive"] = i
        elif "active" in h and "clients" in h:
            idx["active"] = i
    if "num_clients" not in idx:
        idx["num_clients"] = next((i for i, h in enumerate(headers_m) if h and "clients" in str(h).lower()), -1)
    if "active" not in idx:
        idx["active"] = next((i for i, h in enumerate(headers_m) if h and "active" in str(h).lower() and "clients" in str(h).lower() and "inactive" not in str(h).lower()), -1)
    if "inactive" not in idx:
        idx["inactive"] = next((i for i, h in enumerate(headers_m) if h and "inactive" in str(h).lower() and "clients" in str(h).lower()), -1)

    branch_col_m = idx.get("branch", 0)
    num_col1 = idx.get("num_clients", -1) + 1
    active_col1 = idx.get("active", -1) + 1
    inactive_col1 = idx.get("inactive", -1) + 1

    rows_info = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row) if row else []
        if len(row) <= branch_col_m:
            rows_info.append((row_idx, None, False))
            continue
        val = row[branch_col_m]
        if val is None or (isinstance(val, str) and not val.strip()):
            rows_info.append((row_idx, None, False))
            continue
        branch = str(val).strip()
        leaf = is_leaf(branch)
        rows_info.append((row_idx, branch, leaf))

    PRODUCT_ROW_NAMES = ("CS", "LBF", "SME", "Agrifinance", "Maziwa")
    values_by_row = {}

    for row_idx, branch, is_leaf_row in rows_info:
        if not branch:
            continue
        if branch == "Country":
            continue
        if branch in PRODUCT_ROW_NAMES:
            if branch in product_sums:
                v = product_sums[branch]
                values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}
            continue
        if is_leaf_row:
            counts = get_counts(branch)
            if counts:
                total = counts["active"] + counts["inactive"]
                values_by_row[row_idx] = {"num": total, "active": counts["active"], "inactive": counts["inactive"]}
        elif branch in zone_sums:
            v = zone_sums[branch]
            values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}
        elif branch in cluster_sums:
            v = cluster_sums[branch]
            values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}

    for row_idx, branch, _ in rows_info:
        if branch != "Country":
            continue
        c_num = c_active = c_inactive = 0
        for v in product_sums.values():
            c_num += v["active"] + v["inactive"]
            c_active += v["active"]
            c_inactive += v["inactive"]
        values_by_row[row_idx] = {"num": c_num, "active": c_active, "inactive": c_inactive}

    for row_idx, vals in values_by_row.items():
        if num_col1 > 0:
            ws.cell(row=row_idx, column=num_col1).value = vals.get("num", 0)
        if active_col1 > 0:
            ws.cell(row=row_idx, column=active_col1).value = vals.get("active", 0)
        if inactive_col1 > 0:
            ws.cell(row=row_idx, column=inactive_col1).value = vals.get("inactive", 0)

    wb_mgmt.save(mgmt_file)
    wb_mgmt.close()
    print(f"Updated: {mgmt_file.name} ({len(values_by_row)} rows corrected)")

    print("\n" + "=" * 60)
    print("Done.")
    print("=" * 60)


if __name__ == "__main__":
    main()
