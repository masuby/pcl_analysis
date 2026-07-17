"""
Combined Management Processor
=============================
This script combines the functionality of:
1. Loan Management Processor
2. Settlement Management Processor  
3. Client Management Processor

All processes run sequentially in order.
"""

import os
import sys
import zipfile
from datetime import datetime, time
import pandas as pd
from openpyxl import load_workbook
import openpyxl


def read_excel_safe(path, **kwargs):
    """Read an .xlsx after confirming it is a complete, valid workbook.

    A real .xlsx is a ZIP archive; an incomplete download, a half-written file,
    or a OneDrive online-only placeholder fails deep inside openpyxl with a
    cryptic ``BadZipFile: File is not a zip file``. This validates first and
    raises a clear, actionable message instead.
    """
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Excel file not found: {path}")
    if not zipfile.is_zipfile(path):
        size = os.path.getsize(path)
        raise ValueError(
            f"'{os.path.basename(path)}' is not a valid .xlsx file (incomplete or corrupt, "
            f"size {size:,} bytes). It was probably still downloading/saving, was interrupted "
            f"mid-write, or is a cloud (OneDrive) placeholder. Re-export/re-download it fully, "
            f"make sure it is closed in Excel, then run again.\n  Path: {path}"
        )
    return pd.read_excel(path, engine="openpyxl", **kwargs)

# ============================================================================
# SHARED CONFIGURATION
# ============================================================================

folder = os.environ.get("PCL_MANAGEMENT_ROW_FILES") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "Management", "ROW_FILES")
    if os.environ.get("PCL_AUTOMATION_ROOT")
    else os.path.join(os.path.dirname(os.path.abspath(__file__)), "ROW_FILES"))
today = datetime.today().strftime("%Y-%m-%d")

# ============================================================================
# SHARED HELPER FUNCTION
# ============================================================================

def read_excel_preserve_dates(filepath):
    wb = load_workbook(filename=filepath, data_only=True)
    ws = wb.active

    # Read headers
    columns = [cell.value for cell in ws[1]]

    data = []

    # Read rows WITHOUT modifying anything
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    return pd.DataFrame(data, columns=columns)

# ============================================================================
# LOAN MANAGEMENT PROCESSOR
# ============================================================================

print("\n" + "="*80)
print("STARTING LOAN MANAGEMENT PROCESSOR")
print("="*80 + "\n")

# ----------------------------------------------------------
# STEP 1: Collect VALID Loan_Accounts files
#   - Must contain "Loan_Accounts" in filename
#   - Must contain column "Activation Date"
# ----------------------------------------------------------

loan_account_files = []

for file in os.listdir(folder):
    if "Loan_Accounts" in file:
        file_path = os.path.join(folder, file)

        # Check if Activation Date column exists
        try:
            df_check = read_excel_preserve_dates(file_path)
            if "Activation Date" in df_check.columns:
                loan_account_files.append(file_path)
            else:
                print(f"⚠️ Skipped (no Activation Date): {file}")
        except Exception as e:
            print(f"⚠️ Failed to read {file}: {e}")

print("\nValid Loan_Accounts files detected:")
for f in loan_account_files:
    print(" -", os.path.basename(f))

# ----------------------------------------------------------
# Guard: STEP 2 sorts these into exactly TWO buckets (Closed +
# Active_In_Arrears) and renames each onto a fixed name. With more than two,
# the second file to land in a bucket hits an existing target and dies with a
# cryptic FileExistsError (WinError 183); with fewer, later steps get no data.
# Fail fast with something readable instead.
# ----------------------------------------------------------
if len(loan_account_files) != 2:
    print("\n" + "=" * 80)
    print(f"ERROR: found {len(loan_account_files)} valid Loan_Accounts file(s) - expected exactly 2.")
    print("=" * 80)
    print("  This step needs exactly two: the CLOSED export and the ACTIVE/IN-ARREARS export.")
    if len(loan_account_files) > 2:
        print("\n  Remove the extra one(s) from ROW_FILES (open the file list in the")
        print("  Start popup and click the X, or use 'Clean row files'), then run again:")
        for f in loan_account_files:
            print(f"    - {os.path.basename(f)}")
    else:
        print("\n  Upload the missing Loan_Accounts export(s) and run again.")
        print("  Note: a file only counts when it is a complete .xlsx AND has an")
        print("  'Activation Date' column - partial downloads are skipped above.")
    sys.exit(1)

closed_file_path = None
active_file_path = None


# ----------------------------------------------------------
# STEP 2: Identify & rename Active vs Closed files
# ----------------------------------------------------------

for file_path in loan_account_files:

    df = read_excel_preserve_dates(file_path)

    # Check if file is empty (no rows) or contains "Closed" in Account State
    is_empty = len(df) == 0
    has_closed = "Closed" in df["Account State"].astype(str).values if not is_empty else False
    
    if is_empty or has_closed:
        new_name = f"Closed_Loan_Accounts_{today}.xlsx"
        closed_file_path = os.path.join(folder, new_name)
        os.rename(file_path, closed_file_path)
        print(f"Renamed → {new_name}")

    else:
        new_name = f"Active_In_Arrears_Loan_Accounts_{today}.xlsx"
        active_file_path = os.path.join(folder, new_name)

        os.rename(file_path, active_file_path)
        print(f"Renamed → {new_name}")


# ----------------------------------------------------------
# STEP 3: Append Closed + Active
# ----------------------------------------------------------
output_path = None  # Initialize output_path
if closed_file_path and active_file_path:

    df_active = read_excel_preserve_dates(active_file_path)
    df_closed = read_excel_preserve_dates(closed_file_path)

    merged_df = pd.concat([df_active, df_closed], ignore_index=True)


    output_name = f"Closed_Active_In_Arrears_Loan_Accounts_{today}.xlsx"
    output_path = os.path.join(folder, output_name)

    merged_df.to_excel(output_path, index=False)
    print(len(merged_df))

# ----------------------------------------------------------
# STEP 4: Delete the two original files and remove Staff Loans
# ----------------------------------------------------------

# 4.1 Delete the original Active and Closed files
try:
    if closed_file_path and os.path.exists(closed_file_path):
        os.remove(closed_file_path)
    if active_file_path and os.path.exists(active_file_path):
        os.remove(active_file_path)
except Exception as e:
    print("Error while deleting temp files:", e)

# 4.2 Open the combined file
final_df = read_excel_preserve_dates(output_path)

# 4.3 Remove all rows where "Loan Name" contains "Staff Loan"
final_df = final_df[~final_df["Loan Name"].astype(str).str.contains("Staff", case=False, na=False)]

# 4.4 Save the cleaned version (overwrite same file)
final_df.to_excel(output_path, index=False)
print("Current Length of the file is : ",len(final_df))

print("Step 4 completed: Removed Staff Loan rows and deleted old files.")

# ---------------------------------------------------------------------------------------
# FIXED STEP 5: correctly remove rows where "Sales Reps (Client)" is blank/NaN/whitespace
# ---------------------------------------------------------------------------------------

# load the file
df_step5 = read_excel_preserve_dates(output_path)

col = "Sales Reps (Client)"
if col not in df_step5.columns:
    raise KeyError(f'Column "{col}" not found in the dataframe.')

# create a Series view
s = df_step5[col]

# mask: keep rows where value is not NaN and after stripping is not empty
mask = s.notna() & s.astype(str).str.strip().ne("")

before = len(df_step5)
df_step5 = df_step5[mask].copy()
after = len(df_step5)

# save back
df_step5.to_excel(output_path, index=False)

print(f'Step 5 completed: removed {before - after} rows. Remaining rows: {after}.')

# ----------------------------------------------------------
# STEP 6: Delete rows where "Branch" contains "Deactivated"
# ----------------------------------------------------------

df_step6 = read_excel_preserve_dates(output_path)

col = "Branch"
if col not in df_step6.columns:
    raise KeyError(f'Column "{col}" not found in the dataframe.')

# Remove rows where Branch contains "Deactivated" (case-insensitive)
mask = ~df_step6[col].astype(str).str.contains("Deactivated", case=False, na=False)

before = len(df_step6)
df_step6 = df_step6[mask].copy()
after = len(df_step6)

df_step6.to_excel(output_path, index=False)

print(f'Step 6 completed: removed {before - after} rows with Branch = Deactivated.')
print(after)

# ---------------------------------------------------
# STEP 7 & 8 (MERGED): Assign Sales Reps based on condition
# ---------------------------------------------------

df = read_excel_preserve_dates(output_path)

# Ensure required columns exist
required_cols = [
    "Branch",
    "Sales Reps (Client)",
    "Call Centre Agent (Refinance) (Client)"
]

for col in required_cols:
    if col not in df.columns:
        raise KeyError(f'Column "{col}" not found in the dataframe.')

# Clean columns as string for safe checks
branch_col = df["Branch"].astype(str)
sales_reps_col = df["Sales Reps (Client)"].astype(str)
agent_col = df["Call Centre Agent (Refinance) (Client)"]

# Check agent NOT blank
agent_not_blank = agent_col.notna() & (agent_col.astype(str).str.strip() != "")

# Combined condition
mask_assign = (
    branch_col.str.contains("LBF Call Center", case=False, na=False) &
    sales_reps_col.str.contains("Unallocate LBF Callcenter", case=False, na=False) &
    agent_not_blank
)

# Names to be assigned
assigned_names = df.loc[mask_assign, "Call Centre Agent (Refinance) (Client)"]

# Perform assignment
df.loc[mask_assign, "Sales Reps (Client)"] = assigned_names

# Reporting
num_assigned = mask_assign.sum()

print("\nStep 7 & 8 (Merged)")
print(f"Total rows where 'Sales Reps (Client)' was assigned: {num_assigned}")

if num_assigned > 0:
    print("\nNames assigned to 'Sales Reps (Client)':")
    for name in assigned_names:
        print(" -", name)

# Save back to file
df.to_excel(output_path, index=False)

print("\nStep completed: Sales Reps updated correctly based on combined condition.")

# ----------------------------------------------------------
# STEP 9: Add "Net disbursement" column (treat empty Pay Off as 0)
# ----------------------------------------------------------

df_step9 = read_excel_preserve_dates(output_path)

# Ensure required columns exist
cols_needed = ["Loan Amount", "Pay Off Amount"]
for col in cols_needed:
    if col not in df_step9.columns:
        raise KeyError(f'Column "{col}" not found in the dataframe.')

# Replace NaN in Pay Off Amount with 0
df_step9["Pay Off Amount"] = df_step9["Pay Off Amount"].fillna(0)

# Calculate Net disbursement
df_step9["Net disbursement"] = df_step9["Loan Amount"] - df_step9["Pay Off Amount"]

# Save back to the same file
df_step9.to_excel(output_path, index=False)

print('Step 9 completed: "Net disbursement" column added, empty Pay Off treated as 0.')

# -------------------------------------------------------------------
# STEP 10: Transform "Completed Loan Cycles (Client)" to NEW or REPEAT
# -------------------------------------------------------------------

df_step10 = read_excel_preserve_dates(output_path)

col = "Completed Loan Cycles (Client)"
if col not in df_step10.columns:
    raise KeyError(f'Column "{col}" not found in the dataframe.')

# Convert to numeric, treat NaN as 0
df_step10[col] = pd.to_numeric(df_step10[col], errors='coerce').fillna(0)

# Apply transformation
df_step10[col] = df_step10[col].apply(lambda x: "NEW" if x == 0 else "REPEAT")

# Save back to the same file
df_step10.to_excel(output_path, index=False)

print('Step 10 completed: "Completed Loan Cycles (Client)" updated to NEW / REPEAT.')

# ----------------------------------------------------------
# STEP 11: Add "Type of business" column
# ----------------------------------------------------------

df_step11 = read_excel_preserve_dates(output_path)

# Ensure required columns exist
cols_needed = ["Completed Loan Cycles (Client)", "Loan Amount", "Net disbursement"]
for col in cols_needed:
    if col not in df_step11.columns:
        raise KeyError(f'Column "{col}" not found in the dataframe.')

# Apply the logic: Empty for NEW, REACTIVATION when Loan Amount == Net disbursement (REPEAT), else REFINANCE
def type_of_business(row):
    client_type = row["Completed Loan Cycles (Client)"]
    # Empty for NEW (after Step 10 transform, values are "NEW" or "REPEAT")
    if client_type == "NEW" or client_type == 0:
        return ""
    # For REPEAT: REACTIVATION if Loan Amount equals Net disbursement, else REFINANCE
    la = row["Loan Amount"]
    net = row["Net disbursement"]
    try:
        la_val = float(la) if pd.notna(la) else None
        net_val = float(net) if pd.notna(net) else None
        if la_val is not None and net_val is not None:
            if abs(la_val - net_val) < 0.01:  # tolerant comparison for floats
                return "REACTIVATION"
    except (TypeError, ValueError):
        pass
    return "REFINANCE"
 
df_step11["Type of business"] = df_step11.apply(type_of_business, axis=1)

# Save back to the same file
df_step11.to_excel(output_path, index=False)

print('Step 11 completed: "Type of business" column added.')


# ---------------------------------------------------------------------------
# STEP 12: Copy "Net disbursement" to "Loan Amount" and delete Pay Off columns
# ---------------------------------------------------------------------------

df_step12 = read_excel_preserve_dates(output_path)

# 12.1 Ensure "Net disbursement" exists
if "Net disbursement" not in df_step12.columns:
    raise KeyError('"Net disbursement" column not found in the dataframe.')

# 12.2 Copy Net disbursement to Loan Amount
df_step12["Loan Amount"] = df_step12["Net disbursement"]

# 12.3 Delete "Pay Off Amount" and "Pay-Off Amount" columns if they exist
cols_to_drop = [col for col in ["Pay Off Amount", "Pay-Off Amount"] if col in df_step12.columns]
df_step12.drop(columns=cols_to_drop, inplace=True)

# Save back to the same file
df_step12.to_excel(output_path, index=False)

print(f'Step 12 completed: "Loan Amount" updated and columns {cols_to_drop} deleted.')

# ----------------------------------------------------------
# STEP 13: Update Loan.xlsx with processed data
# ----------------------------------------------------------

# Paths
processed_file = output_path
loan_file = os.path.join(folder, "Loan.xlsx")

# --- Load processed file
df_processed = read_excel_preserve_dates(processed_file)

# --- Rename columns
rename_mapping = {
    "Activation Date": "Activation Date (Loan)",
    "Last Payment Date": "Date of Drop off/partials",
    "Loan performing Status": "Status",
    "Completed Loan Cycles (Client)": "CLIENT TYPE",
    "Loan Type": "LOAN TYPE",
    "Account Sub-State": "Check"
}
df_processed.rename(columns=rename_mapping, inplace=True)

# --- Save processed file with renamed columns
df_processed.to_excel(processed_file, index=False, engine="openpyxl")
print(f"Processed file saved with renamed columns: {list(df_processed.columns)}")

# --- Load Loan.xlsx
df_loan = read_excel_preserve_dates(loan_file)

# --- Find similar columns
similar_columns = [col for col in df_processed.columns if col in df_loan.columns]
print("Similar columns to update:", similar_columns)

# --- Create new Loan DataFrame
new_loan_df = pd.DataFrame(columns=df_loan.columns)  # preserve original column structure

# --- Copy similar columns from processed file
for col in similar_columns:
    new_loan_df[col] = df_processed[col]

# --- Fill remaining columns with None
for col in new_loan_df.columns:
    if col not in similar_columns:
        new_loan_df[col] = [None] * len(df_processed)

# --- Save updated Loan.xlsx
new_loan_df.to_excel(loan_file, index=False, engine="openpyxl")

print("Step 13 completed: Similar columns updated, rows now match processed file, timestamps preserved.")
print("Rows in processed file:", len(df_processed))
print("Rows in updated Loan.xlsx:", len(new_loan_df))

# -------------------------------------------------------------
# NEW: Rename worksheet to "Loan Accounts"
# -------------------------------------------------------------
wb = openpyxl.load_workbook(loan_file)
ws = wb.active
ws.title = "Loan Accounts"
wb.save(loan_file)

# -------------------------------------------------------------
# Delete the temporary merged file after processing
# -------------------------------------------------------------
if 'output_path' in locals() and output_path and os.path.exists(output_path):
    try:
        os.remove(output_path)
        print(f"Deleted temporary file: {os.path.basename(output_path)}")
    except Exception as e:
        print(f"⚠️ Error deleting temporary file: {e}")

print("\n" + "="*80)
print("LOAN MANAGEMENT PROCESSOR COMPLETED")
print("="*80 + "\n")

# ============================================================================
# SETTLEMENT MANAGEMENT PROCESSOR
# ============================================================================

print("\n" + "="*80)
print("STARTING SETTLEMENT MANAGEMENT PROCESSOR")
print("="*80 + "\n")

settlement_file_path = None

print("----- Checking for Settlement Loan_Accounts file -----")

# First, check if settlement file already exists (from previous run)
settlement_file = os.path.join(folder, f"Settlement_Loan_Accounts_{today}.xlsx")
if os.path.exists(settlement_file):
    settlement_file_path = settlement_file
    print(f"✅ Settlement file already exists: Settlement_Loan_Accounts_{today}.xlsx")
    print("✔ Step 1 completed: Using existing settlement file.")
else:
    # Look for original settlement file to rename
    # Settlement file must:
    # 1. Start with "Loan_Accounts" (original file, not processed)
    # 2. NOT have "Activation Date" column
    # Exclude files that have already been processed (renamed by loan processor)
    excluded_prefixes = [
        "Closed_Loan_Accounts_", 
        "Active_In_Arrears_Loan_Accounts_", 
        "Settlement_Loan_Accounts_",
        "Closed_Active_In_Arrears_Loan_Accounts_"
    ]
    
    for file in os.listdir(folder):

        # Must start with "Loan_Accounts" (original file pattern)
        if not file.startswith("Loan_Accounts"):
            continue
            
        # Skip files that have already been processed/renamed
        if any(file.startswith(prefix) for prefix in excluded_prefixes):
            print(f"Skipped already processed file: {file}")
            continue

        file_path = os.path.join(folder, file)

        try:
            # normal pandas read – no date preservation needed
            df_check = pd.read_excel(file_path, engine="openpyxl")
        except Exception as e:
            print(f"⚠️ Could not read {file}: {e}")
            continue

        # Identify Settlement file → No "Activation Date" column
        if "Activation Date" not in df_check.columns:
            print(f"Found settlement file: {file} (rows: {len(df_check)})")

            new_name = f"Settlement_Loan_Accounts_{today}.xlsx"
            settlement_file_path = os.path.join(folder, new_name)

            os.rename(file_path, settlement_file_path)

            print(f"✅ Settlement file found & renamed → {new_name}")
            break  # Found and renamed, no need to continue
        else:
            print(f"Skipped normal loan file: {file} (has Activation Date column)")

    if settlement_file_path is None:
        print("⚠️ No settlement file found.")
        print("⚠️ Skipping settlement processing - file not available.")

# ----------------------------------------------------------
# Step 2: Load Settlement file and rename columns
# ----------------------------------------------------------

# Only proceed if settlement file exists
if settlement_file_path is not None and os.path.exists(settlement_file_path):
    
    # Use the settlement file we found/created
    settlement_file = settlement_file_path
    
    # File exists, proceed with processing
    df_settlement = pd.read_excel(settlement_file, engine="openpyxl")
    
    # Clean up "Unnamed" column headers - replace with empty string
    df_settlement.columns = [col if not str(col).startswith("Unnamed") else "" for col in df_settlement.columns]
    
    print(f"Loaded settlement file: {os.path.basename(settlement_file)}")
    print(f"Settlement file has {len(df_settlement)} rows and {len(df_settlement.columns)} columns")

    # Column renaming mapping
    rename_mapping = {
        "Sales Reps (Client)": "Sales Officer",
        "Settlement Principle balance": "Principal-Balance",
        "Source of settlement Payment": "source of settlement pyt"
    }

    df_settlement.rename(columns=rename_mapping, inplace=True)
    df_settlement.to_excel(settlement_file, index=False, engine="openpyxl")
    print("Step 2 completed: Columns renamed.")
    print("Current columns:", list(df_settlement.columns))

    # ----------------------------------------------------------
    # Step 3: Copy similar columns into Settlements.xlsx
    # ----------------------------------------------------------

    settlements_file = os.path.join(folder, "Settlements.xlsx")
    if not os.path.exists(settlements_file):
        print(f"⚠️ Settlements.xlsx not found! Skipping Step 3.")
    else:
        df_settlements = pd.read_excel(settlements_file, engine="openpyxl")
        
        # Clean up "Unnamed" column headers in Settlements.xlsx - replace with empty string
        df_settlements.columns = [col if not str(col).startswith("Unnamed") else "" for col in df_settlements.columns]

        # Identify similar columns
        similar_columns = sorted(set(df_settlement.columns).intersection(set(df_settlements.columns)))

        print("\nSimilar columns to be updated:")
        for col in similar_columns:
            print(" -", col)

        # Create new Settlements DataFrame with same columns as original
        new_settlements = pd.DataFrame(columns=df_settlements.columns)

        # Resize to match settlement file rows
        new_settlements = new_settlements.reindex(range(len(df_settlement)))

        # Copy similar columns
        for col in similar_columns:
            new_settlements[col] = df_settlement[col].values

        # Clean up "Unnamed" column headers in new_settlements - replace with empty string
        new_settlements.columns = [col if not str(col).startswith("Unnamed") else "" for col in new_settlements.columns]

        # Save updated Settlements.xlsx
        new_settlements.to_excel(settlements_file, index=False, engine="openpyxl")

        # Rename worksheet to "Settlements"
        wb = openpyxl.load_workbook(settlements_file)
        ws = wb.active
        ws.title = "Settlements"
        wb.save(settlements_file)

        print("\nStep 3 completed: Similar columns updated in Settlements.xlsx.")
        print(f"Rows in settlement file: {len(df_settlement)}")
        print(f"Rows in updated Settlements.xlsx: {len(new_settlements)}")
else:
    print("⚠️ Settlement processing skipped - no settlement file available.")

print("\n" + "="*80)
print("SETTLEMENT MANAGEMENT PROCESSOR COMPLETED")
print("="*80 + "\n")

# ============================================================================
# CLIENT MANAGEMENT PROCESSOR
# ============================================================================

print("\n" + "="*80)
print("STARTING CLIENT MANAGEMENT PROCESSOR")
print("="*80 + "\n")

# --- Step 1.1: Define folder and find the client file
# (folder already defined above)

# Find the client file (starts with 'Clients-platinumtanzania')
client_file = None
for file in os.listdir(folder):
    if file.startswith("Clients-platinumtanzania") and file.endswith((".xlsx", ".xls")):
        client_file = os.path.join(folder, file)
        break

if not client_file:
    raise FileNotFoundError("No client file found in folder starting with 'Clients-platinumtanzania'")

# --- Step 1.2: Load the client file
df_clients = pd.read_excel(client_file, engine="openpyxl")

print(f"Client file loaded: {client_file}")
print(f"Initial number of rows: {len(df_clients)}")

# --- Step 1.3: Filter out rows with unwanted Branch keywords
keywords = ["Head Office","Head Office", "HR", "Deactivated", "Collection", "Stock24"]

# Build a boolean mask for rows to drop
mask = df_clients["Branch"].astype(str).str.contains("|".join(keywords), case=False, na=False)

# Print number of rows to be deleted
num_deleted = mask.sum()
num_remaining = (~mask).sum()
print(f"Rows to delete: {num_deleted}")
print(f"Rows remaining: {num_remaining}")

# Delete the rows
df_clients_filtered = df_clients[~mask].copy()

# Optional: reset index
df_clients_filtered.reset_index(drop=True, inplace=True)

# --- Step 1.4: Save filtered client file (optional)
filtered_file = os.path.join(folder, "Clients-platinumtanzania_filtered.xlsx")
df_clients_filtered.to_excel(filtered_file, index=False, engine="openpyxl")

print(f"Filtered client file saved: {filtered_file}")

# -------------------------------------------------------------
# STEP 2.1 — Keep only filtered client file
# -------------------------------------------------------------

filtered_client_file = os.path.join(folder, "Clients-platinumtanzania_filtered.xlsx")
print("client file:", filtered_client_file)

df_filtered = read_excel_safe(filtered_client_file)

# -------------------------------------------------------------
# STEP 2.2 — Load Clientstz.xlsx
# -------------------------------------------------------------

clientstz_file = os.path.join(folder, "Clientstz.xlsx")
df_clientstz = read_excel_safe(clientstz_file)

print(f"Loaded Clientstz.xlsx with {len(df_clientstz)} rows.")
print(f"Filtered file has {len(df_filtered)} rows.")

# -------------------------------------------------------------
# STEP 2.3 — Compare column names
# -------------------------------------------------------------

filtered_cols = set(df_filtered.columns)
clientstz_cols = set(df_clientstz.columns)

similar_columns = sorted(filtered_cols.intersection(clientstz_cols))

print("\nSimilar columns to be copied:")
for col in similar_columns:
    print(" -", col)

# -------------------------------------------------------------
# STEP 2.4 — Build a NEW dataframe for Clientstz
# -------------------------------------------------------------
new_clientstz = pd.DataFrame(columns=df_clientstz.columns)
new_clientstz = new_clientstz.reindex(range(len(df_filtered)))

# Copy similar columns
for col in similar_columns:
    new_clientstz[col] = df_filtered[col].values

# -------------------------------------------------------------
# STEP 2.5 — Save back to Clientstz.xlsx
# -------------------------------------------------------------
new_clientstz.to_excel(clientstz_file, index=False, engine="openpyxl")

print("\nStep completed successfully.")
print(f"Final rows in Clientstz.xlsx: {len(new_clientstz)}")
print("Rows now match the filtered client file exactly.")

# -------------------------------------------------------------
# NEW: Rename the worksheet to "Clients"
# -------------------------------------------------------------
wb = openpyxl.load_workbook(clientstz_file)
ws = wb.active
ws.title = "Clients"
wb.save(clientstz_file)

print("\n" + "="*80)
print("CLIENT MANAGEMENT PROCESSOR COMPLETED")
print("="*80 + "\n")

print("\n" + "="*80)
print("ALL PROCESSORS COMPLETED SUCCESSFULLY!")
print("="*80 + "\n")

# ============================================================================
# POST LOAN MANAGEMENT PROCESSOR
# ============================================================================

print("\n" + "="*80)
print("STARTING POST LOAN MANAGEMENT PROCESSOR")
print("="*80 + "\n")

loan_file = os.path.join(folder, "Loan.xlsx")

# Load the Loan.xlsx file
print(f"Loading {loan_file}...")
df_post = pd.read_excel(loan_file, engine="openpyxl")

print(f"Initial number of rows: {len(df_post)}")

# Check if required columns exist
required_cols = ["Account Holder Name", "Loan Name"]
for col in required_cols:
    if col not in df_post.columns:
        raise KeyError(f'Column "{col}" not found in the dataframe.')

# Filter for specific Account Holder names (CL Asset Finance list)
target_names = [
    "SAIDI J. JAWAZI",
    "MUSA SAMWELY NDIBATO",
    "AYUBU EFESO CHIPULILO",
    "KORNEL MTATIRO MAISA",
    "INSPA IAN SINYINZA",
    "BARNABA SAMSON SAHINGA",
    "SENSA MUSSA CHINOGA",
    "HAPPYNESS KASUKU MUSSA",
    "MARIAM J MASOTA",
    "STEPHANO LUKWARO MKAZENI",
    "BRIGHT JAPHET KENNEDY",
    "EDDA HUBERT MWAKANYAMALE",
    "REVOCATUS CHAMKAGA SEBASTIAN",
    "SELLAH NORBERT NTAMPELA",
    "PHILOMENA CLAUDE HAMISI",
    "MASHI HASSAN MASHI",
    "EXAVERY MASANJA STEVEN",
    "BRIGHT JAPHET KENNEDY",
    "STEPHANO CHARLES MAGOMA",
    "SUZAN MARTIN CHUWA",
    "DIOSKORI EPHRAIM HILBAJOJO",
    "HAMIS CHACHA MAKENGE",
    "FESTO JEREMIAH MIDAHO",
    "ASNATH NESTORY ANTHONY",
    "COSMAS MARCO KAPERA",
    "GIDION MUSSA GWANCHELE",
    "MARIAM MWITA JACOB",
    "STEPHANO CHARLES MAGOMA",
    "KASHIHORI RASHID KASHIHORI",
    "CHRISTER CHARLES HAULE",
    "FAUSTIN MKINA MKINA",
    "JUMA MASOUD MUSA",
    "NUHU ABDALLAH SAIDI",
]
account_names = df_post["Account Holder Name"].astype(str).str.strip().str.upper()
mask = account_names.isin(target_names)

# Count how many rows will be updated
num_rows_to_update = mask.sum()
print(
    "Found "
    f"{num_rows_to_update} row(s) with Account Holder Name in: "
    f"{', '.join(target_names)}"
)

if num_rows_to_update > 0:
    # Replace Loan Name with "CL - Asset Finance" for matching rows
    df_post.loc[mask, "Loan Name"] = "CL - Asset Finance"

    print(f"Updated {num_rows_to_update} row(s): Loan Name set to 'CL - Asset Finance'")

    # Save the updated file
    df_post.to_excel(loan_file, index=False, engine="openpyxl")
    print(f"File saved successfully: {loan_file}")
else:
    print("No rows found with Account Holder Name in the target list. No changes made.")

# Ensure worksheet is renamed to "Loan Accounts"
wb_post = openpyxl.load_workbook(loan_file)
ws_post = wb_post.active
ws_post.title = "Loan Accounts"
wb_post.save(loan_file)
print("Worksheet renamed to 'Loan Accounts'")

print("\n" + "="*80)
print("POST LOAN MANAGEMENT PROCESSOR COMPLETED")
print("="*80 + "\n")

# ============================================================================
# CLEANUP: Delete all files except required ones
# ============================================================================

print("\n" + "="*80)
print("STARTING CLEANUP PROCESS")
print("="*80 + "\n")

# Files to keep
files_to_keep = ["Loan.xlsx", "Settlements.xlsx", "Clientstz.xlsx", "Users.xlsx"]

print(f"Files to keep: {', '.join(files_to_keep)}")

# Get all files in the folder
all_files = os.listdir(folder)

# Filter for xlsx files only (and other temporary files if any)
deleted_count = 0
kept_count = 0

for file_name in all_files:
    file_path = os.path.join(folder, file_name)
    
    # Skip directories
    if os.path.isdir(file_path):
        continue
    
    # Keep the required files
    if file_name in files_to_keep:
        kept_count += 1
        print(f"  ✅ Kept: {file_name}")
    else:
        # Delete other files
        try:
            os.remove(file_path)
            deleted_count += 1
            print(f"  🗑️  Deleted: {file_name}")
        except Exception as e:
            print(f"  ⚠️  Error deleting {file_name}: {e}")

print(f"\nCleanup completed: {kept_count} file(s) kept, {deleted_count} file(s) deleted")

print("\n" + "="*80)
print("ALL PROCESSES COMPLETED SUCCESSFULLY!")
print("="*80 + "\n")
print(f"Remaining files in {folder}:")
for file_name in files_to_keep:
    file_path = os.path.join(folder, file_name)
    if os.path.exists(file_path):
        print(f"  ✅ {file_name}")
    else:
        print(f"  ⚠️  {file_name} (not found)")
print("="*80 + "\n")

