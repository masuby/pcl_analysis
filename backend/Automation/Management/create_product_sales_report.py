"""
Create Product Sales Report
============================
Creates an Excel file with:
- Columns: Loan Name, Loan Amount, Products (from Users), Branch
- Summary: Total sales per Product (accounting number format)
Data from Loan.xlsx and Users.xlsx (product_mapping).
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font

# ============================================================================
# CONFIGURATION
# ============================================================================

ROW_FILES = os.environ.get("PCL_MANAGEMENT_ROW_FILES") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "Management", "ROW_FILES")
    if os.environ.get("PCL_AUTOMATION_ROOT")
    else os.path.join(os.path.dirname(os.path.abspath(__file__)), "ROW_FILES"))
USERS_FILE = os.path.join(ROW_FILES, "Users.xlsx")
LOAN_FILE = os.path.join(ROW_FILES, "Loan.xlsx")
ZONE_CLUSTER_FILE = os.path.join(os.path.dirname(__file__), "Zone and cluster.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "Product_Sales_Report.xlsx")

# Accounting format: thousand separators, 2 decimals
ACCOUNTING_FORMAT = "#,##0.00"


def _normalize(val):
    if val is None:
        return ""
    return str(val).strip()


def _find_column_index(ws, header_row, target_names):
    headers = [cell.value for cell in ws[header_row]]
    target_lower = [t.lower() for t in target_names]
    for col_idx, h in enumerate(headers, 1):
        if h is None:
            continue
        h_lower = str(h).strip().lower()
        for t in target_lower:
            if t in h_lower or h_lower in t:
                return col_idx
    return None


def _build_breakdown(rows, key_field):
    """Sum New/Repeat (CLIENT TYPE) and Reactivation/Refinance (Type of business)
    plus Total, grouped by ``key_field`` (e.g. 'Products', 'Zone', 'Cluster').
    Rows with a blank key are skipped."""
    breakdown = {}
    for r in rows:
        key = _normalize(r.get(key_field, ""))
        if not key:
            continue
        if key not in breakdown:
            breakdown[key] = {"New Business": 0, "Repeat Business": 0,
                              "Reactivation": 0, "Refinance": 0, "Total": 0}
        amt = r["Loan Amount"]
        ct = _normalize(r.get("CLIENT TYPE", ""))
        tob = _normalize(r.get("Type of business", ""))
        if ct == "NEW":
            breakdown[key]["New Business"] += amt
        elif ct == "REPEAT":
            breakdown[key]["Repeat Business"] += amt
        if tob == "REACTIVATION":
            breakdown[key]["Reactivation"] += amt
        elif tob == "REFINANCE":
            breakdown[key]["Refinance"] += amt
        breakdown[key]["Total"] += amt
    return breakdown


def _write_summary_block(ws, start_row, title, key_label, breakdown):
    """Write a '<title>' summary block (key + 5 amount columns + grand total).
    Returns the next free row."""
    ws.cell(start_row, 1, title).font = Font(bold=True, size=12)
    start_row += 1
    headers = [key_label, "New Business", "Repeat Business", "Reactivation", "Refinance", "Total"]
    for c, h in enumerate(headers, 1):
        ws.cell(start_row, c, h).font = Font(bold=True)
    start_row += 1
    for key in sorted(breakdown.keys()):
        b = breakdown[key]
        ws.cell(start_row, 1, key)
        for c, k in enumerate(["New Business", "Repeat Business", "Reactivation", "Refinance", "Total"], 2):
            ws.cell(start_row, c, b[k]).number_format = ACCOUNTING_FORMAT
        start_row += 1
    start_row += 1
    grand = {"New Business": 0, "Repeat Business": 0, "Reactivation": 0, "Refinance": 0, "Total": 0}
    for b in breakdown.values():
        for k in grand:
            grand[k] += b.get(k, 0)
    ws.cell(start_row, 1, f"Grand Total ({key_label})").font = Font(bold=True)
    for c, k in enumerate(["New Business", "Repeat Business", "Reactivation", "Refinance", "Total"], 2):
        cell = ws.cell(start_row, c, grand[k])
        cell.number_format = ACCOUNTING_FORMAT
        cell.font = Font(bold=True)
    return start_row + 2


def main():
    print("=" * 60)
    print("Create Product Sales Report")
    print("=" * 60)

    if not os.path.exists(USERS_FILE):
        print(f"[ERROR] Users file not found: {USERS_FILE}")
        return
    if not os.path.exists(LOAN_FILE):
        print(f"[ERROR] Loan file not found: {LOAN_FILE}")
        return

    # Load Users - product_mapping
    users_wb = openpyxl.load_workbook(USERS_FILE, data_only=True)
    if "product_mapping" not in users_wb.sheetnames:
        print("[ERROR] product_mapping sheet not found in Users.xlsx")
        return

    ws_pm = users_wb["product_mapping"]
    ln_col = _find_column_index(ws_pm, 1, ["Loan Name", "LoanName"])
    prod_col = _find_column_index(ws_pm, 1, ["Products", "Product"])
    if not ln_col or not prod_col:
        print("[ERROR] Could not find Loan Name or Products in product_mapping")
        return

    loan_name_to_product = {}
    for row_idx in range(2, ws_pm.max_row + 1):
        ln = _normalize(ws_pm.cell(row_idx, ln_col).value)
        prod = ws_pm.cell(row_idx, prod_col).value
        if ln and ln not in loan_name_to_product:
            loan_name_to_product[ln] = prod if prod is not None else ""

    print(f"   Loaded {len(loan_name_to_product)} Loan Name -> Products mappings")

    # Load Zone & Cluster mapping: Branch -> (Zone, Cluster)
    branch_to_zone_cluster = {}
    if os.path.exists(ZONE_CLUSTER_FILE):
        zc_wb = openpyxl.load_workbook(ZONE_CLUSTER_FILE, data_only=True)
        zc_ws = zc_wb[zc_wb.sheetnames[0]]
        zc_hdr = {str(zc_ws.cell(1, c).value).strip().lower(): c
                  for c in range(1, zc_ws.max_column + 1) if zc_ws.cell(1, c).value is not None}
        zc_branch = zc_hdr.get("branch")
        zc_zone = zc_hdr.get("zone")
        zc_cluster = zc_hdr.get("cluster")
        for r in range(2, zc_ws.max_row + 1):
            b = _normalize(zc_ws.cell(r, zc_branch).value) if zc_branch else ""
            if not b:
                continue
            z = _normalize(zc_ws.cell(r, zc_zone).value) if zc_zone else ""
            cl = _normalize(zc_ws.cell(r, zc_cluster).value) if zc_cluster else ""
            branch_to_zone_cluster.setdefault(b.lower(), (z, cl))
        zc_wb.close()
        print(f"   Loaded {len(branch_to_zone_cluster)} Branch -> (Zone, Cluster) mappings")
    else:
        print(f"   [WARN] Zone and cluster file not found: {ZONE_CLUSTER_FILE}")

    # Load Loan - Loan Accounts
    loan_wb = openpyxl.load_workbook(LOAN_FILE, data_only=True)
    sheet_name = "Loan Accounts" if "Loan Accounts" in loan_wb.sheetnames else loan_wb.sheetnames[0]
    ws_loan = loan_wb[sheet_name]

    ln_loan_col = _find_column_index(ws_loan, 1, ["Loan Name", "LoanName"])
    # Prefer Net disbursement (expected total is based on this)
    amt_col = _find_column_index(ws_loan, 1, ["Net disbursement", "Loan Amount"])
    branch_col = _find_column_index(ws_loan, 1, ["Branch"])
    activation_col = _find_column_index(ws_loan, 1, ["Activation Date (Loan)", "Activation Date"])
    account_id_col = _find_column_index(ws_loan, 1, ["Account ID"])
    client_type_col = _find_column_index(ws_loan, 1, ["CLIENT TYPE", "Completed Loan Cycles (Client)"])
    type_of_business_col = _find_column_index(ws_loan, 1, ["Type of business"])

    if not ln_loan_col or not amt_col:
        print("[ERROR] Could not find Loan Name or Loan Amount in Loan file")
        return
    if not activation_col:
        print("[ERROR] Could not find 'Activation Date (Loan)' column in Loan file")
        return

    now = datetime.now()
    current_year, current_month = now.year, now.month

    def _is_current_month(val):
        """Return True if the date value is in the current month (and year)."""
        if val is None:
            return False
        if hasattr(val, "month") and hasattr(val, "year"):
            return val.month == current_month and val.year == current_year
        try:
            s = str(val).strip()
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    dt = datetime.strptime(s[:10], fmt)
                    return dt.month == current_month and dt.year == current_year
                except ValueError:
                    continue
            if len(s) >= 10 and s[4] == "-":
                return s[:4] == str(current_year) and s[5:7] == f"{current_month:02d}"
            return False
        except (TypeError, ValueError):
            return False

    # Build rows: Loan Name, Loan Amount, Products, Branch (only current month data)
    # Deduplicate by Account ID (or Loan Name+Amount+Branch+Date) to avoid double-counting
    seen_keys = set()
    rows = []
    for row_idx in range(2, ws_loan.max_row + 1):
        activation_val = ws_loan.cell(row_idx, activation_col).value
        if not _is_current_month(activation_val):
            continue

        ln = _normalize(ws_loan.cell(row_idx, ln_loan_col).value)
        amt = ws_loan.cell(row_idx, amt_col).value
        branch_val = ws_loan.cell(row_idx, branch_col).value if branch_col else ""
        branch = _normalize(branch_val) if branch_val is not None else ""
        products = loan_name_to_product.get(ln, "")

        # CLIENT TYPE: NEW or REPEAT (Completed Loan Cycles: 0->NEW, >0->REPEAT)
        client_type_val = ws_loan.cell(row_idx, client_type_col).value if client_type_col is not None else None
        if client_type_val is not None and str(client_type_val).strip().upper() in ("NEW", "REPEAT"):
            client_type = str(client_type_val).strip().upper()
        elif client_type_val is not None:
            try:
                n = float(client_type_val)
                client_type = "NEW" if n == 0 else "REPEAT"
            except (TypeError, ValueError):
                client_type = _normalize(client_type_val) or ""
        else:
            client_type = ""

        # Type of business: REACTIVATION, REFINANCE, or empty (for NEW)
        type_of_business_val = ws_loan.cell(row_idx, type_of_business_col).value if type_of_business_col is not None else None
        type_of_business = _normalize(type_of_business_val) if type_of_business_val else ""

        try:
            amt_val = float(amt) if amt is not None else 0
        except (TypeError, ValueError):
            amt_val = 0

        # Deduplication key: Account ID if available, else composite (Loan+Amount+Branch+Date)
        if account_id_col:
            aid = ws_loan.cell(row_idx, account_id_col).value
            if aid is not None and str(aid).strip():
                key = (str(aid).strip(),)
            else:
                key = (ln, amt_val, branch, str(activation_val)[:10] if activation_val else "")
        else:
            key = (ln, amt_val, branch, str(activation_val)[:10] if activation_val else "")

        if key in seen_keys:
            continue
        seen_keys.add(key)

        zone, cluster = branch_to_zone_cluster.get(branch.lower(), ("", ""))

        rows.append({
            "Loan Name": ln,
            "Loan Amount": amt_val,
            "Products": products,
            "Branch": branch,
            "CLIENT TYPE": client_type,
            "Type of business": type_of_business,
            "Zone": zone,
            "Cluster": cluster,
        })

    print(f"   Filtered {datetime(current_year, current_month, 1).strftime('%B %Y')} data (deduplicated): {len(rows)} loan rows")

    # Split mapped vs unmapped
    mapped_rows = [r for r in rows if _normalize(r["Products"])]
    unmapped_rows = [r for r in rows if not _normalize(r["Products"])]

    # Breakdowns (New Business, Repeat Business, Reactivation, Refinance, Total)
    product_breakdown = _build_breakdown(mapped_rows, "Products")
    zone_breakdown = _build_breakdown(mapped_rows, "Zone")
    cluster_breakdown = _build_breakdown(mapped_rows, "Cluster")

    unmapped_total = sum(r["Loan Amount"] for r in unmapped_rows)

    # Create Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Sales"

    # Headers
    headers = ["Loan Name", "Loan Amount", "Products", "Branch", "CLIENT TYPE",
               "Type of business", "Zone", "Cluster"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True)

    # Data (mapped only)
    for r_idx, r in enumerate(mapped_rows, 2):
        ws.cell(r_idx, 1, r["Loan Name"])
        cell_amt = ws.cell(r_idx, 2, r["Loan Amount"])
        cell_amt.number_format = ACCOUNTING_FORMAT
        ws.cell(r_idx, 3, r["Products"])
        ws.cell(r_idx, 4, r["Branch"])
        ws.cell(r_idx, 5, r.get("CLIENT TYPE", ""))
        ws.cell(r_idx, 6, r.get("Type of business", ""))
        ws.cell(r_idx, 7, r.get("Zone", ""))
        ws.cell(r_idx, 8, r.get("Cluster", ""))

    # Empty row then Summary
    summary_start = len(mapped_rows) + 3
    ws.cell(summary_start, 1, "Summary by Product")
    ws.cell(summary_start, 1).font = Font(bold=True, size=12)
    summary_start += 1

    summary_headers = ["Product", "New Business", "Repeat Business", "Reactivation", "Refinance", "Total"]
    for c, h in enumerate(summary_headers, 1):
        ws.cell(summary_start, c, h).font = Font(bold=True)
    summary_start += 1

    for prod in sorted(product_breakdown.keys()):
        pb = product_breakdown[prod]
        ws.cell(summary_start, 1, prod)
        ws.cell(summary_start, 2, pb["New Business"]).number_format = ACCOUNTING_FORMAT
        ws.cell(summary_start, 3, pb["Repeat Business"]).number_format = ACCOUNTING_FORMAT
        ws.cell(summary_start, 4, pb["Reactivation"]).number_format = ACCOUNTING_FORMAT
        ws.cell(summary_start, 5, pb["Refinance"]).number_format = ACCOUNTING_FORMAT
        ws.cell(summary_start, 6, pb["Total"]).number_format = ACCOUNTING_FORMAT
        summary_start += 1

    # Grand total (mapped)
    summary_start += 1
    grand_totals = {"New Business": 0, "Repeat Business": 0, "Reactivation": 0, "Refinance": 0, "Total": 0}
    for pb in product_breakdown.values():
        for k in grand_totals:
            grand_totals[k] += pb.get(k, 0)
    ws.cell(summary_start, 1, "Grand Total (Mapped)").font = Font(bold=True)
    for c, key in enumerate(["New Business", "Repeat Business", "Reactivation", "Refinance", "Total"], 2):
        cell = ws.cell(summary_start, c, grand_totals[key])
        cell.number_format = ACCOUNTING_FORMAT
        cell.font = Font(bold=True)

    # Summary by Zone, then by Cluster (so they can update the Management zone/cluster rows)
    next_row = summary_start + 3
    next_row = _write_summary_block(ws, next_row, "Summary by Zone", "Zone", zone_breakdown)
    next_row = _write_summary_block(ws, next_row, "Summary by Cluster", "Cluster", cluster_breakdown)

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 16

    # Unmapped sheet
    ws_unmapped = wb.create_sheet("Unmapped")
    unmapped_headers = ["Loan Name", "Loan Amount", "Branch", "CLIENT TYPE", "Type of business"]
    for c, h in enumerate(unmapped_headers, 1):
        ws_unmapped.cell(1, c, h).font = Font(bold=True)
    for r_idx, r in enumerate(unmapped_rows, 2):
        ws_unmapped.cell(r_idx, 1, r["Loan Name"])
        cell_amt = ws_unmapped.cell(r_idx, 2, r["Loan Amount"])
        cell_amt.number_format = ACCOUNTING_FORMAT
        ws_unmapped.cell(r_idx, 3, r["Branch"])
        ws_unmapped.cell(r_idx, 4, r.get("CLIENT TYPE", ""))
        ws_unmapped.cell(r_idx, 5, r.get("Type of business", ""))

    summary_unmapped = len(unmapped_rows) + 3
    ws_unmapped.cell(summary_unmapped, 1, "Total Unmapped").font = Font(bold=True)
    ws_unmapped.cell(summary_unmapped, 2, unmapped_total).number_format = ACCOUNTING_FORMAT
    ws_unmapped.cell(summary_unmapped, 2).font = Font(bold=True)
    ws_unmapped.column_dimensions["A"].width = 35
    ws_unmapped.column_dimensions["B"].width = 16
    ws_unmapped.column_dimensions["C"].width = 25
    ws_unmapped.column_dimensions["D"].width = 14
    ws_unmapped.column_dimensions["E"].width = 16

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Created {OUTPUT_FILE}")
    print(f"     Mapped: {len(mapped_rows)} rows | Unmapped: {len(unmapped_rows)} rows")
    print(f"     Mapped total: {grand_totals['Total']:,.2f} | Unmapped total: {unmapped_total:,.2f}")


if __name__ == "__main__":
    main()
