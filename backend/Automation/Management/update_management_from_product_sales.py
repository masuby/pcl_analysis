"""
Update Management Report from Product Sales Report
==================================================
Standalone script: assumes the Management report already exists.
Opens Product_Sales_Report.xlsx, then updates the Management report's Country sheet
the same way process_management.py does, so numbers align with the Summary by Product.
Uses openpyxl only (preserves formatting).
"""

import os
import glob
import openpyxl

# Paths (same as process_management.py)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PRODUCT_SALES_REPORT_PATH = os.path.join(SCRIPT_DIR, "Product_Sales_Report.xlsx")
ZONE_CLUSTER_PATH = os.path.join(SCRIPT_DIR, "Zone and cluster.xlsx")
# Must match process_management.py's destination_folder.
_MGMT_DIR = os.environ.get("PCL_MANAGEMENT_DIR") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "Management")
    if os.environ.get("PCL_AUTOMATION_ROOT") else SCRIPT_DIR)
DESTINATION_FOLDER = os.environ.get("PCL_MANAGEMENT_OUT", os.path.join(_MGMT_DIR, "OUTPUT"))

HEADER_ROW = 1
START_SUM_ROW = 3
END_SUM_ROW = 16
MAX_COL = 100


def _num(x):
    try:
        return float(x) if x is not None else 0
    except (TypeError, ValueError):
        return 0


def load_product_totals():
    """Load per-product rows and Grand Total (Mapped) from Product_Sales_Report.xlsx."""
    product_totals_from_report = None
    product_totals_by_name = {}
    if not os.path.exists(PRODUCT_SALES_REPORT_PATH):
        print(f"Product_Sales_Report.xlsx not found: {PRODUCT_SALES_REPORT_PATH}")
        return product_totals_from_report, product_totals_by_name

    wb_prod = openpyxl.load_workbook(PRODUCT_SALES_REPORT_PATH, data_only=True)
    if "Product Sales" not in wb_prod.sheetnames:
        wb_prod.close()
        print("Sheet 'Product Sales' not found.")
        return product_totals_from_report, product_totals_by_name

    ws_prod = wb_prod["Product Sales"]
    grand_total_row = None
    summary_header_row = None

    for row in range(2, ws_prod.max_row + 1):
        a1 = ws_prod.cell(row, 1).value
        if a1 is None:
            continue
        name = str(a1).strip()
        if name == "Grand Total (Mapped)":
            grand_total_row = row
            product_totals_from_report = {
                "New Business": ws_prod.cell(row, 2).value or 0,
                "Repeat Business": ws_prod.cell(row, 3).value or 0,
                "Reactivation": ws_prod.cell(row, 4).value or 0,
                "Refinance": ws_prod.cell(row, 5).value or 0,
                "Total": ws_prod.cell(row, 6).value or 0,
            }
            break

    for row in range(2, ws_prod.max_row + 1):
        a1 = ws_prod.cell(row, 1).value
        if a1 is not None and str(a1).strip() == "Product":
            summary_header_row = row
            break

    if summary_header_row is not None and grand_total_row is not None:
        for row in range(summary_header_row + 1, grand_total_row):
            name = ws_prod.cell(row, 1).value
            if name is None or not str(name).strip():
                continue
            name = str(name).strip()
            try:
                b = ws_prod.cell(row, 2).value or 0
                c = ws_prod.cell(row, 3).value or 0
                d = ws_prod.cell(row, 4).value or 0
                e = ws_prod.cell(row, 5).value or 0
                f = ws_prod.cell(row, 6).value or 0
                for v in (b, c, d, e, f):
                    if not isinstance(v, (int, float)):
                        float(v)
                product_totals_by_name[name] = {
                    "New Business": b,
                    "Repeat Business": c,
                    "Reactivation": d,
                    "Refinance": e,
                    "Total": f,
                }
            except (TypeError, ValueError):
                pass

    wb_prod.close()
    return product_totals_from_report, product_totals_by_name


def load_summary_block(title):
    """Parse a '<title>' block in Product_Sales_Report.xlsx (Zone/Cluster summary):
    returns {name: {New Business, Repeat Business, Reactivation, Refinance, Total}}."""
    out = {}
    if not os.path.exists(PRODUCT_SALES_REPORT_PATH):
        return out
    wb = openpyxl.load_workbook(PRODUCT_SALES_REPORT_PATH, data_only=True)
    if "Product Sales" not in wb.sheetnames:
        wb.close()
        return out
    ws = wb["Product Sales"]
    start = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is not None and str(v).strip() == title:
            start = r
            break
    if start is not None:
        for r in range(start + 2, ws.max_row + 1):  # +2 skips the header row
            name = ws.cell(r, 1).value
            if name is None or not str(name).strip():
                break
            nm = str(name).strip()
            if nm.startswith("Grand Total"):
                break
            out[nm] = {
                "New Business": ws.cell(r, 2).value or 0,
                "Repeat Business": ws.cell(r, 3).value or 0,
                "Reactivation": ws.cell(r, 4).value or 0,
                "Refinance": ws.cell(r, 5).value or 0,
                "Total": ws.cell(r, 6).value or 0,
            }
    wb.close()
    return out


def load_zone_cluster_names():
    """Distinct Zone and Cluster names from 'Zone and cluster.xlsx' (the master list
    of which Country-sheet rows are zones / clusters)."""
    zones, clusters = set(), set()
    if not os.path.exists(ZONE_CLUSTER_PATH):
        print(f"[WARN] Zone and cluster file not found: {ZONE_CLUSTER_PATH}")
        return zones, clusters
    wb = openpyxl.load_workbook(ZONE_CLUSTER_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdr = {str(ws.cell(1, c).value).strip().lower(): c
           for c in range(1, ws.max_column + 1) if ws.cell(1, c).value is not None}
    zi, ci = hdr.get("zone"), hdr.get("cluster")
    for r in range(2, ws.max_row + 1):
        if zi:
            z = ws.cell(r, zi).value
            if z and str(z).strip():
                zones.add(str(z).strip())
        if ci:
            c = ws.cell(r, ci).value
            if c and str(c).strip():
                clusters.add(str(c).strip())
    wb.close()
    return zones, clusters


def find_col(ws, name):
    """Return 1-based column index where header (row 1) matches name (case-insensitive)."""
    for col in range(1, MAX_COL + 1):
        v = ws.cell(HEADER_ROW, col).value
        if v and str(v).strip().lower() == name.lower():
            return col
    return None


def main():
    print("=" * 60)
    print("Update Management from Product Sales Report")
    print("=" * 60)

    product_totals_from_report, product_totals_by_name = load_product_totals()
    if not product_totals_from_report and not product_totals_by_name:
        print("No product totals loaded. Exiting.")
        return

    if product_totals_from_report:
        print(f"Loaded Grand Total (Country).")
    print(f"Loaded {len(product_totals_by_name)} product rows for branch alignment.")

    management_files = glob.glob(os.path.join(DESTINATION_FOLDER, "Management*.xlsx"))
    if not management_files:
        management_files = glob.glob(os.path.join(DESTINATION_FOLDER, "Management*.xlsm"))
    if not management_files:
        print(f"No Management file found in {DESTINATION_FOLDER}")
        return

    management_file = management_files[0]
    print(f"Management file: {os.path.basename(management_file)}")

    # Open with data_only=True so formula cells give numeric values for summing; styles preserved on save
    is_xlsm = management_file.lower().endswith(".xlsm")
    wb = openpyxl.load_workbook(management_file, data_only=True, keep_vba=is_xlsm)
    if "Country" not in wb.sheetnames:
        wb.close()
        print("'Country' sheet not found.")
        return

    ws = wb["Country"]
    branch_col = find_col(ws, "Branch")
    if branch_col is None:
        wb.close()
        print("Branch column not found.")
        return

    last_row = None
    for row in range(10000, HEADER_ROW, -1):
        if ws.cell(row, branch_col).value is not None:
            last_row = row
            break
    if last_row is None:
        last_row = HEADER_ROW + 100

    country_row = None
    for row in range(HEADER_ROW + 1, last_row + 1):
        v = ws.cell(row, branch_col).value
        if v and str(v).strip().lower() == "country":
            country_row = row
            break
    if country_row is None:
        wb.close()
        print("Country row not found.")
        return

    col_new_business = find_col(ws, "New Business")
    col_repeat_business = find_col(ws, "Repeat Business")
    col_reactivation = find_col(ws, "Reactivation")
    col_refinance = find_col(ws, "Refinance")
    col_disbursements = find_col(ws, "Disbursements This Month")
    col_number_of_loans = find_col(ws, "Number of loans")
    col_average_loan_size = find_col(ws, "Average loan size")
    product_cols_ok = all((
        col_new_business, col_repeat_business, col_reactivation,
        col_refinance, col_disbursements
    ))

    # 1) Update branch rows that match a product
    if product_totals_by_name and product_cols_ok:
        updated = 0
        for row in range(HEADER_ROW + 1, last_row + 1):
            if row == country_row:
                continue
            branch_value = ws.cell(row, branch_col).value
            if branch_value is None:
                continue
            key = str(branch_value).strip()
            totals = product_totals_by_name.get(key)
            if totals is None:
                for pk, pv in product_totals_by_name.items():
                    if str(pk).strip().lower() == key.lower():
                        totals = pv
                        break
            if totals is None:
                continue
            ws.cell(row, col_new_business).value = _num(totals.get("New Business"))
            ws.cell(row, col_repeat_business).value = _num(totals.get("Repeat Business"))
            ws.cell(row, col_reactivation).value = _num(totals.get("Reactivation"))
            ws.cell(row, col_refinance).value = _num(totals.get("Refinance"))
            ws.cell(row, col_disbursements).value = _num(totals.get("Total"))
            updated += 1
        print(f"Updated {updated} branch row(s) from product report.")

    # 2) Country row: five columns from Grand Total
    if product_totals_from_report:
        ws.cell(country_row, col_new_business).value = _num(product_totals_from_report.get("New Business"))
        ws.cell(country_row, col_repeat_business).value = _num(product_totals_from_report.get("Repeat Business"))
        ws.cell(country_row, col_reactivation).value = _num(product_totals_from_report.get("Reactivation"))
        ws.cell(country_row, col_refinance).value = _num(product_totals_from_report.get("Refinance"))
        ws.cell(country_row, col_disbursements).value = _num(product_totals_from_report.get("Total"))
        print("Updated Country row: New Business, Repeat Business, Reactivation, Refinance, Disbursements This Month.")

    # 3) Country row: other columns = sum of rows START_SUM_ROW..END_SUM_ROW
    columns_to_process = {
        "Target": lambda s: s - 5,
        "Daily Target": lambda s: s,
        "Number of loans": lambda s: s,
        "In arrears": lambda s: s,
        "Value in arrears": lambda s: s,
        "Portfolio": lambda s: s,
        "Active Reps": lambda s: s,
    }
    for col_name, calculation_func in columns_to_process.items():
        col_num = find_col(ws, col_name)
        if col_num is None:
            continue
        total_sum = 0
        for row in range(START_SUM_ROW, END_SUM_ROW + 1):
            cell_value = ws.cell(row, col_num).value
            if cell_value is not None:
                try:
                    if isinstance(cell_value, (int, float)):
                        total_sum += cell_value
                    elif isinstance(cell_value, str):
                        cleaned = cell_value.replace(",", "").strip()
                        if cleaned:
                            total_sum += float(cleaned)
                except (ValueError, TypeError):
                    pass
        ws.cell(country_row, col_num).value = calculation_func(total_sum)
    print("Updated Country row: Target, Daily Target, Number of loans, In arrears, Value in arrears, Portfolio, Active Reps (from sum).")

    # 4) Country Average loan size = Disbursements This Month / Number of loans
    if col_average_loan_size and col_disbursements and col_number_of_loans:
        disbursements_val = _num(ws.cell(country_row, col_disbursements).value)
        num_loans_val = _num(ws.cell(country_row, col_number_of_loans).value)
        avg_loan_size = (disbursements_val / num_loans_val) if num_loans_val else 0
        ws.cell(country_row, col_average_loan_size).value = avg_loan_size
        print(f"Updated Country row: Average loan size = {avg_loan_size}")

    # 5) Zone & Cluster rows: update from the Product report's Zone/Cluster summaries.
    #    A row is a zone/cluster if its Branch matches a name in 'Zone and cluster.xlsx'.
    #    Rows with no sales data are set to 0.
    if product_cols_ok:
        zones, clusters = load_zone_cluster_names()
        zone_totals = load_summary_block("Summary by Zone")
        cluster_totals = load_summary_block("Summary by Cluster")
        zones_ci = {z.lower() for z in zones}
        clusters_ci = {c.lower() for c in clusters}
        zone_totals_ci = {k.lower(): v for k, v in zone_totals.items()}
        cluster_totals_ci = {k.lower(): v for k, v in cluster_totals.items()}
        product_keys_ci = {str(k).lower() for k in product_totals_by_name}

        def _apply(row, totals):
            ws.cell(row, col_new_business).value = _num(totals.get("New Business")) if totals else 0
            ws.cell(row, col_repeat_business).value = _num(totals.get("Repeat Business")) if totals else 0
            ws.cell(row, col_reactivation).value = _num(totals.get("Reactivation")) if totals else 0
            ws.cell(row, col_refinance).value = _num(totals.get("Refinance")) if totals else 0
            ws.cell(row, col_disbursements).value = _num(totals.get("Total")) if totals else 0

        z_updated = c_updated = 0
        for row in range(HEADER_ROW + 1, last_row + 1):
            if row == country_row:
                continue
            bval = ws.cell(row, branch_col).value
            if bval is None or not str(bval).strip():
                continue
            kl = str(bval).strip().lower()
            if kl in product_keys_ci:
                continue  # already handled as a product row
            if kl in zones_ci:
                _apply(row, zone_totals_ci.get(kl))   # None -> zeros
                z_updated += 1
            elif kl in clusters_ci:
                _apply(row, cluster_totals_ci.get(kl))
                c_updated += 1
        print(f"Updated {z_updated} zone row(s) and {c_updated} cluster row(s) "
              f"(rows with no data set to 0).")

    wb.save(management_file)
    wb.close()
    print(f"\nSaved: {management_file}")


if __name__ == "__main__":
    main()
