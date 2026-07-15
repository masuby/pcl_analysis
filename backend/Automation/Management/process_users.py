"""
Process Users
=============
Updates Users.xlsx based on Loan.xlsx:
1. Users sheet: Adds missing Sales Reps from Loan file, with TEAM (Branch) and Target (from existing team)
2. product_mapping sheet: Adds missing Loan Names from Loan file, with Products from most similar existing Loan Name
3. APPENDS just the newly-added rows to the online Users Google Sheet
   (USERS_SHEET_LINK in .env): only users / loan names that aren't already
   online are added at the BOTTOM of their tab. The header and all existing
   rows are left untouched, so programs that read the sheet keep working.
   Flag:  --no-upload  skip step 3.
Uses OpenPyXL to preserve formatting (colors, filters).
"""

import os
import re
import sys
from difflib import SequenceMatcher

import openpyxl
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

ROW_FILES = os.environ.get("PCL_MANAGEMENT_ROW_FILES") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "Management", "ROW_FILES")
    if os.environ.get("PCL_AUTOMATION_ROOT")
    else os.path.join(os.path.dirname(os.path.abspath(__file__)), "ROW_FILES"))
USERS_FILE = os.path.join(ROW_FILES, "Users.xlsx")
LOAN_FILE = os.path.join(ROW_FILES, "Loan.xlsx")


def _normalize(val):
    """Normalize value for comparison: strip and handle None/NaN."""
    if val is None:
        return ""
    return str(val).strip()


def _name_to_perfect_format(val):
    """
    Convert a name to one consistent, perfect format for Display Name / Sales Rep.
    CRUCIAL for correct reporting: same name must match everywhere.
    1. Strip leading/trailing and collapse all runs of spaces/tabs to a single space.
    2. Title Case: first letter of each word capital, rest lowercase (e.g. "JOHN DOE" -> "John Doe").
    Returns the same string for comparison and for writing back to files.
    """
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    s = re.sub(r"\s+", " ", s)
    return s.title()


def _normalize_name_for_display(val):
    """Alias: same as _name_to_perfect_format for backward compatibility in this module."""
    return _name_to_perfect_format(val)


def _apply_perfect_names_to_column(ws, col_idx, header_name, first_data_row=2):
    """
    Rewrite every cell in the column to perfect format (strip, single space, Title Case).
    Returns a dict mapping original_value -> perfect_value for any cell that was changed.
    """
    changed = {}
    for row_idx in range(first_data_row, ws.max_row + 1):
        val = ws.cell(row_idx, col_idx).value
        perfect = _name_to_perfect_format(val)
        if not perfect:
            continue
        current = str(val).strip() if val is not None else ""
        if current != perfect:
            ws.cell(row_idx, col_idx).value = perfect
            changed[current or "(empty)"] = perfect
    return changed


def _normalize_display_name_column(ws, display_col):
    """Rewrite all Display Name cells to perfect format (Title Case, no extra spaces)."""
    changed = _apply_perfect_names_to_column(ws, display_col, "Display Name")
    if changed:
        print(f"   Normalized {len(changed)} Display Name(s) to consistent format (Title Case, single spaces)")


def _sync_display_name_across_users_file(users_wb, ws_users, display_col):
    """
    Use Display Name as source of truth across the entire Users.xlsx file.
    Any cell anywhere (any sheet, any row, any column) whose value matches (after normalization)
    one of the Display Names is replaced with that Display Name value. So the same person
    is always written the same way everywhere – no same-row-only; every occurrence is corrected.
    """
    # Build set of all Display Name values (already in perfect format)
    display_names_set = set()
    for row_idx in range(2, ws_users.max_row + 1):
        val = ws_users.cell(row_idx, display_col).value
        if val and str(val).strip():
            display_names_set.add(str(val).strip())

    # 1. Users sheet: every cell in every row (except Display Name column) – if value matches
    #    any Display Name (after normalization), replace with that Display Name
    users_sheet_replaced = 0
    for row_idx in range(1, ws_users.max_row + 1):
        for col_idx in range(1, ws_users.max_column + 1):
            if col_idx == display_col:
                continue
            cell_val = ws_users.cell(row_idx, col_idx).value
            if cell_val is None:
                continue
            perfect = _name_to_perfect_format(cell_val)
            if perfect and perfect in display_names_set and str(cell_val).strip() != perfect:
                ws_users.cell(row_idx, col_idx).value = perfect
                users_sheet_replaced += 1

    if users_sheet_replaced:
        print(f"   Replaced {users_sheet_replaced} cell(s) in Users sheet with Display Name value (all rows)")

    # 2. Other sheets: same rule – any cell that normalizes to a Display Name is replaced
    other_sheets_replaced = 0
    for sheet_name in users_wb.sheetnames:
        if sheet_name == "Users":
            continue
        ws = users_wb[sheet_name]
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell_val = ws.cell(row_idx, col_idx).value
                if cell_val is None:
                    continue
                perfect = _name_to_perfect_format(cell_val)
                if perfect and perfect in display_names_set and str(cell_val).strip() != perfect:
                    ws.cell(row_idx, col_idx).value = perfect
                    other_sheets_replaced += 1

    if other_sheets_replaced:
        print(f"   Replaced {other_sheets_replaced} cell(s) in other sheet(s) with Display Name format for perfect matching")


# Sales reps that must be assigned only to this TEAM; any other TEAM assignment is removed (duplicate rows deleted)
WASHA_NESTORY_ONLY_NAMES = (
    "Daniel John Kundy",
    "Happyness John Elias",
    "Dickson Vilon Moshi",
    "Zaituni Mushi",
    "Dickson Mbonea Eliya",
    "Eugenia Paul Simon",
    "Masaga Joseph Busanda",
    "Elia Sanare",
)
WASHA_NESTORY_TEAM = "Washa Nestory"


def _assign_washa_nestory_only_and_remove_duplicates(ws_users, display_col, team_col):
    """
    For the three named sales reps: (1) set TEAM to 'Washa Nestory'; (2) if they appear
    in multiple rows (under other TEAMs), keep only one row per person (TEAM = Washa Nestory)
    and delete the duplicate rows.
    """
    names_set = {_name_to_perfect_format(n) for n in WASHA_NESTORY_ONLY_NAMES}
    # 1. Set TEAM to Washa Nestory for every row with these Display Names
    rows_by_name = {}  # perfect_name -> list of row indices
    for row_idx in range(2, ws_users.max_row + 1):
        dn = _name_to_perfect_format(ws_users.cell(row_idx, display_col).value)
        if dn and dn in names_set:
            ws_users.cell(row_idx, team_col).value = WASHA_NESTORY_TEAM
            rows_by_name.setdefault(dn, []).append(row_idx)
    # 2. For each of the three names, if multiple rows exist, keep first row and delete the rest (bottom to top)
    rows_to_delete = []
    for name, row_list in rows_by_name.items():
        if len(row_list) > 1:
            # Keep the first row, delete the others
            rows_to_delete.extend(sorted(row_list[1:], reverse=True))
    for row_idx in rows_to_delete:
        ws_users.delete_rows(row_idx, 1)
    if rows_by_name:
        total_assigned = sum(len(rows) for rows in rows_by_name.values())
        if rows_to_delete:
            print(f"   Assigned {total_assigned} row(s) for {list(names_set)} to TEAM '{WASHA_NESTORY_TEAM}' and removed {len(rows_to_delete)} duplicate row(s)")
        else:
            print(f"   Assigned {total_assigned} row(s) for {list(names_set)} to TEAM '{WASHA_NESTORY_TEAM}' only")


def _find_column_index(ws, header_row, target_names):
    """Find column index by header name (case-insensitive partial match)."""
    headers = [cell.value for cell in ws[header_row]]
    target_lower = [t.lower() for t in target_names]
    for col_idx, h in enumerate(headers, 1):
        if h is None:
            continue
        h_lower = str(h).strip().lower()
        for t in target_lower:
            if t in h_lower or h_lower in t:
                return col_idx
    return None


def _get_sheet_with_columns(wb, required_cols_map, preferred_sheet=None):
    """
    Find a sheet that has the required columns.
    required_cols_map: dict of {friendly_name: [possible_header_names]}
    preferred_sheet: if given, try this sheet name first (e.g. "Loan Accounts")
    Returns (ws, dict of {friendly_name: col_idx}) or (None, None)
    """
    sheet_order = list(wb.sheetnames)
    if preferred_sheet and preferred_sheet in sheet_order:
        sheet_order = [preferred_sheet] + [s for s in sheet_order if s != preferred_sheet]

    for sheet_name in sheet_order:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue
        found = {}
        headers = [cell.value for cell in ws[1]]
        headers_lower = [str(h).strip().lower() if h else "" for h in headers]
        for friendly, possible in required_cols_map.items():
            idx = None
            for p in possible:
                p_lower = p.lower()
                for i, h in enumerate(headers_lower):
                    if h and (p_lower in h or h == p_lower or h.startswith(p_lower) or p_lower in h):
                        idx = i + 1
                        break
                if idx:
                    break
            if not idx:
                break
            found[friendly] = idx
        if len(found) == len(required_cols_map):
            return ws, found
    return None, None


def _similarity(a, b):
    """Return similarity ratio between two strings (0 to 1)."""
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def _find_most_similar(target, candidates):
    """Find the candidate most similar to target. Returns (best_match, products_value) or (None, None)."""
    if not candidates:
        return None, None
    best = max(candidates, key=lambda x: _similarity(target, x[0]))
    return best[0], best[1]


# Minimum similarity (0-1) to treat a Loan Sales Rep as same person as a Users Display Name (autocorrect)
DISPLAY_NAME_SIMILARITY_THRESHOLD = 0.85


def _find_most_similar_display_name(loan_name, users_display_names_set):
    """
    Find the Display Name in Users that is most similar to loan_name.
    Returns (best_match_from_users, ratio) or (None, 0.0).
    Used as fallback to autocorrect Loan file to match Users when names are slightly different.
    """
    if not loan_name or not users_display_names_set:
        return None, 0.0
    best_name = None
    best_ratio = 0.0
    for uname in users_display_names_set:
        r = _similarity(loan_name, uname)
        if r > best_ratio:
            best_ratio = r
            best_name = uname
    return best_name, best_ratio


# ============================================================================
# STEP 1: PROCESS USERS SHEET
# ============================================================================

def process_users_sheet(users_wb, loan_wb):
    """
    - Find Sales Reps in Loan file but not in Users file
    - Add them at bottom of Users: Display Name, TEAM (= Branch from Loan), Target (= lookup from same TEAM in Users)
    """
    print("\n" + "=" * 80)
    print("STEP 1: Processing Users sheet")
    print("=" * 80)
    print("   Applying consistent name format (Title Case, no extra spaces) in both files first...")

    # Find Users sheet
    if "Users" not in users_wb.sheetnames:
        print("⚠️ 'Users' sheet not found in Users.xlsx")
        return

    ws_users = users_wb["Users"]
    header_row = 1

    # Column mappings for Users sheet
    team_col = _find_column_index(ws_users, header_row, ["TEAM", "Team"])
    display_col = _find_column_index(ws_users, header_row, ["Display Name", "DisplayName"])
    target_col = _find_column_index(ws_users, header_row, ["Target"])

    if not all([team_col, display_col, target_col]):
        print("⚠️ Users sheet: Could not find TEAM, Display Name, or Target columns")
        print(f"   TEAM={team_col}, Display Name={display_col}, Target={target_col}")
        return

    # Build sets from Users
    users_display_names = set()
    team_to_target = {}  # TEAM -> Target (use first occurrence or most common)

    for row_idx in range(2, ws_users.max_row + 1):
        dn_raw = ws_users.cell(row_idx, display_col).value
        dn = _normalize_name_for_display(dn_raw)
        team = _normalize(ws_users.cell(row_idx, team_col).value)
        tgt = ws_users.cell(row_idx, target_col).value
        if dn:
            users_display_names.add(dn)
        if team and tgt is not None:
            if team not in team_to_target:
                team_to_target[team] = tgt

    # Normalize existing Display Name cells to perfect format (Title Case, single spaces)
    _normalize_display_name_column(ws_users, display_col)

    # Sync: replace any matching name elsewhere in Users sheet (same row) and in other sheets with Display Name value
    _sync_display_name_across_users_file(users_wb, ws_users, display_col)

    # Rebuild users_display_names from Display Name column after normalization
    users_display_names = set()
    for row_idx in range(2, ws_users.max_row + 1):
        dn = _name_to_perfect_format(ws_users.cell(row_idx, display_col).value)
        if dn:
            users_display_names.add(dn)

    print(f"   Found {len(users_display_names)} existing Display Names in Users")
    print(f"   Found {len(team_to_target)} TEAM->Target mappings")

    # Assign Daniel John Kundy, Happyness John Elias, Dickson Vilon Moshi only to Washa Nestory; remove duplicate rows
    _assign_washa_nestory_only_and_remove_duplicates(ws_users, display_col, team_col)

    # Rebuild users_display_names after possible row deletions
    users_display_names = set()
    for row_idx in range(2, ws_users.max_row + 1):
        dn = _name_to_perfect_format(ws_users.cell(row_idx, display_col).value)
        if dn:
            users_display_names.add(dn)

    # Find Loan sheet - use "Loan Accounts" sheet explicitly
    loan_required = {
        "branch": ["Branch", "branch"],
        "sales_reps": ["Sales Reps (Client)", "Sales Reps", "Sales Officer", "SalesReps"],
    }
    ws_loan, loan_cols = _get_sheet_with_columns(loan_wb, loan_required, preferred_sheet="Loan Accounts")
    if not ws_loan or not loan_cols:
        print("⚠️ Loan file: Could not find Branch and 'Sales Reps (Client)' columns")
        if "Loan Accounts" in loan_wb.sheetnames:
            headers = [cell.value for cell in loan_wb["Loan Accounts"][1]]
            print(f"   Loan Accounts sheet headers: {headers[:15]}...")
        return

    branch_col = loan_cols["branch"]
    sales_col = loan_cols["sales_reps"]

    # Apply perfect format to entire Sales Reps column in Loan (same as Display Name: Title Case, single spaces)
    loan_changed = _apply_perfect_names_to_column(ws_loan, sales_col, "Sales Reps (Client)")
    if loan_changed:
        print(f"   Normalized {len(loan_changed)} Sales Rep value(s) in Loan file to consistent format (Title Case, single spaces)")

    # Collect unique (Sales Rep, Branch) from Loan - canonical name format (perfect format)
    loan_reps = set()
    rep_to_branch = {}

    for row_idx in range(2, ws_loan.max_row + 1):
        rep_raw = ws_loan.cell(row_idx, sales_col).value
        rep = _normalize_name_for_display(rep_raw)
        branch = _normalize(ws_loan.cell(row_idx, branch_col).value)
        if rep:
            loan_reps.add(rep)
            if rep not in rep_to_branch:
                rep_to_branch[rep] = branch

    # Fallback: for each Loan rep not in Users, try to find a similar Display Name in Users (autocorrect)
    # loan_rep -> name to write (either self or the matching Users name)
    rep_to_canonical = {}
    autocorrected = []

    for rep in loan_reps:
        if not rep:
            continue
        if rep in users_display_names:
            rep_to_canonical[rep] = rep
            continue
        similar_name, ratio = _find_most_similar_display_name(rep, users_display_names)
        if similar_name and ratio >= DISPLAY_NAME_SIMILARITY_THRESHOLD:
            rep_to_canonical[rep] = similar_name
            autocorrected.append((rep, similar_name, ratio))
        else:
            rep_to_canonical[rep] = rep

    if autocorrected:
        print(f"\n   Autocorrect (similarity >= {DISPLAY_NAME_SIMILARITY_THRESHOLD}): Loan Sales Rep -> Users Display Name:")
        for loan_rep, users_name, ratio in sorted(autocorrected, key=lambda x: x[0]):
            print(f"      '{loan_rep}' -> '{users_name}' (ratio={ratio:.2f})")

    # Write back to Loan file: canonical name or autocorrected Users name so Loan and Users match
    for row_idx in range(2, ws_loan.max_row + 1):
        rep_raw = ws_loan.cell(row_idx, sales_col).value
        rep = _normalize_name_for_display(rep_raw)
        if rep and rep in rep_to_canonical:
            ws_loan.cell(row_idx, sales_col).value = rep_to_canonical[rep]

    # Missing = in Loan but not in Users and no similar match (will add to Users)
    missing_reps = [r for r in loan_reps if r and r not in users_display_names and rep_to_canonical.get(r) == r]

    if not missing_reps:
        print("   No missing Sales Reps found. All Loan reps exist in Users.")
        report_users_without_target(ws_users, team_col, display_col, target_col)
        return

    print(f"\n   Sales Reps in Loan file but NOT in Users ({len(missing_reps)}):")
    for r in sorted(missing_reps):
        print(f"      - {r}")

    # Add missing reps at bottom
    next_row = ws_users.max_row + 1
    added = 0

    for rep in sorted(missing_reps):
        branch = rep_to_branch.get(rep, "")
        target_val = team_to_target.get(branch) if branch else None

        ws_users.cell(next_row, display_col).value = rep
        ws_users.cell(next_row, team_col).value = branch
        ws_users.cell(next_row, target_col).value = target_val

        # Copy formatting from previous row if exists
        prev_row = next_row - 1
        if prev_row >= 2:
            for col in [team_col, display_col, target_col]:
                src = ws_users.cell(prev_row, col)
                dst = ws_users.cell(next_row, col)
                if src.has_style:
                    dst.font = src.font.copy()
                    dst.border = src.border.copy()
                    dst.fill = src.fill.copy()
                    dst.number_format = src.number_format
                    dst.alignment = src.alignment.copy()

        next_row += 1
        added += 1

    print(f"\n   [OK] Added {added} new rows to Users sheet")
    report_users_without_target(ws_users, team_col, display_col, target_col)


# ============================================================================
# STEP 2: PROCESS product_mapping SHEET
# ============================================================================

def process_product_mapping_sheet(users_wb, loan_wb):
    """
    - Find Loan Names in Loan file but not in product_mapping
    - Add them at bottom with Products from most similar existing Loan Name
    """
    print("\n" + "=" * 80)
    print("STEP 2: Processing product_mapping sheet")
    print("=" * 80)

    if "product_mapping" not in users_wb.sheetnames:
        print("⚠️ 'product_mapping' sheet not found in Users.xlsx")
        return

    ws_pm = users_wb["product_mapping"]
    header_row = 1

    loan_name_col = _find_column_index(ws_pm, header_row, ["Loan Name", "LoanName"])
    products_col = _find_column_index(ws_pm, header_row, ["Products", "Product"])

    if not loan_name_col or not products_col:
        print("⚠️ product_mapping: Could not find Loan Name or Products columns")
        return

    # Existing Loan Name -> Products
    existing_loans = []
    for row_idx in range(2, ws_pm.max_row + 1):
        ln = _normalize(ws_pm.cell(row_idx, loan_name_col).value)
        prod = ws_pm.cell(row_idx, products_col).value
        if ln:
            existing_loans.append((ln, prod))

    existing_loan_names = {ln for ln, _ in existing_loans}
    print(f"   Found {len(existing_loan_names)} existing Loan Names in product_mapping")

    # Loan Names from Loan file - use "Loan Accounts" sheet explicitly
    loan_required = {"loan_name": ["Loan Name", "LoanName"]}
    ws_loan, loan_cols = _get_sheet_with_columns(loan_wb, loan_required, preferred_sheet="Loan Accounts")
    if not ws_loan or not loan_cols:
        print("⚠️ Loan file: Could not find Loan Name column")
        if "Loan Accounts" in loan_wb.sheetnames:
            headers = [cell.value for cell in loan_wb["Loan Accounts"][1]]
            print(f"   Loan Accounts sheet headers: {headers[:15]}...")
        return

    ln_col = loan_cols["loan_name"]
    loan_file_names = set()
    for row_idx in range(2, ws_loan.max_row + 1):
        ln = _normalize(ws_loan.cell(row_idx, ln_col).value)
        if ln:
            loan_file_names.add(ln)

    # Loan Names in Loan file but NOT in product_mapping
    missing_loans = [ln for ln in loan_file_names if ln and ln not in existing_loan_names]

    if not missing_loans:
        print("   No missing Loan Names found. All Loan file names exist in product_mapping.")
        remove_product_mapping_duplicates(ws_pm, loan_name_col)
        return

    print(f"\n   Loan Names in Loan file but NOT in product_mapping ({len(missing_loans)}):")
    for ln in sorted(missing_loans):
        print(f"      - {ln}")

    # Add missing with Products from most similar existing
    next_row = ws_pm.max_row + 1
    added = 0

    for loan_name in sorted(missing_loans):
        _, products_val = _find_most_similar(loan_name, existing_loans)
        if products_val is None:
            products_val = ""

        ws_pm.cell(next_row, loan_name_col).value = loan_name
        ws_pm.cell(next_row, products_col).value = products_val

        # Copy formatting from previous row
        prev_row = next_row - 1
        if prev_row >= 2:
            for col in [loan_name_col, products_col]:
                src = ws_pm.cell(prev_row, col)
                dst = ws_pm.cell(next_row, col)
                if src.has_style:
                    dst.font = src.font.copy()
                    dst.border = src.border.copy()
                    dst.fill = src.fill.copy()
                    dst.number_format = src.number_format
                    dst.alignment = src.alignment.copy()

        next_row += 1
        added += 1
        existing_loans.append((loan_name, products_val))

    print(f"\n   [OK] Added {added} new rows to product_mapping sheet")

    # Remove duplicate rows (by Loan Name - keep first occurrence)
    remove_product_mapping_duplicates(ws_pm, loan_name_col)


def report_users_without_target(ws_users, team_col, display_col, target_col):
    """Print rows in Users sheet that have no Target (to fill manually)."""
    rows_without_target = []
    for row_idx in range(2, ws_users.max_row + 1):
        tgt = ws_users.cell(row_idx, target_col).value
        is_empty = tgt is None or (isinstance(tgt, str) and not str(tgt).strip())
        if is_empty:
            dn = ws_users.cell(row_idx, display_col).value
            team = ws_users.cell(row_idx, team_col).value
            rows_without_target.append((row_idx, dn, team))

    if rows_without_target:
        print(f"\n   [ACTION] Users rows with NO Target - fill manually ({len(rows_without_target)}):")
        for row_idx, dn, team in rows_without_target:
            print(f"      Row {row_idx}: Display Name='{dn}' | TEAM='{team}'")
    else:
        print("\n   [OK] All Users rows have a Target value.")


def remove_product_mapping_duplicates(ws_pm, loan_name_col):
    """Remove duplicate rows in product_mapping (by Loan Name). Keeps first occurrence."""
    seen_loan_names = set()
    rows_to_delete = []  # row indices (1-based) to delete

    for row_idx in range(2, ws_pm.max_row + 1):
        ln = _normalize(ws_pm.cell(row_idx, loan_name_col).value)
        if ln:
            if ln in seen_loan_names:
                rows_to_delete.append(row_idx)
            else:
                seen_loan_names.add(ln)

    if rows_to_delete:
        # Delete from bottom to top so row indices remain valid
        for row_idx in sorted(rows_to_delete, reverse=True):
            ws_pm.delete_rows(row_idx, 1)
        print(f"   [OK] Removed {len(rows_to_delete)} duplicate row(s) from product_mapping")
    else:
        print("   [OK] No duplicates found in product_mapping")


# ============================================================================
# STEP 3: REPLACE THE GOOGLE SHEET WITH THE PROCESSED FILE
# ============================================================================

ENV_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")


def _read_env(path):
    """Tiny .env reader: KEY=VALUE (quotes and spaces around '=' tolerated)."""
    out = {}
    if not os.path.exists(path):
        return out
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, val = line.partition("=")
            out[key.strip()] = val.strip().strip("'\"")
    return out


def _sheet_id_from_link(link):
    m = re.search(r"/d/([a-zA-Z0-9_-]+)", link or "")
    return m.group(1) if m else ((link or "").strip() or None)


def _worksheet_values(ws):
    """openpyxl worksheet -> list of rows for Google Sheets (None -> '')."""
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([
            "" if v is None else (v if isinstance(v, (int, float)) else str(v))
            for v in row
        ])
    while rows and all(v == "" for v in rows[-1]):  # drop trailing empty rows
        rows.pop()
    return rows


# Sheets this script adds rows to, and the column that identifies a row (so we
# can tell which rows are genuinely new and must be appended).
SYNC_KEYS = {"Users": "Display Name", "product_mapping": "Loan Name"}


def _key_col(header, key_name):
    """Index of the key column in a header row (case-insensitive); else col 0."""
    for i, h in enumerate(header):
        if key_name and str(h).strip().lower() == str(key_name).strip().lower():
            return i
    return 0


def append_new_rows_to_google_sheet(users_wb, sheets_to_sync=None):
    """
    APPEND-ONLY sync. Only rows that are NOT already online (new users / new loan
    names, identified by their key column) are appended at the BOTTOM of each tab.
    The header row and every existing row are left completely untouched, and each
    new row is aligned to the ONLINE column order — so downstream programs that
    read the sheet never break.
    Reads USERS_SHEET_LINK / GOOGLE_APPLICATION_CREDENTIALS from Management/.env.
    Returns True on success.
    """
    print("\n" + "=" * 80)
    print("STEP 3: Appending new rows to the Google Sheet (header + existing rows untouched)")
    print("=" * 80)

    env = _read_env(ENV_FILE)
    link = os.getenv("USERS_SHEET_LINK") or env.get("USERS_SHEET_LINK")
    creds_path = (os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
                  or env.get("GOOGLE_APPLICATION_CREDENTIALS"))
    sheet_id = _sheet_id_from_link(link)

    if not sheet_id:
        print("⚠️ USERS_SHEET_LINK not set in .env - Google Sheet NOT updated")
        return False
    if not creds_path or not os.path.exists(creds_path):
        print(f"⚠️ Google credentials file not found ({creds_path}) - Google Sheet NOT updated")
        return False

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as exc:
        print(f"⚠️ gspread/google-auth not installed ({exc}) - Google Sheet NOT updated")
        return False

    try:
        creds = Credentials.from_service_account_file(
            creds_path, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        sh = gspread.authorize(creds).open_by_key(sheet_id)
    except Exception as exc:
        print(f"❌ Could not open the Google Sheet: {exc}")
        sa = env.get("USERS_SHEET_SERVICE_ACCOUNT_EMAIL", "the service account")
        print(f"   Ensure the sheet is shared as Editor with: {sa}")
        return False

    # By default only sync the sheets this script actually adds rows to.
    if sheets_to_sync is None:
        sheets_to_sync = tuple(t for t in SYNC_KEYS if t in users_wb.sheetnames)

    online = {w.title: w for w in sh.worksheets()}
    all_ok, total = True, 0
    for title in sheets_to_sync:
        if title not in users_wb.sheetnames:
            print(f"   ⚠️ '{title}' not in the processed workbook - skipped")
            continue
        local_vals = _worksheet_values(users_wb[title])
        if len(local_vals) < 2:
            print(f"   '{title}': no data rows - skipped")
            continue
        local_header = local_vals[0]
        key_name = SYNC_KEYS.get(title, local_header[0] if local_header else "")

        try:
            gws = online.get(title)
            # Tab missing or empty online -> write it once (header + all rows).
            if gws is None:
                gws = sh.add_worksheet(title=title, rows=len(local_vals) + 20,
                                       cols=max(len(r) for r in local_vals))
                gws.update(local_vals, value_input_option="USER_ENTERED")
                print(f"   Created '{title}' with {len(local_vals) - 1} row(s)")
                total += len(local_vals) - 1
                continue
            online_vals = gws.get_all_values()
            if not online_vals:
                gws.update(local_vals, value_input_option="USER_ENTERED")
                print(f"   '{title}': was empty - wrote {len(local_vals) - 1} row(s)")
                total += len(local_vals) - 1
                continue

            online_header = online_vals[0]
            oi = _key_col(online_header, key_name)
            li = _key_col(local_header, key_name)
            existing = {str(r[oi]).strip().lower()
                        for r in online_vals[1:] if oi < len(r) and str(r[oi]).strip()}

            # Align each new local row to the ONLINE column order (by header name).
            local_pos = {str(h).strip().lower(): i for i, h in enumerate(local_header)}

            def align(local_row):
                out = []
                for h in online_header:
                    i = local_pos.get(str(h).strip().lower())
                    out.append(local_row[i] if (i is not None and i < len(local_row)) else "")
                return out

            new_rows = []
            for r in local_vals[1:]:
                k = str(r[li]).strip().lower() if li < len(r) else ""
                if k and k not in existing:
                    new_rows.append(align(r))
                    existing.add(k)          # avoid duplicating within this same batch

            if new_rows:
                gws.append_rows(new_rows, value_input_option="USER_ENTERED")
                print(f"   [OK] '{title}': appended {len(new_rows)} new row(s) at the bottom")
                total += len(new_rows)
            else:
                print(f"   [OK] '{title}': no new rows to append (already up to date)")
        except Exception as exc:
            all_ok = False
            print(f"   ❌ '{title}': append failed: {exc}")

    print(f"\n   {'✅' if all_ok else '⚠️'} Appended {total} new row(s) total. "
          f"Header and existing rows were left untouched.")
    return all_ok


# ============================================================================
# MAIN
# ============================================================================

def main():
    print("=" * 80)
    print("PROCESS USERS - Update Users.xlsx from Loan.xlsx")
    print("=" * 80)

    if not os.path.exists(USERS_FILE):
        print(f"❌ Users file not found: {USERS_FILE}")
        print("   Note: Use Users.xlsx (not ~$Users.xlsx). Close Excel if the file is open.")
        return

    # Reject Excel lock files (~$Users.xlsx)
    if os.path.basename(USERS_FILE).startswith("~$"):
        print(f"❌ Cannot use Excel lock file: {USERS_FILE}")
        print("   Close Users.xlsx in Excel and run again.")
        return

    # Check file is a valid xlsx (min size; lock files are tiny)
    users_size = os.path.getsize(USERS_FILE)
    if users_size < 1024:
        print(f"❌ Users file appears invalid or corrupted (size: {users_size} bytes)")
        print("   Ensure Users.xlsx is a real workbook, not a lock file. Close Excel and try again.")
        return

    if not os.path.exists(LOAN_FILE):
        print(f"❌ Loan file not found: {LOAN_FILE}")
        return

    try:
        # Load with formatting preserved (data_only=False)
        print(f"\nLoading {os.path.basename(USERS_FILE)}...")
        users_wb = openpyxl.load_workbook(USERS_FILE, data_only=False, keep_links=False)

        print(f"Loading {os.path.basename(LOAN_FILE)}...")
        loan_wb = openpyxl.load_workbook(LOAN_FILE, data_only=False)

        process_users_sheet(users_wb, loan_wb)
        process_product_mapping_sheet(users_wb, loan_wb)

        users_wb.save(USERS_FILE)
        print(f"\n✅ Saved {USERS_FILE}")
        loan_wb.save(LOAN_FILE)
        print(f"✅ Saved {LOAN_FILE} (Sales Reps column in consistent format: Title Case, single spaces)")

        # STEP 3: append only the newly-added rows to the online Google Sheet
        # (header + existing rows left untouched; skip with --no-upload)
        if "--no-upload" in sys.argv:
            print("\n(--no-upload) Google Sheet update skipped")
        else:
            append_new_rows_to_google_sheet(users_wb)

    except KeyError as e:
        if "Content_Types" in str(e) or "archive" in str(e).lower():
            print(f"\n❌ Users.xlsx could not be read (invalid or corrupted file)")
            print("   • Close Users.xlsx in Excel if it is open")
            print("   • Ensure you are using Users.xlsx, not the lock file ~$Users.xlsx")
            print("   • Try opening Users.xlsx in Excel and saving it again")
        else:
            raise
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback

        traceback.print_exc()


if __name__ == "__main__":
    main()
