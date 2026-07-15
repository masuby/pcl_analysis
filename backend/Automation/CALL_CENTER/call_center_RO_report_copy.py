"""
Call Center daily report — CS / LBF / RO / ERR.

Per product:
  * One Excel with 6 sheets: Summary, All_Call_Data, Agent_Performance,
    Inbound_Summary, Outbound_Summary, Call_Notes_Summary.
    (All_Call_Data has the noisy columns removed: Call From, Call To, Ivr, Queue,
     AI Receptionist, Disconnected by, Call Legs, Ring Group.)
  * Charts drawn as PNG images (matplotlib) and embedded inline in the email.
  * A professional HTML email: greeting + KPI boxes (5 per row) + narrative
    explanation (Calls Summary / Outbound / Inbound) with the charts in their
    sections.
Saved as:  NEW_FILES/<PRODUCT>/<date>/...
After processing it asks "Send to all recipients? [yes/no]":
  yes -> real distribution lists + DB upload.
  no  -> sends every product email to ME only (testing); no upload.
"""

from __future__ import annotations

import os
import re
import smtplib
import sys
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, to_hex
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# =============================================================================
# CONFIG
# =============================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_PATH = os.path.join(SCRIPT_DIR, "ROW_FILES")
EXPORT_ROOT = os.path.join(SCRIPT_DIR, "NEW_FILES")
ENV_PATH = os.path.join(SCRIPT_DIR, ".env")
UPLOAD_SCRIPT = r"C:\Users\Daniel\Desktop\code\pcl\CRM\direct_upload_files_to_db.py"

PRODUCTS = ["RO"]   # RO-specific report (call it as a subprocess from the main report)
MY_EMAIL = "daniel@platinumcredit.co.tz"

# Call-center display name -> official RO Name in the Loan file (case-insensitive).
# Fixes spelling/surname differences so the agent's loan portfolio is matched.
RO_NAME_ALIASES = {
    "Happyness Mwakikuti": "Happiness Mwakikuti",
    "Likoko Zarafi": "Likoko Likoko",
    "Kelvin Michael": "Kelvin Mwakanjuki",
}

DROP_FROM_CDR = ["Call From", "Call To", "Ivr", "Queue", "AI Receptionist",
                 "Disconnected by", "Call Legs", "Ring Group",
                 "Call Leg ID", "Call Flow", "Campaign"]

EXCLUDED_AGENTS = ["Ikrah Ally", "David Kileo", "Aziza Mfanga", "Madina Mohamed",
                   "Jackson Swai", "Thomas Francis", "Conference Call"]

RECIPIENTS = {
    "CS": [
        "raphael@platinumcredit.co.tz", "dorice@platinumcredit.co.tz",
        "sigfrid@platinumcredit.co.tz", "murigi@platinumcredit.co.ke",
        "wayne@platinumcredit.co.ke", "yusuph@platinumcredit.co.tz",
        "allan@platinumcredit.co.tz", "fragrance@platinumcredit.co.tz",
        "vivian@platinumcredit.co.tz", "thomas@platinumcredit.co.tz",
        "mohamedi.omar.platinum@gmail.com", "kelvin.mwasala@platinumcredit.co.tz",
        "daniel@platinumcredit.co.tz", "kelvin.peter.platinum@gmail.com",
    ],
    "LBF": [
        "raphael@platinumcredit.co.tz", "allan@platinumcredit.co.tz",
        "dorice@platinumcredit.co.tz", "fragrance@platinumcredit.co.tz",
        "sigfrid@platinumcredit.co.tz", "murigi@platinumcredit.co.ke",
        "wayne@platinumcredit.co.ke", "yusuph@platinumcredit.co.tz",
        "thomas@platinumcredit.co.tz", "augustine@platinumcredit.co.tz",
        "irene.mmari@platinumcredit.co.tz", "daniel@platinumcredit.co.tz",
        "aziza.mfanga.platinum@gmail.com", "madina.mohamed.platinum@gmail.com",
        "barnabas.ngassa.platinum@gmail.com", "zaituni@platinumcredit.co.tz",
    ],
    "RO": [
        "raphael@platinumcredit.co.tz", "allan@platinumcredit.co.tz",
        "dorice@platinumcredit.co.tz", "fragrance@platinumcredit.co.tz",
        "sigfrid@platinumcredit.co.tz", "daniel@platinumcredit.co.tz",
        "relationshipofficers@platinumcredit.co.tz",
        "jonas@platinumcredit.co.tz"
    ],
    "ERR": ["daniel@platinumcredit.co.tz"],
}

# 8-colour spectrum, red -> violet.
SPECTRUM = ["#E11D48", "#F97316", "#F59E0B", "#EAB308",
            "#22C55E", "#06B6D4", "#3B82F6", "#7C3AED"]
INK, MUTED, HEAD, LINE = "#0F172A", "#475569", "#1E3A5F", "#E2E8F0"

# Continuous VIOLET -> RED colormap (largest bar = violet, smallest = red).
_VR_CMAP = LinearSegmentedColormap.from_list("violet_red", list(reversed(SPECTRUM)))


def gradient_colors(n):
    """n colours spanning violet -> red, recalculated to the count of data points."""
    if n <= 0:
        return []
    if n == 1:
        return [_VR_CMAP(0.0)]
    return [_VR_CMAP(i / (n - 1)) for i in range(n)]


# Whole-box colour per KPI label (semantic).
KPI_COLORS = {
    "Total Calls": "#1E3A5F", "Inbound": "#3B82F6", "Outbound": "#F97316",
    "Internal": "#64748B", "Successful": "#22C55E", "Unsuccessful": "#E11D48",
    "Success Rate": "#16A34A", "Unique Called": "#06B6D4", "Unique Callers": "#0EA5E9",
    "Agents": "#7C3AED", "Agents 50+ Succ": "#15803D", "Avg Out / Agent": "#EA580C",
    "Avg In / Agent": "#2563EB", "Inbound Calls": "#0891B2", "Inbound Called Back": "#D97706",
}


# =============================================================================
# LOAD / CLEAN
# =============================================================================
def _find(prefix, ext):
    for f in os.listdir(FOLDER_PATH):
        if f.lower().startswith(prefix) and f.lower().endswith(ext):
            return os.path.join(FOLDER_PATH, f)
    return None


def extract_report_date(time_series):
    if len(time_series) == 0 or pd.isna(time_series.iloc[0]):
        return (datetime.now() - pd.Timedelta(days=1)).strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(time_series.iloc[0]).split(" ")[0], "%m/%d/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return (datetime.now() - pd.Timedelta(days=1)).strftime("%Y-%m-%d")


def clean_name(name):
    if pd.isna(name):
        return name
    name = re.sub(r"^Voicemail\s+", "", str(name))
    name = re.sub(r"<.*?>", "", name).strip()
    return "Hadija Mohamed" if name == "Khadija Mohamed" else name


def determine_success(row):
    status = str(row["Status"]).lower()
    if status in ["no answer", "busy", "failed", "voicemail", "abandoned"]:
        return "Unsuccessful"
    if status == "answered":
        try:
            return "Successful" if int(row["Talk Duration"]) >= 5 else "Unsuccessful"
        except (TypeError, ValueError):
            return "Unsuccessful"
    return "Unsuccessful"


def clean_notes(row):
    note = str(row["Call Notes"])
    if note.strip() == "" or note.lower() == "nan":
        status = str(row["Status"]).lower()
        return {"no answer": "Not picking", "voicemail": "Not picking",
                "failed": "Failed Connection", "busy": "Busy", "abandoned": "Abandoned",
                "answered": "NO COMMENT WRITTEN"}.get(status, "UNKNOWN")
    return note[:len(note.lower().split("remark")[0])].strip()


def load_data():
    cdr = _find("pse-cdr", ".csv")
    if not cdr:
        raise FileNotFoundError(f"No PSE-Cdr*.csv in {FOLDER_PATH}")
    print(f"CDR file: {os.path.basename(cdr)}")
    df = pd.read_csv(cdr)
    rename = {}
    if "Status" not in df.columns and "Last Status" in df.columns:
        rename["Last Status"] = "Status"
    if "Talk Duration" not in df.columns and "Handling Duration" in df.columns:
        rename["Handling Duration"] = "Talk Duration"
    if rename:
        df = df.rename(columns=rename)
    report_date = extract_report_date(df["Time"])
    print(f"Report date: {report_date}")
    df["Call From"] = df["Call From"].apply(clean_name)
    df["Call To"] = df["Call To"].apply(clean_name)
    df["Successful ?"] = df.apply(determine_success, axis=1)
    df["Call Notes"] = df.apply(clean_notes, axis=1)
    return df, report_date


# =============================================================================
# LBF LOAN ENRICHMENT  (RO report only)
# =============================================================================
LOAN_DROP_COLS = ["REFERRAL CONVERSION", "Pay Off Amount", "Pay-Off Amount",
                  "Loan Name", "Account ID", "Mobile Financing Sales Agent"]
LOAN_PHONE_SRC = ["Mobile Phone (Client)", "Mobile Phone 2 (Client)",
                  "Updated Mobile Number (Client)"]
MAX_PHONE_COLS = 4


def normalize_tz(v):
    """Normalise any phone value to '255XXXXXXXXX' (12 digits); '' if invalid."""
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    d = re.sub(r"\D", "", s)
    if not d:
        return ""
    if d.startswith("255"):
        return d[:12] if len(d) >= 12 else d
    if len(d) == 10 and d.startswith("0"):
        return "255" + d[1:]
    if len(d) == 9 and d[0] in "67":
        return "255" + d
    if d.startswith("0"):
        return "255" + d.lstrip("0")
    return "255" + d if len(d) == 9 else d


def _client_phones(row):
    """All distinct normalised client phones from the loan phone columns (split on /)."""
    phones = []
    for col in LOAN_PHONE_SRC:
        if col in row and pd.notna(row[col]):
            for part in re.split(r"[\/,;]+", str(row[col])):
                n = normalize_tz(part)
                if len(n) == 12 and n not in phones:
                    phones.append(n)
    return phones


def process_lbf_loans():
    """Load + clean the Loan_Accounts file: drop noisy cols, add Zone/Product via
    Branch lookup, keep Product==LBF, split phones into Phone number 1..N (client)."""
    lf = _find("loan_accounts", ".xlsx")
    zf = _find("zone and cluster", ".xlsx")
    if not lf:
        print("  WARN: no Loan_Accounts*.xlsx in ROW_FILES — RO loan enrichment skipped.")
        return None
    ln = pd.read_excel(lf)
    # Detect the known export defect where date columns collapse to ~1970
    # (the export kept only the time-of-day and dropped the date — unrecoverable).
    for dc in ("Activation Date", "Last Payment Date", "Last Modified"):
        if dc in ln.columns:
            yrs = pd.to_datetime(ln[dc], errors="coerce").dt.year
            if len(yrs) and (yrs <= 1971).mean() > 0.5:
                print(f"  WARN: Loan export column '{dc}' is corrupt (all ~1970-01-01). The export "
                      f"dropped the date and kept only the time-of-day — re-download the Loan_Accounts "
                      f"file to get real dates. The report passes the values through unchanged.")
                break
    ln = ln.drop(columns=[c for c in LOAN_DROP_COLS if c in ln.columns])
    zb = {}
    if zf:
        z = pd.read_excel(zf)
        zb = {str(b).strip().lower(): (str(zz).strip(), str(p).strip())
              for b, zz, p in zip(z["Branch"], z["Zone"], z["Product"])}
    ln["Zone"] = ln["Branch"].apply(lambda b: zb.get(str(b).strip().lower(), ("", ""))[0])
    ln["Product"] = ln["Branch"].apply(lambda b: zb.get(str(b).strip().lower(), ("", ""))[1])
    ln = ln[ln["Product"].astype(str).str.upper() == "LBF"].reset_index(drop=True)
    phone_lists = ln.apply(_client_phones, axis=1)
    for i in range(MAX_PHONE_COLS):
        ln[f"Phone number {i + 1} (client)"] = phone_lists.apply(lambda L: L[i] if i < len(L) else "")
    ln["_phones"] = phone_lists
    print(f"  LBF loans loaded for RO enrichment: {len(ln)} rows (Product=LBF).")
    return ln


def build_phone_long(loans_df):
    """phone -> loan row index (first match per phone)."""
    if loans_df is None or loans_df.empty:
        return pd.DataFrame(columns=["_phone", "_loan_idx"])
    rows = [(p, idx) for idx, phones in loans_df["_phones"].items() for p in phones]
    return pd.DataFrame(rows, columns=["_phone", "_loan_idx"]).drop_duplicates("_phone", keep="first")


def _name_key(s):
    """Case/space-insensitive key to match an RO Name to a call-agent name."""
    return re.sub(r"\s+", " ", str(s).strip()).upper() if s is not None and str(s).strip().lower() != "nan" else ""


def ro_portfolio_metrics(loans_df):
    """Per RO Name: portfolio loans, outstanding balance, PAR>30 (balance), amount collected."""
    empty = pd.DataFrame(columns=["_key", "Portfolio Loans", "Portfolio Balance",
                                  "PAR>30", "Total Due"])
    if loans_df is None or loans_df.empty or "RO Name" not in loans_df.columns:
        return empty
    d = loans_df.copy()
    d["_key"] = d["RO Name"].apply(_name_key)
    bal = pd.to_numeric(d.get("Total Balance"), errors="coerce").fillna(0)
    dia = pd.to_numeric(d.get("Days In Arrears"), errors="coerce").fillna(0)
    due = pd.to_numeric(d.get("Total Due"), errors="coerce").fillna(0)
    d["_bal"], d["_par30"], d["_due"] = bal, bal.where(dia > 30, 0), due
    d = d[d["_key"].ne("")]
    if d.empty:
        return empty
    g = d.groupby("_key").agg(
        Portfolio_Loans=("_key", "count"),
        Portfolio_Balance=("_bal", "sum"),
        _par30sum=("_par30", "sum"),
        Total_Due=("_due", "sum")).reset_index()
    g["PAR>30"] = (g["_par30sum"] / g["Portfolio_Balance"]).where(g["Portfolio_Balance"] > 0, 0)
    return g.rename(columns={"Portfolio_Loans": "Portfolio Loans",
                             "Portfolio_Balance": "Portfolio Balance",
                             "Total_Due": "Total Due"})[
        ["_key", "Portfolio Loans", "Portfolio Balance", "PAR>30", "Total Due"]]


def enrich_ro_calls(pdf, phone_long, loans_df):
    """RO All_Call_Data: add normalised Client Phone (inbound caller / outbound callee)
    then attach the full matched loan row (prefixed 'Loan: ')."""
    cdr = pdf.copy()

    def client_phone(r):
        if r["Communication Type"] == "Inbound":
            return normalize_tz(r["Call From"])
        if r["Communication Type"] == "Outbound":
            return normalize_tz(r["Call To"])
        return ""

    cdr["Client Phone"] = cdr.apply(client_phone, axis=1)
    noisy = ["Ivr", "Queue", "AI Receptionist", "Disconnected by", "Call Legs",
             "Ring Group", "Call Leg ID", "Call Flow", "Campaign"]
    cdr = cdr.drop(columns=[c for c in noisy if c in cdr.columns])
    if phone_long is None or phone_long.empty or loans_df is None:
        return cdr
    cdr = cdr.merge(phone_long, left_on="Client Phone", right_on="_phone", how="left")
    loan_cols = [c for c in loans_df.columns if c != "_phones"]
    loan_view = loans_df[loan_cols].add_prefix("Loan: ")
    loan_view["_loan_idx"] = loans_df.index
    cdr = cdr.merge(loan_view, on="_loan_idx", how="left").drop(
        columns=["_phone", "_loan_idx"], errors="ignore")
    # trim redundant / empty enriched columns
    drop_enriched = [
        "Loan: Phone number 3 (client)", "Loan: Phone number 4 (client)",
        "Loan: Mobile Phone 2 (Client)", "Loan: Updated Mobile Number (Client)",
        "Loan: Principal Due", "Loan: Account Sub-State",
        "Loan: Completed Loan Cycles (Client)",
    ]
    cdr = cdr.drop(columns=[c for c in drop_enriched if c in cdr.columns], errors="ignore")
    return cdr


def unallocated_portfolio_row(loans_df):
    """Aggregate analysis of the portfolio with NO RO assigned (blank RO Name)."""
    if loans_df is None or loans_df.empty or "RO Name" not in loans_df.columns:
        return None
    d = loans_df.copy()
    una = d[d["RO Name"].apply(_name_key) == ""]
    if una.empty:
        return None
    bal = pd.to_numeric(una.get("Total Balance"), errors="coerce").fillna(0)
    dia = pd.to_numeric(una.get("Days In Arrears"), errors="coerce").fillna(0)
    due = pd.to_numeric(una.get("Total Due"), errors="coerce").fillna(0)
    tot_bal = float(bal.sum())
    par = (float(bal.where(dia > 30, 0).sum()) / tot_bal) if tot_bal > 0 else 0.0
    return {"Agent Name": "Unallocated portfolio", "outbound_calls": 0, "inbound_calls": 0,
            "Total Calls": 0, "Successful Calls": 0, "Success Rate": 0.0,
            "Portfolio Loans": int(len(una)), "Portfolio Balance": tot_bal,
            "PAR>30": par, "Total Due": float(due.sum())}


def split_ro_scorecards(ap, ro_metrics, master_ro_names, loan_name_by_key, loans_df=None):
    """Split RO call agents into reconciled (main scorecard) and missing.
    Missing = call agents with no loan portfolio + loan ROs not in the RO Master
    + the unallocated portfolio (loans with no RO assigned)."""
    ap = ap.copy()
    if "Portfolio Loans" not in ap.columns:
        ap["Portfolio Loans"] = 0
    has_port = pd.to_numeric(ap["Portfolio Loans"], errors="coerce").fillna(0) > 0
    main_df = ap[has_port].copy()
    miss_call = ap[~has_port].copy()
    master_keys = {_name_key(RO_NAME_ALIASES.get(n, n)) for n in (master_ro_names or [])}
    call_keys = {_name_key(RO_NAME_ALIASES.get(n, n)) for n in ap["Agent Name"]}
    extra = []
    if ro_metrics is not None and not ro_metrics.empty:
        for _, r in ro_metrics.iterrows():
            k = r["_key"]
            if k in master_keys or k in call_keys:
                continue
            extra.append({
                "Agent Name": loan_name_by_key.get(k, k),
                "outbound_calls": 0, "inbound_calls": 0, "Total Calls": 0,
                "Successful Calls": 0, "Success Rate": 0.0,
                "Portfolio Loans": r["Portfolio Loans"], "Portfolio Balance": r["Portfolio Balance"],
                "PAR>30": r["PAR>30"], "Total Due": r["Total Due"]})
    miss_loan = pd.DataFrame(extra)
    missing_df = pd.concat([miss_call, miss_loan], ignore_index=True) if len(miss_loan) else miss_call
    una = unallocated_portfolio_row(loans_df)
    if una is not None:
        missing_df = pd.concat([missing_df, pd.DataFrame([una])], ignore_index=True)
    return main_df.reset_index(drop=True), missing_df.reset_index(drop=True)


def ro_supervisor_status(loans_df):
    """Collection status grouped by RO Supervisor: #ROs, portfolio loans/balance, PAR>30, Total Due."""
    cols = ["RO Supervisor", "# ROs", "Portfolio Loans", "Portfolio Balance", "PAR>30", "Total Due"]
    if loans_df is None or loans_df.empty or "RO Supervisor" not in loans_df.columns:
        return pd.DataFrame(columns=cols)
    d = loans_df.copy()
    d["_sup"] = d["RO Supervisor"].apply(
        lambda s: re.sub(r"\s+", " ", str(s).strip()) if pd.notna(s) and str(s).strip().lower() != "nan" else "")
    bal = pd.to_numeric(d.get("Total Balance"), errors="coerce").fillna(0)
    dia = pd.to_numeric(d.get("Days In Arrears"), errors="coerce").fillna(0)
    due = pd.to_numeric(d.get("Total Due"), errors="coerce").fillna(0)
    d["_bal"], d["_par30"], d["_due"] = bal, bal.where(dia > 30, 0), due
    d = d[d["_sup"] != ""]
    if d.empty:
        return pd.DataFrame(columns=cols)
    g = d.groupby("_sup").agg(
        ros=("RO Name", lambda s: s.dropna().nunique()), loans=("_sup", "count"),
        pbal=("_bal", "sum"), par=("_par30", "sum"), due=("_due", "sum")).reset_index()
    g["PAR>30"] = (g["par"] / g["pbal"]).where(g["pbal"] > 0, 0)
    g = g.rename(columns={"_sup": "RO Supervisor", "ros": "# ROs", "loans": "Portfolio Loans",
                          "pbal": "Portfolio Balance", "due": "Total Due"})
    return g[cols].sort_values("PAR>30").reset_index(drop=True)


def all_call_summary(pdf, phone_long):
    """Summary of RO call data vs the loan file (matched / unmatched to a client)."""
    matched = set(phone_long["_phone"]) if phone_long is not None and not phone_long.empty else set()

    def cphone(r):
        if r["Communication Type"] == "Inbound":
            return normalize_tz(r["Call From"])
        if r["Communication Type"] == "Outbound":
            return normalize_tz(r["Call To"])
        return ""
    ph = pdf.apply(cphone, axis=1)
    is_client = ph.isin(matched) & (ph != "")
    total = len(pdf)
    inb = pdf["Communication Type"] == "Inbound"
    out = pdf["Communication Type"] == "Outbound"
    nm = int((~is_client).sum())
    rows = [
        ("Total RO calls", f"{total:,}"),
        ("Matched to an LBF loan client", f"{int(is_client.sum()):,} ({pct(int(is_client.sum()), total)})"),
        ("Not matched (no client in Loan file)", f"{nm:,} ({pct(nm, total)})"),
        ("Distinct loan clients contacted", f"{ph[is_client].nunique():,}"),
        ("Distinct unmatched numbers", f"{ph[(~is_client) & (ph != '')].nunique():,}"),
        ("Inbound matched / not matched", f"{int((inb & is_client).sum()):,} / {int((inb & ~is_client).sum()):,}"),
        ("Outbound matched / not matched", f"{int((out & is_client).sum()):,} / {int((out & ~is_client).sum()):,}"),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _build_summary_sheet(ws, summary_df, sup_df, call_sum_df):
    """Write the Summary sheet: daily call summary + RO Collection Status (by Supervisor)
    + All_Call_Data summary, each as its own styled block."""
    navy = PatternFill("solid", fgColor="1E3A5F")
    band = PatternFill("solid", fgColor="F1F5F9")
    thin = Side(style="thin", color="D6E2F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    state = {"r": 1}

    def block(title, df, money=(), integer=(), pctc=()):
        if df is None or df.empty:
            return
        r = state["r"]
        ncol = df.shape[1]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        c = ws.cell(r, 1, title)
        c.fill = PatternFill("solid", fgColor="7C3AED")
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        r += 1
        for j, name in enumerate(df.columns, 1):
            cell = ws.cell(r, j, name)
            cell.fill = navy
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        r += 1
        for _, row in df.iterrows():
            for j, name in enumerate(df.columns, 1):
                v = row[name]
                cell = ws.cell(r, j, None if (isinstance(v, float) and pd.isna(v)) else v)
                cell.border = border
                if r % 2 == 0:
                    cell.fill = band
                if name in money or name in integer:
                    cell.number_format = "#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif name in pctc:
                    cell.number_format = "0.00%"
                    cell.alignment = Alignment(horizontal="right")
            r += 1
        state["r"] = r + 2

    block("Daily Call Summary", summary_df)
    block("RO Collection Status (by Supervisor)", sup_df,
          money={"Portfolio Balance", "Total Due"}, integer={"# ROs", "Portfolio Loans"}, pctc={"PAR>30"})
    block("All_Call_Data Summary", call_sum_df)
    ws.column_dimensions["A"].width = 40
    for j in range(2, 8):
        ws.column_dimensions[get_column_letter(j)].width = 18
    ws.sheet_properties.tabColor = "1E3A5F"


def build_agent_products(df):
    inbound = df[(df["Communication Type"] == "Inbound") & (~df["Call To"].isin(EXCLUDED_AGENTS))]
    outbound = df[(df["Communication Type"] == "Outbound") & (~df["Call From"].isin(EXCLUDED_AGENTS))]
    in_df = inbound.groupby("Call To").agg(
        inbound_calls=("Call To", "count"),
        successful_inbound=("Successful ?", lambda x: (x == "Successful").sum()))
    out_df = outbound.groupby("Call From").agg(
        outbound_calls=("Call From", "count"),
        successful_outbound=("Successful ?", lambda x: (x == "Successful").sum()))
    agents = pd.merge(out_df, in_df, left_index=True, right_index=True, how="outer").fillna(0)
    agents.index.name = "Agent Name"
    agents = agents.reset_index()
    agents["Agent Name"] = agents["Agent Name"].replace("Khadija Mohamed", "Hadija Mohamed")
    agents = agents[~agents["Agent Name"].isin(EXCLUDED_AGENTS + ["Barnabas Ngassa"])]
    agents = agents.groupby("Agent Name", as_index=False).sum()
    agents["Total Calls"] = agents["outbound_calls"] + agents["inbound_calls"]
    agents["Successful Calls"] = agents["successful_outbound"] + agents["successful_inbound"]
    agents["Success Rate (%)"] = (agents["Successful Calls"] / 50 * 100).apply(lambda x: f"{x:.2f}%")
    agents = agents[["Agent Name", "outbound_calls", "inbound_calls", "Total Calls",
                     "Successful Calls", "Success Rate (%)"]].sort_values("Successful Calls", ascending=False)

    lookup = {}
    master = _find("master_cdr_call", ".xlsx")
    if master and os.path.exists(master):
        try:
            m = pd.read_excel(master, sheet_name="Agent_Performance")
            if {"Agent Name", "Product"} <= set(m.columns):
                for _, r in m.iterrows():
                    if pd.notna(r["Agent Name"]) and pd.notna(r["Product"]):
                        lookup[r["Agent Name"]] = str(r["Product"]).strip()
        except Exception as e:
            print(f"WARN master: {e}")
    agents["Product"] = agents["Agent Name"].map(lookup)
    agents["Product"] = agents["Product"].where(agents["Product"].isin(["CS", "LBF", "RO"]), "ERR")
    return agents


# =============================================================================
# METRICS
# =============================================================================
def compute_metrics(pdf, agents_df):
    m = {}
    total = len(pdf)
    comm = pdf["Communication Type"].value_counts()
    m["total"] = total
    m["inbound_n"] = int(comm.get("Inbound", 0))
    m["outbound_n"] = int(comm.get("Outbound", 0))
    m["internal_n"] = int(comm.get("Internal", 0))
    m["successful_n"] = int((pdf["Successful ?"] == "Successful").sum())
    m["unsuccessful_n"] = total - m["successful_n"]
    m["distinct_called"] = pdf[pdf["Communication Type"] == "Outbound"]["Call To"].nunique()
    m["distinct_calling"] = pdf[pdf["Communication Type"] == "Inbound"]["Call From"].nunique()
    m["total_agents"] = len(agents_df)
    m["agents_50"] = int((agents_df["Successful Calls"] >= 50).sum())

    out = pdf[(pdf["Communication Type"] == "Outbound") & (~pdf["Call From"].isin(EXCLUDED_AGENTS))]
    out_agents = out["Call From"].nunique()
    m["out_agents"] = out_agents
    m["avg_out"] = round(len(out) / out_agents, 2) if out_agents else 0
    out_succ = out[out["Successful ?"] == "Successful"]
    out_unsucc = out[out["Successful ?"] == "Unsuccessful"]
    m["out_succ_50"] = out_succ.groupby("Call From").filter(lambda x: len(x) >= 50)["Call From"].nunique()
    m["out_unsucc_50"] = out_unsucc.groupby("Call From").filter(lambda x: len(x) >= 50)["Call From"].nunique()

    inb = pdf[(pdf["Communication Type"] == "Inbound") & (~pdf["Call To"].isin(EXCLUDED_AGENTS))]
    inb_total = len(inb)
    m["inb_total"] = inb_total
    in_agents = inb["Call To"].nunique()
    m["avg_in"] = round(inb_total / in_agents, 2) if in_agents else 0
    inb_succ = inb[inb["Successful ?"] == "Successful"]
    inb_unsucc = inb[inb["Successful ?"] == "Unsuccessful"]
    m["inb_succ"] = len(inb_succ)
    m["inb_unsucc"] = len(inb_unsucc)
    out_numbers = out["Call To"].unique()
    called_back = inb_unsucc[inb_unsucc["Call From"].isin(out_numbers)]
    m["called_back"] = len(called_back)
    return m


def pct(n, d):
    return f"{(n/d if d else 0):.0%}"


def summary_rows(m):
    t = m["total"]
    return [
        ("Total calls", f"{t:,}"),
        ("Inbound", f"{m['inbound_n']:,} ({pct(m['inbound_n'], t)})"),
        ("Outbound", f"{m['outbound_n']:,} ({pct(m['outbound_n'], t)})"),
        ("Internal", f"{m['internal_n']:,} ({pct(m['internal_n'], t)})"),
        ("Successful", f"{m['successful_n']:,} ({pct(m['successful_n'], t)})"),
        ("Unsuccessful", f"{m['unsuccessful_n']:,} ({pct(m['unsuccessful_n'], t)})"),
        ("Unique numbers called (outbound)", f"{m['distinct_called']:,}"),
        ("Unique numbers calling in (inbound)", f"{m['distinct_calling']:,}"),
        ("Agents who made calls", f"{m['total_agents']:,}"),
        ("Agents with 50+ successful calls", f"{m['agents_50']} ({pct(m['agents_50'], m['total_agents'])})"),
        ("Avg outbound calls / agent", f"{m['avg_out']}"),
        ("Avg inbound calls / agent", f"{m['avg_in']}"),
    ]


# =============================================================================
# CHARTS (matplotlib PNGs, embedded inline in the email)
# =============================================================================
def _chart(items, title, path, kind="bar"):
    items = [(str(k), int(v)) for k, v in items if int(v) > 0]
    if not items:
        return None
    labels = [k for k, _ in items]
    values = [v for _, v in items]
    # violet (largest) -> red (smallest); items are already sorted descending
    colors = gradient_colors(len(items))
    plt.figure(figsize=(9, max(3.2, len(items) * 0.5)) if kind == "barh" else (9, 5))
    ax = plt.gca()
    ax.set_facecolor("#FFFFFF")
    if kind == "barh":
        bars = plt.barh(range(len(values)), values, color=colors, edgecolor="white")
        for b, v in zip(bars, values):
            plt.text(b.get_width(), b.get_y() + b.get_height() / 2, f" {v:,}",
                     va="center", fontsize=10, fontweight="bold", color=INK)
        plt.yticks(range(len(labels)), labels, fontsize=10)
        plt.gca().invert_yaxis()
    else:
        bars = plt.bar(range(len(values)), values, color=colors, edgecolor="white")
        for b, v in zip(bars, values):
            plt.text(b.get_x() + b.get_width() / 2, b.get_height(), f"{v:,}",
                     ha="center", va="bottom", fontsize=10, fontweight="bold", color=INK)
        plt.xticks(range(len(labels)), labels, rotation=30, ha="right", fontsize=10)
    plt.title(title, fontsize=13, fontweight="bold", color=HEAD, pad=12)
    plt.grid(axis="x" if kind == "barh" else "y", alpha=0.25, linestyle="--", linewidth=0.5)
    for s in ax.spines.values():
        s.set_visible(False)
    plt.tight_layout()
    plt.savefig(path, dpi=160, bbox_inches="tight", facecolor="white")
    plt.close()
    return path


def make_charts(pdf, agents_df, product, out_dir, report_date):
    paths = {}
    p = _chart(pdf["Communication Type"].value_counts().items(),
               f"Communication Type — {product} — {report_date}",
               os.path.join(out_dir, f"communication_type_{product}_{report_date}.png"), "bar")
    if p: paths["communication_type"] = p
    p = _chart(pdf["Successful ?"].value_counts().items(),
               f"Call Success — {product} — {report_date}",
               os.path.join(out_dir, f"success_distribution_{product}_{report_date}.png"), "bar")
    if p: paths["success_distribution"] = p
    p = _chart(pdf["Call Notes"].value_counts().head(10).items(),
               f"Call Notes Distribution — {product} — {report_date}",
               os.path.join(out_dir, f"call_notes_distribution_{product}_{report_date}.png"), "barh")
    if p: paths["call_notes_distribution"] = p
    top = agents_df.nlargest(10, "Successful Calls")
    p = _chart(zip(top["Agent Name"], top["Successful Calls"]),
               f"Top Agents by Successful Calls — {product} — {report_date}",
               os.path.join(out_dir, f"top_agents_{product}_{report_date}.png"), "bar")
    if p: paths["top_agents"] = p
    p = _chart(pdf["Status"].value_counts().items(),
               f"Call Status — {product} — {report_date}",
               os.path.join(out_dir, f"status_distribution_{product}_{report_date}.png"), "barh")
    if p: paths["status_distribution"] = p
    return paths


# =============================================================================
# EXCEL  (6 sheets)
# =============================================================================
def _style_sheet(ws, pct_cols=(), int_cols=(), tab="1E3A5F", first_width=30):
    """Header styling, banded rows, borders, number formats, freeze + tab colour."""
    ws.sheet_properties.tabColor = tab
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    navy = PatternFill(start_color=tab, end_color=tab, fill_type="solid")
    band = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    thin = Side(style="thin", color="D6E2F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = navy
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if r % 2 == 0:
                cell.fill = band
    for name, col in headers.items():
        fmt = "0.00%" if name in pct_cols else ("#,##0" if name in int_cols else None)
        if fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = fmt
                ws.cell(r, col).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = first_width if c == 1 else 16
    ws.row_dimensions[1].height = 22
    if ws.max_row >= 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"


def build_agent_perf(agents_df, is_ro, ro_metrics):
    """Agent performance table; for RO, merge the portfolio metrics (alias-matched)."""
    ap = agents_df.copy()
    ap["Success Rate"] = ap["Successful Calls"] / 50
    ap = ap[["Agent Name", "outbound_calls", "inbound_calls", "Total Calls",
             "Successful Calls", "Success Rate"]]
    if is_ro and ro_metrics is not None and not ro_metrics.empty:
        ap["_key"] = ap["Agent Name"].apply(lambda n: _name_key(RO_NAME_ALIASES.get(n, n)))
        ap = ap.merge(ro_metrics, on="_key", how="left").drop(columns="_key")
        for c in ("Portfolio Loans", "Portfolio Balance", "PAR>30", "Total Due"):
            if c in ap.columns:
                ap[c] = ap[c].fillna(0)
    return ap


def write_excel(path, product, report_date, m, pdf, agents_df,
                loans_df=None, phone_long=None, ro_metrics=None):
    is_ro = product == "RO" and loans_df is not None
    if is_ro:
        cdr_df = enrich_ro_calls(pdf, phone_long, loans_df)
    else:
        cdr_df = pdf.drop(columns=[c for c in DROP_FROM_CDR if c in pdf.columns])
    summary_df = pd.DataFrame(summary_rows(m), columns=["Metric", "Value"])
    ap = build_agent_perf(agents_df, is_ro, ro_metrics)

    out = pdf[(pdf["Communication Type"] == "Outbound") & (~pdf["Call From"].isin(EXCLUDED_AGENTS))]
    inb = pdf[(pdf["Communication Type"] == "Inbound") & (~pdf["Call To"].isin(EXCLUDED_AGENTS))]

    if len(inb):
        ins = inb.groupby("Call To").agg(
            Total_Inbound=("Successful ?", "count"),
            Successful_Inbound=("Successful ?", lambda x: (x == "Successful").sum()))
        ins["Inbound_Success_Rate"] = ins["Successful_Inbound"] / ins["Total_Inbound"]
        ins = ins.reset_index().rename(columns={"Call To": "Agent"})
        ins = ins.sort_values("Inbound_Success_Rate", ascending=False)
    else:
        ins = pd.DataFrame(columns=["Agent", "Total_Inbound", "Successful_Inbound", "Inbound_Success_Rate"])

    if len(out):
        outs = out.groupby("Call From").agg(
            Total_Outbound=("Successful ?", "count"),
            Successful_Outbound=("Successful ?", lambda x: (x == "Successful").sum()))
        outs["Outbound_Success_Rate"] = outs["Successful_Outbound"] / outs["Total_Outbound"]
        outs = outs.reset_index().rename(columns={"Call From": "Agent"})
        outs = outs.sort_values("Outbound_Success_Rate", ascending=False)
    else:
        outs = pd.DataFrame(columns=["Agent", "Total_Outbound", "Successful_Outbound", "Outbound_Success_Rate"])

    notes = pdf["Call Notes"].value_counts().reset_index()
    notes.columns = ["Call_Notes", "Count"]
    notes["Percentage"] = (notes["Count"] / m["total"]) if m["total"] else 0.0

    sup_df = ro_supervisor_status(loans_df) if is_ro else None
    call_sum_df = all_call_summary(pdf, phone_long) if is_ro else None

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        cdr_df.to_excel(writer, sheet_name="All_Call_Data", index=False)
        ap.to_excel(writer, sheet_name="Agent_Performance", index=False)
        ins.to_excel(writer, sheet_name="Inbound_Summary", index=False)
        outs.to_excel(writer, sheet_name="Outbound_Summary", index=False)
        notes.to_excel(writer, sheet_name="Call_Notes_Summary", index=False)

        # Summary sheet built manually (multiple tables) and placed first
        sum_ws = writer.book.create_sheet("Summary", 0)
        _build_summary_sheet(sum_ws, summary_df, sup_df, call_sum_df)

        sheets = writer.sheets
        _style_sheet(sheets["All_Call_Data"],
                     int_cols=("Total Duration", "Routing Duration", "Talk Duration"),
                     tab="3B82F6", first_width=22)
        _style_sheet(sheets["Agent_Performance"],
                     int_cols=("outbound_calls", "inbound_calls", "Total Calls", "Successful Calls",
                               "Portfolio Loans", "Portfolio Balance", "Total Due"),
                     pct_cols=("Success Rate", "PAR>30"), tab="7C3AED", first_width=26)
        _style_sheet(sheets["Inbound_Summary"],
                     int_cols=("Total_Inbound", "Successful_Inbound"),
                     pct_cols=("Inbound_Success_Rate",), tab="06B6D4", first_width=26)
        _style_sheet(sheets["Outbound_Summary"],
                     int_cols=("Total_Outbound", "Successful_Outbound"),
                     pct_cols=("Outbound_Success_Rate",), tab="F97316", first_width=26)
        _style_sheet(sheets["Call_Notes_Summary"],
                     int_cols=("Count",), pct_cols=("Percentage",), tab="22C55E", first_width=34)


# =============================================================================
# EMAIL  (greeting + KPI boxes + narrative with inline chart images)
# =============================================================================
def kpi_cards_html(m):
    t = m["total"]
    succ_rate = f"{(m['successful_n']/t if t else 0):.0%}"
    cb_rate = f"{(m['called_back']/m['inb_unsucc'] if m['inb_unsucc'] else 0):.0%}"
    cards = [
        (f"{t:,}", "Total Calls"), (f"{m['inbound_n']:,}", "Inbound"),
        (f"{m['outbound_n']:,}", "Outbound"), (f"{m['internal_n']:,}", "Internal"),
        (f"{m['successful_n']:,}", "Successful"), (f"{m['unsuccessful_n']:,}", "Unsuccessful"),
        (succ_rate, "Success Rate"), (f"{m['distinct_called']:,}", "Unique Called"),
        (f"{m['distinct_calling']:,}", "Unique Callers"), (f"{m['total_agents']:,}", "Agents"),
        (f"{m['agents_50']:,}", "Agents 50+ Succ"), (f"{m['avg_out']}", "Avg Out / Agent"),
        (f"{m['avg_in']}", "Avg In / Agent"), (f"{m['inb_total']:,}", "Inbound Calls"),
        (cb_rate, "Inbound Called Back"),
    ]
    html = ('<table role="presentation" width="100%" cellpadding="0" cellspacing="0" '
            'style="border-collapse:separate;border-spacing:7px;width:100%;">')
    for i in range(0, len(cards), 5):
        html += "<tr>"
        chunk = cards[i:i + 5]
        for value, label in chunk:
            color = KPI_COLORS.get(label, HEAD)
            html += (f'<td width="20%" style="background:{color};border-radius:8px;'
                     'padding:16px 8px;text-align:center;vertical-align:middle;">'
                     f'<div style="font-size:22px;font-weight:800;color:#FFFFFF;line-height:1;">{value}</div>'
                     f'<div style="font-size:10px;color:#FFFFFF;opacity:.92;text-transform:uppercase;'
                     f'letter-spacing:.4px;margin-top:6px;font-weight:700;">{label}</div></td>')
        for _ in range(5 - len(chunk)):
            html += '<td width="20%"></td>'
        html += "</tr>"
    return html + "</table>"


def _bullets(lines):
    return "".join(
        f'<p style="margin:6px 0;line-height:1.55;font-size:13.5px;color:{INK};padding-left:16px;'
        f'position:relative;">&bull;&nbsp; {ln}</p>' for ln in lines)


def _heading(text, color):
    return (f'<div style="text-align:center;margin:24px 0 10px;">'
            f'<span style="display:inline-block;font-size:16px;font-weight:700;color:{HEAD};'
            f'padding-bottom:6px;border-bottom:3px solid {color};">{text}</span></div>')


SEP = f'<hr style="border:none;border-top:1px solid {LINE};margin:20px 0;">'


def _chart_block(title, cid):
    return (f'<div style="text-align:center;margin:12px 0;">'
            f'<div style="font-size:13px;font-weight:700;color:{MUTED};margin-bottom:6px;">{title}</div>'
            f'<img src="cid:{cid}" width="100%" style="width:100%;max-width:100%;height:auto;'
            f'border:1px solid {LINE};border-radius:6px;"></div>')


def _text_on(hexcolor):
    """Readable text colour (black/white) for a given background hex."""
    h = hexcolor.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    return "#FFFFFF" if (0.299 * r + 0.587 * g + 0.114 * b) < 150 else "#0F172A"


def ro_agent_table_html(ap):
    """Static RO scorecard — every agent ranked best→worst by PAR>30. The WHOLE
    row is coloured along the violet→red spectrum (violet = best PAR, red = worst).
    Excludes Portfolio Loans."""
    if ap is None or "PAR>30" not in ap.columns:
        return ""
    d = ap.copy().sort_values("PAR>30", ascending=True, kind="stable").reset_index(drop=True)
    cols = [("Agent Name", "Agent"), ("outbound_calls", "Out"), ("inbound_calls", "In"),
            ("Total Calls", "Total"), ("Successful Calls", "Succ"),
            ("Success Rate", "Succ %"), ("Portfolio Balance", "Portfolio Bal."),
            ("PAR>30", "PAR&gt;30"), ("Total Due", "Total Due")]
    cols = [(k, lbl) for k, lbl in cols if k in d.columns]
    head = "".join(
        f'<th style="padding:8px 6px;text-align:{"left" if k=="Agent Name" else "right"};'
        f'background:{HEAD};color:#fff;font-size:11px;font-weight:700;">{lbl}</th>'
        for k, lbl in cols)
    row_bg = [to_hex(c) for c in gradient_colors(len(d))]  # violet(best) -> red(worst)
    rows = ""
    for i, r in d.iterrows():
        bg = row_bg[i] if i < len(row_bg) else "#FFFFFF"
        fg = _text_on(bg)
        cells = ""
        for k, _lbl in cols:
            v = r[k]
            align = "left" if k == "Agent Name" else "right"
            if k == "Success Rate":
                txt = f"{float(v or 0):.0%}"
            elif k == "PAR>30":
                txt = f"{float(v or 0):.1%}"
            elif k in ("Portfolio Balance", "Total Due"):
                txt = f"{float(v or 0):,.0f}"
            elif k == "Agent Name":
                txt = str(v)
            else:
                txt = f"{int(v or 0):,}"
            weight = "700" if k in ("Agent Name", "PAR>30") else "500"
            cells += (f'<td style="padding:7px 6px;text-align:{align};color:{fg};'
                      f'font-weight:{weight};border-bottom:1px solid rgba(255,255,255,.35);">{txt}</td>')
        rows += f'<tr style="background:{bg};">{cells}</tr>'
    return (f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" '
            f'style="border-collapse:collapse;width:100%;font-size:12px;'
            f'box-shadow:0 1px 3px rgba(0,0,0,.06);border-radius:6px;overflow:hidden;">'
            f'<tr>{head}</tr>{rows}</table>')


def load_master_ro_names():
    """All RO agent names from the master CDR file (Agent_Performance, Product==RO)."""
    master = _find("master_cdr_call", ".xlsx")
    names = []
    if master and os.path.exists(master):
        try:
            m = pd.read_excel(master, sheet_name="Agent_Performance")
            if {"Agent Name", "Product"} <= set(m.columns):
                names = [str(r["Agent Name"]).strip() for _, r in m.iterrows()
                         if str(r.get("Product", "")).strip().upper() == "RO"
                         and pd.notna(r["Agent Name"])]
        except Exception as e:
            print(f"WARN master RO list: {e}")
    return names


def missing_ro_dataframe(loans_df, master_ro_names):
    """ROs present in the Loan file but not in the RO Master, and vice versa."""
    rows = []
    loan_disp = {}
    if loans_df is not None and "RO Name" in loans_df.columns:
        for n in loans_df["RO Name"].dropna():
            k = _name_key(n)
            if k:
                loan_disp.setdefault(k, str(n).strip())
    master_disp = {}
    for n in (master_ro_names or []):
        k = _name_key(RO_NAME_ALIASES.get(n, n))
        if k:
            master_disp.setdefault(k, str(n).strip())
    for k, disp in sorted(loan_disp.items()):
        if k not in master_disp:
            rows.append((disp, "In Loan file, not in RO Master"))
    for k, disp in sorted(master_disp.items()):
        if k not in loan_disp:
            rows.append((disp, "In RO Master, not in Loan file"))
    return pd.DataFrame(rows, columns=["RO Name", "Status"])


def missing_ro_table_html(missing_df):
    if missing_df is None or missing_df.empty:
        return ('<p style="margin:6px 0;font-size:13px;color:#16A34A;">'
                "All ROs reconcile between the Loan file and the RO Master &mdash; none missing.</p>")
    head = ('<tr>'
            f'<th style="padding:8px 6px;text-align:left;background:{HEAD};color:#fff;font-size:11px;font-weight:700;">RO Name</th>'
            f'<th style="padding:8px 6px;text-align:left;background:{HEAD};color:#fff;font-size:11px;font-weight:700;">Status</th></tr>')
    rows = ""
    for i, r in missing_df.reset_index(drop=True).iterrows():
        loan_side = "Loan" in r["Status"] and "not in RO Master" in r["Status"]
        bg = "#FEF3C7" if loan_side else "#E0E7FF"      # amber vs indigo tint
        rows += (f'<tr style="background:{bg};">'
                 f'<td style="padding:7px 6px;text-align:left;color:{INK};font-weight:600;border-bottom:1px solid {LINE};">{r["RO Name"]}</td>'
                 f'<td style="padding:7px 6px;text-align:left;color:{INK};border-bottom:1px solid {LINE};">{r["Status"]}</td></tr>')
    return (f'<table role="presentation" width="100%" cellpadding="0" cellspacing="0" '
            f'style="border-collapse:collapse;width:100%;font-size:12px;'
            f'box-shadow:0 1px 3px rgba(0,0,0,.06);border-radius:6px;overflow:hidden;">'
            f'{head}{rows}</table>')


def build_email_html(product, report_date, m, cids, agent_perf=None, missing_df=None):
    t = m["total"]
    comm_line = (f"Out of the total {t:,} calls, {m['inbound_n']:,} ({pct(m['inbound_n'],t)}) were inbound calls, "
                 f"{m['outbound_n']:,} ({pct(m['outbound_n'],t)}) were outbound calls and "
                 f"{m['internal_n']:,} ({pct(m['internal_n'],t)}) were internal calls.")
    succ_line = (f"Out of the total {t:,} calls, {m['successful_n']:,} ({pct(m['successful_n'],t)}) were successful and "
                 f"{m['unsuccessful_n']:,} ({pct(m['unsuccessful_n'],t)}) were unsuccessful.")
    body = []
    body.append(_heading("Calls Summary", SPECTRUM[0]))
    body.append(_bullets([
        f"Total calls made for the day were {t:,}, with {m['distinct_called']:,} unique phone numbers being "
        f"called (outbound) and {m['distinct_calling']:,} unique phone numbers that called in (inbound).",
        comm_line,
    ]))
    if "communication_type" in cids:
        body.append(_chart_block("Communication Type Distribution", "communication_type"))
    body.append(_bullets([succ_line]))
    if "success_distribution" in cids:
        body.append(_chart_block("Call Success Distribution", "success_distribution"))
    body.append(_bullets(["The distribution of the calls disposition is visualized below."]))
    if "call_notes_distribution" in cids:
        body.append(_chart_block("Call Notes Distribution", "call_notes_distribution"))
    body.append(SEP)

    body.append(_heading("Agents Performance Highlights", SPECTRUM[4]))
    body.append(_bullets([
        f"Of the total {m['total_agents']} agents who made calls, {m['agents_50']} "
        f"({pct(m['agents_50'],m['total_agents'])}) had 50 or more successful calls for the day "
        f"(both inbound &amp; outbound).",
    ]))
    if "top_agents" in cids:
        body.append(_chart_block("Top Agents by Successful Calls", "top_agents"))
    ro_tbl = ro_agent_table_html(agent_perf)
    if ro_tbl:
        body.append(_heading("RO Collections Scorecard (ranked best → worst by PAR&gt;30)", SPECTRUM[7]))
        body.append(_bullets([
            "Agents are ranked from the strongest portfolio quality (lowest PAR&gt;30) to the weakest. "
            "Figures cover calls (out/in, successful, success rate), outstanding portfolio balance, "
            "PAR&gt;30 and amount collected.",
        ]))
        body.append(ro_tbl)
        miss_tbl = ro_agent_table_html(missing_df)
        if miss_tbl:
            body.append(_heading("Missing RO Collections Scorecard", SPECTRUM[0]))
            body.append(_bullets([
                "ROs that are not active, ranked by PAR&gt;30.",
            ]))
            body.append(miss_tbl)
    body.append(SEP)

    body.append(_heading("For Outbound Calls", SPECTRUM[1]))
    body.append(_bullets([
        f"Average outbound calls made per agent was {m['avg_out']}.",
        f"Of the total {m['out_agents']} agents who made outbound calls, {m['out_succ_50']} "
        f"({pct(m['out_succ_50'],m['out_agents'])}) agents had 50 or more successful outbound calls for the day.",
        f"{m['out_unsucc_50']} ({pct(m['out_unsucc_50'],m['out_agents'])}) agents had 50 or more "
        f"unsuccessful outbound calls for the day.",
    ]))
    body.append(SEP)

    body.append(_heading("For Inbound Calls", SPECTRUM[6]))
    if m["inb_total"]:
        body.append(_bullets([
            f"Average inbound calls received per agent was {m['avg_in']}.",
            f"Of the total {m['inb_total']:,} inbound calls, {m['inb_succ']:,} ({pct(m['inb_succ'],m['inb_total'])}) "
            f"were successful and {m['inb_unsucc']:,} ({pct(m['inb_unsucc'],m['inb_total'])}) were unsuccessful.",
            f"Of the {m['inb_unsucc']:,} unsuccessful inbound calls, {m['called_back']:,} "
            f"({pct(m['called_back'],m['inb_unsucc'])}) were called back.",
        ]))
    else:
        body.append(_bullets(["No inbound calls recorded for the day."]))
    if "status_distribution" in cids:
        body.append(_chart_block("Call Status Distribution", "status_distribution"))

    body_html = "".join(body)
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#F1F5F9;font-family:'Segoe UI',Tahoma,Arial,sans-serif;color:{INK};">
  <div style="max-width:960px;width:100%;margin:0 auto;background:#FFFFFF;">
    <div style="background:{HEAD};padding:24px 28px;text-align:center;">
      <div style="color:#FFFFFF;font-size:22px;font-weight:800;">Call Center Daily Report</div>
      <div style="color:#CBD5E1;font-size:13px;margin-top:5px;">{product} team &nbsp;&bull;&nbsp; {report_date}</div>
    </div>
    <div style="height:4px;background:linear-gradient(90deg,{','.join(SPECTRUM)});"></div>
    <div style="padding:24px 28px;">
      <p style="font-size:14px;color:{INK};margin:0 0 12px;">Hi team,</p>
      <p style="font-size:13.5px;color:{MUTED};margin:0 0 18px;line-height:1.55;">
        Below is the call center summary report for <b>{product}</b> on <b>{report_date}</b>.
        Key numbers are shown first, followed by the breakdown and charts.</p>
      {kpi_cards_html(m)}
      {SEP}
      {body_html}
    </div>
    <div style="background:{HEAD};padding:14px 28px;color:#94A3B8;font-size:11px;">
      The full report (Summary, All Call Data, Agent Performance, Inbound/Outbound &amp; Call Notes
      summaries) is attached as an Excel file.
    </div>
  </div>
</body></html>"""


def send_email(product, report_date, html, excel_path, chart_paths, recipients, sender, password):
    subject = f"CALL CENTER {product} REPORT FOR {report_date}"
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587, timeout=60)
        server.starttls()
        server.login(sender, password)
        msg = MIMEMultipart("related")
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        alt = MIMEMultipart("alternative")
        msg.attach(alt)
        alt.attach(MIMEText(html, "html"))
        # inline chart images (cid keyed by chart type)
        for ctype, path in chart_paths.items():
            if path and os.path.exists(path):
                with open(path, "rb") as fh:
                    img = MIMEImage(fh.read(), name=os.path.basename(path))
                img.add_header("Content-ID", f"<{ctype}>")
                img.add_header("Content-Disposition", "inline", filename=os.path.basename(path))
                msg.attach(img)
        if excel_path and os.path.exists(excel_path):
            with open(excel_path, "rb") as fh:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(fh.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(excel_path)}"')
            msg.attach(part)
        server.sendmail(sender, recipients, msg.as_string())
        server.quit()
        print(f"  [SENT] {product} -> {', '.join(recipients)}")
        return True
    except Exception as e:
        print(f"  [FAIL] {product}: {e}")
        return False


# =============================================================================
# MAIN
# =============================================================================
def main():
    os.makedirs(EXPORT_ROOT, exist_ok=True)
    df, report_date = load_data()
    agents = build_agent_products(df)
    print("Product distribution:", agents["Product"].value_counts().to_dict())

    # LBF loan context for the RO report (phone -> loan row, RO portfolio metrics)
    loans_df = process_lbf_loans()
    phone_long = build_phone_long(loans_df)
    ro_metrics = ro_portfolio_metrics(loans_df)
    master_ro_names = load_master_ro_names()
    loan_name_by_key = ({_name_key(n): str(n).strip() for n in loans_df["RO Name"].dropna()}
                        if loans_df is not None and "RO Name" in loans_df.columns else {})

    # Standalone LBF loan report (processed) for manual inspection
    if loans_df is not None and not loans_df.empty:
        lbf_dir = os.path.join(EXPORT_ROOT, "RO", report_date)
        os.makedirs(lbf_dir, exist_ok=True)
        lbf_out = os.path.join(lbf_dir, f"LBF_Loan_Report_{report_date}.xlsx")
        loans_df.drop(columns=["_phones"], errors="ignore").to_excel(lbf_out, index=False)
        print(f"  LBF loan report saved for inspection: {lbf_out}")

    built = {}
    for product in PRODUCTS:
        adf = agents[agents["Product"] == product]
        if len(adf) == 0:
            print(f"  {product}: no agents — skipped")
            continue
        names = adf["Agent Name"].tolist()
        pdf = df[(df["Call From"].isin(names)) | (df["Call To"].isin(names))].copy()
        if len(pdf) == 0:
            print(f"  {product}: no calls — skipped")
            continue
        out_dir = os.path.join(EXPORT_ROOT, product, report_date)
        os.makedirs(out_dir, exist_ok=True)
        m = compute_metrics(pdf, adf)
        is_ro = product == "RO" and loans_df is not None
        ap = build_agent_perf(adf, is_ro, ro_metrics)
        if is_ro:
            main_sc, missing_sc = split_ro_scorecards(ap, ro_metrics, master_ro_names, loan_name_by_key, loans_df)
            print(f"  RO scorecard: {len(main_sc)} reconciled, {len(missing_sc)} missing")
        else:
            main_sc, missing_sc = ap, None
        excel_path = os.path.join(out_dir, f"CALL_REPORT_{product}_{report_date}.xlsx")
        write_excel(excel_path, product, report_date, m, pdf, adf,
                    loans_df=loans_df, phone_long=phone_long, ro_metrics=ro_metrics)
        chart_paths = make_charts(pdf, adf, product, out_dir, report_date)
        built[product] = (m, excel_path, chart_paths, main_sc, missing_sc)
        print(f"  {product}: {len(adf)} agents, {len(pdf):,} calls -> {os.path.basename(excel_path)} "
              f"({len(chart_paths)} charts)")

    if not built:
        print("Nothing to send.")
        return

    # Send mode: from CLI flag when launched by the main report (--yes / --no),
    # otherwise prompt. When called as a child, the parent handles the DB upload.
    cli = set(a.lower() for a in sys.argv[1:])
    invoked_by_parent = bool(cli & {"--yes", "--no"})
    if "--yes" in cli:
        send_all = True
    elif "--no" in cli:
        send_all = False
    else:
        try:
            ans = input("\nSend to all recipients? [yes/no] (no = send to Daniel only for testing): ").strip().lower()
        except EOFError:
            ans = "no"
        send_all = ans in ("y", "yes")

    load_dotenv(ENV_PATH)
    sender = os.getenv("EMAIL_USERNAME")
    password = os.getenv("EMAIL_PASSWORD")
    if not sender or not password:
        print(f"EMAIL_USERNAME / EMAIL_PASSWORD not found in {ENV_PATH}. Email skipped.")
        return

    print(f"\nSending emails ({'ALL recipients' if send_all else 'Daniel only — TEST'}):")
    for product, (m, excel_path, chart_paths, main_sc, missing_sc) in built.items():
        html = build_email_html(product, report_date, m, chart_paths,
                                agent_perf=main_sc, missing_df=missing_sc)
        recipients = RECIPIENTS.get(product, [MY_EMAIL]) if send_all else [MY_EMAIL]
        send_email(product, report_date, html, excel_path, chart_paths, recipients, sender, password)

    if send_all and not invoked_by_parent:
        print("\nUploading new report(s) to the live PCL system ...")
        try:
            import subprocess
            subprocess.run([sys.executable, UPLOAD_SCRIPT, "--commit"], check=False)
        except Exception as e:
            print(f"  Upload step could not run: {e}")
    elif invoked_by_parent:
        print("\nRO report finished (launched by main report — parent handles DB upload).")
    else:
        print("\nTest run (no) — skipping DB upload.")


if __name__ == "__main__":
    main()
