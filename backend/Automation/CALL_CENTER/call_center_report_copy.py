"""
Call Center daily report — CS / LBF / RO / ERR.

Per product:
  * One Excel with 6 sheets: Summary, All_Call_Data, Agent_Performance,
    Inbound_Summary, Outbound_Summary, Call_Notes_Summary.
    (All_Call_Data has the noisy columns removed: Call From, Call To, Ivr, Queue,
     AI Receptionist, Disconnected by, Call Legs, Ring Group.)
  * Charts drawn as PNG images (matplotlib) and embedded inline in the email.
  * A professional HTML email: greeting + KPI boxes (5 per row) + narrative
    explanation (Calls Summary / Outbound / Inbound) with the charts in their
    sections.
Saved as:  NEW_FILES/<PRODUCT>/<date>/...
After processing it asks "Send to all recipients? [yes/no]":
  yes -> real distribution lists + DB upload.
  no  -> sends every product email to ME only (testing); no upload.
"""

from __future__ import annotations

import argparse
import os
import re
import smtplib
import sys
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# =============================================================================
# CONFIG
# =============================================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_PATH = os.path.join(SCRIPT_DIR, "ROW_FILES")
EXPORT_ROOT = os.path.join(SCRIPT_DIR, "NEW_FILES")
ENV_PATH = os.path.join(SCRIPT_DIR, ".env")
# Dynamic: the DB-upload helper lives in <Automation>/CRM. Honour the orchestrator's
# root, else fall back to this script's grandparent (…/Automation)/CRM.
_AUTOMATION_ROOT = os.environ.get("PCL_AUTOMATION_ROOT") or os.path.dirname(SCRIPT_DIR)
UPLOAD_SCRIPT = os.path.join(_AUTOMATION_ROOT, "CRM", "direct_upload_files_to_db.py")

PRODUCTS = ["CS", "LBF", "ERR"]   # RO is handled by the dedicated RO subprocess (call_center_RO_report_copy.py)
MY_EMAIL = "daniel@platinumcredit.co.tz"

DROP_FROM_CDR = ["Call From", "Call To", "Ivr", "Queue", "AI Receptionist",
                 "Disconnected by", "Call Legs", "Ring Group",
                 "Call Leg ID", "Call Flow", "Campaign"]

EXCLUDED_AGENTS = ["Ikrah Ally", "David Kileo", "Aziza Mfanga", "Madina Mohamed",
                   "Jackson Swai", "Thomas Francis", "Conference Call"]

RECIPIENTS = {
    "CS": [
        "raphael@platinumcredit.co.tz", "dorice@platinumcredit.co.tz",
        "sigfrid@platinumcredit.co.tz", "murigi@platinumcredit.co.ke",
        "wayne@platinumcredit.co.ke", "yusuph@platinumcredit.co.tz",
        "allan@platinumcredit.co.tz", "fragrance@platinumcredit.co.tz",
        "vivian@platinumcredit.co.tz", "thomas@platinumcredit.co.tz",
        "mohamedi.omar.platinum@gmail.com", "kelvin.mwasala@platinumcredit.co.tz",
        "daniel@platinumcredit.co.tz", "kelvin.peter.platinum@gmail.com",
    ],
    "LBF": [
        "raphael@platinumcredit.co.tz", "allan@platinumcredit.co.tz",
        "dorice@platinumcredit.co.tz", "fragrance@platinumcredit.co.tz",
        "sigfrid@platinumcredit.co.tz", "murigi@platinumcredit.co.ke",
        "wayne@platinumcredit.co.ke", "yusuph@platinumcredit.co.tz",
        "thomas@platinumcredit.co.tz", "augustine@platinumcredit.co.tz",
        "irene.mmari@platinumcredit.co.tz", "daniel@platinumcredit.co.tz",
        "aziza.mfanga.platinum@gmail.com", "madina.mohamed.platinum@gmail.com",
        "barnabas.ngassa.platinum@gmail.com", "zaituni@platinumcredit.co.tz",
    ],
    "RO": [
        "raphael@platinumcredit.co.tz", "allan@platinumcredit.co.tz",
        "dorice@platinumcredit.co.tz", "fragrance@platinumcredit.co.tz",
        "sigfrid@platinumcredit.co.tz", "daniel@platinumcredit.co.tz",
        "relationshipofficers@platinumcredit.co.tz",
        "jonas@platinumcredit.co.tz"
    ],
    "ERR": ["daniel@platinumcredit.co.tz"],
}

# Recipients can be managed from the web via master_email_call_center.xlsx. If it
# exists it OVERRIDES the defaults above (flow-only change — logic untouched).
def _load_recipient_master():
    try:
        from openpyxl import load_workbook
        mp = os.path.join(SCRIPT_DIR, "master_email_call_center.xlsx")
        if not os.path.exists(mp):
            return
        wb = load_workbook(mp, read_only=True, data_only=True)
        ws = wb["Recipients"] if "Recipients" in wb.sheetnames else wb.active
        headers = {str(c.value).strip().upper(): c.column for c in ws[1] if c.value}
        for dept, col in headers.items():
            emails = [str(ws.cell(row=r, column=col).value).strip()
                      for r in range(2, ws.max_row + 1)
                      if ws.cell(row=r, column=col).value and "@" in str(ws.cell(row=r, column=col).value)]
            if emails:
                RECIPIENTS[dept] = emails
        wb.close()
    except Exception as e:
        print(f"  (recipient master not loaded: {e})")

_load_recipient_master()

# 8-colour spectrum, red -> violet.
SPECTRUM = ["#E11D48", "#F97316", "#F59E0B", "#EAB308",
            "#22C55E", "#06B6D4", "#3B82F6", "#7C3AED"]
INK, MUTED, HEAD, LINE = "#0F172A", "#475569", "#1E3A5F", "#E2E8F0"

# Continuous VIOLET -> RED colormap (largest bar = violet, smallest = red).
_VR_CMAP = LinearSegmentedColormap.from_list("violet_red", list(reversed(SPECTRUM)))


def gradient_colors(n):
    """n colours spanning violet -> red, recalculated to the count of data points."""
    if n <= 0:
        return []
    if n == 1:
        return [_VR_CMAP(0.0)]
    return [_VR_CMAP(i / (n - 1)) for i in range(n)]


# Whole-box colour per KPI label (semantic).
KPI_COLORS = {
    "Total Calls": "#1E3A5F", "Inbound": "#3B82F6", "Outbound": "#F97316",
    "Internal": "#64748B", "Successful": "#22C55E", "Unsuccessful": "#E11D48",
    "Success Rate": "#16A34A", "Unique Called": "#06B6D4", "Unique Callers": "#0EA5E9",
    "Agents": "#7C3AED", "Agents 50+ Succ": "#15803D", "Avg Out / Agent": "#EA580C",
    "Avg In / Agent": "#2563EB", "Inbound Calls": "#0891B2", "Inbound Called Back": "#D97706",
}


# =============================================================================
# LOAD / CLEAN
# =============================================================================
def _find(prefix, ext):
    for f in os.listdir(FOLDER_PATH):
        if f.lower().startswith(prefix) and f.lower().endswith(ext):
            return os.path.join(FOLDER_PATH, f)
    return None


def extract_report_date(time_series):
    if len(time_series) == 0 or pd.isna(time_series.iloc[0]):
        return (datetime.now() - pd.Timedelta(days=1)).strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(time_series.iloc[0]).split(" ")[0], "%m/%d/%Y").strftime("%Y-%m-%d")
    except ValueError:
        return (datetime.now() - pd.Timedelta(days=1)).strftime("%Y-%m-%d")


def clean_name(name):
    if pd.isna(name):
        return name
    name = re.sub(r"^Voicemail\s+", "", str(name))
    name = re.sub(r"<.*?>", "", name).strip()
    return "Hadija Mohamed" if name == "Khadija Mohamed" else name


def determine_success(row):
    status = str(row["Status"]).lower()
    if status in ["no answer", "busy", "failed", "voicemail", "abandoned"]:
        return "Unsuccessful"
    if status == "answered":
        try:
            return "Successful" if int(row["Talk Duration"]) >= 5 else "Unsuccessful"
        except (TypeError, ValueError):
            return "Unsuccessful"
    return "Unsuccessful"


def clean_notes(row):
    note = str(row["Call Notes"])
    if note.strip() == "" or note.lower() == "nan":
        status = str(row["Status"]).lower()
        return {"no answer": "Not picking", "voicemail": "Not picking",
                "failed": "Failed Connection", "busy": "Busy", "abandoned": "Abandoned",
                "answered": "NO COMMENT WRITTEN"}.get(status, "UNKNOWN")
    return note[:len(note.lower().split("remark")[0])].strip()


def load_data():
    cdr = _find("pse-cdr", ".csv")
    if not cdr:
        raise FileNotFoundError(f"No PSE-Cdr*.csv in {FOLDER_PATH}")
    print(f"CDR file: {os.path.basename(cdr)}")
    df = pd.read_csv(cdr)
    rename = {}
    if "Status" not in df.columns and "Last Status" in df.columns:
        rename["Last Status"] = "Status"
    if "Talk Duration" not in df.columns and "Handling Duration" in df.columns:
        rename["Handling Duration"] = "Talk Duration"
    if rename:
        df = df.rename(columns=rename)
    report_date = extract_report_date(df["Time"])
    print(f"Report date: {report_date}")
    df["Call From"] = df["Call From"].apply(clean_name)
    df["Call To"] = df["Call To"].apply(clean_name)
    df["Successful ?"] = df.apply(determine_success, axis=1)
    df["Call Notes"] = df.apply(clean_notes, axis=1)
    return df, report_date


def build_agent_products(df):
    inbound = df[(df["Communication Type"] == "Inbound") & (~df["Call To"].isin(EXCLUDED_AGENTS))]
    outbound = df[(df["Communication Type"] == "Outbound") & (~df["Call From"].isin(EXCLUDED_AGENTS))]
    in_df = inbound.groupby("Call To").agg(
        inbound_calls=("Call To", "count"),
        successful_inbound=("Successful ?", lambda x: (x == "Successful").sum()))
    out_df = outbound.groupby("Call From").agg(
        outbound_calls=("Call From", "count"),
        successful_outbound=("Successful ?", lambda x: (x == "Successful").sum()))
    agents = pd.merge(out_df, in_df, left_index=True, right_index=True, how="outer").fillna(0)
    agents.index.name = "Agent Name"
    agents = agents.reset_index()
    agents["Agent Name"] = agents["Agent Name"].replace("Khadija Mohamed", "Hadija Mohamed")
    agents = agents[~agents["Agent Name"].isin(EXCLUDED_AGENTS + ["Barnabas Ngassa"])]
    agents = agents.groupby("Agent Name", as_index=False).sum()
    agents["Total Calls"] = agents["outbound_calls"] + agents["inbound_calls"]
    agents["Successful Calls"] = agents["successful_outbound"] + agents["successful_inbound"]
    agents["Success Rate (%)"] = (agents["Successful Calls"] / 50 * 100).apply(lambda x: f"{x:.2f}%")
    agents = agents[["Agent Name", "outbound_calls", "inbound_calls", "Total Calls",
                     "Successful Calls", "Success Rate (%)"]].sort_values("Successful Calls", ascending=False)

    lookup = {}
    master = _find("master_cdr_call", ".xlsx")
    if master and os.path.exists(master):
        try:
            m = pd.read_excel(master, sheet_name="Agent_Performance")
            if {"Agent Name", "Product"} <= set(m.columns):
                for _, r in m.iterrows():
                    if pd.notna(r["Agent Name"]) and pd.notna(r["Product"]):
                        lookup[r["Agent Name"]] = str(r["Product"]).strip()
        except Exception as e:
            print(f"WARN master: {e}")
    agents["Product"] = agents["Agent Name"].map(lookup)
    agents["Product"] = agents["Product"].where(agents["Product"].isin(["CS", "LBF", "RO"]), "ERR")
    return agents


# =============================================================================
# METRICS
# =============================================================================
def compute_metrics(pdf, agents_df):
    m = {}
    total = len(pdf)
    comm = pdf["Communication Type"].value_counts()
    m["total"] = total
    m["inbound_n"] = int(comm.get("Inbound", 0))
    m["outbound_n"] = int(comm.get("Outbound", 0))
    m["internal_n"] = int(comm.get("Internal", 0))
    m["successful_n"] = int((pdf["Successful ?"] == "Successful").sum())
    m["unsuccessful_n"] = total - m["successful_n"]
    m["distinct_called"] = pdf[pdf["Communication Type"] == "Outbound"]["Call To"].nunique()
    m["distinct_calling"] = pdf[pdf["Communication Type"] == "Inbound"]["Call From"].nunique()
    m["total_agents"] = len(agents_df)
    m["agents_50"] = int((agents_df["Successful Calls"] >= 50).sum())

    out = pdf[(pdf["Communication Type"] == "Outbound") & (~pdf["Call From"].isin(EXCLUDED_AGENTS))]
    out_agents = out["Call From"].nunique()
    m["out_agents"] = out_agents
    m["avg_out"] = round(len(out) / out_agents, 2) if out_agents else 0
    out_succ = out[out["Successful ?"] == "Successful"]
    out_unsucc = out[out["Successful ?"] == "Unsuccessful"]
    m["out_succ_50"] = out_succ.groupby("Call From").filter(lambda x: len(x) >= 50)["Call From"].nunique()
    m["out_unsucc_50"] = out_unsucc.groupby("Call From").filter(lambda x: len(x) >= 50)["Call From"].nunique()

    inb = pdf[(pdf["Communication Type"] == "Inbound") & (~pdf["Call To"].isin(EXCLUDED_AGENTS))]
    inb_total = len(inb)
    m["inb_total"] = inb_total
    in_agents = inb["Call To"].nunique()
    m["avg_in"] = round(inb_total / in_agents, 2) if in_agents else 0
    inb_succ = inb[inb["Successful ?"] == "Successful"]
    inb_unsucc = inb[inb["Successful ?"] == "Unsuccessful"]
    m["inb_succ"] = len(inb_succ)
    m["inb_unsucc"] = len(inb_unsucc)
    out_numbers = out["Call To"].unique()
    called_back = inb_unsucc[inb_unsucc["Call From"].isin(out_numbers)]
    m["called_back"] = len(called_back)
    return m


def pct(n, d):
    return f"{(n/d if d else 0):.0%}"


def summary_rows(m):
    t = m["total"]
    return [
        ("Total calls", f"{t:,}"),
        ("Inbound", f"{m['inbound_n']:,} ({pct(m['inbound_n'], t)})"),
        ("Outbound", f"{m['outbound_n']:,} ({pct(m['outbound_n'], t)})"),
        ("Internal", f"{m['internal_n']:,} ({pct(m['internal_n'], t)})"),
        ("Successful", f"{m['successful_n']:,} ({pct(m['successful_n'], t)})"),
        ("Unsuccessful", f"{m['unsuccessful_n']:,} ({pct(m['unsuccessful_n'], t)})"),
        ("Unique numbers called (outbound)", f"{m['distinct_called']:,}"),
        ("Unique numbers calling in (inbound)", f"{m['distinct_calling']:,}"),
        ("Agents who made calls", f"{m['total_agents']:,}"),
        ("Agents with 50+ successful calls", f"{m['agents_50']} ({pct(m['agents_50'], m['total_agents'])})"),
        ("Avg outbound calls / agent", f"{m['avg_out']}"),
        ("Avg inbound calls / agent", f"{m['avg_in']}"),
    ]


# =============================================================================
# CHARTS (matplotlib PNGs, embedded inline in the email)
# =============================================================================
def _chart(items, title, path, kind="bar"):
    items = [(str(k), int(v)) for k, v in items if int(v) > 0]
    if not items:
        return None
    labels = [k for k, _ in items]
    values = [v for _, v in items]
    # violet (largest) -> red (smallest); items are already sorted descending
    colors = gradient_colors(len(items))
    plt.figure(figsize=(9, max(3.2, len(items) * 0.5)) if kind == "barh" else (9, 5))
    ax = plt.gca()
    ax.set_facecolor("#FFFFFF")
    if kind == "barh":
        bars = plt.barh(range(len(values)), values, color=colors, edgecolor="white")
        for b, v in zip(bars, values):
            plt.text(b.get_width(), b.get_y() + b.get_height() / 2, f" {v:,}",
                     va="center", fontsize=10, fontweight="bold", color=INK)
        plt.yticks(range(len(labels)), labels, fontsize=10)
        plt.gca().invert_yaxis()
    else:
        bars = plt.bar(range(len(values)), values, color=colors, edgecolor="white")
        for b, v in zip(bars, values):
            plt.text(b.get_x() + b.get_width() / 2, b.get_height(), f"{v:,}",
                     ha="center", va="bottom", fontsize=10, fontweight="bold", color=INK)
        plt.xticks(range(len(labels)), labels, rotation=30, ha="right", fontsize=10)
    plt.title(title, fontsize=13, fontweight="bold", color=HEAD, pad=12)
    plt.grid(axis="x" if kind == "barh" else "y", alpha=0.25, linestyle="--", linewidth=0.5)
    for s in ax.spines.values():
        s.set_visible(False)
    plt.tight_layout()
    plt.savefig(path, dpi=160, bbox_inches="tight", facecolor="white")
    plt.close()
    return path


def make_charts(pdf, agents_df, product, out_dir, report_date):
    paths = {}
    p = _chart(pdf["Communication Type"].value_counts().items(),
               f"Communication Type — {product} — {report_date}",
               os.path.join(out_dir, f"communication_type_{product}_{report_date}.png"), "bar")
    if p: paths["communication_type"] = p
    p = _chart(pdf["Successful ?"].value_counts().items(),
               f"Call Success — {product} — {report_date}",
               os.path.join(out_dir, f"success_distribution_{product}_{report_date}.png"), "bar")
    if p: paths["success_distribution"] = p
    p = _chart(pdf["Call Notes"].value_counts().head(10).items(),
               f"Call Notes Distribution — {product} — {report_date}",
               os.path.join(out_dir, f"call_notes_distribution_{product}_{report_date}.png"), "barh")
    if p: paths["call_notes_distribution"] = p
    top = agents_df.nlargest(10, "Successful Calls")
    p = _chart(zip(top["Agent Name"], top["Successful Calls"]),
               f"Top Agents by Successful Calls — {product} — {report_date}",
               os.path.join(out_dir, f"top_agents_{product}_{report_date}.png"), "bar")
    if p: paths["top_agents"] = p
    p = _chart(pdf["Status"].value_counts().items(),
               f"Call Status — {product} — {report_date}",
               os.path.join(out_dir, f"status_distribution_{product}_{report_date}.png"), "barh")
    if p: paths["status_distribution"] = p
    return paths


# =============================================================================
# EXCEL  (6 sheets)
# =============================================================================
def _style_sheet(ws, pct_cols=(), int_cols=(), tab="1E3A5F", first_width=30):
    """Header styling, banded rows, borders, number formats, freeze + tab colour."""
    ws.sheet_properties.tabColor = tab
    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    navy = PatternFill(start_color=tab, end_color=tab, fill_type="solid")
    band = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    thin = Side(style="thin", color="D6E2F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(1, c)
        cell.fill = navy
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.border = border
            if r % 2 == 0:
                cell.fill = band
    for name, col in headers.items():
        fmt = "0.00%" if name in pct_cols else ("#,##0" if name in int_cols else None)
        if fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = fmt
                ws.cell(r, col).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = first_width if c == 1 else 16
    ws.row_dimensions[1].height = 22
    if ws.max_row >= 1:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"


def write_excel(path, product, report_date, m, pdf, agents_df):
    cdr_df = pdf.drop(columns=[c for c in DROP_FROM_CDR if c in pdf.columns])
    summary_df = pd.DataFrame(summary_rows(m), columns=["Metric", "Value"])

    # Agent performance with a numeric success rate (fraction) for % formatting
    ap = agents_df.copy()
    ap["Success Rate"] = ap["Successful Calls"] / 50
    ap = ap[["Agent Name", "outbound_calls", "inbound_calls", "Total Calls",
             "Successful Calls", "Success Rate"]]

    out = pdf[(pdf["Communication Type"] == "Outbound") & (~pdf["Call From"].isin(EXCLUDED_AGENTS))]
    inb = pdf[(pdf["Communication Type"] == "Inbound") & (~pdf["Call To"].isin(EXCLUDED_AGENTS))]

    if len(inb):
        ins = inb.groupby("Call To").agg(
            Total_Inbound=("Successful ?", "count"),
            Successful_Inbound=("Successful ?", lambda x: (x == "Successful").sum()))
        ins["Inbound_Success_Rate"] = ins["Successful_Inbound"] / ins["Total_Inbound"]
        ins = ins.reset_index().rename(columns={"Call To": "Agent"})
        ins = ins.sort_values("Inbound_Success_Rate", ascending=False)
    else:
        ins = pd.DataFrame(columns=["Agent", "Total_Inbound", "Successful_Inbound", "Inbound_Success_Rate"])

    if len(out):
        outs = out.groupby("Call From").agg(
            Total_Outbound=("Successful ?", "count"),
            Successful_Outbound=("Successful ?", lambda x: (x == "Successful").sum()))
        outs["Outbound_Success_Rate"] = outs["Successful_Outbound"] / outs["Total_Outbound"]
        outs = outs.reset_index().rename(columns={"Call From": "Agent"})
        outs = outs.sort_values("Outbound_Success_Rate", ascending=False)
    else:
        outs = pd.DataFrame(columns=["Agent", "Total_Outbound", "Successful_Outbound", "Outbound_Success_Rate"])

    notes = pdf["Call Notes"].value_counts().reset_index()
    notes.columns = ["Call_Notes", "Count"]
    notes["Percentage"] = (notes["Count"] / m["total"]) if m["total"] else 0.0

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        cdr_df.to_excel(writer, sheet_name="All_Call_Data", index=False)
        ap.to_excel(writer, sheet_name="Agent_Performance", index=False)
        ins.to_excel(writer, sheet_name="Inbound_Summary", index=False)
        outs.to_excel(writer, sheet_name="Outbound_Summary", index=False)
        notes.to_excel(writer, sheet_name="Call_Notes_Summary", index=False)

        sheets = writer.sheets
        _style_sheet(sheets["Summary"], tab="1E3A5F", first_width=38)
        _style_sheet(sheets["All_Call_Data"],
                     int_cols=("Total Duration", "Routing Duration", "Talk Duration"),
                     tab="3B82F6", first_width=22)
        _style_sheet(sheets["Agent_Performance"],
                     int_cols=("outbound_calls", "inbound_calls", "Total Calls", "Successful Calls"),
                     pct_cols=("Success Rate",), tab="7C3AED", first_width=26)
        _style_sheet(sheets["Inbound_Summary"],
                     int_cols=("Total_Inbound", "Successful_Inbound"),
                     pct_cols=("Inbound_Success_Rate",), tab="06B6D4", first_width=26)
        _style_sheet(sheets["Outbound_Summary"],
                     int_cols=("Total_Outbound", "Successful_Outbound"),
                     pct_cols=("Outbound_Success_Rate",), tab="F97316", first_width=26)
        _style_sheet(sheets["Call_Notes_Summary"],
                     int_cols=("Count",), pct_cols=("Percentage",), tab="22C55E", first_width=34)


# =============================================================================
# EMAIL  (greeting + KPI boxes + narrative with inline chart images)
# =============================================================================
def kpi_cards_html(m):
    t = m["total"]
    succ_rate = f"{(m['successful_n']/t if t else 0):.0%}"
    cb_rate = f"{(m['called_back']/m['inb_unsucc'] if m['inb_unsucc'] else 0):.0%}"
    cards = [
        (f"{t:,}", "Total Calls"), (f"{m['inbound_n']:,}", "Inbound"),
        (f"{m['outbound_n']:,}", "Outbound"), (f"{m['internal_n']:,}", "Internal"),
        (f"{m['successful_n']:,}", "Successful"), (f"{m['unsuccessful_n']:,}", "Unsuccessful"),
        (succ_rate, "Success Rate"), (f"{m['distinct_called']:,}", "Unique Called"),
        (f"{m['distinct_calling']:,}", "Unique Callers"), (f"{m['total_agents']:,}", "Agents"),
        (f"{m['agents_50']:,}", "Agents 50+ Succ"), (f"{m['avg_out']}", "Avg Out / Agent"),
        (f"{m['avg_in']}", "Avg In / Agent"), (f"{m['inb_total']:,}", "Inbound Calls"),
        (cb_rate, "Inbound Called Back"),
    ]
    html = ('<table role="presentation" width="100%" cellpadding="0" cellspacing="0" '
            'style="border-collapse:separate;border-spacing:7px;width:100%;">')
    for i in range(0, len(cards), 5):
        html += "<tr>"
        chunk = cards[i:i + 5]
        for value, label in chunk:
            color = KPI_COLORS.get(label, HEAD)
            html += (f'<td width="20%" style="background:{color};border-radius:8px;'
                     'padding:16px 8px;text-align:center;vertical-align:middle;">'
                     f'<div style="font-size:22px;font-weight:800;color:#FFFFFF;line-height:1;">{value}</div>'
                     f'<div style="font-size:10px;color:#FFFFFF;opacity:.92;text-transform:uppercase;'
                     f'letter-spacing:.4px;margin-top:6px;font-weight:700;">{label}</div></td>')
        for _ in range(5 - len(chunk)):
            html += '<td width="20%"></td>'
        html += "</tr>"
    return html + "</table>"


def _bullets(lines):
    return "".join(
        f'<p style="margin:6px 0;line-height:1.55;font-size:13.5px;color:{INK};padding-left:16px;'
        f'position:relative;">&bull;&nbsp; {ln}</p>' for ln in lines)


def _heading(text, color):
    return (f'<div style="text-align:center;margin:24px 0 10px;">'
            f'<span style="display:inline-block;font-size:16px;font-weight:700;color:{HEAD};'
            f'padding-bottom:6px;border-bottom:3px solid {color};">{text}</span></div>')


SEP = f'<hr style="border:none;border-top:1px solid {LINE};margin:20px 0;">'


def _chart_block(title, cid):
    return (f'<div style="text-align:center;margin:12px 0;">'
            f'<div style="font-size:13px;font-weight:700;color:{MUTED};margin-bottom:6px;">{title}</div>'
            f'<img src="cid:{cid}" width="100%" style="width:100%;max-width:100%;height:auto;'
            f'border:1px solid {LINE};border-radius:6px;"></div>')


def build_email_html(product, report_date, m, cids):
    t = m["total"]
    comm_line = (f"Out of the total {t:,} calls, {m['inbound_n']:,} ({pct(m['inbound_n'],t)}) were inbound calls, "
                 f"{m['outbound_n']:,} ({pct(m['outbound_n'],t)}) were outbound calls and "
                 f"{m['internal_n']:,} ({pct(m['internal_n'],t)}) were internal calls.")
    succ_line = (f"Out of the total {t:,} calls, {m['successful_n']:,} ({pct(m['successful_n'],t)}) were successful and "
                 f"{m['unsuccessful_n']:,} ({pct(m['unsuccessful_n'],t)}) were unsuccessful.")
    body = []
    body.append(_heading("Calls Summary", SPECTRUM[0]))
    body.append(_bullets([
        f"Total calls made for the day were {t:,}, with {m['distinct_called']:,} unique phone numbers being "
        f"called (outbound) and {m['distinct_calling']:,} unique phone numbers that called in (inbound).",
        comm_line,
    ]))
    if "communication_type" in cids:
        body.append(_chart_block("Communication Type Distribution", "communication_type"))
    body.append(_bullets([succ_line]))
    if "success_distribution" in cids:
        body.append(_chart_block("Call Success Distribution", "success_distribution"))
    body.append(_bullets(["The distribution of the calls disposition is visualized below."]))
    if "call_notes_distribution" in cids:
        body.append(_chart_block("Call Notes Distribution", "call_notes_distribution"))
    body.append(SEP)

    body.append(_heading("Agents Performance Highlights", SPECTRUM[4]))
    body.append(_bullets([
        f"Of the total {m['total_agents']} agents who made calls, {m['agents_50']} "
        f"({pct(m['agents_50'],m['total_agents'])}) had 50 or more successful calls for the day "
        f"(both inbound &amp; outbound).",
    ]))
    if "top_agents" in cids:
        body.append(_chart_block("Top Agents by Successful Calls", "top_agents"))
    body.append(SEP)

    body.append(_heading("For Outbound Calls", SPECTRUM[1]))
    body.append(_bullets([
        f"Average outbound calls made per agent was {m['avg_out']}.",
        f"Of the total {m['out_agents']} agents who made outbound calls, {m['out_succ_50']} "
        f"({pct(m['out_succ_50'],m['out_agents'])}) agents had 50 or more successful outbound calls for the day.",
        f"{m['out_unsucc_50']} ({pct(m['out_unsucc_50'],m['out_agents'])}) agents had 50 or more "
        f"unsuccessful outbound calls for the day.",
    ]))
    body.append(SEP)

    body.append(_heading("For Inbound Calls", SPECTRUM[6]))
    if m["inb_total"]:
        body.append(_bullets([
            f"Average inbound calls received per agent was {m['avg_in']}.",
            f"Of the total {m['inb_total']:,} inbound calls, {m['inb_succ']:,} ({pct(m['inb_succ'],m['inb_total'])}) "
            f"were successful and {m['inb_unsucc']:,} ({pct(m['inb_unsucc'],m['inb_total'])}) were unsuccessful.",
            f"Of the {m['inb_unsucc']:,} unsuccessful inbound calls, {m['called_back']:,} "
            f"({pct(m['called_back'],m['inb_unsucc'])}) were called back.",
        ]))
    else:
        body.append(_bullets(["No inbound calls recorded for the day."]))
    if "status_distribution" in cids:
        body.append(_chart_block("Call Status Distribution", "status_distribution"))

    body_html = "".join(body)
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#F1F5F9;font-family:'Segoe UI',Tahoma,Arial,sans-serif;color:{INK};">
  <div style="max-width:960px;width:100%;margin:0 auto;background:#FFFFFF;">
    <div style="background:{HEAD};padding:24px 28px;text-align:center;">
      <div style="color:#FFFFFF;font-size:22px;font-weight:800;">Call Center Daily Report</div>
      <div style="color:#CBD5E1;font-size:13px;margin-top:5px;">{product} team &nbsp;&bull;&nbsp; {report_date}</div>
    </div>
    <div style="height:4px;background:linear-gradient(90deg,{','.join(SPECTRUM)});"></div>
    <div style="padding:24px 28px;">
      <p style="font-size:14px;color:{INK};margin:0 0 12px;">Hi team,</p>
      <p style="font-size:13.5px;color:{MUTED};margin:0 0 18px;line-height:1.55;">
        Below is the call center summary report for <b>{product}</b> on <b>{report_date}</b>.
        Key numbers are shown first, followed by the breakdown and charts.</p>
      {kpi_cards_html(m)}
      {SEP}
      {body_html}
    </div>
    <div style="background:{HEAD};padding:14px 28px;color:#94A3B8;font-size:11px;">
      The full report (Summary, All Call Data, Agent Performance, Inbound/Outbound &amp; Call Notes
      summaries) is attached as an Excel file.
    </div>
  </div>
</body></html>"""


def send_email(product, report_date, html, excel_path, chart_paths, recipients, sender, password):
    subject = f"CALL CENTER {product} REPORT FOR {report_date}"
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587, timeout=60)
        server.starttls()
        server.login(sender, password)
        msg = MIMEMultipart("related")
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        alt = MIMEMultipart("alternative")
        msg.attach(alt)
        alt.attach(MIMEText(html, "html"))
        # inline chart images (cid keyed by chart type)
        for ctype, path in chart_paths.items():
            if path and os.path.exists(path):
                with open(path, "rb") as fh:
                    img = MIMEImage(fh.read(), name=os.path.basename(path))
                img.add_header("Content-ID", f"<{ctype}>")
                img.add_header("Content-Disposition", "inline", filename=os.path.basename(path))
                msg.attach(img)
        if excel_path and os.path.exists(excel_path):
            with open(excel_path, "rb") as fh:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(fh.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(excel_path)}"')
            msg.attach(part)
        server.sendmail(sender, recipients, msg.as_string())
        server.quit()
        print(f"  [SENT] {product} -> {', '.join(recipients)}")
        return True
    except Exception as e:
        print(f"  [FAIL] {product}: {e}")
        return False


# =============================================================================
# MAIN
# =============================================================================
def main():
    os.makedirs(EXPORT_ROOT, exist_ok=True)
    df, report_date = load_data()
    agents = build_agent_products(df)
    print("Product distribution:", agents["Product"].value_counts().to_dict())

    built = {}
    for product in PRODUCTS:
        adf = agents[agents["Product"] == product]
        if len(adf) == 0:
            print(f"  {product}: no agents — skipped")
            continue
        names = adf["Agent Name"].tolist()
        pdf = df[(df["Call From"].isin(names)) | (df["Call To"].isin(names))].copy()
        if len(pdf) == 0:
            print(f"  {product}: no calls — skipped")
            continue
        out_dir = os.path.join(EXPORT_ROOT, product, report_date)
        os.makedirs(out_dir, exist_ok=True)
        m = compute_metrics(pdf, adf)
        excel_path = os.path.join(out_dir, f"CALL_REPORT_{product}_{report_date}.xlsx")
        write_excel(excel_path, product, report_date, m, pdf, adf)
        chart_paths = make_charts(pdf, adf, product, out_dir, report_date)
        built[product] = (m, excel_path, chart_paths)
        print(f"  {product}: {len(adf)} agents, {len(pdf):,} calls -> {os.path.basename(excel_path)} "
              f"({len(chart_paths)} charts)")

    if not built:
        print("Nothing to send.")
        return

    # Send choice: --send yes|no (from the web orchestrator), else prompt, else no.
    _ap = argparse.ArgumentParser(add_help=False)
    _ap.add_argument("--send", choices=["yes", "no"], default=None)
    _args, _ = _ap.parse_known_args()
    if _args.send is not None:
        ans = _args.send
    else:
        try:
            ans = input("\nSend to all recipients? [yes/no] (no = send to Daniel only for testing): ").strip().lower()
        except EOFError:
            ans = "no"
    send_all = ans in ("y", "yes")

    load_dotenv(ENV_PATH)
    sender = os.getenv("EMAIL_USERNAME")
    password = os.getenv("EMAIL_PASSWORD")
    if not sender or not password:
        print(f"EMAIL_USERNAME / EMAIL_PASSWORD not found in {ENV_PATH}. Email skipped.")
        return

    print(f"\nSending emails ({'ALL recipients' if send_all else 'Daniel only — TEST'}):")
    for product, (m, excel_path, chart_paths) in built.items():
        html = build_email_html(product, report_date, m, chart_paths)
        recipients = RECIPIENTS.get(product, [MY_EMAIL]) if send_all else [MY_EMAIL]
        send_email(product, report_date, html, excel_path, chart_paths, recipients, sender, password)

    # Also run the RO-specific report, matching this run's send choice (yes/no).
    # It generates + emails the RO report; the DB upload below covers its files too.
    print("\n" + "=" * 60)
    print("Running the RO report ...")
    try:
        import subprocess
        ro_script = os.path.join(SCRIPT_DIR, "call_center_RO_report_copy.py")
        subprocess.run([sys.executable, ro_script, "--yes" if send_all else "--no"], check=False)
    except Exception as e:
        print(f"  RO report could not run: {e}")
    print("=" * 60)

    if send_all:
        print("\nUploading new report(s) to the live PCL system ...")
        try:
            import subprocess
            subprocess.run([sys.executable, UPLOAD_SCRIPT, "--commit"], check=False)
        except Exception as e:
            print(f"  Upload step could not run: {e}")
    else:
        print("\nTest run (no) — skipping DB upload.")


if __name__ == "__main__":
    main()
