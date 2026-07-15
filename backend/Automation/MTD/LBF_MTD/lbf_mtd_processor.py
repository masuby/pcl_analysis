import pandas as pd
import numpy as np
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import glob
from datetime import datetime

# Configuration
# Define the base paths (dynamic: orchestrator sets PCL_MTD_DIR / PCL_AUTOMATION_ROOT;
# fallback = this file's grandparent, i.e. .../Automation/MTD).
from pathlib import Path as _Path
base_dir = os.environ.get("PCL_MTD_DIR") or (
    os.path.join(os.environ["PCL_AUTOMATION_ROOT"], "MTD")
    if os.environ.get("PCL_AUTOMATION_ROOT") else str(_Path(__file__).resolve().parents[1]))
input_folder = os.path.join(base_dir, "ROW_FILES")

# Output directory will be inside LBF_MTD folder
output_base = os.path.join(base_dir, "LBF_MTD")
output_dir = os.path.join(output_base, "split_supervisions_lbf")

# Automatically find the LBF MTD file
lbf_files = glob.glob(os.path.join(input_folder, "*LBF*MTD*.xlsx"))

if not lbf_files:
    # Try alternative patterns
    lbf_files = glob.glob(os.path.join(input_folder, "*LBF*.xlsx"))
    
if not lbf_files:
    raise FileNotFoundError(f"No LBF MTD file found in {input_folder}")

# Use the first matching file
input_file = lbf_files[0]
print(f"Found input file: {input_file}")

# Create output directories if they don't exist
if not os.path.exists(output_base):
    os.makedirs(output_base)
    print(f"Created base output directory: {output_base}")

if not os.path.exists(output_dir):
    os.makedirs(output_dir)
    print(f"Created split files output directory: {output_dir}")
else:
    print(f"Output directory already exists: {output_dir}")

# Create output directory
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# ========== DATE EXTRACTION FROM FILENAME ==========
# Extract date from input filename (parent file in ROW_FILES)
input_filename = os.path.basename(input_file)
print(f"Input filename (parent file): {input_filename}")

# Try multiple date patterns to extract date from parent filename
# Priority: patterns that match the actual filename format first
date_patterns = [
    # Pattern for "20th December 2025" or "20th DECEMBER 2025" (with spaces, case insensitive)
    r'(\d+)(?:st|nd|rd|th)\s+([A-Za-z]+)\s+(\d{4})',  # 20th December 2025
    # Pattern for "20TH-DECEMBER-2025" (with dashes, uppercase)
    r'(\d+)(?:ST|ND|RD|TH)-([A-Z]+)-(\d{4})',  # 22TH-DECEMBER-2025
    # Pattern for "20-DECEMBER-2025" (with dashes, no ordinal)
    r'(\d+)-([A-Z]+)-(\d{4})',                  # 22-DECEMBER-2025
    # Pattern for numeric dates
    r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',       # 22/12/2025 or 22-12-2025
    r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})',       # 2025/12/22 or 2025-12-22
]

day = None
month = None
year = None
date_info = None

for pattern_idx, pattern in enumerate(date_patterns):
    # Use case-insensitive search for first pattern, case-sensitive for others
    flags = re.IGNORECASE if pattern_idx == 0 else 0
    match = re.search(pattern, input_filename, flags)
    if match:
        if pattern_idx == 0:  # "20th December 2025" format (with spaces)
            day = match.group(1)
            month = match.group(2).upper()  # Convert to uppercase
            year = match.group(3)
        elif pattern_idx < 3:  # First three patterns: day-month-year (with dashes or spaces)
            day = match.group(1)
            month = match.group(2).upper()  # Ensure month is uppercase
            year = match.group(3)
        elif pattern_idx == 3:  # day/month/year (numeric)
            day = match.group(1)
            month_num = int(match.group(2))
            year = match.group(3)
            # Convert month number to month name
            month_names = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
                          "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]
            if 1 <= month_num <= 12:
                month = month_names[month_num - 1]
        else:  # year/month/day (numeric)
            year = match.group(1)
            month_num = int(match.group(2))
            day = match.group(3)
            # Convert month number to month name
            month_names = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
                          "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]
            if 1 <= month_num <= 12:
                month = month_names[month_num - 1]
        
        if day and month and year:
            # Format date info for filename: DAYTH-MONTH-YEAR-MTD
            date_info = f"{day}TH-{month}-{year}-MTD"
            print(f"✓ Extracted date from parent file: {day} {month} {year}")
            print(f"✓ Date format for filename: {date_info}")
            break

# If no date found, use current date as fallback
if not date_info:
    now = datetime.now()
    day = now.strftime('%d')
    month = now.strftime('%B').upper()
    year = now.strftime('%Y')
    date_info = f"{day}TH-{month}-{year}-MTD"
    print(f"⚠ Date not found in parent filename, using current date: {date_info}")

# Function to clean and normalize names for matching
def normalize_name(name):
    """Normalize name for better matching"""
    if pd.isna(name):
        return ""
    name_str = str(name).upper()
    # Remove common variations
    name_str = re.sub(r'\s+', ' ', name_str)  # Normalize spaces
    name_str = re.sub(r'LBF\s*', '', name_str)  # Remove 'LBF' prefix
    name_str = re.sub(r'\([^)]*\)', '', name_str)  # Remove text in parentheses
    name_str = name_str.replace('BRANCH', '').replace('REGION', '').strip()
    name_str = name_str.replace('  ', ' ').strip()
    return name_str

# Function to format numbers in DataFrame
def format_numbers(df):
    """Format numbers in DataFrame: 2 decimal places, percentage for values < 1, accounting format"""
    df_formatted = df.copy()
    
    for col in df_formatted.columns:
        # Check if column contains numeric data
        if df_formatted[col].dtype in ['float64', 'int64'] or any(isinstance(x, (int, float)) for x in df_formatted[col].dropna() if pd.notna(x)):
            for idx, value in df_formatted[col].items():
                if pd.notna(value) and isinstance(value, (int, float)):
                    # Convert to percentage if value is between 0 and 1 (exclusive)
                    if 0 < value < 1:
                        df_formatted.at[idx, col] = f"{value:.2%}"
                    else:
                        # Format with 2 decimal places and accounting format (comma separated)
                        df_formatted.at[idx, col] = f"{value:,.2f}"
    
    return df_formatted

# Step 1: Read and identify sheets
print("\nReading Excel file...")
xls = pd.ExcelFile(input_file)
sheet_names = xls.sheet_names

print(f"All sheets found: {sheet_names}")

# Find sheets (automatically)
mtd_sheet_name = None
ranking_sheet_name = None
listing_sheet_name = None

for sheet in sheet_names:
    sheet_upper = sheet.upper()
    if 'MTD' in sheet_upper:
        mtd_sheet_name = sheet
    elif 'RANKING' in sheet_upper:
        ranking_sheet_name = sheet
    elif 'LISTING' in sheet_upper:
        listing_sheet_name = sheet

# If not found by keywords, use positional approach
if not mtd_sheet_name and len(sheet_names) >= 1:
    mtd_sheet_name = sheet_names[0]
if not ranking_sheet_name and len(sheet_names) >= 2:
    ranking_sheet_name = sheet_names[1]
if not listing_sheet_name and len(sheet_names) >= 3:
    listing_sheet_name = sheet_names[2]

print(f"Using sheets: MTD='{mtd_sheet_name}', RANKING='{ranking_sheet_name}', LISTING='{listing_sheet_name}'")

# Step 2: Read all sheets
# Read MTD sheet - read without header first to get all raw data
try:
    mtd_raw = pd.read_excel(input_file, sheet_name=mtd_sheet_name, header=None)
    print(f"MTD sheet loaded successfully with {len(mtd_raw)} rows and {len(mtd_raw.columns)} columns")
    
except Exception as e:
    print(f"Error loading MTD sheet: {e}")
    exit()

# Read Sales Listing sheet
if listing_sheet_name:
    try:
        listing_raw = pd.read_excel(input_file, sheet_name=listing_sheet_name, header=None)
        # Find supervision column header row
        listing_header_row = 0
        for i in range(min(10, len(listing_raw))):
            if any('SUPERVISION' in str(val).upper() for val in listing_raw.iloc[i].astype(str).tolist()):
                listing_header_row = i
                break
        
        sales_listing_df = pd.read_excel(input_file, sheet_name=listing_sheet_name, header=listing_header_row)
        print(f"Sales Listing sheet loaded successfully with {len(sales_listing_df)} rows")
    except Exception as e:
        print(f"Error loading Sales Listing sheet: {e}")
        exit()
else:
    print("ERROR: No Sales Listing sheet found!")
    exit()

# Read Sales Ranking sheet
sales_ranking_df = None
ranking_raw = None
if ranking_sheet_name:
    try:
        ranking_raw = pd.read_excel(input_file, sheet_name=ranking_sheet_name, header=None)
        # Find header row
        ranking_header_row = 0
        for i in range(min(10, len(ranking_raw))):
            if any('SALES' in str(val).upper() and 'REPS' in str(val).upper() for val in ranking_raw.iloc[i].astype(str).tolist()):
                ranking_header_row = i
                break
        
        sales_ranking_df = pd.read_excel(input_file, sheet_name=ranking_sheet_name, header=ranking_header_row)
        print(f"Sales Ranking sheet loaded successfully with {len(sales_ranking_df)} rows")
    except Exception as e:
        print(f"Error loading Sales Ranking sheet: {e}")
        sales_ranking_df = None

# Step 3: Get unique supervisions from Sales Listing
supervision_column = None
for col in sales_listing_df.columns:
    if 'SUPERVISION' in str(col).upper():
        supervision_column = col
        break

if not supervision_column:
    print("ERROR: No 'Supervision' column found in sales listing!")
    # Try to find by position (usually one of the first columns)
    if len(sales_listing_df.columns) >= 9:  # Based on your data structure
        supervision_column = sales_listing_df.columns[8]  # Column I
        print(f"Using column {supervision_column} as supervision column")
    else:
        exit()

unique_supervisions = sales_listing_df[supervision_column].dropna().unique()
print(f"\nFound {len(unique_supervisions)} unique supervisions: {list(unique_supervisions)}")

# Step 5: STRICT SUPERVISION BOUNDARY DETECTION
print("\nIdentifying supervision boundaries in MTD sheet...")

# Find where the actual data starts (after headers)
data_start_row = None
for i in range(len(mtd_raw)):
    row_vals = mtd_raw.iloc[i].astype(str).tolist()
    # Look for the first row that has typical branch/team leader names
    if any('LBF' in val.upper() or 'CALL CENTER' in val.upper() for val in row_vals):
        data_start_row = i
        break

if data_start_row is None:
    data_start_row = 4  # Default to row 4 if not found

print(f"Data starts at row: {data_start_row}")

# STRICT APPROACH: First, identify all potential supervision rows
potential_supervision_rows = []
for i in range(data_start_row, len(mtd_raw)):
    first_col_value = str(mtd_raw.iloc[i, 0]) if pd.notna(mtd_raw.iloc[i, 0]) else ""
    
    if first_col_value.strip() == "":
        continue
        
    # Check if this matches any of our unique supervisions
    normalized_first_col = normalize_name(first_col_value)
    for supervision in unique_supervisions:
        normalized_supervision = normalize_name(supervision)
        # STRICT MATCHING: Only consider exact or very close matches
        if (normalized_first_col == normalized_supervision or 
            (normalized_supervision in normalized_first_col and len(normalized_supervision) > 5) or
            (normalized_first_col in normalized_supervision and len(normalized_first_col) > 5)):
            
            potential_supervision_rows.append({
                'row_index': i,
                'supervision': supervision,
                'original_name': first_col_value
            })
            print(f"Found potential supervision '{supervision}' at row {i} (original: '{first_col_value}')")
            break

# Remove duplicates - keep only the first occurrence of each supervision
unique_supervision_rows = {}
for item in potential_supervision_rows:
    supervision = item['supervision']
    if supervision not in unique_supervision_rows:
        unique_supervision_rows[supervision] = item
    else:
        # If we find a duplicate, keep the one with better match
        current_match = unique_supervision_rows[supervision]
        current_original = current_match['original_name']
        new_original = item['original_name']
        
        # Prefer exact matches
        if current_original.upper() != supervision.upper() and new_original.upper() == supervision.upper():
            unique_supervision_rows[supervision] = item
            print(f"  Replaced with better match for '{supervision}'")

# Create supervision sections with STRICT boundaries
supervision_sections = {}
supervision_row_indices = sorted([item['row_index'] for item in unique_supervision_rows.values()])

for i, supervision_item in enumerate(unique_supervision_rows.values()):
    supervision = supervision_item['supervision']
    start_row = supervision_item['row_index']
    
    # Determine end row: next supervision or end of data
    if i < len(supervision_row_indices) - 1:
        end_row = supervision_row_indices[i + 1]
    else:
        end_row = len(mtd_raw)
    
    # Collect team leaders in this section
    team_leaders = []
    for j in range(start_row + 1, end_row):
        first_col_value = str(mtd_raw.iloc[j, 0]) if pd.notna(mtd_raw.iloc[j, 0]) else ""
        if first_col_value.strip() != "" and j not in supervision_row_indices:
            team_leaders.append({
                'name': first_col_value,
                'row_index': j
            })
    
    supervision_sections[supervision] = {
        'start_row': start_row,
        'end_row': end_row,
        'team_leaders': team_leaders
    }

# Print summary with STRICT boundaries
print("\n=== FINAL SUPERVISION BOUNDARIES ===")
for supervision, section in supervision_sections.items():
    print(f"Supervision '{supervision}': rows {section['start_row']}-{section['end_row']-1}, {len(section['team_leaders'])} team leaders")

# Step 6: Process each supervision
for supervision in unique_supervisions:
    print(f"\nProcessing supervision: {supervision}")
    
    try:
        if supervision not in supervision_sections:
            print(f"  ⚠ No MTD data found for supervision: {supervision}")
            continue
        
        section = supervision_sections[supervision]
        
        # Verify the section is valid
        if section['start_row'] >= section['end_row']:
            print(f"  ⚠ Invalid section boundaries: {section['start_row']}-{section['end_row']}")
            continue
        
        # Get all row indices for this supervision section
        supervision_rows = list(range(section['start_row'], section['end_row']))
        
        # EXCLUDE GRAND TOTAL ROWS
        filtered_supervision_rows = []
        for row_idx in supervision_rows:
            first_col_value = str(mtd_raw.iloc[row_idx, 0]) if pd.notna(mtd_raw.iloc[row_idx, 0]) else ""
            if 'GRAND TOTAL' not in first_col_value.upper():
                filtered_supervision_rows.append(row_idx)
            else:
                print(f"  ⚠ Excluded Grand Total row: {row_idx}")
        
        if not filtered_supervision_rows:
            print(f"  No data remaining after excluding Grand Total for {supervision}")
            continue
        
        # Filter Sales Listing data for this supervision
        supervision_sales_data = sales_listing_df[sales_listing_df[supervision_column] == supervision].copy()
        
        if len(supervision_sales_data) == 0:
            print(f"  No sales data found for {supervision}")
            continue
        
        # Filter Sales Ranking data for this supervision (if available)
        supervision_ranking_data = pd.DataFrame()
        if sales_ranking_df is not None:
            # Find branch column in ranking
            branch_column_ranking = None
            for col in sales_ranking_df.columns:
                if 'BRANCH' in str(col).upper():
                    branch_column_ranking = col
                    break
            
            if branch_column_ranking:
                supervision_ranking_data = sales_ranking_df[sales_ranking_df[branch_column_ranking] == supervision].copy()
        
        print(f"  Including MTD rows: {filtered_supervision_rows[0]}-{filtered_supervision_rows[-1]}")
        print(f"  Team leaders found: {[tl['name'] for tl in section['team_leaders']]}")
        
        # FORMAT NUMBERS in all data
        # Create formatted MTD data (keep original values for column H to format as percentage)
        formatted_mtd_data = []
        original_mtd_data = []  # Store original numeric values for column H
        for row_idx in filtered_supervision_rows:
            row_data = mtd_raw.iloc[row_idx].values
            formatted_row = []
            original_row = []
            for col_idx, value in enumerate(row_data):
                if pd.notna(value) and isinstance(value, (int, float)):
                    # For column H (index 7, 0-based), keep original value for percentage formatting
                    if col_idx == 7:  # Column H is index 7 (0-based)
                        formatted_row.append(value)  # Keep original numeric value
                        original_row.append(value)
                    else:
                        # Convert to percentage if value is between 0 and 1 (exclusive)
                        if 0 < value < 1:
                            formatted_row.append(f"{value:.2%}")
                        else:
                            # Format with 2 decimal places and accounting format
                            formatted_row.append(f"{value:,.2f}")
                        original_row.append(value)
                else:
                    formatted_row.append(value)
                    original_row.append(value)
            formatted_mtd_data.append(formatted_row)
            original_mtd_data.append(original_row)
        
        # Format Sales Listing data
        supervision_sales_data_formatted = format_numbers(supervision_sales_data)
        
        # Format Sales Ranking data
        if len(supervision_ranking_data) > 0:
            supervision_ranking_data_formatted = format_numbers(supervision_ranking_data)
        else:
            supervision_ranking_data_formatted = pd.DataFrame()
        
        # Step 7: Create filename and save with STYLING
        supervision_clean = re.sub(r'[^\w\s-]', '', str(supervision))
        supervision_clean = supervision_clean.replace(' ', '-').replace('/', '-')
        
        # Clean the date_info for filename
        date_clean = date_info.replace(' ', '-').replace('/', '-')
        
        # Create filename with extracted date
        filename = f"LBF_{supervision_clean}_{date_clean}.xlsx"
        file_path = os.path.join(output_dir, filename)
        
        # Create Excel file with all available sheets and STYLING
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Create workbook
            workbook = writer.book
            
            # Define styles
            bold_font = Font(bold=True)
            orange_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Write MTD sheet with original header structure (rows 0-3) + supervision data
            mtd_sheet = workbook.create_sheet(mtd_sheet_name, 0)
            current_row = 1
            
            # Write header rows for MTD (first 4 rows) with WHITE background
            for i in range(4):  # Include rows 0, 1, 2, 3 as headers
                for col_idx, value in enumerate(mtd_raw.iloc[i].values, 1):
                    cell = mtd_sheet.cell(row=current_row, column=col_idx, value=value)
                    cell.font = bold_font
                    cell.fill = white_fill
                current_row += 1
            
            # Write the filtered MTD data (supervision section) with ORANGE background and FORMATTED NUMBERS
            for row_idx, row_data in enumerate(formatted_mtd_data):
                original_row = original_mtd_data[row_idx] if row_idx < len(original_mtd_data) else row_data
                for col_idx, value in enumerate(row_data, 1):
                    # For column H, use original numeric value
                    if col_idx == 8 and row_idx < len(original_mtd_data):
                        original_value = original_row[col_idx - 1]  # col_idx is 1-based, array is 0-based
                        cell = mtd_sheet.cell(row=current_row, column=col_idx, value=original_value)
                    else:
                        cell = mtd_sheet.cell(row=current_row, column=col_idx, value=value)
                    cell.font = bold_font
                    cell.fill = orange_fill
                    
                    # Format column H (column 8) as percentage
                    if col_idx == 8:
                        if row_idx < len(original_mtd_data):
                            original_value = original_row[col_idx - 1]
                            if isinstance(original_value, (int, float)) and pd.notna(original_value):
                                cell.number_format = "0.00%"
                current_row += 1
        
            
            # Write Sales Listing sheet
            listing_sheet = workbook.create_sheet(listing_sheet_name, 2)
            current_row_listing = 1
            
            # Write header rows for listing with WHITE background
            for i in range(listing_header_row + 1):
                for col_idx, value in enumerate(listing_raw.iloc[i].values, 1):
                    cell = listing_sheet.cell(row=current_row_listing, column=col_idx, value=value)
                    cell.font = bold_font
                    cell.fill = white_fill
                current_row_listing += 1
            
            # Write the filtered listing data with ORANGE background and FORMATTED NUMBERS
            for r_idx, row in enumerate(dataframe_to_rows(supervision_sales_data_formatted, index=False, header=False), current_row_listing):
                for col_idx, value in enumerate(row, 1):
                    cell = listing_sheet.cell(row=r_idx, column=col_idx, value=value)

                    cell.fill = orange_fill
            
            # Auto-adjust column widths for all sheets
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"  ✓ Created: {filename}")
        
    except Exception as e:
        print(f"  ✗ Error processing {supervision}: {str(e)}")
        import traceback
        traceback.print_exc()

print(f"\n✅ LBF Processing complete! Files saved to: {output_dir}")
print(f"Total supervisions processed: {len(unique_supervisions)}")


# =============================================================================
# END PROCESS: upload the freshly generated MTD report(s) to the live PCL system.
# Runs direct_upload_files_to_db.py --commit, which logs in over SSH, dedups by
# (department, type, date) and uploads only files not already in the DB.
# =============================================================================
import subprocess as _up_subprocess
import sys as _up_sys
_UPLOAD_SCRIPT = os.path.join(
    os.environ.get("PCL_AUTOMATION_ROOT")
    or os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "CRM", "direct_upload_files_to_db.py")
print("\n" + "=" * 78)
print("END PROCESS: uploading new report(s) to the live PCL system ...")
print("=" * 78)
try:
    _up_subprocess.run([_up_sys.executable, _UPLOAD_SCRIPT, "--commit"], check=False)
except Exception as _up_err:
    print(f"  Upload step could not run: {_up_err}")
    print(f'  You can upload manually:  python "{_UPLOAD_SCRIPT}" --commit')