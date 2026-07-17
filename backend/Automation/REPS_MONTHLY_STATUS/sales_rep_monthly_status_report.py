"""
Build Sales Rep monthly sales status report from Loan.xlsx and Users.xlsx (product_mapping).

Input files (default):
  Loan:  .../Management/ROW_FILES/Loan.xlsx  (sheet: Loan Accounts)
  Users: .../Management/ROW_FILES/Users.xlsx  (sheet: product_mapping)
  Zone & cluster: .../Management/Zone and cluster.xlsx — maps Branch → Product (used as **Main product**)

Output:
  C:\\Users\\Daniel\\Desktop\\Management\\Sales_Reps_Monthly_Status_YYYY-MM.xlsx

Optional email (after save): prompts to send to all managers or developer only.
  Configure ``Management/.env`` with ``EMAIL_USERNAME`` and ``APP_PASSWORD`` (or ``EMAIL_PASSWORD`` /
  ``app_password``). Optional: ``SMTP_HOST``, ``SMTP_PORT``, ``DEVELOPER_EMAIL``.

Weeks are **calendar weeks of the month**: week 1 = days 1–7, week 2 = 8–14, etc.

**Eligible weeks** (which “Sold week N” sheets exist, and summary week columns):
  - Selected month **before** today’s month → all week buckets in that month.
  - Selected month **equals** today’s month → only weeks that have **started** on or before
    today’s calendar day (e.g. 20 Mar 2026 → weeks 1, 2, 3).
  - Selected month **after** today’s month → no eligible weeks (no per-week sheets).

**“Sold all weeks” qualification** (``Sold all weeks`` sheet & related counts) uses only **completed**
eligible weeks: a week counts only after **today ≥ last day** of that week’s calendar bucket in the month.
The **in-progress** week (e.g. week 4 before it ends) is **not** required yet, so reps who **sold** in all **past**
completed weeks still qualify.

Sheets:
  - Summary — first sheet; slate→teal banner; distinct main products table with weekly
    columns, **Sold all-weeks** reps, **Active reps (as of date)** (distinct reps with ≥1 sale in that main product in the month), and totals.
  - Sold all weeks — reps with ≥1 activation in **every completed** eligible week (see above).
  - Sold week 1, Sold week 2, … — one sheet per eligible week (filtered by Activation Date).
  - All month — all loans in the selected calendar month.
  - By product — sections by **main product** (no freeze: multiple blocks per sheet).
"""

from __future__ import annotations

import calendar
import html
import os
import re
import smtplib
import ssl
from collections import defaultdict
from datetime import date
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Any, Dict, List, Sequence, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, GradientFill, PatternFill, Side
from openpyxl.utils import get_column_letter
# --- Paths (dynamic: orchestrator sets PCL_AUTOMATION_ROOT; fallbacks derive
#     from this script's location so it also runs standalone) ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_AUTOMATION_ROOT = os.environ.get("PCL_AUTOMATION_ROOT") or os.path.dirname(SCRIPT_DIR)


def _first_existing(*paths):
    for p in paths:
        if p and os.path.isfile(p):
            return p
    return paths[0]


DEFAULT_LOAN_PATH = os.path.join(SCRIPT_DIR, "ROW_FILES", "Loan.xlsx")
DEFAULT_USERS_PATH = os.path.join(SCRIPT_DIR, "ROW_FILES", "Users.xlsx")
# Zone & cluster + .env may live beside this script OR be shared from Management/CRM.
DEFAULT_ZONE_CLUSTER_PATH = _first_existing(
    os.path.join(SCRIPT_DIR, "Zone and cluster.xlsx"),
    os.path.join(_AUTOMATION_ROOT, "Management", "Zone and cluster.xlsx"))
ENV_FILE_PATH = _first_existing(
    os.path.join(SCRIPT_DIR, ".env"),
    os.path.join(_AUTOMATION_ROOT, "CRM", ".env"))
OUTPUT_DIR = os.environ.get("PCL_REPS_OUT", os.path.join(SCRIPT_DIR, "OUTPUT"))

# --- Email: distribution list (edit freely — add/remove one line per address) ---
MANAGER_RECIPIENTS: List[str] = [
    "daniel@platinumcredit.co.tz",
    "sigfrid@platinumcredit.co.tz",
    "augustine@platinumcredit.co.tz",
    "doris@platinumcredit.co.tz",
    "abdulhakim.khalfan@platinumcredit.co.tz",
    "fragrance@platinumcredit.co.tz",
    "mohamedi.omar.platinum@gmail.com",
    "allan@platinumcredit.co.tz",
    "raphael@platinumcredit.co.tz",
    "kelvin.mwasala@platinumcredit.co.tz",
    "denis.albert@platinumcredit.co.tz",
    "vivian.karatta.platinum@gmail.com",
    "madina.mohamed.platinum@gmail.com",
    "irene.mmari@platinumcredit.co.tz",
]


# Recipients can be managed from the web via master_emails_reps.xlsx (column
# "Managers"). If it exists it OVERRIDES the list above (flow-only change).
def _load_recipient_master():
    try:
        from openpyxl import load_workbook
        mp = os.path.join(SCRIPT_DIR, "master_emails_reps.xlsx")
        if not os.path.isfile(mp):
            return
        wb = load_workbook(mp, read_only=True, data_only=True)
        ws = wb["Recipients"] if "Recipients" in wb.sheetnames else wb.active
        col = next((c.column for c in ws[1] if str(c.value).strip().lower() == "managers"), None)
        if col:
            emails = [str(ws.cell(row=r, column=col).value).strip()
                      for r in range(2, ws.max_row + 1)
                      if ws.cell(row=r, column=col).value and "@" in str(ws.cell(row=r, column=col).value)]
            if emails:
                MANAGER_RECIPIENTS[:] = emails
        wb.close()
    except Exception as _e:
        print(f"  (reps recipient master not loaded: {_e})")


_load_recipient_master()
DEFAULT_SMTP_HOST = "smtp.gmail.com"
DEFAULT_SMTP_PORT = 587

OUTPUT_COLUMNS = [
    "Sales Rep",
    "Team Leader",
    "Client name",
    "Branch",
    "Activation Date (Loan)",
    "CLIENT TYPE",
    "Product",
    "Main product",
    "Loan Amount",
]

# 1-based column positions, derived so they auto-adjust if OUTPUT_COLUMNS changes.
OUT_AMT_COL = OUTPUT_COLUMNS.index("Loan Amount") + 1          # in OUTPUT-only layout
OUT_DATE_COL = OUTPUT_COLUMNS.index("Activation Date (Loan)") + 1
GRP_AMT_COL = OUT_AMT_COL + 1                                  # grouped layout has a leading '#'
GRP_DATE_COL = OUT_DATE_COL + 1

HEADER_ROW_HEIGHT = 22
DATA_ROW_HEIGHT = 18

# --- Visual theme: one simple blue everywhere (banners, headers, totals). ---
# Only the Summary data tables use different colours per column (see
# _SUMMARY_TABLE_HEADER_FILLS) — that multi-colour header is intentional.
BLUE = "1F4E79"          # the single blue used across the report
BLUE_DARK = "163A5C"
BLUE_LIGHT = "DBEAFE"    # light fill (totals / zebra-a)
BLUE_TINT = "EFF6FF"
BLUE_GRID = "BFDBFE"

SUMMARY_TITLE_FILL = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
SUMMARY_TABLE_BAND = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
SUMMARY_LABEL_FONT = Font(bold=True, size=10, color="334155")
SUMMARY_VALUE_FONT = Font(size=10, color="475569")
SUMMARY_SECTION_FONT = Font(bold=True, size=11, color="FFFFFF")

# Distinct categories: teal / slate / amber accents (readable, not rainbow)
_SUMMARY_TABLE_HEADER_FILLS: List[Tuple[str, bool]] = [
    ("0F766E", False),
    ("115E59", False),
    ("134E4A", False),
    ("0E7490", False),
    ("155E75", False),
    ("164E63", False),
    ("0C4A6E", False),
    ("1E3A5F", False),
    ("B45309", True),
    ("D97706", True),
    ("0F172A", False),
    ("1E293B", False),
]

HEADER_BOTTOM_BORDER = Side(style="medium", color="0F172A")
HEADER_INNER_BORDER = Side(style="thin", color="E2E8F0")

# Per-sheet grouped header rows: (bg_hex, dark_font). Each sheet uses one theme.
_GROUPED_HEADERS_OCEAN: List[Tuple[str, bool]] = [
    ("0C4A6E", False),
    ("0E7490", False),
    ("0D9488", False),
    ("14B8A6", False),
    ("CCFBF1", True),
    ("115E59", False),
    ("0F766E", False),
    ("134E4A", False),
    ("FBBF24", True),
]
_GROUPED_HEADERS_INDIGO: List[Tuple[str, bool]] = [
    ("312E81", False),
    ("3730A3", False),
    ("4338CA", False),
    ("4F46E5", False),
    ("E0E7FF", True),
    ("3730A3", False),
    ("4338CA", False),
    ("4F46E5", False),
    ("F59E0B", True),
]
_GROUPED_HEADERS_SLATE: List[Tuple[str, bool]] = [
    ("1E293B", False),
    ("334155", False),
    ("475569", False),
    ("64748B", False),
    ("F1F5F9", True),
    ("334155", False),
    ("475569", False),
    ("64748B", False),
    ("EA580C", True),
]
_OUTPUT_HEADERS_BY_PRODUCT: List[Tuple[str, bool]] = [
    ("78350F", False),
    ("92400E", False),
    ("B45309", False),
    ("D97706", False),
    ("FFFBEB", True),
    ("9A3412", False),
    ("C2410C", False),
    ("EA580C", False),
    ("FCD34D", True),
]

GROUPED_COLUMNS = ["#"] + OUTPUT_COLUMNS  # left count column + detail columns

# Zebra, grid, totals, rep separators — one simple blue palette for every sheet.
_BLUE_THEME: Dict[str, Any] = {
    "zebra_a": BLUE_TINT,
    "zebra_b": "FFFFFF",
    "grid": "CBD5E1",
    "total_fill": BLUE_LIGHT,
    "total_border": BLUE,
    "total_font": BLUE_DARK,
    "sep": BLUE,
    "rep_font": BLUE,
}
_SHEET_THEME: Dict[str, Dict[str, Any]] = {
    "ocean": dict(_BLUE_THEME),
    "indigo": dict(_BLUE_THEME),
    "slate": dict(_BLUE_THEME),
    "by_product": dict(_BLUE_THEME),
}


def _main_products_sorted_list(display_base: pd.DataFrame) -> List[str]:
    """Sorted distinct **main product** values (for Summary and weekly breakdown)."""
    return sorted(
        display_base["Main product"].fillna("").replace("", "(Unmapped)").unique(),
        key=lambda x: str(x).lower(),
    )


def _norm_key(val: Any) -> str:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return " ".join(str(val).split())


def _parse_month_year(text: str) -> tuple[int, int]:
    text = text.strip()
    m = re.match(r"^(\d{1,2})\s*[/\-.]\s*(\d{4})$", text)
    if not m:
        raise ValueError(f"Invalid format: {text!r}. Use MM/YYYY e.g. 02/2026")
    month, year = int(m.group(1)), int(m.group(2))
    if not 1 <= month <= 12:
        raise ValueError("Month must be 1-12")
    return year, month


def _calendar_week_of_month(day: int) -> int:
    """1-based week index within the month: days 1–7 → 1, 8–14 → 2, …"""
    return (int(day) - 1) // 7 + 1


def _max_calendar_weeks_in_month(year: int, month: int) -> int:
    _, last = calendar.monthrange(year, month)
    return (last - 1) // 7 + 1


def _eligible_calendar_weeks(year: int, month: int, today: date) -> list[int]:
    """
    Week numbers (calendar-of-month) to include in this report run.
    Future selected month → no weeks.
    Past month → all weeks in that month.
    Current month → weeks whose first day (1, 8, 15, …) is on or before today's calendar day.
    """
    if (year, month) > (today.year, today.month):
        return []
    max_w = _max_calendar_weeks_in_month(year, month)
    if (year, month) < (today.year, today.month):
        return list(range(1, max_w + 1))

    _, last = calendar.monthrange(year, month)
    cutoff_day = min(today.day, last)
    eligible: list[int] = []
    for w in range(1, max_w + 1):
        first_day_of_bucket = (w - 1) * 7 + 1
        if first_day_of_bucket <= cutoff_day:
            eligible.append(w)
    return eligible


def _week_bucket_end_date(year: int, month: int, week_num: int) -> date:
    """Last calendar day of week bucket ``week_num`` in that month (week 1 = days 1–7, …)."""
    _, last = calendar.monthrange(year, month)
    d_end = min(week_num * 7, last)
    return date(year, month, d_end)


def _completed_eligible_weeks_for_sold(
    year: int, month: int, today: date, eligible_list: list[int]
) -> list[int]:
    """Eligible weeks that are **finished** as of ``today`` — used to qualify “sold all weeks”.

    While the current week is still open (e.g. week 4 before its last day), reps are not required to have
    a sale in that week yet; they qualify if they have a sale in every **earlier** eligible week whose
    bucket has ended. For a **past** report month, all eligible weeks count as completed.
    """
    if not eligible_list:
        return []
    if (year, month) < (today.year, today.month):
        return list(eligible_list)
    if (year, month) > (today.year, today.month):
        return []
    out: list[int] = []
    for w in eligible_list:
        end_d = _week_bucket_end_date(year, month, w)
        if today >= end_d:
            out.append(w)
    return out


def _load_product_map(users_path: str) -> Tuple[dict[str, str], dict[str, str]]:
    df = pd.read_excel(users_path, sheet_name="product_mapping", engine="openpyxl")
    if "Loan Name" not in df.columns or "Products" not in df.columns:
        raise KeyError("product_mapping must have columns 'Loan Name' and 'Products'")

    exact: dict[str, str] = {}
    lower: dict[str, str] = {}

    for _, row in df.iterrows():
        k = _norm_key(row.get("Loan Name"))
        if not k:
            continue
        prod = row.get("Products")
        pv = "" if prod is None or (isinstance(prod, float) and pd.isna(prod)) else str(prod).strip()
        exact[k] = pv
        lower[k.lower()] = pv
    return exact, lower


def _load_team_leader_map(users_path: str) -> dict[str, str]:
    """From Users.xlsx sheet 'Users': map a sales agent (Display Name) -> their
    Team (TEAM column = the team leader the agent reports to). Keyed on the
    normalised, lower-cased agent name for a VLOOKUP-style match against Sales Rep."""
    df = pd.read_excel(users_path, sheet_name="Users", engine="openpyxl")
    cols = {str(c).strip().lower(): c for c in df.columns}
    team_col = cols.get("team")
    name_col = cols.get("display name") or cols.get("sales agent") or cols.get("name")
    if team_col is None or name_col is None:
        raise KeyError("Users sheet must have 'TEAM' and 'Display Name' columns")
    out: dict[str, str] = {}
    for _, row in df.iterrows():
        agent = _norm_key(row.get(name_col))
        if not agent:
            continue
        tl = row.get(team_col)
        tl = "" if tl is None or (isinstance(tl, float) and pd.isna(tl)) else str(tl).strip()
        out.setdefault(agent.lower(), tl)
    return out


def _lookup_product(loan_name: Any, exact: dict[str, str], lower: dict[str, str]) -> str:
    k = _norm_key(loan_name)
    if not k:
        return ""
    if k in exact:
        return exact[k]
    lk = k.lower()
    if lk in lower:
        return lower[lk]
    return ""


def _load_branch_main_product_map(zone_path: str) -> Tuple[dict[str, str], dict[str, str]]:
    """Branch → main product from Zone and cluster.xlsx (columns Branch, Product). First match wins."""
    df = pd.read_excel(zone_path, sheet_name=0, engine="openpyxl")
    if "Branch" not in df.columns or "Product" not in df.columns:
        raise KeyError("Zone and cluster file must have columns 'Branch' and 'Product'")

    exact: dict[str, str] = {}
    lower: dict[str, str] = {}
    for _, row in df.iterrows():
        b = _norm_key(row.get("Branch"))
        if not b:
            continue
        raw_p = row.get("Product")
        pv = (
            ""
            if raw_p is None or (isinstance(raw_p, float) and pd.isna(raw_p))
            else str(raw_p).strip()
        )
        if b not in exact:
            exact[b] = pv
        bl = b.lower()
        if bl not in lower:
            lower[bl] = pv
    return exact, lower


def _lookup_main_product(
    branch: Any,
    exact: dict[str, str],
    lower: dict[str, str],
) -> str:
    b = _norm_key(branch)
    if not b:
        return ""
    if b in exact:
        return exact[b]
    bl = b.lower()
    if bl in lower:
        return lower[bl]
    return ""


def _prepare_dataframe(
    loan_path: str,
    users_path: str,
    zone_cluster_path: str,
    year: int,
    month: int,
) -> pd.DataFrame:
    df = pd.read_excel(loan_path, sheet_name=0, engine="openpyxl")

    need = [
        "Loan Name",
        "Account Holder Name",
        "Branch",
        "Activation Date (Loan)",
        "CLIENT TYPE",
        "Sales Reps (Client)",
        "Loan Amount",
    ]
    missing = [c for c in need if c not in df.columns]
    if missing:
        raise KeyError(f"Loan sheet missing columns: {missing}")

    exact, lower = _load_product_map(users_path)
    br_exact, br_lower = _load_branch_main_product_map(zone_cluster_path)
    tl_map = _load_team_leader_map(users_path)

    df = df.copy()
    df["_act_dt"] = pd.to_datetime(df["Activation Date (Loan)"], errors="coerce")
    df = df[df["_act_dt"].notna()]
    df = df[(df["_act_dt"].dt.year == year) & (df["_act_dt"].dt.month == month)]

    out = pd.DataFrame(
        {
            "Sales Rep": df["Sales Reps (Client)"].apply(lambda x: _norm_key(x) or x),
            "Team Leader": df["Sales Reps (Client)"].apply(
                lambda x: tl_map.get(_norm_key(x).lower(), "")
            ),
            "Client name": df["Account Holder Name"],
            "Branch": df["Branch"],
            "Activation Date (Loan)": df["_act_dt"].dt.strftime("%d/%m/%Y"),
            "CLIENT TYPE": df["CLIENT TYPE"],
            "Product": df["Loan Name"].apply(lambda ln: _lookup_product(ln, exact, lower)),
            "Main product": df["Branch"].apply(
                lambda br: _lookup_main_product(br, br_exact, br_lower)
            ),
            "Loan Amount": df["Loan Amount"],
        }
    )

    out["_cal_week"] = df["_act_dt"].apply(
        lambda ts: _calendar_week_of_month(ts.day) if pd.notna(ts) else None
    )
    out["_rep_key"] = out["Sales Rep"].apply(lambda x: _norm_key(x))

    return out


def _reps_all_weeks(out: pd.DataFrame, eligible_weeks: set[int]) -> set[str]:
    """Reps with at least one sale in every week in ``eligible_weeks`` (typically completed weeks only)."""
    if not eligible_weeks:
        return set()
    rep_weeks: dict[str, set[int]] = defaultdict(set)
    for _, row in out.iterrows():
        rk = row["_rep_key"]
        if not rk:
            continue
        cw = row["_cal_week"]
        if cw is not None and int(cw) in eligible_weeks:
            rep_weeks[rk].add(int(cw))
    qualified: set[str] = set()
    for rk, wset in rep_weeks.items():
        if eligible_weeks <= wset:
            qualified.add(rk)
    return qualified


def _fill_header_row(ws, row: int, num_cols: int) -> None:
    """One simple blue header row across ``num_cols`` columns."""
    blue = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = blue
        cell.border = Border(
            left=HEADER_INNER_BORDER,
            right=HEADER_INNER_BORDER,
            top=HEADER_INNER_BORDER,
            bottom=HEADER_BOTTOM_BORDER,
        )


def _style_header_grouped(ws, row: int, *, theme: str = "ocean") -> None:
    """Column headers for sheets using GROUPED_COLUMNS (# + OUTPUT_COLUMNS)."""
    _fill_header_row(ws, row, len(GROUPED_COLUMNS))


def _style_header_output(ws, row: int) -> None:
    """Column headers for sheets using OUTPUT_COLUMNS only (e.g. By product)."""
    _fill_header_row(ws, row, len(OUTPUT_COLUMNS))


def _style_summary_table_headers(ws, row: int, num_cols: int) -> None:
    """Distinct colours per column for Summary data tables (cycles if many columns)."""
    pal = _SUMMARY_TABLE_HEADER_FILLS
    for col in range(1, num_cols + 1):
        bg, dark_font = pal[(col - 1) % len(pal)]
        cell = ws.cell(row=row, column=col)
        fg = "263238" if dark_font else "FFFFFF"
        cell.font = Font(bold=True, color=fg, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        cell.border = Border(
            left=HEADER_INNER_BORDER,
            right=HEADER_INNER_BORDER,
            top=HEADER_INNER_BORDER,
            bottom=HEADER_BOTTOM_BORDER,
        )


def _style_data_area(
    ws,
    start_row: int,
    end_row: int,
    max_col: int,
    *,
    loan_amount_col: int = 8,
    date_cols: Tuple[int, ...] = (4,),
    theme: str = "ocean",
) -> None:
    t = _SHEET_THEME[theme]
    thin = Side(style="thin", color=t["grid"])
    alt_fill = PatternFill(start_color=t["zebra_a"], end_color=t["zebra_a"], fill_type="solid")
    white = PatternFill(start_color=t["zebra_b"], end_color=t["zebra_b"], fill_type="solid")
    for r in range(start_row, end_row + 1):
        fill = alt_fill if (r - start_row) % 2 == 1 else white
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.font = Font(size=10, color="334155")
            if c == loan_amount_col:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif c in date_cols:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")


def _auto_width(ws, max_col: int, min_width: float = 10, max_width: float = 48) -> None:
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_len + 2, max_width))


def _apply_rep_separator_row(ws, row: int, max_col: int, *, theme: str) -> None:
    t = _SHEET_THEME[theme]
    thin = Side(style="thin", color=t["grid"])
    bottom = Side(style="medium", color=t["sep"])
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=bottom)


def _write_grouped_rep_sheet(
    wb: openpyxl.Workbook,
    title: str,
    df_display: pd.DataFrame,
    *,
    freeze: str = "A2",
    theme: str = "ocean",
) -> None:
    """Rows sorted by Sales Rep A–Z; col # shows rep ordinal on first row of each rep; separator between reps."""
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = True
    max_col = len(GROUPED_COLUMNS)

    for col, name in enumerate(GROUPED_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)

    _style_header_grouped(ws, 1, theme=theme)
    ws.freeze_panes = freeze
    ws.row_dimensions[1].height = max(HEADER_ROW_HEIGHT, 26)
    th = _SHEET_THEME[theme]

    if df_display.empty:
        _auto_width(ws, max_col)
        return

    df2 = df_display.sort_values(
        "Sales Rep",
        key=lambda s: s.fillna("").astype(str).str.lower(),
        kind="stable",
    )
    rep_list = sorted(df2["Sales Rep"].fillna("").astype(str).unique(), key=lambda x: str(x).lower())
    rep_index = {name: i + 1 for i, name in enumerate(rep_list)}

    row_idx = 2
    prev_rep: str | None = None
    amounts: List[float] = []
    for _, rec in df2.iterrows():
        rep = str(rec["Sales Rep"])
        num_cell = rep_index[rep] if rep != prev_rep else ""
        prev_rep = rep
        ws.cell(row=row_idx, column=1, value=num_cell)
        for c_idx, col_name in enumerate(OUTPUT_COLUMNS, start=2):
            ws.cell(row=row_idx, column=c_idx, value=rec[col_name])
        raw_amt = rec["Loan Amount"]
        try:
            amounts.append(float(raw_amt) if raw_amt is not None and str(raw_amt) != "" else 0.0)
        except (TypeError, ValueError):
            amounts.append(0.0)
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT
        row_idx += 1

    last_data = row_idx - 1
    if last_data >= 2:
        _style_data_area(
            ws, 2, last_data, max_col, loan_amount_col=GRP_AMT_COL, date_cols=(GRP_DATE_COL,), theme=theme
        )
        rep_c = th["rep_font"]
        for r in range(2, last_data + 1):
            for c in (2, 3):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(size=10, color=rep_c, bold=(c == 2))
            c1 = ws.cell(row=r, column=1)
            if c1.value not in ("", None):
                c1.font = Font(bold=True, size=11, color=rep_c)
                c1.alignment = Alignment(horizontal="center", vertical="center")

        boundaries: List[int] = []
        r = 2
        while r <= last_data:
            rv = ws.cell(row=r, column=2).value
            while r <= last_data and ws.cell(row=r, column=2).value == rv:
                r += 1
            boundaries.append(r - 1)
        for br in boundaries:
            _apply_rep_separator_row(ws, br, max_col, theme=theme)

    total_amt = sum(amounts)
    tr = last_data + 1
    ws.row_dimensions[tr].height = 22
    tot_fill = PatternFill(start_color=th["total_fill"], end_color=th["total_fill"], fill_type="solid")
    thin_tot = Side(style="thin", color=th["total_border"])
    bold_border = Border(
        left=thin_tot,
        right=thin_tot,
        top=Side(style="medium", color=th["total_border"]),
        bottom=thin_tot,
    )
    ws.cell(row=tr, column=1, value="")
    ws.cell(row=tr, column=2, value="TOTALS")
    for c in range(3, GRP_AMT_COL):
        ws.cell(row=tr, column=c, value="")
    ws.cell(row=tr, column=GRP_AMT_COL, value=total_amt)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=tr, column=c)
        cell.fill = tot_fill
        cell.border = bold_border
        cell.font = Font(bold=True, size=11, color=th["total_font"])
        if c == GRP_AMT_COL:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    note = ws.cell(row=tr + 1, column=2, value=f"Loan rows in sheet: {len(df2)}")
    note.font = Font(italic=True, size=9, color="666666")

    _auto_width(ws, max_col)


def _write_all_month_sheet(wb: openpyxl.Workbook, df_display: pd.DataFrame) -> None:
    """All month: A–Z by rep, # column, separators between reps, totals row (slate theme)."""
    theme = "slate"
    ws = wb.create_sheet("All month")
    ws.sheet_view.showGridLines = True
    max_col = len(GROUPED_COLUMNS)

    for col, name in enumerate(GROUPED_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)

    _style_header_grouped(ws, 1, theme=theme)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = max(HEADER_ROW_HEIGHT, 26)
    th = _SHEET_THEME[theme]

    if df_display.empty:
        _auto_width(ws, max_col)
        return

    df2 = df_display.sort_values(
        "Sales Rep",
        key=lambda s: s.fillna("").astype(str).str.lower(),
        kind="stable",
    )
    rep_list = sorted(df2["Sales Rep"].fillna("").astype(str).unique(), key=lambda x: str(x).lower())
    rep_index = {name: i + 1 for i, name in enumerate(rep_list)}

    row_idx = 2
    prev_rep: str | None = None
    amounts: List[float] = []
    for _, rec in df2.iterrows():
        rep = str(rec["Sales Rep"])
        num_cell = rep_index[rep] if rep != prev_rep else ""
        prev_rep = rep
        ws.cell(row=row_idx, column=1, value=num_cell)
        for c_idx, col_name in enumerate(OUTPUT_COLUMNS, start=2):
            ws.cell(row=row_idx, column=c_idx, value=rec[col_name])
        raw_amt = rec["Loan Amount"]
        try:
            amounts.append(float(raw_amt) if raw_amt is not None and str(raw_amt) != "" else 0.0)
        except (TypeError, ValueError):
            amounts.append(0.0)
        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT
        row_idx += 1

    last_data = row_idx - 1
    if last_data >= 2:
        _style_data_area(
            ws, 2, last_data, max_col, loan_amount_col=GRP_AMT_COL, date_cols=(GRP_DATE_COL,), theme=theme
        )
        rep_c = th["rep_font"]
        for r in range(2, last_data + 1):
            for c in (2, 3):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(size=10, color=rep_c, bold=(c == 2))
            c1 = ws.cell(row=r, column=1)
            if c1.value not in ("", None):
                c1.font = Font(bold=True, size=11, color=rep_c)
                c1.alignment = Alignment(horizontal="center", vertical="center")

        boundaries: List[int] = []
        r = 2
        while r <= last_data:
            rv = ws.cell(row=r, column=2).value
            while r <= last_data and ws.cell(row=r, column=2).value == rv:
                r += 1
            boundaries.append(r - 1)
        for br in boundaries:
            _apply_rep_separator_row(ws, br, max_col, theme=theme)

    total_amt = sum(amounts)
    tr = last_data + 1
    ws.row_dimensions[tr].height = 22
    tot_fill = PatternFill(start_color=th["total_fill"], end_color=th["total_fill"], fill_type="solid")
    thin = Side(style="thin", color=th["total_border"])
    bold_border = Border(
        left=thin, right=thin, top=Side(style="medium", color=th["total_border"]), bottom=thin
    )
    ws.cell(row=tr, column=1, value="")
    ws.cell(row=tr, column=2, value="TOTALS")
    for c in range(3, GRP_AMT_COL):
        ws.cell(row=tr, column=c, value="")
    ws.cell(row=tr, column=GRP_AMT_COL, value=total_amt)
    for c in range(1, max_col + 1):
        cell = ws.cell(row=tr, column=c)
        cell.fill = tot_fill
        cell.border = bold_border
        cell.font = Font(bold=True, size=11, color=th["total_font"])
        if c == GRP_AMT_COL:
            cell.number_format = "#,##0.00"
            cell.alignment = Alignment(horizontal="right", vertical="center")
        elif c == 2:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    note = ws.cell(row=tr + 1, column=2, value=f"Loan rows in month: {len(df2)}")
    note.font = Font(italic=True, size=9, color="666666")

    _auto_width(ws, max_col)


def _week_day_span_label(year: int, month: int, week_num: int) -> str:
    _, last = calendar.monthrange(year, month)
    d0 = (week_num - 1) * 7 + 1
    d1 = min(week_num * 7, last)
    return f"Days {d0}–{d1}"


CRM_NEW_EXCEL_DIR = os.environ.get(
    "PCL_CRM_CS_NEW_EXCEL",
    os.path.join(_AUTOMATION_ROOT, "CRM", "CS", "NEW_EXCEL"))
CRM_ROLES = ["loan officer", "call center agent", "branch loan officer", "branch manager"]
_CRM_FILE_RE = re.compile(r"cs_crm_(\d{2})_(\d{2})_(\d{4})\.xlsx$", re.IGNORECASE)
_CRM_COUNT_CACHE: Dict[str, Dict[str, int]] = {}


def _crm_files_by_date() -> List[Tuple[date, str]]:
    """All CS_CRM_DD_MM_YYYY.xlsx files in CRM_NEW_EXCEL_DIR with their parsed
    report date, sorted ascending by date."""
    out: List[Tuple[date, str]] = []
    try:
        names = os.listdir(CRM_NEW_EXCEL_DIR)
    except FileNotFoundError:
        return out
    for f in names:
        if f.startswith("~$"):
            continue
        m = _CRM_FILE_RE.search(f)
        if not m:
            continue
        dd, mm, yyyy = (int(x) for x in m.groups())
        try:
            out.append((date(yyyy, mm, dd), os.path.join(CRM_NEW_EXCEL_DIR, f)))
        except ValueError:
            continue
    out.sort(key=lambda t: t[0])
    return out


def _crm_counts_from_file(path: str) -> Dict[str, int]:
    """Distinct reps per PRODUCT (upper-case) in one CRM file's 'user_data'
    sheet, counting names whose Role matches CRM_ROLES (substring). Cached."""
    if path in _CRM_COUNT_CACHE:
        return _CRM_COUNT_CACHE[path]
    try:
        df = pd.read_excel(path, sheet_name="user_data")
    except Exception as e:
        print(f"WARN reading CRM user_data ({os.path.basename(path)}): {e}")
        _CRM_COUNT_CACHE[path] = {}
        return {}
    if "Role" not in df.columns or "Product" not in df.columns:
        _CRM_COUNT_CACHE[path] = {}
        return {}
    role_l = df["Role"].astype(str).str.lower()
    sub = df[role_l.apply(lambda s: isinstance(s, str) and any(k in s for k in CRM_ROLES))].copy()
    name_col = "Name" if "Name" in sub.columns else None
    counts: Dict[str, int] = {}
    for prod, g in sub.groupby(sub["Product"].astype(str).str.strip().str.upper()):
        counts[prod] = int(g[name_col].nunique()) if name_col else int(len(g))
    _CRM_COUNT_CACHE[path] = counts
    return counts


def load_crm_total_reps(as_of: date | None = None) -> Dict[str, int]:
    """Total reps per product (keyed by PRODUCT upper-case) from the most
    recent CRM file dated on/before `as_of`. With as_of=None, uses the latest
    file available. This makes each weekly table reflect the CRM snapshot that
    was current for that week rather than one static file."""
    files = _crm_files_by_date()
    if not files:
        print("WARN: no CS_CRM_*.xlsx found in CRM NEW_EXCEL for rep counts")
        return {}
    if as_of is None:
        chosen = files[-1]
    else:
        on_or_before = [t for t in files if t[0] <= as_of]
        # fall back to the earliest file if none precede the target date
        chosen = on_or_before[-1] if on_or_before else files[0]
    print(f"CRM user_data source (as_of {as_of or 'latest'}): {os.path.basename(chosen[1])}")
    return _crm_counts_from_file(chosen[1])


def _product_breakdown(sub: pd.DataFrame, products_sorted: List[str], crm_total: Dict[str, int]):
    """Per-product rows for a subset of `out`:
    [product, loan_count, new, repeat, amount, sold_reps, total_reps_crm]."""
    prodc = sub["Main product"].fillna("").replace("", "(Unmapped)")
    ct = sub["CLIENT TYPE"].astype(str).str.upper() if "CLIENT TYPE" in sub.columns else None
    rk = sub["_rep_key"]
    rows = []
    for p in products_sorted:
        mp = prodc == p
        loan = int(mp.sum())
        new = int((mp & (ct == "NEW")).sum()) if ct is not None else 0
        rep = int((mp & (ct == "REPEAT")).sum()) if ct is not None else 0
        amt = float(pd.to_numeric(sub.loc[mp, "Loan Amount"], errors="coerce").fillna(0).sum())
        sold = int(sub.loc[mp & (rk.astype(str).str.len() > 0), "_rep_key"].nunique())
        total = int(crm_total.get(str(p).upper(), 0))
        rows.append([p, loan, new, rep, amt, sold, total])
    return rows


def _build_summary_stats(
    year: int,
    month: int,
    month_name: str,
    today: date,
    eligible_list: List[int],
    completed_weeks_for_sold: List[int],
    display_base: pd.DataFrame,
    out: pd.DataFrame,
    qualified_reps: set[str],
    df_all_weeks: pd.DataFrame,
) -> Dict[str, Any]:
    rep_col = out["_rep_key"]
    active_reps = int(rep_col[rep_col.astype(str).str.len() > 0].nunique())
    products_sorted = _main_products_sorted_list(display_base)
    prod_norm = out["Main product"].fillna("").replace("", "(Unmapped)")
    week_rows_detail: List[Dict[str, Any]] = []
    for w in eligible_list:
        mask = out["_cal_week"] == w
        n_loans = int(mask.sum())
        n_reps = int(
            out.loc[
                mask & (rep_col.astype(str).str.len() > 0),
                "_rep_key",
            ].nunique()
        )
        m_sold_mask = mask & out["_rep_key"].isin(qualified_reps)
        n_sold_any_product = int(out.loc[m_sold_mask, "_rep_key"].nunique())
        amt_week = float(
            pd.to_numeric(out.loc[mask, "Loan Amount"], errors="coerce").fillna(0).sum()
        )
        week_rows_detail.append(
            {
                "week": w,
                "span": _week_day_span_label(year, month, w),
                "activation_count": n_loans,
                "distinct_reps": n_reps,
                "sold_reps_all_products": n_sold_any_product,
                "amount": amt_week,
            }
        )

    # Per main product: monthly totals + for each week: (activations, distinct reps w/ ≥1 sale that week for this product).
    # Weekly rep counts are NOT limited to "sold all weeks" reps — only that calendar week + main product.
    product_category_rows: List[
        Tuple[str, int, float, int, Tuple[Tuple[int, int], ...], int]
    ] = []
    for pname in products_sorted:
        mp = prod_norm == pname
        cnt = int(mp.sum())
        amt = float(
            pd.to_numeric(out.loc[mp, "Loan Amount"], errors="coerce").fillna(0).sum()
        )
        sold_cnt = int(
            out.loc[mp & out["_rep_key"].isin(qualified_reps), "_rep_key"].nunique()
        )
        weekly_pairs: List[Tuple[int, int]] = []
        for w in eligible_list:
            m_week = (out["_cal_week"] == w) & mp
            n_act = int(m_week.sum())
            m_rep = m_week & (rep_col.astype(str).str.len() > 0)
            n_distinct_reps = int(out.loc[m_rep, "_rep_key"].nunique())
            weekly_pairs.append((n_act, n_distinct_reps))
        # Active reps (as of report date): distinct reps with ≥1 sale in this main product in the month.
        active_asof = int(
            out.loc[mp & (rep_col.astype(str).str.len() > 0), "_rep_key"].nunique()
        )
        product_category_rows.append((pname, cnt, amt, sold_cnt, tuple(weekly_pairs), active_asof))

    total = len(display_base)

    # CRM total reps per product + product breakdowns (monthly and per week).
    # Monthly table uses the latest CRM snapshot on/before today (or month end);
    # each weekly table uses the latest CRM snapshot dated within that week.
    last_dom = calendar.monthrange(year, month)[1]
    month_asof = min(today, date(year, month, last_dom))
    crm_total = load_crm_total_reps(as_of=month_asof)
    month_product_rows = _product_breakdown(out, products_sorted, crm_total)
    weekly_tables: List[Dict[str, Any]] = []
    for w in eligible_list:
        sub = out[out["_cal_week"] == w]
        week_end_day = min(w * 7, last_dom)
        week_asof = date(year, month, week_end_day)
        week_crm = load_crm_total_reps(as_of=week_asof)
        rows = _product_breakdown(sub, products_sorted, week_crm)
        week_sold = int(sub.loc[rep_col.astype(str).str.len() > 0, "_rep_key"].nunique())
        weekly_tables.append({
            "week": w,
            "span": _week_day_span_label(year, month, w),
            "rows": rows,
            "sold_reps": week_sold,
            "crm_total": week_crm,
            "total_crm_reps": int(sum(week_crm.get(str(p).upper(), 0) for p in products_sorted)),
        })

    return {
        "year": year,
        "month": month,
        "month_name": month_name,
        "today": today,
        "eligible_list": eligible_list,
        "completed_weeks_for_sold": completed_weeks_for_sold,
        "total_loans": total,
        "active_reps": active_reps,
        "reps_sold_all_weeks": len(qualified_reps),
        "loans_sold_all_weeks": len(df_all_weeks),
        "week_rows_detail": week_rows_detail,
        "products_sorted": products_sorted,
        "product_category_rows": product_category_rows,
        "num_products": len(product_category_rows),
        "crm_total": crm_total,
        # total over the products shown in the tables (so TOTAL = sum of product rows)
        "total_crm_reps": int(sum(crm_total.get(str(p).upper(), 0) for p in products_sorted)),
        "month_product_rows": month_product_rows,
        "weekly_tables": weekly_tables,
    }


def _write_summary_block_table(ws, r, border, headers, rows, total_row=None, money_cols=(), pct_cols=()):
    """Write a small styled table (coloured header + banded rows + optional TOTAL).
    money_cols -> #,##0.00 ; pct_cols -> 0.0% (1-based indices). Returns next free row."""
    hdr_row = r
    for col, h in enumerate(headers, start=1):
        ws.cell(row=hdr_row, column=col, value=h)
    _style_summary_table_headers(ws, hdr_row, len(headers))
    ws.row_dimensions[hdr_row].height = 24
    r += 1
    for i, row_vals in enumerate(rows):
        for c, val in enumerate(row_vals, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.font = SUMMARY_VALUE_FONT
            cell.fill = (SUMMARY_TABLE_BAND if i % 2 == 1
                         else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"))
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right",
                                       vertical="center", indent=1 if c == 1 else 0)
            if c in money_cols:
                cell.number_format = "#,##0.00"
            elif c in pct_cols:
                cell.number_format = "0.0%"
        ws.row_dimensions[r].height = 18
        r += 1
    if total_row is not None:
        tot_fill = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
        for c, val in enumerate(total_row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.font = Font(bold=True, size=11, color=BLUE_DARK)
            cell.fill = tot_fill
            cell.alignment = Alignment(horizontal="left" if c == 1 else "right", vertical="center",
                                       indent=1 if c == 1 else 0)
            if c in money_cols:
                cell.number_format = "#,##0.00"
            elif c in pct_cols:
                cell.number_format = "0.0%"
        ws.row_dimensions[r].height = 20
        r += 1
    return r


def _summary_pair_row(ws, r: int, border: Border, label: str, value: Any) -> int:
    a = ws.cell(row=r, column=1, value=label)
    b = ws.cell(row=r, column=2, value=value)
    for cell in (a, b):
        cell.border = border
        cell.fill = SUMMARY_TABLE_BAND if r % 2 == 0 else PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
    a.font = SUMMARY_LABEL_FONT
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    b.font = SUMMARY_VALUE_FONT
    b.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[r].height = 18
    return r + 1


def _write_summary_sheet_first(wb: openpyxl.Workbook, stats: Dict[str, Any]) -> None:
    """Insert numeric-only summary as the first sheet (index 0)."""
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = True
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    eligible: List[int] = stats["eligible_list"]
    today_s = stats["today"].strftime("%d/%m/%Y")
    NC = 8  # product-summary tables have 8 columns
    merge_banner = NC

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=merge_banner)
    top = ws.cell(row=r, column=1, value="Sales rep monthly status — executive summary")
    top.font = Font(bold=True, size=14, color="FFFFFF")
    top.fill = SUMMARY_TITLE_FILL
    top.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 28
    r += 2

    def section_title(text: str, span_cols: int) -> None:
        nonlocal r
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span_cols)
        c = ws.cell(row=r, column=1, value=text)
        c.font = SUMMARY_SECTION_FONT
        c.fill = SUMMARY_TITLE_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[r].height = 22
        r += 1

    # --- Table 1: Report scope ---
    section_title("Table 1 — Report scope", 4)
    r = _summary_pair_row(
        ws, r, border, "Reporting period", f"{stats['month_name']} {stats['year']}"
    )
    r = _summary_pair_row(ws, r, border, "Summary generated (as of date)", stats["today"].strftime("%d/%m/%Y"))
    r = _summary_pair_row(ws, r, border, "Eligible calendar weeks in this run", len(stats["eligible_list"]))
    r = _summary_pair_row(
        ws,
        r,
        border,
        "Eligible week numbers",
        ", ".join(str(w) for w in stats["eligible_list"]) if stats["eligible_list"] else "— (none)",
    )
    r = _summary_pair_row(
        ws,
        r,
        border,
        "Weeks required for sold (all completed weeks to date)",
        ", ".join(str(w) for w in stats["completed_weeks_for_sold"])
        if stats["completed_weeks_for_sold"]
        else "— (none)",
    )
    r = _summary_pair_row(ws, r, border, "Total loan activations (month)", stats["total_loans"])
    r = _summary_pair_row(ws, r, border, "Distinct sales reps (≥1 activation)", stats["active_reps"])
    r += 1

    prod_headers = ["Main product", "Loan count", "New loan", "Repeat loan",
                    "Loan amount total", "Sold reps", "Total reps (CRM)", "% sold"]
    total_crm = stats["total_crm_reps"]

    def _rows_with_pct(prod_rows):
        out_rows = []
        for p, loan, new, rep, amt, sold, total in prod_rows:
            out_rows.append([p, loan, new, rep, amt, sold, total,
                             (sold / total) if total else 0.0])
        return out_rows

    def _total_row(prod_rows, sold_total, total_crm_n=total_crm):
        return [
            "TOTAL",
            sum(x[1] for x in prod_rows), sum(x[2] for x in prod_rows),
            sum(x[3] for x in prod_rows), sum(x[4] for x in prod_rows),
            sold_total, total_crm_n, (sold_total / total_crm_n) if total_crm_n else 0.0,
        ]

    # --- Monthly product summary ---
    section_title(f"Monthly summary — {stats['month_name']} {stats['year']} (by main product)", NC)
    month_rows = stats["month_product_rows"]
    if month_rows:
        r = _write_summary_block_table(
            ws, r, border, prod_headers, _rows_with_pct(month_rows),
            total_row=_total_row(month_rows, stats["active_reps"]),
            money_cols=(5,), pct_cols=(8,))
    else:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        cell = ws.cell(row=r, column=1, value="No activations in selected month.")
        cell.font = Font(italic=True, color="666666")
        cell.border = border
        r += 1
    r += 1

    # --- One summary table per week (same shape as the monthly table) ---
    for wk in stats["weekly_tables"]:
        section_title(f"Week {wk['week']} summary — {wk['span']} (by main product)", NC)
        r = _write_summary_block_table(
            ws, r, border, prod_headers, _rows_with_pct(wk["rows"]),
            total_row=_total_row(wk["rows"], wk["sold_reps"], wk["total_crm_reps"]),
            money_cols=(5,), pct_cols=(8,))
        r += 1

    # --- Sold all weeks (reps by main product) ---
    section_title("Sold all weeks — reps by main product", min(4, NC))
    pcat = stats["product_category_rows"]
    saw_headers = ["Main product", "Sold all-weeks reps", "Total reps (CRM)", "% sold all weeks"]
    saw_rows = []
    for row in pcat:
        pname, sold_n = row[0], row[3]
        tot = int(stats["crm_total"].get(str(pname).upper(), 0))
        saw_rows.append([pname, sold_n, tot, (sold_n / tot) if tot else 0.0])
    saw_total = ["TOTAL (distinct reps, all products)", stats["reps_sold_all_weeks"], total_crm,
                 (stats["reps_sold_all_weeks"] / total_crm) if total_crm else 0.0]
    r = _write_summary_block_table(ws, r, border, saw_headers, saw_rows,
                                   total_row=saw_total, pct_cols=(4,))

    ws.column_dimensions["A"].width = 24
    for col_idx in range(2, NC + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _write_by_product_sheet(wb: openpyxl.Workbook, df_display: pd.DataFrame) -> None:
    ws = wb.create_sheet("By product")
    max_col = len(OUTPUT_COLUMNS)
    current_row = 1

    if df_display.empty:
        ws.cell(row=1, column=1, value="No loan rows for the selected month.")
        return

    products = sorted(
        df_display["Main product"].fillna("").replace("", "(Unmapped)").unique(),
        key=lambda x: str(x).lower(),
    )

    for prod in products:
        if prod == "(Unmapped)":
            sub = df_display[
                (df_display["Main product"].isna()) | (df_display["Main product"] == "")
            ]
        else:
            sub = df_display[df_display["Main product"] == prod]

        # Section band
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_col)
        band = ws.cell(row=current_row, column=1, value=f"MAIN PRODUCT: {prod}")
        band.font = Font(bold=True, size=12, color="FFFFFF")
        band.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
        band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        band.border = Border(bottom=Side(style="medium", color=BLUE))
        ws.row_dimensions[current_row].height = 24
        current_row += 1

        for col, name in enumerate(OUTPUT_COLUMNS, start=1):
            ws.cell(row=current_row, column=col, value=name)
        _style_header_output(ws, current_row)
        ws.row_dimensions[current_row].height = max(HEADER_ROW_HEIGHT, 26)
        current_row += 1

        body = sub[OUTPUT_COLUMNS]
        start_body = current_row
        for _, row in body.iterrows():
            for c_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
                ws.cell(row=current_row, column=c_idx, value=row[col_name])
            ws.row_dimensions[current_row].height = DATA_ROW_HEIGHT
            current_row += 1

        end_body = current_row - 1
        if end_body >= start_body:
            _style_data_area(
                ws,
                start_body,
                end_body,
                max_col,
                loan_amount_col=OUT_AMT_COL,
                date_cols=(OUT_DATE_COL,),
                theme="by_product",
            )
            tp = _SHEET_THEME["by_product"]
            total_amt = float(
                pd.to_numeric(sub["Loan Amount"], errors="coerce").fillna(0).sum()
            )
            tr = end_body + 1
            ws.row_dimensions[tr].height = 22
            tot_fill = PatternFill(start_color=tp["total_fill"], end_color=tp["total_fill"], fill_type="solid")
            thin_tot = Side(style="thin", color=tp["total_border"])
            bold_border = Border(
                left=thin_tot,
                right=thin_tot,
                top=Side(style="medium", color=tp["total_border"]),
                bottom=thin_tot,
            )
            ws.cell(row=tr, column=1, value="TOTALS")
            for c in range(2, OUT_AMT_COL):
                ws.cell(row=tr, column=c, value="")
            ws.cell(row=tr, column=OUT_AMT_COL, value=total_amt)
            for c in range(1, max_col + 1):
                cell = ws.cell(row=tr, column=c)
                cell.fill = tot_fill
                cell.border = bold_border
                cell.font = Font(bold=True, size=11, color=tp["total_font"])
                if c == OUT_AMT_COL:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif c == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            pnote = ws.cell(row=tr + 1, column=1, value=f"Loan rows — {prod}: {len(sub)}")
            pnote.font = Font(italic=True, size=9, color="666666")
            current_row = tr + 2
        else:
            current_row += 1
        current_row += 1  # blank row before next MAIN PRODUCT block

    # No freeze panes — multiple product blocks would confuse frozen headers
    _auto_width(ws, max_col)


def _load_management_env() -> None:
    """Load Management/.env into os.environ. Values from this file override the process env.

    Prevents 535 auth failures when EMAIL_USERNAME comes from .env but APP_PASSWORD was set
    globally to a different app's secret (wrong password paired with your inbox).
    """
    if not os.path.isfile(ENV_FILE_PATH):
        return
    try:
        with open(ENV_FILE_PATH, encoding="utf-8-sig") as f:  # utf-8-sig strips BOM
            for raw in f:
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if line.startswith("export "):
                    line = line[7:].strip()
                m = re.match(r"^([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.*)$", line)
                if not m:
                    continue
                key, val = m.group(1).strip(), m.group(2).strip()
                if (val.startswith('"') and val.endswith('"')) or (
                    val.startswith("'") and val.endswith("'")
                ):
                    val = val[1:-1]
                os.environ[key] = val
    except OSError:
        return


def _smtp_password(raw: str) -> str:
    """Gmail-style app passwords may contain spaces; SMTP needs them removed."""
    return re.sub(r"\s+", "", raw.strip())


def _unique_emails(addresses: Sequence[str]) -> List[str]:
    seen: set[str] = set()
    out: List[str] = []
    for a in addresses:
        e = a.strip()
        if not e:
            continue
        k = e.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(e)
    return out


def _get_email_smtp_config() -> Tuple[str, str, str, int]:
    """Returns (from_email, password, smtp_host, smtp_port).

    Uses *matched* credentials only — same bug as Gmail 535 when USER is daniel@… but
    PASS is picked from a stray global APP_PASSWORD meant for another service.
    """
    _load_management_env()
    host = os.environ.get("SMTP_HOST", DEFAULT_SMTP_HOST).strip()
    port = int(os.environ.get("SMTP_PORT", str(DEFAULT_SMTP_PORT)))

    email_user = (os.environ.get("EMAIL_USERNAME") or os.environ.get("SMTP_USER") or "").strip()
    email_pass = _smtp_password(os.environ.get("EMAIL_PASSWORD", ""))
    app_pass = _smtp_password(
        os.environ.get("APP_PASSWORD")
        or os.environ.get("app_password")
        or ""
    )
    mgmt_user = (os.environ.get("MANAGEMENT_EMAIL") or "").strip()
    mgmt_pass = _smtp_password(os.environ.get("MANAGEMENT_EMAIL_PASSWORD", ""))

    # 1) Primary: inbox + its app password (your .env: EMAIL_USERNAME + EMAIL_PASSWORD)
    if email_user and email_pass:
        return email_user, email_pass, host, port
    # 2) Daniel's inbox + APP_PASSWORD only (if EMAIL_PASSWORD empty but app password key set)
    if email_user and app_pass:
        return email_user, app_pass, host, port
    # 3) Shared reporting mailbox + its password
    if mgmt_user and mgmt_pass:
        return mgmt_user, mgmt_pass, host, port

    return "", "", host, port


def _developer_email() -> str:
    _load_management_env()
    return (
        os.environ.get("DEVELOPER_EMAIL")
        or os.environ.get("EMAIL_USERNAME")
        or (MANAGER_RECIPIENTS[0] if MANAGER_RECIPIENTS else "")
    ).strip()


def _html_report_email_body(
    stats: Dict[str, Any],
    year: int,
    attachment_filename: str,
) -> str:
    """HTML email: instructions + Table 1 — Report scope (inline CSS for clients)."""
    mn = html.escape(str(stats["month_name"]))
    yr = year
    eligible_n = stats["eligible_list"]
    fname = html.escape(attachment_filename)

    row = """
    <tr style="background:{bg};">
      <td style="padding:12px 16px;border:1px solid #e2e8f0;font-weight:600;color:#0f766e;width:48%;">{label}</td>
      <td style="padding:12px 16px;border:1px solid #e2e8f0;color:#334155;text-align:right;">{value}</td>
    </tr>
    """
    rows = []
    cw_sold = stats.get("completed_weeks_for_sold") or []
    specs = [
        ("#f1f5f9", "Reporting period", f"{stats['month_name']} {yr}"),
        ("#ffffff", "Summary generated (as of date)", stats["today"].strftime("%d/%m/%Y")),
        ("#f1f5f9", "Eligible calendar weeks in this run", str(len(eligible_n))),
        ("#ffffff", "Eligible week numbers", ", ".join(str(w) for w in eligible_n) or "— (none)"),
        (
            "#f1f5f9",
            "Weeks required for sold (all completed weeks to date)",
            ", ".join(str(w) for w in cw_sold) or "— (none)",
        ),
        ("#ffffff", "Total loan activations (month)", str(stats["total_loans"])),
        ("#f1f5f9", "Distinct sales reps (≥1 activation)", str(stats["active_reps"])),
    ]
    for bg, lab, val in specs:
        rows.append(
            row.format(
                bg=bg,
                label=html.escape(lab),
                value=html.escape(str(val)),
            )
        )
    table_rows = "\n".join(rows)

    return f"""<!DOCTYPE html>
<html lang="en">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#f1f5f9;font-family:'Segoe UI',Roboto,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:24px 12px;">
    <tr><td align="center">
      <table role="presentation" width="640" cellpadding="0" cellspacing="0" style="max-width:640px;background:#ffffff;border-radius:12px;overflow:hidden;box-shadow:0 8px 24px rgba(15,23,42,0.1);">
        <tr>
          <td style="background:linear-gradient(90deg,#0f172a 0%,#0f766e 100%);padding:28px 24px;text-align:center;">
            <h1 style="margin:0;font-size:22px;color:#ffffff;letter-spacing:0.5px;">Sales rep monthly status</h1>
            <p style="margin:10px 0 0;font-size:15px;color:#ccfbf1;opacity:0.95;">{mn} {yr} &nbsp;·&nbsp; Executive report</p>
          </td>
        </tr>
        <tr><td style="padding:28px 24px 8px;">
          <p style="margin:0 0 16px;font-size:15px;line-height:1.6;color:#334155;">
            Hello — please find the workbook <strong>{fname}</strong> attached. Below is a quick <strong>Table 1 — Report scope</strong>
            snapshot; full detail is on the <strong>Summary</strong> sheet and other tabs in Excel.
          </p>
          <h2 style="margin:24px 0 12px;font-size:16px;color:#0f766e;border-bottom:2px solid #d97706;padding-bottom:8px;">How to use this report</h2>
          <ul style="margin:0 0 20px;padding-left:20px;color:#475569;font-size:14px;line-height:1.65;">
            <li>Open the file in <strong>Microsoft Excel</strong> (desktop recommended for freeze panes and formatting).</li>
            <li>Start on <strong>Summary</strong>: report period, week logic, and the main product table with weekly activations.</li>
            <li><strong>Sold all weeks</strong> = reps with at least one sale in <em>every completed</em> eligible week (the current in-progress week is not required until that week has ended).</li>
            <li><strong>Sold week 1, 2, …</strong> = loans whose activation falls in that week bucket only.</li>
            <li><strong>All month / By product</strong> = full listings; use <strong>By product</strong> to review by main product blocks.</li>
            <li>Headers stay visible when you scroll: freeze is on row 2 for most data sheets.</li>
          </ul>
          <h2 style="margin:8px 0 12px;font-size:16px;color:#0f766e;border-bottom:2px solid #d97706;padding-bottom:8px;">Table 1 — Report scope</h2>
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;border-radius:8px;overflow:hidden;">
            {table_rows}
          </table>
          <p style="margin:24px 0 0;font-size:12px;color:#78909c;line-height:1.5;">
            This message was generated automatically. Reply to the sender if you need changes to the distribution list or report layout.
          </p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body>
</html>"""


def send_sales_report_email(
    *,
    attachment_path: str,
    stats: Dict[str, Any],
    year: int,
    send_to_all_managers: bool,
) -> None:
    """Send HTML email with Excel attachment. Requires EMAIL_USERNAME + APP_PASSWORD (or fallbacks) in .env."""
    if not os.path.isfile(attachment_path):
        print(f"Email skipped: attachment not found: {attachment_path}")
        return

    from_addr, password, smtp_host, smtp_port = _get_email_smtp_config()
    if not from_addr or not password:
        print(
            "Email skipped: configure Management/.env with EMAIL_USERNAME and "
            "APP_PASSWORD (or EMAIL_PASSWORD / app_password)."
        )
        return

    dev = _developer_email()
    if not dev:
        print("Email skipped: set DEVELOPER_EMAIL or EMAIL_USERNAME in .env.")
        return

    if send_to_all_managers:
        recipients = _unique_emails(MANAGER_RECIPIENTS)
        if not recipients:
            print("Email skipped: MANAGER_RECIPIENTS list is empty.")
            return
    else:
        recipients = _unique_emails([dev])

    month_name = stats["month_name"]
    subject = f"Sales rep monthly status — {month_name} {year}"
    safe_name = os.path.basename(attachment_path)

    msg_root = MIMEMultipart("mixed")
    msg_root["Subject"] = subject
    msg_root["From"] = from_addr
    msg_root["To"] = ", ".join(recipients)

    body_html = _html_report_email_body(stats, year, safe_name)
    alt = MIMEMultipart("alternative")
    alt.attach(
        MIMEText(
            "Sales rep monthly status report attached. Please use an HTML-capable email client.",
            "plain",
            "utf-8",
        )
    )
    alt.attach(MIMEText(body_html, "html", "utf-8"))
    msg_root.attach(alt)

    with open(attachment_path, "rb") as f:
        part = MIMEBase(
            "application",
            "vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f'attachment; filename="{safe_name}"',
    )
    msg_root.attach(part)

    use_ssl = os.environ.get("SMTP_SSL", "").strip().lower() in ("1", "true", "yes")
    ssl_port = int(os.environ.get("SMTP_SSL_PORT", "465"))
    print(f"SMTP: signing in as {from_addr!r} via {smtp_host} (SSL={use_ssl})")
    try:
        ctx = ssl.create_default_context()
        if use_ssl:
            with smtplib.SMTP_SSL(smtp_host, ssl_port, timeout=120, context=ctx) as server:
                server.login(from_addr, password)
                server.sendmail(from_addr, recipients, msg_root.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=120) as server:
                server.starttls(context=ctx)
                server.login(from_addr, password)
                server.sendmail(from_addr, recipients, msg_root.as_string())
        print(f"Email sent successfully to: {', '.join(recipients)}")
    except smtplib.SMTPException as e:
        print(f"Email failed (SMTP): {e}")
        print(
            "Tip: For Google Workspace use an **App password** in EMAIL_PASSWORD (16 letters, spaces optional). "
            "If login still fails, add SMTP_SSL=1 to Management/.env (port 465)."
        )
    except OSError as e:
        print(f"Email failed (network): {e}")


def main() -> None:
    print("Sales Rep monthly status report")
    print("-" * 50)
    # month/year + send choice come from args/env (web orchestrator) or prompts.
    import argparse as _ap
    _p = _ap.ArgumentParser(add_help=False)
    _p.add_argument("--month", default=os.environ.get("PCL_REPORT_MONTH"))
    _p.add_argument("--send", choices=["yes", "no"], default=None)
    _cli, _ = _p.parse_known_args()
    if _cli.month:
        raw = _cli.month.strip()
    else:
        try:
            raw = input("Enter month/year (MM/YYYY, e.g. 02/2026): ").strip()
        except EOFError:
            raise SystemExit("No --month provided and no interactive input available.")
    year, month = _parse_month_year(raw)
    today = date.today()

    loan_path = DEFAULT_LOAN_PATH
    users_path = DEFAULT_USERS_PATH
    zone_path = DEFAULT_ZONE_CLUSTER_PATH
    if not os.path.isfile(loan_path):
        raise FileNotFoundError(loan_path)
    if not os.path.isfile(users_path):
        raise FileNotFoundError(users_path)
    if not os.path.isfile(zone_path):
        raise FileNotFoundError(zone_path)

    out = _prepare_dataframe(loan_path, users_path, zone_path, year, month)
    display_base = out.drop(columns=["_cal_week", "_rep_key"], errors="ignore")

    eligible_list = _eligible_calendar_weeks(year, month, today)
    eligible_set = set(eligible_list)
    completed_for_sold_list = _completed_eligible_weeks_for_sold(
        year, month, today, eligible_list
    )
    completed_for_sold_set = set(completed_for_sold_list)

    qualified_reps = _reps_all_weeks(out, completed_for_sold_set)
    df_all_weeks = display_base.loc[out["_rep_key"].isin(qualified_reps)].copy()

    month_name = calendar.month_name[month]
    out_name = f"Sales_Reps_Monthly_Status_{year:04d}-{month:02d}.xlsx"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, out_name)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_grouped_rep_sheet(wb, "Sold all weeks", df_all_weeks, theme="ocean")
    for w in eligible_list:
        mask = out["_cal_week"] == w
        df_w = display_base.loc[mask].copy()
        _write_grouped_rep_sheet(wb, f"Sold week {w}", df_w, theme="indigo")

    _write_all_month_sheet(wb, display_base.copy())
    _write_by_product_sheet(wb, display_base.copy())

    summary_stats = _build_summary_stats(
        year,
        month,
        month_name,
        today,
        eligible_list,
        completed_for_sold_list,
        display_base,
        out,
        qualified_reps,
        df_all_weeks,
    )
    _write_summary_sheet_first(wb, summary_stats)

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Rows (all month): {len(display_base)}")
    print(f"  Eligible calendar weeks (as of {today.isoformat()}): {eligible_list or '— (none)'}")
    print(
        f"  Weeks used for sold qualification (completed to date): "
        f"{completed_for_sold_list or '— (none)'}"
    )
    print(
        f"  Reps with sale in every **completed** week (sold) for {month_name} {year}: "
        f"{len(qualified_reps)}"
    )

    if _cli.send is not None:
        send_all = _cli.send == "yes"
    else:
        try:
            send_all = input("Send email to all managers? [y/N]: ").strip().lower() in ("y", "yes")
        except EOFError:
            send_all = False
    send_sales_report_email(
        attachment_path=out_path,
        stats=summary_stats,
        year=year,
        send_to_all_managers=send_all,
    )


if __name__ == "__main__":
    main()
