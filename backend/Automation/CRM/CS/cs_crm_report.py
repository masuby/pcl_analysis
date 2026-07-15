import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import xlwings as xw

# =============================================================
# INPUT CONFIGURATION
# =============================================================
FOLDER = fr"C:\Users\Daniel\Desktop\code\pcl\CRM\ROW_EXCEL"
MANUAL_DATE = input("Enter the MOST RECENT login date (dd/mm/yyyy): ").strip()

# =============================================================
# STEP 1: AUTO-DETECT FILES
# =============================================================
print("Scanning for Excel files...")
excel_files = [
    os.path.join(FOLDER, f)
    for f in os.listdir(FOLDER)
    if f.lower().endswith(".xlsx") and not f.startswith("~$")
]

if not excel_files:
    raise ValueError("No Excel files found in ROW_EXCEL folder")

# Classify files
activities_files = []
agent_activities_file = None
lead_report_file = None
crm_file = None
user_list_file = None

for f in excel_files:
    name = os.path.basename(f).lower()
    if name.startswith("activities_report"):
        activities_files.append(f)
    elif name.startswith("agent_activities_report"):
        agent_activities_file = f
    elif name.startswith("lead_report"):
        lead_report_file = f
    elif name.startswith("cs_crm"):
        crm_file = f
    elif name.startswith("user_list_report"):
        user_list_file = f

# Determine today's activities file (most recent Expected_End_Date)
# =============================================================
# STEP 1: AUTO-DETECT FILES
# =============================================================
print("Scanning for Excel files...")
excel_files = [
    os.path.join(FOLDER, f)
    for f in os.listdir(FOLDER)
    if f.lower().endswith(".xlsx") and not f.startswith("~$")
]

if not excel_files:
    raise ValueError("No Excel files found in ROW_EXCEL folder")

# Classify files
activities_files = []
agent_activities_file = None
lead_report_file = None
crm_file = None
user_list_file = None

for f in excel_files:
    name = os.path.basename(f).lower()
    if name.startswith("activities_report"):
        activities_files.append(f)
    elif name.startswith("agent_activities_report"):
        agent_activities_file = f
    elif name.startswith("lead_report"):
        lead_report_file = f
    elif name.startswith("cs_crm"):
        crm_file = f
    elif name.startswith("user_list_report"):
        user_list_file = f

# Determine today's activities file (most recent Expected_Start_Date)
def extract_latest_start_date(file_path):
    try:
        df = pd.read_excel(file_path)
        if "Expected_Start_Date" not in df.columns:
            print(f"Warning: 'Expected_Start_Date' column not found in {os.path.basename(file_path)}")
            return None
        
        # Try multiple date parsing strategies
        date_series = df["Expected_Start_Date"]
        
        # First, try to parse with dayfirst=True (for dd/mm/yyyy)
        parsed_dates = pd.to_datetime(date_series, errors='coerce', dayfirst=True)
        
        # If that doesn't work well, try without dayfirst
        if parsed_dates.isna().all():
            parsed_dates = pd.to_datetime(date_series, errors='coerce')
        
        # Convert to datetime.date to avoid timezone issues
        parsed_dates = parsed_dates.dt.date
        
        # Return the maximum date
        if not parsed_dates.empty:
            max_date = parsed_dates.max()
            return max_date
        return None
    except Exception as e:
        print(f"Error extracting date from {file_path}: {e}")
        return None

today_activities_file = None
today_activities_date = None
for f in activities_files:
    max_date = extract_latest_start_date(f)
    if max_date:
        print(f"File: {os.path.basename(f)}, Max Expected_Start_Date: {max_date}")
        if today_activities_date is None or max_date > today_activities_date:
            today_activities_date = max_date
            today_activities_file = f

if today_activities_file:
    print(f"Selected today's activities file: {os.path.basename(today_activities_file)}")
    print(f"Most recent Expected_Start_Date: {today_activities_date}")
else:
    print("Warning: Could not determine today's activities file. Using first file.")
    if activities_files:
        today_activities_file = activities_files[0]

print("Files auto-detected successfully")

# =============================================================
# STEP 2: LOAD BASE FILES
# =============================================================
print("Loading base files...")

# Load CRM file
crm_wb = openpyxl.load_workbook(crm_file)
crm_sheets_data = {}
for sheet_name in crm_wb.sheetnames:
    ws = crm_wb[sheet_name]
    data = [row for row in ws.iter_rows(values_only=True)]
    crm_sheets_data[sheet_name] = pd.DataFrame(data[1:], columns=data[0]) if data else pd.DataFrame()

# Load user_list
users_wb = openpyxl.load_workbook(user_list_file)
users_ws = users_wb.active
users_data = [row for row in users_ws.iter_rows(values_only=True)]
users = pd.DataFrame(users_data[1:], columns=users_data[0]) if users_data else pd.DataFrame()

crm_user_data = crm_sheets_data.get('user_data', pd.DataFrame())
print("Base files loaded")

# =============================================================
# STEP 3: ENHANCE USERS DATA
# =============================================================
print("Enhancing users data...")

if "Teams" in users.columns and "Tenant" in users.columns:
    insert_pos = users.columns.get_loc("Tenant")
    users.insert(insert_pos, "Zone", "")

if "Last_Login" in users.columns and "Products" in users.columns:
    insert_pos = users.columns.get_loc("Products")
    users.insert(insert_pos, "Product", "")

if "Login_Location" in users.columns and "Last_Login" in users.columns:
    insert_pos = users.columns.get_loc("Last_Login")
    users.insert(insert_pos, "Month", "")

# Assign Zone and Product from CRM user_data
if "Tenant" in users.columns and "Zone" in users.columns:
    users["Zone"] = users["Tenant"].apply(
        lambda x: crm_user_data.loc[crm_user_data["Tenant"] == x, "Zone"].values[0]
        if (crm_user_data["Tenant"] == x).any() else "ERR"
    )

if "Tenant" in users.columns and "Product" in users.columns:
    users["Product"] = users["Tenant"].apply(
        lambda x: crm_user_data.loc[crm_user_data["Tenant"] == x, "Product"].values[0]
        if (crm_user_data["Tenant"] == x).any() else "ERR"
    )

def extract_month(value):
    try:
        if "/" in str(value):
            date_part = str(value).split(" ")[0]
            month = date_part.split("/")[1]
            return month if len(month) == 2 else "ERR"
        return "ERR"
    except:
        return "ERR"

if "Month" in users.columns:
    users["Month"] = users["Last_Login"].apply(extract_month)

print("Users data enhanced")

# =============================================================
# STEP 4: FILTER USERS DATA
# =============================================================
print("Filtering users data...")

current_month = datetime.now().month
prev_month = 12 if current_month == 1 else current_month - 1
current_month_str = f"{current_month:02d}"
prev_month_str = f"{prev_month:02d}"
allowed_months = ["ERR", current_month_str, prev_month_str]

bad_months = users[~users["Month"].isin(allowed_months)].copy()

blocked_tenants = [
    "Head Office", "Head Office,Head_Office",
    "Head Office,Platinum Credit Tanzania",
    "Head Office,Platinum Credit Tanzania,SME Arusha,SME MBEYA,SME MOROGORO,SME MOSHI,SME Tazara,SME TEGETA",
    "Platinum Credit Tanzania"
]
bad_tenants = bad_months[bad_months["Tenant"].isin(blocked_tenants)].copy()

df_removed_bad_months = bad_months.merge(bad_tenants, how="left", indicator=True)
df_removed_bad_months = df_removed_bad_months[df_removed_bad_months["_merge"] == "left_only"].drop(columns=["_merge"])

users["User_Name"] = users["User_Name"].astype(str).str.strip()
df_removed_bad_months["User_Name"] = df_removed_bad_months["User_Name"].astype(str).str.strip()

df_final = users[~users["User_Name"].isin(df_removed_bad_months["User_Name"])].copy()
print("Users data filtered")

# =============================================================
# STEP 5: ADD LOGIN STATUS
# =============================================================
print("Adding login status...")

manual_target_date = pd.to_datetime(MANUAL_DATE, format="%d/%m/%Y")

df_final["Last_Login_dt"] = pd.to_datetime(
    df_final["Last_Login"].str.split().str[0],
    format="%d/%m/%Y",
    errors="coerce"
)

def evaluate_status(row):
    if pd.isna(row["Last_Login_dt"]):
        return "Not Logged In"
    if (row["Last_Login_dt"] == manual_target_date) and (row["Login_Result"] == "Success"):
        return "Logged In"
    return "Not Logged In"

df_final["Login_Result"] = df_final.apply(evaluate_status, axis=1)
df_final = df_final.drop(columns=["Last_Login_dt"])
print("Login status added")

# =============================================================
# STEP 6: PREPARE ALL DATA FOR SAVING
# =============================================================
print("Preparing all data for saving...")

# Create a dictionary to store all processed data
processed_data = {}

# Store user_data
processed_data['user_data'] = df_final

# Load and process other sheets
lookup = df_final[["Tenant", "Zone", "Product"]].drop_duplicates()

# Today's activity_data
if today_activities_file:
    df_today = pd.read_excel(today_activities_file)
    for col in ["Zone", "Product"]:
        if col not in df_today.columns:
            df_today[col] = ""
    if "Branch" in df_today.columns:
        df_today["Zone"] = df_today["Branch"].map(lambda x: lookup.loc[lookup["Tenant"] == x, "Zone"].iloc[0] if (lookup["Tenant"] == x).any() else "ERR")
        df_today["Product"] = df_today["Branch"].map(lambda x: lookup.loc[lookup["Tenant"] == x, "Product"].iloc[0] if (lookup["Tenant"] == x).any() else "ERR")
    processed_data["Today's activity_data"] = df_today

# agent_activity
if agent_activities_file:
    df_agent = pd.read_excel(agent_activities_file)
    for col in ["Zone", "Product"]:
        if col not in df_agent.columns:
            df_agent[col] = ""
    if "Branch" in df_agent.columns:
        df_agent["Zone"] = df_agent["Branch"].map(lambda x: lookup.loc[lookup["Tenant"] == x, "Zone"].iloc[0] if (lookup["Tenant"] == x).any() else "ERR")
        df_agent["Product"] = df_agent["Branch"].map(lambda x: lookup.loc[lookup["Tenant"] == x, "Product"].iloc[0] if (lookup["Tenant"] == x).any() else "ERR")
    processed_data["agent_activity"] = df_agent

# Lead_Report (Raw lead report): use Created_By → user_data sheet "Name" → Zone, Product
if lead_report_file:
    df_leads = pd.read_excel(lead_report_file)
    for col in ["Zone", "PRODUCT"]:
        if col not in df_leads.columns:
            df_leads[col] = ""
    if not crm_user_data.empty and "Name" in crm_user_data.columns and "Zone" in crm_user_data.columns and "Product" in crm_user_data.columns:
        lookup_lead = crm_user_data[["Name", "Zone", "Product"]].drop_duplicates(subset=["Name"], keep="first")
        if "Created_By" in df_leads.columns:
            df_leads["Zone"] = df_leads["Created_By"].astype(str).str.strip().map(
                lambda x: lookup_lead.loc[lookup_lead["Name"].astype(str).str.strip() == x, "Zone"].iloc[0]
                if (lookup_lead["Name"].astype(str).str.strip() == x).any() else "ERR"
            )
            df_leads["PRODUCT"] = df_leads["Created_By"].astype(str).str.strip().map(
                lambda x: lookup_lead.loc[lookup_lead["Name"].astype(str).str.strip() == x, "Product"].iloc[0]
                if (lookup_lead["Name"].astype(str).str.strip() == x).any() else "ERR"
            )
    processed_data["Lead_Report"] = df_leads

# activity_data (main activities)
if activities_files:
    df_activities = pd.read_excel(activities_files[0])
    for col in ["Zone", "Product"]:
        if col not in df_activities.columns:
            df_activities[col] = ""
    if "Branch" in df_activities.columns:
        df_activities["Zone"] = df_activities["Branch"].map(lambda x: lookup.loc[lookup["Tenant"] == x, "Zone"].iloc[0] if (lookup["Tenant"] == x).any() else "ERR")
        df_activities["Product"] = df_activities["Branch"].map(lambda x: lookup.loc[lookup["Tenant"] == x, "Product"].iloc[0] if (lookup["Tenant"] == x).any() else "ERR")
    processed_data["activity_data"] = df_activities

print("All data prepared for saving")

# =============================================================
# STEP 7: SAVE PROCESSED DATA USING PANDAS (CLEAN EXCEL FILE)
# =============================================================
print("Saving processed data using pandas...")

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
processed_file = fr"C:\Users\Daniel\Desktop\code\pcl\CRM\CS\NEW_EXCEL\CRM_PROCESSED_{timestamp}.xlsx"
os.makedirs(os.path.dirname(processed_file), exist_ok=True)

# Save all sheets using pandas (creates clean Excel file)
with pd.ExcelWriter(processed_file, engine='openpyxl') as writer:
    for sheet_name, data in processed_data.items():
        data.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"Processed file saved: {processed_file}")

# =============================================================
# STEP 8: COPY TO MASTER CRM USING XLWINGS (WITH COLUMN MATCHING)
# =============================================================
print("Starting xlwings copy process to master CRM...")

# Define master file path with date from user input
date_obj = pd.to_datetime(MANUAL_DATE, format="%d/%m/%Y")
date_str = date_obj.strftime("%d_%b_%Y").lower()
master_file = crm_file

def get_master_columns(xl_ws):
    """Get column headers from master worksheet"""
    header_range = xl_ws.range('1:1')
    headers = []
    for cell in header_range:
        if cell.value is not None:
            headers.append(cell.value)
        else:
            break
    return headers

def copy_matching_columns(master_ws, processed_ws, sheet_name):
    """Copy only matching columns from processed to master, preserving master column order"""
    # Get column headers from both sheets
    master_headers = get_master_columns(master_ws)
    processed_headers = get_master_columns(processed_ws)
    
    print(f"  Master columns ({len(master_headers)}): {master_headers}")
    print(f"  Processed columns ({len(processed_headers)}): {processed_headers}")
    
    # Find matching columns (only those present in both sheets)
    matching_cols = [col for col in master_headers if col in processed_headers]
    
    if not matching_cols:
        print(f"  Warning: No matching columns found for {sheet_name}")
        return
    
    print(f"  Matching columns ({len(matching_cols)}): {matching_cols}")
    
    # Get column indices for matching columns in processed sheet
    processed_col_indices = []
    for col in matching_cols:
        if col in processed_headers:
            processed_col_indices.append(processed_headers.index(col) + 1)  # +1 for 1-based index
    
    # Get all data from processed sheet
    processed_used_range = processed_ws.used_range
    if processed_used_range is None:
        print(f"  Warning: {sheet_name} in processed file is empty")
        return
    
    processed_data = processed_used_range.value
    if not processed_data:
        print(f"  Warning: No data found in {sheet_name}")
        return
    
    # Clear the master sheet first
    master_ws.clear()
    
    # Copy headers to master sheet (all master headers)
    master_ws.range('A1').value = [master_headers]
    
    # Copy matching data rows
    if len(processed_data) > 1:  # If there are data rows beyond header
        for row_idx, row in enumerate(processed_data[1:], start=2):  # Start from row 2 (after header)
            if row:  # Check if row is not None or empty
                new_row = []
                for master_col in master_headers:
                    if master_col in matching_cols:
                        # Find the value from processed data
                        processed_idx = processed_headers.index(master_col)
                        if processed_idx < len(row):
                            new_row.append(row[processed_idx])
                        else:
                            new_row.append(None)
                    else:
                        new_row.append(None)  # Empty for non-matching columns
                
                # Write the row to master sheet
                master_ws.range(f'A{row_idx}').value = [new_row]

def copy_sheets_to_master():
    print("Starting xlwings copy process...")
    print(f"Master file: {master_file}")
    print(f"Processed file: {processed_file}")
    
    # Check if files exist
    if not os.path.exists(master_file):
        raise FileNotFoundError(f"Master file not found: {master_file}")
    if not os.path.exists(processed_file):
        raise FileNotFoundError(f"Processed file not found: {processed_file}")
    
    # Run Excel in the background
    app = xw.App(visible=False)
    
    try:
        # Open workbooks
        print("Opening workbooks...")
        master_wb = app.books.open(master_file)
        processed_wb = app.books.open(processed_file)
        
        # Define sheets to copy
        sheets_to_copy = [
            'user_data', 
            'activity_data', 
            'agent_activity', 
            "Today's activity_data",
            'Lead_Report'
        ]
        
        print("Copying sheets with column matching...")
        
        for sheet_name in sheets_to_copy:
            if sheet_name in master_wb.sheet_names and sheet_name in processed_wb.sheet_names:
                print(f"  Processing {sheet_name}...")
                
                # Get sheets
                master_ws = master_wb.sheets[sheet_name]
                processed_ws = processed_wb.sheets[sheet_name]
                
                # Copy only matching columns
                copy_matching_columns(master_ws, processed_ws, sheet_name)
                
                print(f"  {sheet_name} copied successfully!")
            else:
                print(f"  Warning: {sheet_name} not found in one of the workbooks")
        
        print("All sheets copied successfully!")
        
        # AUTO REFRESH AND RECALCULATION SECTION
        print("Refreshing and recalculating workbook...")
        
        # Refresh all data connections and pivot tables
        master_wb.api.RefreshAll()
        
        # Force full calculation of the entire workbook
        master_wb.app.calculate()
        
        # Wait a moment for refresh to complete
        import time
        time.sleep(3)
        
        print("Workbook refreshed and recalculated successfully!")

        new_master_file = fr"C:\Users\Daniel\Desktop\code\pcl\CRM\CS\NEW_EXCEL/CS_CRM_{MANUAL_DATE.replace('/', '_')}.xlsx"
        
        # Save master file with new name
        master_wb.save(new_master_file)
        print(f"New master file saved successfully: {new_master_file}")
        print(f"Master file saved successfully: {new_master_file}")
        
    except Exception as e:
        print(f"Error during xlwings copy process: {e}")
        # Print more detailed error information
        import traceback
        traceback.print_exc()
        raise
    finally:
        # Close workbooks and quit app
        try:
            processed_wb.close()
        except:
            pass
        try:
            master_wb.close()
        except:
            pass
        app.quit()
        print("Excel application closed")

# Execute the copy process
copy_sheets_to_master()

# Clean up processed file (optional)
try:
    os.remove(processed_file)
    print(f"Temporary processed file removed: {processed_file}")
except:
    print(f"Could not remove temporary file: {processed_file}")

print("========================================================")
print("PROCESS COMPLETED 100% SUCCESSFULLY!")
print(f"MASTER FILE UPDATED → {master_file}")
print("All data copied using xlwings - preserves pivot tables and formulas")
print("Only matching columns copied - preserving master sheet structure")
