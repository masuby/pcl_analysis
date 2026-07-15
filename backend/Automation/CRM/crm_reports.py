"""
crm_reports.py — build ALL CRM department reports in ONE run, in ONE file.
================================================================================
No Excel pivot tables, no VBA (BranchFunctions.xlam), no xlwings. Every summary
table is computed directly from the raw data in pandas, with clean, documented
definitions (SECTION 2), validated against the live CS master.

Pipeline (clearly separated sections below):
  SECTION 1  Source detection + shared processing (read each export ONCE)
  SECTION 2  Metric definitions — the validated, pivot-free calculations
             + per-Team-Leader analysis (agent_activity.Sales_Team)
             + dynamic narrative (plain-English explanations, data-driven daily)
  SECTION 3  Build the per-department report workbook (openpyxl; each analysis
             on its own sheet, header + left column FROZEN, violet->red spectrum)
  SECTION 4  Email — professional HTML with prose + compact colour-graded tables
  SECTION 5  Orchestration (main)

One <DEPT>_CRM_<date>.xlsx per department into its NEW_EXCEL folder, plus a
<...>_email_preview.html you can open before sending.

Usage:
    python crm_reports.py                          # prompts for the date
    python crm_reports.py 06/07/2026 --only CS
    python crm_reports.py 06/07/2026 --send test   # test|specific|actual|no
At the end it ASKS whether to update (replace) the ROW_EXCEL row files.
"""
from __future__ import annotations

import argparse
import importlib.util
import os
import re
import smtplib
import ssl
import subprocess
import sys
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# --------------------------------------------------------------------------- #
# CONFIGURATION
# --------------------------------------------------------------------------- #
BASE_DIR = Path(__file__).resolve().parent
ROW_DIR = BASE_DIR / "ROW_EXCEL"
EMAIL_MASTER = BASE_DIR / "Master_crm_emails.xlsx"
CRM_ENV = BASE_DIR / ".env"

# email_group = which grouping the email agent/TL tables use.
# lead_groups  = the lead breakdown tables to show (label, Lead_Report column).
_CS_LEADS = [("By Zone", "Zone"), ("By Branch", "Branch"), ("By Team Leader", "Team_Name")]
_LBF_LEADS = [("By Branch", "Branch"), ("By Team Leader", "Team_Name")]
DEPARTMENTS = [
    {"code": "CS",  "prefix": "cs_crm",  "out_dir": BASE_DIR / "CS" / "NEW_EXCEL",  "out_prefix": "CS_CRM",  "email_col": "CS",          "product": "CS",  "email_group": "zone",   "lead_groups": _CS_LEADS},
    {"code": "LBF", "prefix": "lbf_crm", "out_dir": BASE_DIR / "LBF" / "NEW_EXCEL", "out_prefix": "LBF_CRM", "email_col": "LBF",         "product": "LBF", "email_group": "branch", "lead_groups": _LBF_LEADS},
    {"code": "SME", "prefix": "sme_crm", "out_dir": BASE_DIR / "SME" / "NEW_EXCEL", "out_prefix": "SME_CRM", "email_col": "SME",         "product": "SME", "email_group": "branch", "lead_groups": _LBF_LEADS},
    # {"code": "AGRI", "prefix": "agri_crm", "out_dir": BASE_DIR / "AGRI" / "NEW_EXCEL", "out_prefix": "AGRI_CRM", "email_col": "Agrifinance", "product": "Agrifinance", "email_group": "branch", "lead_groups": _LBF_LEADS},
]

DATA_SHEETS = ["user_data", "activity_data", "agent_activity",
               "Today's activity_data", "Lead_Report"]

# Columns to DROP from each raw data sheet before writing (display only — the
# metric computations in SECTION 2 already ran, so dropping here is safe).
DROP_COLUMNS = {
    "user_data": ["Month", "Products", "Login_Location", "Last_Logout", "Logout_Location"],
    "activity_data": ["Lead_Name", "Client_Name", "Description", "Activity_Subunit",
                      "Activity_Latitude", "Activity_Longitude", "Expected_Duration",
                      "Created_Date", "comments"],
    "Today's activity_data": ["Lead_Name", "Client_Name", "Description", "Activity_Subunit",
                              "Activity_Latitude", "Activity_Longitude", "Expected_Duration",
                              "Created_Date", "comments"],
    "agent_activity": ["Accuracy_of_Measurement", "Status_Lat", "Status_Lng",
                       "IP_Address", "Assigned_Lat", "Assigned_Lng"],
    "Lead_Report": ["ID_Number", "Emp_Number", "Affordability_Outcome", "Total_Affordability",
                    "Installment_Amount", "Affordability_DateCreated", "Assignment_Type",
                    "Consent_Type"],
}

BLOCKED_TENANTS = [
    "Head Office", "Head Office,Head_Office",
    "Head Office,Platinum Credit Tanzania",
    "Head Office,Platinum Credit Tanzania,SME Arusha,SME MBEYA,SME MOROGORO,SME MOSHI,SME Tazara,SME TEGETA",
    "Platinum Credit Tanzania",
]

ROW_UPDATE_SCRIPT = ROW_DIR / "update_master_crm_files.py"

# Styling / colour spectrum (violet -> red across the visible spectrum) used to
# grade a table's key metric from best (violet) to worst (red).
# One consistent brand blue (medium, not navy-black) is used throughout.
FONT = "Calibri"
BRAND_BLUE = "2F5597"   # consistent medium blue for titles + table headers
NAVY = BRAND_BLUE
HEAD_FILL = BRAND_BLUE
ALT_FILL = "EBF1F9"     # very light blue (alternating rows)
TOTAL_FILL = "D6E0F0"   # light blue (total rows)
SPECTRUM_STOPS = [(124, 58, 237), (37, 99, 235), (22, 163, 74),
                  (234, 179, 8), (234, 88, 12), (220, 38, 38)]  # violet->blue->green->yellow->orange->red


# =========================================================================== #
# SECTION 1 — SOURCE DETECTION + SHARED PROCESSING (read each export once)
# =========================================================================== #
def detect_source_files():
    """Classify the raw ROW_EXCEL exports and locate each department master."""
    files = [p for p in ROW_DIR.iterdir()
             if p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")]
    if not files:
        raise ValueError(f"No Excel files found in {ROW_DIR}")
    src = {"activities": [], "agent": None, "lead": None, "user_list": None, "masters": {}}
    for p in sorted(files):
        name = p.name.lower()
        if name.startswith("activities_report"):
            src["activities"].append(p)
        elif name.startswith("agent_activities_report"):
            src["agent"] = p
        elif name.startswith("lead_report"):
            src["lead"] = p
        elif name.startswith("user_list_report"):
            src["user_list"] = p
        else:
            for dept in DEPARTMENTS:
                if name.startswith(dept["prefix"]):
                    src["masters"][dept["code"]] = p
    if src["user_list"] is None:
        raise ValueError("No User_List_Report file found in ROW_EXCEL")
    return src


def latest_start_date(df):
    if "Expected_Start_Date" not in df.columns:
        return None
    parsed = pd.to_datetime(df["Expected_Start_Date"], errors="coerce", dayfirst=True)
    if parsed.isna().all():
        parsed = pd.to_datetime(df["Expected_Start_Date"], errors="coerce")
    parsed = parsed.dt.date.dropna()
    return parsed.max() if not parsed.empty else None


def _extract_month(value):
    try:
        if "/" in str(value):
            month = str(value).split(" ")[0].split("/")[1]
            return month if len(month) == 2 else "ERR"
        return "ERR"
    except Exception:
        return "ERR"


def build_base_users(users: pd.DataFrame, manual_date: pd.Timestamp) -> pd.DataFrame:
    """Department-independent user prep (done ONCE): insert Zone/Product/Month,
    drop stale-month users (unless blocked tenant), set Login_Result = 'Logged In'
    only when they logged in successfully ON the report date."""
    users = users.copy()
    if "Teams" in users.columns and "Tenant" in users.columns:
        users.insert(users.columns.get_loc("Tenant"), "Zone", "")
    if "Last_Login" in users.columns and "Products" in users.columns:
        users.insert(users.columns.get_loc("Products"), "Product", "")
    if "Login_Location" in users.columns and "Last_Login" in users.columns:
        users.insert(users.columns.get_loc("Last_Login"), "Month", "")
    users["Month"] = users["Last_Login"].apply(_extract_month)

    current_month = datetime.now().month
    prev_month = 12 if current_month == 1 else current_month - 1
    allowed = ["ERR", f"{current_month:02d}", f"{prev_month:02d}"]

    bad_months = users[~users["Month"].isin(allowed)].copy()
    bad_tenants = bad_months[bad_months["Tenant"].isin(BLOCKED_TENANTS)].copy()
    removed = bad_months.merge(bad_tenants, how="left", indicator=True)
    removed = removed[removed["_merge"] == "left_only"].drop(columns=["_merge"])

    users["User_Name"] = users["User_Name"].astype(str).str.strip()
    removed["User_Name"] = removed["User_Name"].astype(str).str.strip()
    df = users[~users["User_Name"].isin(removed["User_Name"])].copy()

    login_dt = pd.to_datetime(df["Last_Login"].astype(str).str.split().str[0],
                              format="%d/%m/%Y", errors="coerce")
    df["Login_Result"] = [
        "Logged In" if (pd.notna(d) and d == manual_date and r == "Success") else "Not Logged In"
        for d, r in zip(login_dt, df["Login_Result"])
    ]
    return df


def _map_series(series, mapping):
    return series.map(lambda x: mapping.get(x, "ERR"))


def build_department_frames(base_users, crm_user_data, shared):
    """Attach Zone/Product to each of the 5 data sheets using the department
    master's user_data as the Tenant->Zone/Product lookup (a plain lookup table)."""
    tz, tp = {}, {}
    for _, r in crm_user_data.drop_duplicates(subset=["Tenant"], keep="first").iterrows():
        tz[r["Tenant"]] = r["Zone"]
        tp[r["Tenant"]] = r["Product"]

    users = base_users.copy()
    users["Zone"] = _map_series(users["Tenant"], tz)
    users["Product"] = _map_series(users["Tenant"], tp)
    frames = {"user_data": users}

    lk = users[["Tenant", "Zone", "Product"]].drop_duplicates(subset=["Tenant"], keep="first")
    bz = dict(zip(lk["Tenant"], lk["Zone"]))
    bp = dict(zip(lk["Tenant"], lk["Product"]))

    for sheet, df in [("Today's activity_data", shared.get("today")),
                      ("agent_activity", shared.get("agent")),
                      ("activity_data", shared.get("activities"))]:
        if df is None:
            continue
        df = df.copy()
        for col in ["Zone", "Product"]:
            if col not in df.columns:
                df[col] = ""
        if "Branch" in df.columns:
            df["Zone"] = _map_series(df["Branch"], bz)
            df["Product"] = _map_series(df["Branch"], bp)
        frames[sheet] = df

    if shared.get("lead") is not None:
        df = shared["lead"].copy()
        for col in ["Zone", "PRODUCT"]:
            if col not in df.columns:
                df[col] = ""
        if not crm_user_data.empty and {"Name", "Zone", "Product"} <= set(crm_user_data.columns):
            lu = crm_user_data[["Name", "Zone", "Product"]].drop_duplicates(subset=["Name"], keep="first")
            lu = lu.assign(_key=lu["Name"].astype(str).str.strip())
            nz = dict(zip(lu["_key"], lu["Zone"]))
            np_ = dict(zip(lu["_key"], lu["Product"]))
            if "Created_By" in df.columns:
                created = df["Created_By"].astype(str).str.strip()
                df["Zone"] = _map_series(created, nz)
                df["PRODUCT"] = _map_series(created, np_)
        frames["Lead_Report"] = df

    return frames


# =========================================================================== #
# SECTION 2 — METRIC DEFINITIONS  (validated, pivot-free)
# --------------------------------------------------------------------------- #
#   Agent  = Role contains 'loan officer' OR 'call center agent' (covers 'branch
#            loan officers'), NOT 'team leader'.   TL = Role contains 'team leader'.
#   D logins=user_data Logged In | L actual=user_data count | E users assigned=
#   distinct User_Name in activity_data | H locations planned=distinct
#   Activity_Location in activity_data | F completing=distinct User_Name in
#   agent_activity COMPLETED&AT_LOCATION | I reached=distinct Assigned_Location
#   AT_LOCATION | J total completed=rows COMPLETED&AT_LOCATION | M assigned=
#   distinct User_Name in agent_activity | O planned locations=distinct
#   Activity_Location in Today's activity_data | G=F/E K=I/H N=M/L
# =========================================================================== #
BLANK_GROUPS = {"", "ERR", "NAN", "NONE"}

# NOTE: leads columns are intentionally NOT in the Summary block (leads live in
# the Lead Summary section with their own breakdown tables).
METRIC_COLUMNS = [
    ("logins", "CRM Users (Logins)"),
    ("users_assigned", "Users Assigned Activities"),
    ("users_completing", "Users Completing @ Location"),
    ("completion_rate", "Activity Completion Rate"),
    ("locations_planned", "Locations Planned"),
    ("locations_reached", "Locations Reached"),
    ("total_completed", "Total Completed @ Location"),
    ("pct_reached", "% Locations Reached"),
    ("actual", "Actual"),
    ("assigned_today", "Assigned (Today's Plan)"),
    ("pct_assigned", "% Assigned"),
    ("planned_locations", "Planned Locations"),
]
METRIC_LABELS = dict(METRIC_COLUMNS)
PCT_KEYS = {"completion_rate", "pct_reached", "pct_assigned", "pct_consenting", "pct_completed"}
LEAD_LABELS = {"total": "Total Leads", "consented": "Consented", "not_provided": "Not Provided",
               "rejected": "Rejected", "prospects": "Prospects"}
LEAD_KEYS = ["total", "consented", "not_provided", "rejected", "prospects"]


def is_pct_key(k):
    k = str(k).lower()
    return (k in PCT_KEYS or k.startswith("pct_") or "rate" in k
            or "movement" in k or k.startswith(("today", "yesterday")) or k == "% reached")

# Compact set of columns shown in the EMAIL agent/TL tables (kept short).
EMAIL_METRIC_KEYS = ["logins", "users_assigned", "users_completing", "completion_rate",
                     "locations_planned", "locations_reached", "pct_reached"]


def _agent_mask(role_series):
    r = role_series.fillna("").astype(str).str.lower()
    return (r.str.contains("loan officer") | r.str.contains("call center agent")) & ~r.str.contains("team leader")


def _tl_mask(role_series):
    return role_series.fillna("").astype(str).str.lower().str.contains("team leader")


def _safe_div(a, b):
    return (a / b).where(b != 0, 0.0)


def _clean_groups(index):
    return [g for g in index if str(g).strip().upper() not in BLANK_GROUPS and pd.notna(g)]


def compute_block(frames, product, role, group):
    """One summary block as a DataFrame indexed by group ('zone' or 'branch').
    Leads columns are added for the agents/branch block."""
    ud = frames["user_data"]
    act = frames.get("activity_data", pd.DataFrame())
    ag = frames.get("agent_activity", pd.DataFrame())
    today = frames.get("Today's activity_data", pd.DataFrame())
    lead = frames.get("Lead_Report", pd.DataFrame())

    def prod(df):
        return df[df["Product"].astype(str).str.upper() == product.upper()] if len(df) and "Product" in df.columns else df

    pud, pact, pag, ptoday = prod(ud), prod(act), prod(ag), prod(today)
    mask = _agent_mask if role == "agent" else _tl_mask
    ud_r = pud[mask(pud["Role"])]
    act_r = pact[mask(pact["Role"])] if len(pact) else pact
    ag_r = pag[mask(pag["Role"])] if len(pag) else pag
    today_r = ptoday[mask(ptoday["Role"])] if len(ptoday) else ptoday

    g_ud = "Zone" if group == "zone" else "Tenant"
    g_ac = "Zone" if group == "zone" else "Branch"

    def gsize(df, col):
        return df.groupby(col).size() if len(df) and col in df.columns else pd.Series(dtype=int)

    def gnun(df, col, val):
        return df.groupby(col)[val].nunique() if len(df) and col in df.columns and val in df.columns else pd.Series(dtype=int)

    logins = gsize(ud_r[ud_r["Login_Result"] == "Logged In"], g_ud)
    # 'Actual' = distinct reps who are in the CRM user list OR active in today's
    # plan (the union), so 'Assigned (Today's Plan)' can never exceed 'Actual'
    # and '% Assigned' never goes over 100%.
    def _ug(df, gcol):
        if not len(df) or gcol not in df.columns or "User_Name" not in df.columns:
            return pd.DataFrame(columns=["_g", "_u"])
        out = df[[gcol, "User_Name"]].copy()
        out.columns = ["_g", "_u"]
        out["_u"] = out["_u"].astype(str).str.strip()
        return out
    _union = pd.concat([_ug(ud_r, g_ud), _ug(today_r, g_ac)], ignore_index=True)
    actual = _union.groupby("_g")["_u"].nunique() if len(_union) else pd.Series(dtype=int)
    users_assigned = gnun(act_r, g_ac, "User_Name")
    locations_planned = gnun(act_r, g_ac, "Activity_Location")
    atloc = ag_r[ag_r["Target_Met"] == "AT_LOCATION"] if len(ag_r) else ag_r
    done = atloc[atloc["Status"] == "COMPLETED"] if len(atloc) else atloc
    users_completing = gnun(done, g_ac, "User_Name")
    locations_reached = gnun(atloc, g_ac, "Assigned_Location")
    total_completed = gsize(done, g_ac)
    # 'Assigned (Today's Plan)' = distinct sales reps (agents) in TODAY'S plan
    # (Today's activity_data), NOT the acted-on subset in agent_activity — this
    # is the true count of reps assigned for the day, and matches Planned Locations.
    assigned_today = gnun(today_r, g_ac, "User_Name")
    planned_locations = gnun(today_r, g_ac, "Activity_Location")

    groups = sorted(set(_clean_groups(actual.index)) | set(_clean_groups(users_assigned.index))
                    | set(_clean_groups(assigned_today.index)))
    df = pd.DataFrame(index=groups)
    for name, s in [("logins", logins), ("users_assigned", users_assigned),
                    ("users_completing", users_completing), ("locations_planned", locations_planned),
                    ("locations_reached", locations_reached), ("total_completed", total_completed),
                    ("actual", actual), ("assigned_today", assigned_today),
                    ("planned_locations", planned_locations)]:
        df[name] = s.reindex(groups).fillna(0).astype(int)
    df["completion_rate"] = _safe_div(df["users_completing"], df["users_assigned"])
    df["pct_reached"] = _safe_div(df["locations_reached"], df["locations_planned"])
    df["pct_assigned"] = _safe_div(df["assigned_today"], df["actual"])

    if role == "agent" and group == "branch" and len(lead):
        plead = lead[lead["PRODUCT"].astype(str).str.upper() == product.upper()]
        if "Branch" in plead.columns:
            captured = plead.groupby("Branch").size()
            accepted = plead[plead["Consent_Status"].astype(str).str.upper() == "ACCEPTED"].groupby("Branch").size()
            df["leads_captured"] = captured.reindex(groups).fillna(0).astype(int)
            df["leads_accepted"] = accepted.reindex(groups).fillna(0).astype(int)
            df["pct_consenting"] = _safe_div(df["leads_accepted"], df["leads_captured"])

    df.index.name = "Zone" if group == "zone" else "Branch"
    return df.sort_index()


def build_summary_blocks(frames, product):
    return {
        "zone_agents": compute_block(frames, product, "agent", "zone"),
        "zone_tls": compute_block(frames, product, "tl", "zone"),
        "branch_agents": compute_block(frames, product, "agent", "branch"),
        "branch_tls": compute_block(frames, product, "tl", "branch"),
    }


def _has_comment(df):
    """A lead is a PROSPECT when its Comment is not empty."""
    if "Comment" not in df.columns:
        return pd.Series(False, index=df.index)
    return df["Comment"].fillna("").astype(str).str.strip() != ""


def build_leads_summary(frames, product):
    lead = frames.get("Lead_Report", pd.DataFrame())
    plead = lead[lead["PRODUCT"].astype(str).str.upper() == product.upper()] if len(lead) else lead
    status = plead["Consent_Status"].fillna("").astype(str).str.upper() if len(plead) else pd.Series(dtype=str)
    counts = {s: int((status == s).sum()) for s in ["ACCEPTED", "NOT PROVIDED", "REJECTED"]}
    # PROSPECTS = leads that have a comment (from the Lead Report), NOT a consent status
    counts["PROSPECTS"] = int(_has_comment(plead).sum()) if len(plead) else 0
    total = int(len(plead))
    row = {"Product": product, **{s.title(): counts[s] for s in counts}, "Grand Total": total}
    pct = {s: (counts[s] / total if total else 0.0) for s in counts}
    return pd.DataFrame([row]), pct, counts, total


def build_lead_breakdowns(frames, product, groupings):
    """Lead counts by each grouping (Zone / Branch / Team). Each table: Total,
    Consented, Not Provided, Rejected, Prospects — sorted most leads -> fewest.
    For 'By Team Leader', leads with no Team_Name fall back to Assigned_To (these
    are usually cluster/zonal managers' own leads)."""
    lead = frames.get("Lead_Report", pd.DataFrame())
    plead = lead[lead["PRODUCT"].astype(str).str.upper() == product.upper()] if len(lead) else lead
    out = {}
    if not len(plead):
        return out
    plead = plead.copy()
    plead["_status"] = plead["Consent_Status"].fillna("").astype(str).str.upper()
    for label, col in groupings:
        if col == "Team_Name":
            g = plead["Team_Name"].fillna("").astype(str).str.strip() if "Team_Name" in plead.columns else pd.Series("", index=plead.index)
            if "Assigned_To" in plead.columns:
                g = g.where(g != "", plead["Assigned_To"].fillna("").astype(str).str.strip())
            plead["_grp"] = g
            gcol = "_grp"
        elif col in plead.columns:
            gcol = col
        else:
            continue
        idx = sorted(_clean_groups(plead.groupby(gcol).size().index))
        if not idx:
            continue
        df = pd.DataFrame(index=idx)
        df["total"] = plead.groupby(gcol).size().reindex(idx).fillna(0).astype(int)
        for key, name in [("ACCEPTED", "consented"), ("NOT PROVIDED", "not_provided"),
                          ("REJECTED", "rejected")]:
            df[name] = plead[plead["_status"] == key].groupby(gcol).size().reindex(idx).fillna(0).astype(int)
        # PROSPECTS = leads with a comment (not a consent status)
        df["prospects"] = plead[_has_comment(plead)].groupby(gcol).size().reindex(idx).fillna(0).astype(int)
        df.index.name = label.replace("By ", "")
        out[label] = df.sort_values("total", ascending=False)
    return out


def build_team_leader_analysis(frames, product):
    """Per-team analysis from agent_activity.Sales_Team — so each Team Leader's
    team is measured. Metrics per team: agents (distinct users), activities,
    completed at location, % completed, locations reached."""
    ag = frames.get("agent_activity", pd.DataFrame())
    if not len(ag) or "Sales_Team" not in ag.columns:
        return pd.DataFrame()
    pag = ag[ag["Product"].astype(str).str.upper() == product.upper()] if "Product" in ag.columns else ag
    pag = pag[pag["Sales_Team"].fillna("").astype(str).str.strip() != ""]
    if not len(pag):
        return pd.DataFrame()
    done = pag[(pag["Target_Met"] == "AT_LOCATION") & (pag["Status"] == "COMPLETED")]
    g = pag.groupby("Sales_Team")
    df = pd.DataFrame(index=sorted(_clean_groups(g.size().index)))
    df["agents"] = g["User_Name"].nunique().reindex(df.index).fillna(0).astype(int)
    df["activities"] = g.size().reindex(df.index).fillna(0).astype(int)
    df["completed"] = done.groupby("Sales_Team").size().reindex(df.index).fillna(0).astype(int)
    df["loc_reached"] = done.groupby("Sales_Team")["Assigned_Location"].nunique().reindex(df.index).fillna(0).astype(int)
    df["pct_completed"] = _safe_div(df["completed"], df["activities"])
    df.index.name = "Sales Team"
    return df.sort_values("pct_completed", ascending=False)


TL_LABELS = {"agents": "Agents", "activities": "Activities", "completed": "Completed @ Location",
             "pct_completed": "% Completed", "loc_reached": "Locations Reached"}


def _totals(df, num_cols):
    t = {c: int(df[c].sum()) for c in num_cols if c in df.columns}
    if "users_assigned" in t:
        t["completion_rate"] = (t.get("users_completing", 0) / t["users_assigned"]) if t["users_assigned"] else 0.0
    if "locations_planned" in t:
        t["pct_reached"] = (t.get("locations_reached", 0) / t["locations_planned"]) if t["locations_planned"] else 0.0
    if "actual" in t:
        t["pct_assigned"] = (t.get("assigned_today", 0) / t["actual"]) if t["actual"] else 0.0
    if "leads_captured" in t:
        t["pct_consenting"] = (t.get("leads_accepted", 0) / t["leads_captured"]) if t["leads_captured"] else 0.0
    return t


NUM_KEYS = ["logins", "users_assigned", "users_completing", "locations_planned",
            "locations_reached", "total_completed", "actual", "assigned_today",
            "planned_locations", "leads_captured", "leads_accepted"]


def _branches_no_visited(blk):
    """Branches whose locations were planned but none reached (per branch block)."""
    return [str(b) for b, r in blk.iterrows()
            if r["locations_planned"] > 0 and r["locations_reached"] == 0]


def _branches_no_activity(blk):
    """Branches with no activities assigned and no locations planned at all."""
    return [str(b) for b, r in blk.iterrows()
            if r["users_assigned"] == 0 and r["locations_planned"] == 0]


def _today_plan_stats(frames, block, product, role):
    """Today's planned locations, distinct assigned people, and avg locations
    per person — the same computation build_today_narrative uses."""
    actual = int(block["actual"].sum())
    today = frames.get("Today's activity_data", pd.DataFrame())
    if len(today) and "Product" in today.columns:
        today = today[today["Product"].astype(str).str.upper() == product.upper()]
    mask = _agent_mask if role == "agent" else _tl_mask
    tr = today[mask(today["Role"])] if len(today) else today
    planned_locs = tr["Activity_Location"].nunique() if len(tr) else 0
    assigned = tr["User_Name"].nunique() if len(tr) else 0
    visits = len(tr)
    avg = round(visits / assigned, 1) if assigned else 0
    return planned_locs, assigned, avg


def build_email_summary(blocks, leads_counts, leads_total, frames=None, product=None):
    """The headline numbers (the old hidden 'Email' sheet), computed from the
    blocks so the workbook and the email can never disagree. Returns rows of
    (Section, Metric, Value, IsPercent) for a readable 'Email Summary' sheet.

    Also carries the extended detail the legacy 'Email' sheet exposed: the
    branch lists (no location visited / no activities assigned) and the today
    plan figures (locations planned, avg locations per person). These are
    computed straight from the branch blocks + Today's activity data — the same
    values the email narrative prints — so the sheet stays the single source."""
    a = _totals(blocks["branch_agents"], NUM_KEYS)
    t = _totals(blocks["branch_tls"], NUM_KEYS)
    rows = []

    def add(section, metric, value, pct=False):
        rows.append({"Section": section, "Metric": metric, "Value": value, "IsPercent": pct})

    add("Leads", "Leads generated", leads_total)
    add("Leads", "Accepted consent", leads_counts["ACCEPTED"])
    add("Leads", "Not provided", leads_counts["NOT PROVIDED"])
    add("Leads", "Rejected", leads_counts["REJECTED"])
    add("Leads", "Prospects", leads_counts["PROSPECTS"])
    add("Leads", "% accepted", (leads_counts["ACCEPTED"] / leads_total if leads_total else 0.0), True)

    for label, blk, block_df, role in [("Sales Agents", a, blocks["branch_agents"], "agent"),
                                       ("Team Leaders", t, blocks["branch_tls"], "tl")]:
        add(label, "Actual on CRM", blk.get("actual", 0))
        add(label, "Logged in", blk.get("logins", 0))
        add(label, "Assigned activities", blk.get("users_assigned", 0))
        add(label, "Completing at location", blk.get("users_completing", 0))
        add(label, "Activity completion rate", blk.get("completion_rate", 0.0), True)
        add(label, "Locations planned", blk.get("locations_planned", 0))
        add(label, "Locations reached", blk.get("locations_reached", 0))
        add(label, "% locations reached", blk.get("pct_reached", 0.0), True)
        add(label, "Assigned in today's plan", blk.get("assigned_today", 0))
        add(label, "% assigned today", blk.get("pct_assigned", 0.0), True)

        # ── extended detail (branch lists + today plan) ──────────────────────
        no_visited  = _branches_no_visited(block_df)
        no_activity = _branches_no_activity(block_df)
        add(label, "Branches with no location visited (count)", len(no_visited))
        add(label, "Branches with no location visited", ", ".join(no_visited) if no_visited else "None")
        add(label, "Branches with no activities assigned (count)", len(no_activity))
        add(label, "Branches with no activities assigned", ", ".join(no_activity) if no_activity else "None")
        if frames is not None and product is not None:
            planned_locs, _assigned, avg = _today_plan_stats(frames, block_df, product, role)
            add(label, "Today locations planned", planned_locs)
            add(label, "Avg locations visited", avg)

    return pd.DataFrame(rows)


# ---- Dynamic narrative (plain-English, recomputed from the data every day) --- #
def _pct2(x):
    return f"{x * 100:.2f}%"


def build_role_narrative(branch_block, noun_plural, noun_singular):
    """The example-style paragraphs for Agents / Team Leaders, data-driven."""
    t = _totals(branch_block, NUM_KEYS)
    no_visited = [b for b, r in branch_block.iterrows()
                  if r["locations_planned"] > 0 and r["locations_reached"] == 0]
    no_activity = [b for b, r in branch_block.iterrows()
                   if r["users_assigned"] == 0 and r["locations_planned"] == 0]
    lines = [
        f"Total count of {noun_plural} in CRM stood at {t['actual']}, and only "
        f"{t['logins']} {noun_plural} logged in for the day.",
        f"Out of {t['users_assigned']} {noun_plural} assigned activities for the day, "
        f"{t['users_completing']} ({_pct2(t['completion_rate'])}) {noun_plural} completed at "
        f"least one activity at the assigned location.",
        f"{t['locations_planned']} locations were planned for the day. Only "
        f"{t['locations_reached']} ({_pct2(t['pct_reached'])}) locations were reached on the day.",
        (("None" if not no_visited else ", ".join(no_visited))
         + f" had no planned location visited by a {noun_singular}."),
        (("None" if not no_activity else ", ".join(no_activity))
         + f" had not assigned activities or planned locations to be visited by a {noun_singular}."),
    ]
    return lines


def build_today_narrative(frames, block, product, role, noun_plural, noun_singular, date_str):
    """The 'Today's Plan' analysis, computed purely from Today's activity_data:
      - locations planned  = distinct Activity_Location today
      - <role> assigned    = distinct agents/TLs with a plan today (of all Actual)
      - avg per <role>     = total planned visits / those assigned agents
    (This is why it is a *today* report — it is the plan for the day ahead.)"""
    actual = int(block["actual"].sum())
    today = frames.get("Today's activity_data", pd.DataFrame())
    if len(today) and "Product" in today.columns:
        today = today[today["Product"].astype(str).str.upper() == product.upper()]
    mask = _agent_mask if role == "agent" else _tl_mask
    tr = today[mask(today["Role"])] if len(today) else today

    planned_locs = tr["Activity_Location"].nunique() if len(tr) else 0
    assigned = tr["User_Name"].nunique() if len(tr) else 0
    visits = len(tr)
    pct = assigned / actual if actual else 0.0
    avg = round(visits / assigned, 1) if assigned else 0
    # date is rendered as a header line above these points (see _today_block)
    return [
        f"{planned_locs} locations have been planned.",
        f"{assigned} ({_pct2(pct)}) {noun_plural} have been assigned activities.",
        f"Average locations to be visited per {noun_singular} is {avg}.",
    ]


def build_leads_narrative(counts, total, product):
    if total == 0:
        return [f"No leads were generated in the system across all {product} branches for the day."]

    def p(n):
        return f"{_pct2(n / total)} ({n})"

    return [
        f"{total} leads were generated in the system across all {product} branches.",
        f"{p(counts['ACCEPTED'])} of leads generated were consented, "
        f"{p(counts['NOT PROVIDED'])} were not provided and {p(counts['REJECTED'])} were rejected.",
        f"Out of {total} leads, {counts['PROSPECTS']} are prospects.",
    ]


def build_tl_team_narrative(tl_df):
    if not len(tl_df):
        return ["No team activity was recorded for the day."]
    best = tl_df.iloc[0]
    worst = tl_df.iloc[-1]
    lines = [f"{len(tl_df)} sales teams were active for the day."]
    lines.append(
        f"{best.name} led with {_pct2(best['pct_completed'])} of activities completed at location, "
        f"while {worst.name} was lowest at {_pct2(worst['pct_completed'])}.")
    zero = [t for t, r in tl_df.iterrows() if r["completed"] == 0]
    if zero:
        lines.append(f"{', '.join(zero)} completed no activity at any assigned location.")
    return lines


def build_comparison_narrative(cmp_df, noun):
    if cmp_df is None or not len(cmp_df):
        return ["No comparison data available."]
    improved = cmp_df[cmp_df["Completion Rate Movement"] > 0]
    declined = cmp_df[cmp_df["Completion Rate Movement"] < 0]
    lines = [f"Compared with yesterday, {len(improved)} zone(s) improved their {noun} activity "
             f"completion rate and {len(declined)} declined."]
    if len(improved):
        top = improved.sort_values("Completion Rate Movement", ascending=False).iloc[0]
        lines.append(f"{top.name} improved the most (+{_pct2(top['Completion Rate Movement'])}).")
    if len(declined):
        low = declined.sort_values("Completion Rate Movement").iloc[0]
        lines.append(f"{low.name} declined the most ({_pct2(low['Completion Rate Movement'])}).")
    return lines


def build_comparison(today_blocks, master_path, product):
    """Zone-level today (this run) vs yesterday (existing master) for Agents/TLs."""
    try:
        sheets = pd.ExcelFile(master_path).sheet_names
        master = {s: pd.read_excel(master_path, sheet_name=s) for s in DATA_SHEETS if s in sheets}
    except Exception as exc:
        print(f"    (comparison skipped — could not read master: {exc})")
        return None, None
    y_blocks = build_summary_blocks(master, product)

    def cmp(role):
        td, yd = today_blocks[f"zone_{role}"], y_blocks[f"zone_{role}"]
        zones = sorted(set(td.index) | set(yd.index))
        out = pd.DataFrame(index=zones)
        out["Today Completion Rate"] = td["completion_rate"].reindex(zones).fillna(0.0)
        out["Yesterday Completion Rate"] = yd["completion_rate"].reindex(zones).fillna(0.0)
        out["Completion Rate Movement"] = out["Today Completion Rate"] - out["Yesterday Completion Rate"]
        out["Today % Reached"] = td["pct_reached"].reindex(zones).fillna(0.0)
        out["Yesterday % Reached"] = yd["pct_reached"].reindex(zones).fillna(0.0)
        out["% Reached Movement"] = out["Today % Reached"] - out["Yesterday % Reached"]
        out.index.name = "Zone"
        return out

    return cmp("agents"), cmp("tls")


# --------------------------------------------------------------------------- #
# Colour-spectrum helper (violet best -> red worst), used by Excel & email.
# --------------------------------------------------------------------------- #
def spectrum_hex(frac):
    """frac in [0,1]; 0 -> violet (best), 1 -> red (worst). Multi-stop lerp."""
    frac = max(0.0, min(1.0, frac))
    seg = frac * (len(SPECTRUM_STOPS) - 1)
    i = min(int(seg), len(SPECTRUM_STOPS) - 2)
    a, b, t = SPECTRUM_STOPS[i], SPECTRUM_STOPS[i + 1], seg - i
    rgb = tuple(round(a[k] + (b[k] - a[k]) * t) for k in range(3))
    return "%02X%02X%02X" % rgb


# =========================================================================== #
# SECTION 3 — WRITE THE DEPARTMENT REPORT WORKBOOK
# The workbook keeps the ORIGINAL simple shape: ONE Summary sheet (all four
# blocks stacked together), a Lead Summary sheet, an Email Summary sheet, plus
# the processed data sheets. Nothing is split across extra tabs. (Freeze panes
# / sticky headers are an EMAIL feature, not used in the workbook.)
# =========================================================================== #
_thin = Side(style="thin", color="D0D7E5")


def _hdr(cell):
    cell.font = Font(name=FONT, bold=True, size=10, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(bottom=Side(style="medium", color=NAVY), left=_thin, right=_thin)


def _write_block(ws, top, title, df, index_label):
    """Write one titled metric block (full columns + TOTAL) starting at row `top`,
    with a violet->red colour scale on the completion-rate column. Returns the
    next free row (a couple of blank rows below the block)."""
    ws.cell(top, 1, title).font = Font(name=FONT, bold=True, size=12, color=NAVY)
    if df is None or not len(df):
        ws.cell(top + 1, 1, "(no data for the day)").font = Font(name=FONT, italic=True, color="808080")
        return top + 3

    col_keys = [k for k, _ in METRIC_COLUMNS if k in df.columns]
    headers = [index_label] + [METRIC_LABELS[k] for k in col_keys]
    hr = top + 1
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(hr, j, h))

    # order rows best -> worst by completion rate (matches the violet->red scale)
    df = df.sort_values("completion_rate", ascending=False) if "completion_rate" in df.columns else df
    first = hr + 1
    rr = first
    for i, (idx, row) in enumerate(df.iterrows()):
        lc = ws.cell(rr, 1, idx)
        lc.font = Font(name=FONT, size=10)
        lc.border = Border(bottom=_thin)
        if i % 2 == 1:
            lc.fill = PatternFill("solid", fgColor=ALT_FILL)
        for j, k in enumerate(col_keys, 2):
            v = row[k]
            cell = ws.cell(rr, j, float(v) if isinstance(v, float) else v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=_thin)
            if k in PCT_KEYS:
                cell.number_format = "0.00%"
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=ALT_FILL)
        rr += 1
    last = rr - 1

    t = _totals(df, NUM_KEYS)
    ws.cell(rr, 1, "TOTAL").font = Font(name=FONT, bold=True, size=10)
    for j, k in enumerate(col_keys, 2):
        if k in t:
            cell = ws.cell(rr, j, float(t[k]) if isinstance(t[k], float) else t[k])
            cell.font = Font(name=FONT, bold=True, size=10)
            cell.alignment = Alignment(horizontal="center")
            if k in PCT_KEYS:
                cell.number_format = "0.00%"
    for j in range(1, len(headers) + 1):
        ws.cell(rr, j).fill = PatternFill("solid", fgColor=TOTAL_FILL)
        ws.cell(rr, j).border = Border(top=Side(style="medium", color=NAVY))

    if "completion_rate" in col_keys and last >= first:
        col_letter = get_column_letter(col_keys.index("completion_rate") + 2)
        ws.conditional_formatting.add(f"{col_letter}{first}:{col_letter}{last}", ColorScaleRule(
            start_type="min", start_color="DC2626",
            mid_type="percentile", mid_value=50, mid_color="FDE047",
            end_type="max", end_color="7C3AED"))
    return last + 4


def _write_named_block(ws, top, title, df, index_label, col_keys, label_map, color_col):
    """A titled table (arbitrary columns) with a violet->red colour scale on
    color_col. Used for the By-Team-Leader and Comparison blocks. Returns next row."""
    ws.cell(top, 1, title).font = Font(name=FONT, bold=True, size=12, color=NAVY)
    if df is None or not len(df):
        ws.cell(top + 1, 1, "(no data for the day)").font = Font(name=FONT, italic=True, color="808080")
        return top + 3
    hr = top + 1
    headers = [index_label] + [label_map.get(k, k) for k in col_keys]
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(hr, j, h))
    first = hr + 1
    rr = first
    for i, (idx, row) in enumerate(df.iterrows()):
        ws.cell(rr, 1, idx).font = Font(name=FONT, size=10)
        for j, k in enumerate(col_keys, 2):
            v = row[k]
            c = ws.cell(rr, j, float(v) if isinstance(v, float) else v)
            c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(horizontal="center")
            if is_pct_key(k):
                c.number_format = "0.00%"
        if i % 2 == 1:
            for j in range(1, len(headers) + 1):
                ws.cell(rr, j).fill = PatternFill("solid", fgColor=ALT_FILL)
        rr += 1
    last = rr - 1
    if color_col in col_keys and last >= first:
        cl = get_column_letter(col_keys.index(color_col) + 2)
        ws.conditional_formatting.add(f"{cl}{first}:{cl}{last}", ColorScaleRule(
            start_type="min", start_color="DC2626",
            mid_type="percentile", mid_value=50, mid_color="FDE047",
            end_type="max", end_color="7C3AED"))
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 30)
    for j in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    return last + 3


def _lead_break_block(ws, top, title, df):
    """Write one lead-breakdown table (Total + statuses) with a colour scale on
    Total (most leads = violet). Returns next free row."""
    ws.cell(top, 1, title).font = Font(name=FONT, bold=True, size=12, color=NAVY)
    hr = top + 1
    headers = [df.index.name or "Group"] + [LEAD_LABELS[k] for k in LEAD_KEYS]
    for j, h in enumerate(headers, 1):
        _hdr(ws.cell(hr, j, h))
    first = hr + 1
    rr = first
    for i, (idx, row) in enumerate(df.iterrows()):
        ws.cell(rr, 1, idx).font = Font(name=FONT, size=10)
        for j, k in enumerate(LEAD_KEYS, 2):
            c = ws.cell(rr, j, int(row[k])); c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(horizontal="center")
        if i % 2 == 1:
            for j in range(1, len(headers) + 1):
                ws.cell(rr, j).fill = PatternFill("solid", fgColor=ALT_FILL)
        rr += 1
    last = rr - 1
    if last >= first:
        ws.conditional_formatting.add(f"B{first}:B{last}", ColorScaleRule(
            start_type="min", start_color="DC2626",
            mid_type="percentile", mid_value=50, mid_color="FDE047",
            end_type="max", end_color="7C3AED"))
    # TOTAL row (sum of each column)
    ws.cell(rr, 1, "TOTAL").font = Font(name=FONT, bold=True, size=10)
    for j, k in enumerate(LEAD_KEYS, 2):
        c = ws.cell(rr, j, int(df[k].sum())); c.font = Font(name=FONT, bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
    for j in range(1, len(headers) + 1):
        ws.cell(rr, j).fill = PatternFill("solid", fgColor=TOTAL_FILL)
        ws.cell(rr, j).border = Border(top=Side(style="medium", color=NAVY))
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 26 if j == 1 else 14
    return rr + 3


# Login_Result colours (Excel "good/bad" palette)
GREEN_FILL, GREEN_FONT = "C6EFCE", "006100"
RED_FILL, RED_FONT = "FFC7CE", "9C0006"


def _write_raw_sheet(wb, name, df):
    """Raw data sheet with the configured columns dropped, header row frozen, and
    an AutoFilter on the header so any column can be filtered. For user_data,
    Last_Login and Login_Result are moved to the end and Login_Result is coloured
    green (Logged In) / red (Not Logged In)."""
    df = df.drop(columns=[c for c in DROP_COLUMNS.get(name, []) if c in df.columns], errors="ignore")
    if name == "user_data":
        tail = [c for c in ["Last_Login", "Login_Result"] if c in df.columns]
        df = df[[c for c in df.columns if c not in tail] + tail]
    login_col = (df.columns.get_loc("Login_Result") + 1) if name == "user_data" and "Login_Result" in df.columns else None

    ws = wb.create_sheet(name[:31])
    ws.sheet_view.showGridLines = False
    for j, col in enumerate(df.columns, 1):
        cell = ws.cell(1, j, str(col))
        cell.font = Font(name=FONT, bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEAD_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(j)].width = 18
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for j, v in enumerate(row, 1):
            cell = ws.cell(i, j, v)
            cell.font = Font(name=FONT, size=9)
            if j == login_col:
                val = str(v).strip().lower()
                if val == "logged in":
                    cell.fill = PatternFill("solid", fgColor=GREEN_FILL)
                    cell.font = Font(name=FONT, size=9, bold=True, color=GREEN_FONT)
                elif val == "not logged in":
                    cell.fill = PatternFill("solid", fgColor=RED_FILL)
                    cell.font = Font(name=FONT, size=9, bold=True, color=RED_FONT)
    n_rows = len(df) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(max(1, len(df.columns)))}{max(1, n_rows)}"
    ws.freeze_panes = "A2"


def write_report_workbook(out_path, frames, blocks, leads_df, leads_pct, leads_total,
                          lead_breaks, email_summary, tl_df, comparison, cur_date, prev_date):
    """Original shape: Summary (4 blocks + By Team Leader) + Lead Summary +
    Email Summary + Comparison + data sheets."""
    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary: the four blocks + By Team Leader, all on ONE sheet ---
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    r = 1
    r = _write_block(ws, r, "SALES AGENTS  —  by ZONE", blocks["zone_agents"], "Zone")
    r = _write_block(ws, r, "TEAM LEADERS  —  by ZONE", blocks["zone_tls"], "Zone")
    r = _write_block(ws, r, "SALES AGENTS  —  by BRANCH", blocks["branch_agents"], "Branch")
    r = _write_block(ws, r, "TEAM LEADERS  —  by BRANCH", blocks["branch_tls"], "Branch")
    tl_keys = ["agents", "activities", "completed", "pct_completed", "loc_reached"]
    _write_named_block(ws, r, "BY TEAM LEADER  (each Team Leader's team)", tl_df,
                       "Sales Team", tl_keys, TL_LABELS, "pct_completed")
    ws.column_dimensions["A"].width = 30
    for j in range(2, 18):
        ws.column_dimensions[get_column_letter(j)].width = 14

    # --- Lead Summary: totals + % row, then the breakdown tables ---
    ws = wb.create_sheet("Lead Summary")
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "LEAD SUMMARY").font = Font(name=FONT, bold=True, size=13, color=NAVY)
    for j, h in enumerate(leads_df.columns, 1):
        _hdr(ws.cell(3, j, str(h)))
    for j, v in enumerate(leads_df.iloc[0].tolist(), 1):
        ws.cell(4, j, v).font = Font(name=FONT, size=10)
        ws.cell(4, j).alignment = Alignment(horizontal="center")
    # percentage row (share of total) under the counts
    ws.cell(5, 1, "% of Total").font = Font(name=FONT, bold=True, size=10, italic=True)
    status_by_col = {"Accepted": "ACCEPTED", "Not Provided": "NOT PROVIDED",
                     "Rejected": "REJECTED", "Prospects": "PROSPECTS"}
    for j, col in enumerate(leads_df.columns, 1):
        if col in status_by_col:
            c = ws.cell(5, j, leads_pct[status_by_col[col]]); c.number_format = "0.00%"
            c.font = Font(name=FONT, size=10, italic=True, color="1F3864")
        elif col == "Grand Total":
            c = ws.cell(5, j, 1.0 if leads_total else 0.0); c.number_format = "0.00%"
            c.font = Font(name=FONT, size=10, italic=True, color="1F3864")
    for j in range(1, len(leads_df.columns) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    rr = 7
    for label, bdf in lead_breaks.items():
        rr = _lead_break_block(ws, rr, f"LEADS  —  {label}", bdf)

    # --- Email Summary (headline numbers) ---
    ws = wb.create_sheet("Email Summary")
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, "EMAIL SUMMARY").font = Font(name=FONT, bold=True, size=13, color=NAVY)
    for j, h in enumerate(["Section", "Metric", "Value"], 1):
        _hdr(ws.cell(3, j, h))
    rr = 4
    for _, row in email_summary.iterrows():
        ws.cell(rr, 1, row["Section"]).font = Font(name=FONT, size=10, italic=True, color="808080")
        ws.cell(rr, 2, row["Metric"]).font = Font(name=FONT, size=10)
        cell = ws.cell(rr, 3, float(row["Value"]) if row["IsPercent"] else row["Value"])
        cell.font = Font(name=FONT, size=10, bold=True)
        cell.alignment = Alignment(horizontal="center")
        if row["IsPercent"]:
            cell.number_format = "0.00%"
        rr += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 60

    # --- Comparison (today's run vs the previous master) ---
    cmp_agents, cmp_tls = comparison if comparison else (None, None)
    if cmp_agents is not None:
        ws = wb.create_sheet("Comparison")
        ws.sheet_view.showGridLines = False
        cmp_cols = list(cmp_agents.columns)
        cmp_labels = {
            "Today Completion Rate": f"{cur_date} Completion Rate",
            "Yesterday Completion Rate": f"{prev_date} Completion Rate",
            "Completion Rate Movement": "Completion Rate Movement",
            "Today % Reached": f"{cur_date} % Reached",
            "Yesterday % Reached": f"{prev_date} % Reached",
            "% Reached Movement": "% Reached Movement",
        }
        r = 1
        r = _write_named_block(ws, r, f"SALES AGENTS  —  {cur_date} vs {prev_date}",
                               cmp_agents, "Zone", cmp_cols, cmp_labels, "Completion Rate Movement")
        r = _write_named_block(ws, r, f"TEAM LEADERS  —  {cur_date} vs {prev_date}",
                               cmp_tls, "Zone", cmp_cols, cmp_labels, "Completion Rate Movement")
        ws.column_dimensions["A"].width = 26
        for j in range(2, len(cmp_cols) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 20

    # --- The processed data sheets ---
    for name in DATA_SHEETS:
        if name in frames:
            _write_raw_sheet(wb, name, frames[name])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# =========================================================================== #
# SECTION 4 — EMAIL  (prose + compact colour-graded HTML tables)
# =========================================================================== #
def load_recipients(mode, email_col):
    sheet = {"actual": "Actual", "specific": "Specific", "test": "Test"}[mode]
    df = pd.read_excel(EMAIL_MASTER, sheet_name=sheet, header=3)
    if email_col not in df.columns:
        return []
    return [str(e).strip() for e in df[email_col].dropna() if "@" in str(e)]


# --------------------------------------------------------------------------- #
# Email visual language.  Cell styling is INLINE on every cell (email clients —
# Gmail especially — strip <style> blocks, so inline is the only reliable way).
# We keep the email small enough to avoid Gmail's ~102 KB clip by NOT duplicating
# tables (no expand copy) and by showing the long By-Team-Leader table only in
# the Excel workbook, not the email.
# --------------------------------------------------------------------------- #
SPECTRUM_CSS = "linear-gradient(90deg,#7c3aed,#2563eb,#16a34a,#eab308,#ea580c,#dc2626)"
BLUE = "#1d4ed8"
DARK = "#1e3a8a"

_BOX = ("max-width:100%;overflow-x:auto;border:1px solid #cbd5e1;"
        "box-shadow:0 2px 8px rgba(30,58,138,0.10);margin:8px auto 18px;")
_TABLE = "border-collapse:separate;border-spacing:0;width:max-content;min-width:100%;"
_THBASE = ("padding:9px 13px;background:#2f5597;color:#fff;font-size:12px;font-weight:600;"
           "border-bottom:2px solid #24466f;white-space:nowrap;")
_TH_TOP = _THBASE + "text-align:center;position:-webkit-sticky;position:sticky;top:0;z-index:2;"
_TH_CORNER = _THBASE + "text-align:left;position:-webkit-sticky;position:sticky;top:0;left:0;z-index:3;"
_TD = "padding:7px 13px;border-bottom:1px solid #eef1f6;font-size:12px;color:#222;text-align:center;white-space:nowrap;"
_TD_LEFT = ("padding:7px 13px;border-bottom:1px solid #eef1f6;font-size:12px;color:#1f2937;text-align:left;"
            "font-weight:600;white-space:nowrap;background:#ffffff;position:-webkit-sticky;position:sticky;left:0;z-index:1;")


def _fmt(v, pct=False):
    if pct:
        return f"{v * 100:.2f}%"
    if isinstance(v, float):
        return f"{v:.1f}" if v % 1 else f"{int(v)}"
    return str(v)


def _hl(text):
    """Bold + blue the important details (whole dates, numbers, percentages).
    Single pass over plain text so it never corrupts inserted markup."""
    pat = r"(\d{1,2}[-/]\d{2}[-/]\d{4}|\d[\d,]*(?:\.\d+)?%?)"
    return re.sub(pat, r'<b style="color:#1d4ed8">\1</b>', text)


def _header_banner(product, date_str):
    """Centered header banner framed top and bottom by full-width spectrum lines
    (no rounded corners — a clean, sharp masthead)."""
    rep = pd.to_datetime(date_str, format="%d/%m/%Y").strftime("%d %B %Y")
    gen = datetime.now().strftime("%d-%m-%Y")
    return f"""
  <div style="height:6px;background:{SPECTRUM_CSS};"></div>
  <div style="background:linear-gradient(135deg,#24466f 0%,#2f5597 55%,#3a67b0 100%);padding:30px 22px;text-align:center;">
    <div style="font-size:12px;letter-spacing:4px;color:#cfe0f5;text-transform:uppercase;margin-bottom:9px;">Platinum Credit &nbsp;&bull;&nbsp; CRM</div>
    <div style="font-size:33px;font-weight:800;color:#ffffff;letter-spacing:0.5px;">{product} CRM User Report</div>
    <div style="font-size:14px;color:#dbeafe;margin-top:11px;">Report Date: <b style="color:#fff">{rep}</b> &nbsp;&bull;&nbsp; Generated on: <b style="color:#fff">{gen}</b></div>
  </div>
  <div style="height:6px;background:{SPECTRUM_CSS};margin-bottom:20px;"></div>"""


def _subheader(text):
    return f"""
  <h3 style="text-align:center;color:{BLUE};font-size:21px;font-weight:800;margin:30px 0 6px;letter-spacing:0.3px;">{text}</h3>
  <div style="height:3px;width:250px;margin:0 auto 14px;background:{SPECTRUM_CSS};"></div>"""


def _prose(lines, icon="◆"):
    """One paragraph per point, each led by a spectrum-blue icon, with the key
    figures bolded in blue and generous spacing between points."""
    items = "".join(
        f'<p style="font-size:13.5px;line-height:1.6;margin:10px 0;color:#333;">'
        f'<span style="color:{BLUE};font-size:12px;margin-right:9px;">{icon}</span>{_hl(ln)}</p>'
        for ln in lines)
    return f'<div style="max-width:820px;margin:0 auto 10px;">{items}</div>'


def _today_block(date_str, lines):
    """'For today <date>' header line, then the three Today's-Plan points."""
    d = pd.to_datetime(date_str, format="%d/%m/%Y").strftime("%d-%m-%Y")
    hdr = (f'<p style="text-align:center;font-weight:700;color:{DARK};font-size:14.5px;'
           f'margin:16px 0 2px;">For today <b style="color:{BLUE}">{d}</b></p>')
    return hdr + _prose(lines, icon="★")


def _graded_table(df, index_label, key_metric, col_keys, label_map, higher_better=True):
    """One inline-styled table, sorted by key_metric, the key cell shaded across
    the violet->red spectrum (best/most -> worst/least). Full table (no expand
    copy) so the email stays small and renders reliably in Gmail."""
    if df is None or not len(df):
        return "<p style='color:#888;font-size:12px;text-align:center;'>(no data for the day)</p>"
    dfx = df.sort_values(key_metric, ascending=not higher_better) if key_metric in df.columns else df
    n = len(dfx)
    heads = "".join(f'<th style="{_TH_TOP}">{label_map.get(k, k)}</th>' for k in col_keys)
    body = ""
    for rank, (idx, row) in enumerate(dfx.iterrows()):
        frac = rank / (n - 1) if n > 1 else 0.0     # 0 = best/most (violet), 1 = worst/least (red)
        tds = f'<td style="{_TD_LEFT}">{idx}</td>'
        for k in col_keys:
            style = _TD + (f"background:#{spectrum_hex(frac)};color:#fff;font-weight:700;" if k == key_metric else "")
            tds += f'<td style="{style}">{_fmt(row[k], is_pct_key(k))}</td>'
        body += f"<tr>{tds}</tr>"
    return (f'<div style="{_BOX}"><table style="{_TABLE}">'
            f'<thead><tr><th style="{_TH_CORNER}">{index_label}</th>{heads}</tr></thead>'
            f'<tbody>{body}</tbody></table></div>')


def _leads_table(leads_df, leads_pct, leads_total):
    """Counts row + a % row (share of total) beneath it."""
    heads = "".join(f'<th style="{_TH_TOP}">{c}</th>' for c in leads_df.columns)
    cells = "".join(f'<td style="{_TD}">{_fmt(v)}</td>' for v in leads_df.iloc[0].tolist())
    status_by_col = {"Accepted": "ACCEPTED", "Not Provided": "NOT PROVIDED",
                     "Rejected": "REJECTED", "Prospects": "PROSPECTS"}
    pct_cells = ""
    for c in leads_df.columns:
        pct_style = _TD + "color:#1f3864;font-style:italic;"
        if c in status_by_col:
            pct_cells += f'<td style="{pct_style}">{_fmt(leads_pct[status_by_col[c]], True)}</td>'
        elif c == "Grand Total":
            pct_cells += f'<td style="{pct_style}">{_fmt(1.0 if leads_total else 0.0, True)}</td>'
        else:
            pct_cells += f'<td style="{pct_style}">% of Total</td>'
    return (f'<div style="{_BOX}"><table style="{_TABLE}">'
            f'<thead><tr>{heads}</tr></thead>'
            f'<tbody><tr>{cells}</tr><tr>{pct_cells}</tr></tbody></table></div>')


def render_email_html(product, date_str, blocks, tl_df, narr, leads_df, leads_pct,
                      leads_total, lead_breaks, group, cur_date, prev_date):
    idx_label = "Zone" if group == "zone" else "Branch"
    agents = blocks[f"{group}_agents"]
    tls = blocks[f"{group}_tls"]

    cmp_html = ""
    if narr.get("cmp_agents") is not None:
        cmp_cols = ["Today Completion Rate", "Yesterday Completion Rate", "Completion Rate Movement",
                    "Today % Reached", "Yesterday % Reached", "% Reached Movement"]
        cmp_labels = {"Today Completion Rate": f"{cur_date} Completion", "Yesterday Completion Rate": f"{prev_date} Completion",
                      "Completion Rate Movement": "Completion Movement", "Today % Reached": f"{cur_date} % Reached",
                      "Yesterday % Reached": f"{prev_date} % Reached", "% Reached Movement": "% Reached Movement"}
        cmp_html = (
            _subheader(f"Comparison — Sales Agents ({cur_date} vs {prev_date})")
            + _prose(narr["cmp_agents"])
            + _graded_table(narr["_cmp_agents"], "Zone", "Completion Rate Movement", cmp_cols, cmp_labels)
            + _subheader(f"Comparison — Team Leaders ({cur_date} vs {prev_date})")
            + _prose(narr["cmp_tls"])
            + _graded_table(narr["_cmp_tls"], "Zone", "Completion Rate Movement", cmp_cols, cmp_labels))

    return f"""
<div style="font-family:Arial,Helvetica,sans-serif;max-width:920px;margin:0 auto;color:#222;background:#ffffff;padding:8px 12px;">
  {_header_banner(product, date_str)}

  {_subheader("Leads")}
  {_prose(narr['leads'])}
  {_leads_table(leads_df, leads_pct, leads_total)}

  {_subheader("Sales Agents")}
  {_prose(narr['agents'])}
  {_today_block(date_str, narr['agents_today'])}
  {_graded_table(agents, idx_label, "completion_rate", EMAIL_METRIC_KEYS, METRIC_LABELS)}

  {_subheader("Team Leaders")}
  {_prose(narr['tls'])}
  {_today_block(date_str, narr['tls_today'])}
  {_graded_table(tls, idx_label, "completion_rate", EMAIL_METRIC_KEYS, METRIC_LABELS)}

  {cmp_html}

  <div style="border-top:1px solid #e5e7eb;margin-top:28px;padding-top:10px;text-align:center;">
    <p style="font-size:11px;color:#9ca3af;margin:0;">Automated CRM report &nbsp;&bull;&nbsp; Platinum Credit Limited &nbsp;&bull;&nbsp; generated {datetime.now():%d %b %Y %H:%M}</p>
  </div>
</div>"""


def _email_creds():
    try:
        from dotenv import load_dotenv
        load_dotenv(CRM_ENV)
        load_dotenv()
    except Exception:
        pass
    return os.getenv("EMAIL_USERNAME"), os.getenv("EMAIL_PASSWORD")


def send_email(recipients, subject, html, attachment):
    sender, password = _email_creds()
    if not sender or not password:
        print("    [email] EMAIL_USERNAME/EMAIL_PASSWORD not set — not sent.")
        return False
    if not recipients:
        print("    [email] no recipients — not sent.")
        return False
    msg = MIMEMultipart("mixed")
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html, "html"))
    msg.attach(alt)
    if attachment and Path(attachment).exists():
        with open(attachment, "rb") as fh:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(fh.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{Path(attachment).name}"')
        msg.attach(part)
    with smtplib.SMTP("smtp.gmail.com", 587, timeout=120) as server:
        server.starttls(context=ssl.create_default_context())
        server.login(sender, password)
        server.sendmail(sender, recipients, msg.as_string())
    print(f"    [email] sent to {len(recipients)} recipient(s): {', '.join(recipients[:3])}"
          + (" ..." if len(recipients) > 3 else ""))
    return True


# =========================================================================== #
# SECTION 5 — ORCHESTRATION
# =========================================================================== #
def run_lead_correction():
    script = ROW_DIR / "correct_lead_report.py"
    if not script.exists():
        print("  (lead correction script not found — skipped)")
        return
    spec = importlib.util.spec_from_file_location("correct_lead_report", script)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    mod.main()


def main():
    ap = argparse.ArgumentParser(description="Build all CRM department reports in one run (no pivots/xlwings).")
    ap.add_argument("date", nargs="?", help="report/login date dd/mm/yyyy (prompted if omitted)")
    ap.add_argument("--only", help="comma-separated department codes, e.g. CS,SME")
    ap.add_argument("--skip-lead-fix", action="store_true", help="skip the lead correction step")
    ap.add_argument("--send", choices=["no", "test", "specific", "actual"], default=None,
                    help="who to email (default: ask). Previews are always written.")
    ap.add_argument("--update-rows", choices=["yes", "no"], default=None,
                    help="replace the ROW_EXCEL row files at the end (default: ask)")
    args = ap.parse_args()

    date_str = args.date or input("Enter the report date (dd/mm/yyyy): ").strip()
    manual_date = pd.to_datetime(date_str, format="%d/%m/%Y")
    date_tag = date_str.replace("/", "_")

    departments = DEPARTMENTS
    if args.only:
        wanted = {c.strip().upper() for c in args.only.split(",")}
        departments = [d for d in DEPARTMENTS if d["code"] in wanted]

    if not args.skip_lead_fix:
        print("=" * 74, "\nSTEP 1  Correcting the lead report\n", "=" * 74, sep="")
        run_lead_correction()

    print("\n" + "=" * 74, "\nSTEP 2  Shared processing (reading sources once)\n", "=" * 74, sep="")
    src = detect_source_files()
    activities_dfs = [(p, pd.read_excel(p)) for p in src["activities"]]
    shared = {
        "agent": pd.read_excel(src["agent"]) if src["agent"] else None,
        "lead": pd.read_excel(src["lead"]) if src["lead"] else None,
        "activities": activities_dfs[0][1] if activities_dfs else None,
        "today": None,
    }
    today_df, today_date = None, None
    for p, df in activities_dfs:
        d = latest_start_date(df)
        if d and (today_date is None or d > today_date):
            today_date, today_df = d, df
    shared["today"] = today_df if today_df is not None else (activities_dfs[0][1] if activities_dfs else None)
    base_users = build_base_users(pd.read_excel(src["user_list"]), manual_date)
    print(f"  users after filtering: {len(base_users)}")

    print("\n" + "=" * 74, "\nSTEP 3  Building reports (computed summaries + narratives)\n", "=" * 74, sep="")
    results, made = {}, []
    for dept in departments:
        code, product = dept["code"], dept["product"]
        master = src["masters"].get(code)
        print(f"\n[{code}]")
        if master is None:
            print(f"  no {dept['prefix']}* master in ROW_EXCEL — skipped")
            continue
        try:
            crm_user_data = pd.read_excel(master, sheet_name="user_data")
            frames = build_department_frames(base_users, crm_user_data, shared)
            blocks = build_summary_blocks(frames, product)
            leads_df, leads_pct, leads_counts, leads_total = build_leads_summary(frames, product)
            lead_breaks = build_lead_breakdowns(frames, product, dept["lead_groups"])
            tl_df = build_team_leader_analysis(frames, product)
            email_summary = build_email_summary(blocks, leads_counts, leads_total, frames, product)
            cmp_agents, cmp_tls = build_comparison(blocks, master, product)

            narr = {
                "agents": build_role_narrative(blocks["branch_agents"], "agents", "Sales Agent"),
                "agents_today": build_today_narrative(frames, blocks["branch_agents"], product, "agent", "agents", "Sales Agent", date_str),
                "tls": build_role_narrative(blocks["branch_tls"], "TLs", "TL"),
                "tls_today": build_today_narrative(frames, blocks["branch_tls"], product, "tl", "TLs", "TL", date_str),
                "leads": build_leads_narrative(leads_counts, leads_total, product),
                "tl_team": build_tl_team_narrative(tl_df),
                "cmp_agents": build_comparison_narrative(cmp_agents, "agents") if cmp_agents is not None else None,
                "cmp_tls": build_comparison_narrative(cmp_tls, "team leaders") if cmp_tls is not None else None,
                "_cmp_agents": cmp_agents, "_cmp_tls": cmp_tls,
            }

            # dates for the comparison labels: current run vs the previous master
            cur_date = manual_date.strftime("%d-%m-%Y")
            m = re.search(r"(\d{2})_(\d{2})_(\d{4})", master.name)
            prev_date = f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else "previous"

            # Workbook keeps the original shape (Summary + Lead Summary + Email Summary + data)
            out_path = dept["out_dir"] / f"{dept['out_prefix']}_{date_tag}.xlsx"
            write_report_workbook(out_path, frames, blocks, leads_df, leads_pct, leads_total,
                                  lead_breaks, email_summary, tl_df, (cmp_agents, cmp_tls),
                                  cur_date, prev_date)

            html = render_email_html(product, date_str, blocks, tl_df, narr, leads_df, leads_pct,
                                     leads_total, lead_breaks, dept["email_group"], cur_date, prev_date)
            preview = out_path.with_name(f"{dept['out_prefix']}_{date_tag}_email_preview.html")
            preview.write_text(html, encoding="utf-8")

            za = blocks["zone_agents"]
            print(f"  agents actual={int(za['actual'].sum())} logins={int(za['logins'].sum())} | "
                  f"leads={leads_total} | teams={len(tl_df)}")
            print(f"  saved: {out_path.name}  (+ preview {preview.name})")
            results[code] = {"dept": dept, "out": out_path, "html": html}
            made.append(code)
        except Exception as exc:
            import traceback
            print(f"  FAILED: {exc}")
            traceback.print_exc()

    print("\n" + "=" * 74, f"\nReports built: {', '.join(made) or 'none'}\n", "=" * 74, sep="")

    send_mode = args.send or input("\nSend emails? [no/test/specific/actual] (default no): ").strip().lower() or "no"
    if send_mode != "no":
        # The 'specific' send is the MID-DAY distribution — flag it in both the
        # subject line and the email masthead so recipients can tell it apart
        # from the end-of-day report (e.g. "LBF CRM MID-DAY REPORT - 11 July 2026").
        mid_day = send_mode == "specific"
        label = "MID-DAY " if mid_day else ""
        print(f"\nSending emails ({send_mode})...")
        for code in made:
            r = results[code]
            product = r["dept"]["product"]
            recips = load_recipients(send_mode, r["dept"]["email_col"])
            html = r["html"]
            if mid_day:
                html = html.replace(f"{product} CRM User Report", f"{product} CRM MID-DAY Report")
            subject = f"{product} CRM {label}REPORT - {manual_date:%d %B %Y}"
            print(f"  [{code}]")
            send_email(recips, subject, html, r["out"])
    else:
        print("Emails not sent. Preview .html files are in each NEW_EXCEL folder.")

    upd = args.update_rows or input(
        "\nUpdate (replace) the ROW_EXCEL row files with these new files? [yes/no] (default no): ").strip().lower() or "no"
    if upd in ("yes", "y"):
        if ROW_UPDATE_SCRIPT.exists():
            print("Updating row files...")
            subprocess.run([sys.executable, str(ROW_UPDATE_SCRIPT)], check=False)
        else:
            print(f"  row-update script not found: {ROW_UPDATE_SCRIPT}")
    else:
        print("Row files left unchanged.")

    print("\nDone.")


if __name__ == "__main__":
    main()
