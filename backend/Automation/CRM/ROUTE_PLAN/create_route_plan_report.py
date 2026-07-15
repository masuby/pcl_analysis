"""
create_route_plan_report.py
============================
Build polished ROUTE PLAN reports per product (CS, SME, AGRI, LBF) from the
CRM Activities export, enriched with Zone + Product looked up from the latest
CS CRM file's `user_data` sheet.

PIPELINE
--------
1. Ensure folders exist under ROUTE_PLAN/ :  CS, SME, AGRI, LBF, ROW_FILES
2. Stage source files into ROUTE_PLAN/ROW_FILES/:
       - the Activities_Report*.xlsx
       - the latest CS_CRM*.xlsx from CRM/ROW_EXCEL
3. VLOOKUP-style enrichment:
       Activities.Branch  ->  user_data.Tenant  =>  (Zone, Product)
4. Split activity rows by Product into CS / LBF / SME / AGRI.
5. For each product with data, write a 3-sheet workbook:
       Sheet 1  Summary       — KPI cards + Type / Status / Zone breakdowns + top TLs/Agents
       Sheet 2  TL Summary    — one row per Team Leader (Created_By)
       Sheet 3  Agent Summary — one row per (Team Leader, Agent)
   Saved as:  ROUTE_PLAN/<PRODUCT>/ROUTE_PLAN_FROM_<low>_<max>.xlsx
              (dates = Expected_Start_Date range, DD-MM-YYYY)

Definitions
-----------
  Route        = one planned activity row.
  Team Leader  = Created_By  (the person who planned the route).
  Agent        = Agent_Name  (the person who runs it).
  Plan Days    = distinct Expected_Start_Date calendar days.
"""

from __future__ import annotations

import os
import re
import shutil
import smtplib
import ssl
from collections import Counter, defaultdict
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover
    load_dotenv = None

# --------------------------------------------------------------------------
# PATHS
# --------------------------------------------------------------------------
CRM_DIR        = Path(r"C:\Users\Daniel\Desktop\code\pcl\CRM")
ROUTE_PLAN_DIR = CRM_DIR / "ROUTE_PLAN"
ROW_EXCEL_DIR  = CRM_DIR / "ROW_EXCEL"
ROW_FILES_DIR  = ROUTE_PLAN_DIR / "ROW_FILES"
PRODUCT_FOLDERS = ["CS", "SME", "AGRI", "LBF"]

# Activity types — known ones first, extras appended dynamically
KNOWN_TYPES = ["General Sales", "Phone Call", "Site Visit", "Data Capture", "Event", "Collection"]

# --------------------------------------------------------------------------
# PALETTE per product (dark header / mid / light / very-light tint / accent)
# --------------------------------------------------------------------------
PALETTES = {
    "CS":   {"dark": "1E3A8A", "mid": "2563EB", "light": "DBEAFE", "tint": "EFF6FF", "accent": "3B82F6"},
    "LBF":  {"dark": "0F766E", "mid": "0D9488", "light": "CCFBF1", "tint": "F0FDFA", "accent": "14B8A6"},
    "SME":  {"dark": "B45309", "mid": "D97706", "light": "FEF3C7", "tint": "FFFBEB", "accent": "F59E0B"},
    "AGRI": {"dark": "166534", "mid": "16A34A", "light": "DCFCE7", "tint": "F0FDF4", "accent": "22C55E"},
}
SLATE = "0F172A"
MUTED = "475569"
GRIDC = "CBD5E1"
ALT_ROW = "F8FAFC"
WHITE = "FFFFFF"

INT_FMT = "#,##0"
PCT_FMT = "0.0%"
DATE_FMT = "DD-MM-YYYY"

# daily-coverage status colours (fill + text)
GOOD_FILL, GOOD_TXT = "DCFCE7", "166534"   # >= 2 activities -> on target
WARN_FILL, WARN_TXT = "FEF3C7", "92400E"   # exactly 1 activity -> below target
BAD_FILL,  BAD_TXT  = "FEE2E2", "991B1B"   # 0 activities -> no plan
DAILY_TARGET = 2                            # min activities per agent per day

PRODUCT_FULL = {"CS": "Civil Servant (CS)", "LBF": "Logbook Finance (LBF)",
                "SME": "Small and Medium Enterprise (SME)", "AGRI": "Agrifinance (AGRI)"}

# --------------------------------------------------------------------------
# EMAIL recipients  (dorice + daniel always; product owners added per product)
# --------------------------------------------------------------------------
ENV_PATH = CRM_DIR / ".env"
EMAIL_BASE = ["dorice@platinumcredit.co.tz", "daniel@platinumcredit.co.tz"]
EMAIL_EXTRA = {
    "CS":   ["vivian@platinumcredit.co.tz", "kelvin.mwasala@platinumcredit.co.tz"],
    "LBF":  ["augustine@platinumcredit.co.tz", "denis.albert@platinumcredit.co.tz",
             "irene.mmari@platinumcredit.co.tz"],
    "SME":  ["abdulhakim.khalfan@platinumcredit.co.tz",
             "abdulhakim.khalfan.platinum@gmail.com"],
    "AGRI": ["allan@platinumcredit.co.tz"],
}


def recipients_for(product: str) -> list[str]:
    seen, out = set(), []
    for e in EMAIL_BASE + EMAIL_EXTRA.get(product, []):
        if e.lower() not in seen:
            seen.add(e.lower())
            out.append(e)
    return out


# --------------------------------------------------------------------------
# small style helpers
# --------------------------------------------------------------------------
def solid(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def side(color: str = GRIDC, style: str = "thin") -> Side:
    return Side(style=style, color=color)


def box(color: str = GRIDC) -> Border:
    s = side(color)
    return Border(top=s, bottom=s, left=s, right=s)


def clean(v) -> str:
    if v is None:
        return ""
    s = re.sub(r"\s+", " ", str(v)).strip()
    if s.lower() in ("none", "null", "n/a", "nan"):
        return ""
    return s


def parse_dt(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("none", "null", "n/a", "nan"):
        return None
    for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
                "%d/%m/%Y %H:%M", "%d/%m/%Y", "%m/%d/%Y %H:%M", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def canon_product(p):
    if p is None:
        return None
    s = str(p).strip().upper()
    if s.startswith("CS"):
        return "CS"
    if s.startswith("LBF"):
        return "LBF"
    if s.startswith("SME"):
        return "SME"
    if "AGRI" in s:
        return "AGRI"
    return None


# --------------------------------------------------------------------------
# STEP 1-2  folders + staging
# --------------------------------------------------------------------------
def setup_folders():
    ROW_FILES_DIR.mkdir(parents=True, exist_ok=True)
    for p in PRODUCT_FOLDERS:
        (ROUTE_PLAN_DIR / p).mkdir(parents=True, exist_ok=True)


def stage_files() -> tuple[Path, Path]:
    # Activities file: look in ROUTE_PLAN root and ROW_FILES; pick newest
    acts = [f for f in ROUTE_PLAN_DIR.glob("Activities_Report*.xlsx") if not f.name.startswith("~$")]
    acts += [f for f in ROW_FILES_DIR.glob("Activities_Report*.xlsx") if not f.name.startswith("~$")]
    if not acts:
        raise FileNotFoundError("No Activities_Report*.xlsx found in ROUTE_PLAN or ROW_FILES.")
    act_src = max(acts, key=lambda f: f.stat().st_mtime)
    act_dst = ROW_FILES_DIR / act_src.name
    if act_src.resolve() != act_dst.resolve():
        shutil.copy2(act_src, act_dst)
        print(f"Staged activities -> {act_dst.name}")

    # Latest CS CRM file
    cs_files = [f for f in ROW_EXCEL_DIR.glob("CS_CRM*.xlsx") if not f.name.startswith("~$")]
    if not cs_files:
        raise FileNotFoundError("No CS_CRM*.xlsx found in ROW_EXCEL.")
    cs_src = max(cs_files, key=lambda f: f.stat().st_mtime)
    cs_dst = ROW_FILES_DIR / cs_src.name
    shutil.copy2(cs_src, cs_dst)
    print(f"Staged CS CRM   -> {cs_dst.name}")
    return act_dst, cs_dst


# --------------------------------------------------------------------------
# STEP 3  lookup + activities
# --------------------------------------------------------------------------
def load_lookup(cs_path: Path) -> dict[str, tuple]:
    wb = openpyxl.load_workbook(cs_path, read_only=True, data_only=True)
    if "user_data" not in wb.sheetnames:
        wb.close()
        raise KeyError("user_data sheet not found in CS CRM file.")
    ws = wb["user_data"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ti, zi, pi = hdr.index("Tenant"), hdr.index("Zone"), hdr.index("Product")
    m: dict[str, tuple] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        t = r[ti]
        if t is None:
            continue
        key = re.sub(r"\s+", " ", str(t)).strip()
        if key and key not in m:
            m[key] = (clean(r[zi]), r[pi])
    wb.close()
    print(f"user_data lookup: {len(m)} tenants")
    return m


def norm_name(s) -> str:
    """Normalised person name for matching (collapse spaces, upper-case)."""
    return re.sub(r"\s+", " ", str(s or "")).strip().upper()


def load_tl_roster(cs_path: Path) -> dict[str, list[dict]]:
    """From user_data, list every person whose Role contains 'Team Leader',
    grouped by canonical Product.  Used to flag TLs who created no route plan."""
    wb = openpyxl.load_workbook(cs_path, read_only=True, data_only=True)
    ws = wb["user_data"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ni, ri, pi = hdr.index("Name"), hdr.index("Role"), hdr.index("Product")
    ti, zi = hdr.index("Tenant"), hdr.index("Zone")
    roster: dict[str, list[dict]] = defaultdict(list)
    seen: dict[str, set] = defaultdict(set)
    for r in ws.iter_rows(min_row=2, values_only=True):
        if "team leader" not in str(r[ri]).lower():
            continue
        prod = canon_product(r[pi])
        if prod is None:
            continue
        nm = clean(r[ni])
        key = norm_name(nm)
        if not key or key in seen[prod]:
            continue
        seen[prod].add(key)
        roster[prod].append({"name": nm, "norm": key,
                             "tenant": clean(r[ti]), "zone": clean(r[zi])})
    wb.close()
    counts = ", ".join(f"{p}:{len(v)}" for p, v in sorted(roster.items()))
    print(f"TL roster (user_data): {counts}")
    return roster


def load_activities(act_path: Path, lookup: dict) -> list[dict]:
    wb = openpyxl.load_workbook(act_path, read_only=True, data_only=True)
    ws = wb["Activities_Report"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    idx = {h: i for i, h in enumerate(hdr)}
    recs: list[dict] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        branch = re.sub(r"\s+", " ", str(r[idx["Branch"]])).strip() if r[idx["Branch"]] else ""
        zone, prod_raw = lookup.get(branch, ("", None))
        recs.append({
            "tl": clean(r[idx["Created_By"]]) or "(unassigned)",
            "agent": clean(r[idx["Agent_Name"]]) or "(unknown)",
            "role": clean(r[idx["Role"]]),
            "branch": branch,
            "zone": zone or "Unknown",
            "product": canon_product(prod_raw),
            "atype": clean(r[idx["Type"]]) or "Other",
            "status": (clean(r[idx["Status"]]) or "UNKNOWN").upper(),
            "location": clean(r[idx["Activity_Location"]]),
            "start": parse_dt(r[idx["Expected_Start_Date"]]),
            "created": parse_dt(r[idx["Created_Date"]]),
        })
    wb.close()
    return recs


# --------------------------------------------------------------------------
# title + section + table helpers
# --------------------------------------------------------------------------
def write_title(ws, product: str, lo: datetime | None, hi: datetime | None,
                n_routes: int, pal: dict, ncols: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    t = ws.cell(1, 1, f"ROUTE PLAN  ·  {PRODUCT_FULL.get(product, product)}")
    t.font = Font(bold=True, size=20, color=WHITE)
    t.fill = solid(pal["dark"])
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    rng = "—"
    if lo and hi:
        rng = f"{lo:%d-%m-%Y}  to  {hi:%d-%m-%Y}"
    sub = ws.cell(2, 1, f"Planned period: {rng}        Total routes: {n_routes:,}        "
                        f"Generated: {datetime.now():%d-%m-%Y %H:%M}")
    sub.font = Font(size=10, color=WHITE, bold=True)
    sub.fill = solid(pal["mid"])
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20
    return 4


def kpi_card(ws, r: int, c: int, value, label: str, pal: dict, fmt=INT_FMT):
    """A 2-col wide, 2-row tall KPI card (value on top, label below)."""
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + 1)
    ws.merge_cells(start_row=r + 1, start_column=c, end_row=r + 1, end_column=c + 1)
    v = ws.cell(r, c, value)
    v.font = Font(bold=True, size=20, color=pal["dark"])
    v.alignment = Alignment(horizontal="center", vertical="center")
    if isinstance(value, (int, float)):
        v.number_format = fmt
    lab = ws.cell(r + 1, c, label)
    lab.font = Font(size=9, bold=True, color=MUTED)
    lab.alignment = Alignment(horizontal="center", vertical="center")
    top = Side(style="thick", color=pal["accent"])
    for rr in (r, r + 1):
        for cc in (c, c + 1):
            cell = ws.cell(rr, cc)
            cell.fill = solid(pal["tint"])
            b = box(pal["light"])
            if rr == r:
                cell.border = Border(top=top, left=b.left, right=b.right, bottom=b.bottom)
            else:
                cell.border = b
    ws.row_dimensions[r].height = 30
    ws.row_dimensions[r + 1].height = 16


def section(ws, r: int, c: int, span: int, text: str, pal: dict) -> int:
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c + span - 1)
    cell = ws.cell(r, c, text)
    cell.font = Font(bold=True, size=11, color=WHITE)
    cell.fill = solid(pal["mid"])
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 22
    return r + 1


def write_table(ws, r: int, c: int, headers: list[str], rows: list[list],
                fmts: list[str], pal: dict, databar_col: int | None = None,
                row_fill: str | None = None, text_color: str | None = None,
                total_row: bool = False) -> int:
    """Generic styled table. fmts entries: text/int/pct/date/money. Returns next free row."""
    ncol = len(headers)
    # header
    for j, h in enumerate(headers):
        cell = ws.cell(r, c + j, h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = solid(pal["dark"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box(pal["dark"])
    ws.row_dimensions[r].height = 26

    first_data = r + 1
    for i, row in enumerate(rows):
        rr = first_data + i
        if row_fill:
            fill = solid(row_fill)
        else:
            fill = solid(ALT_ROW) if i % 2 else solid(WHITE)
        for j, val in enumerate(row):
            cell = ws.cell(rr, c + j, val)
            cell.fill = fill
            cell.border = box()
            cell.font = Font(color=text_color or SLATE, size=10,
                             bold=bool(text_color))
            f = fmts[j]
            if f == "text":
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if f == "int":
                cell.number_format = INT_FMT
            elif f == "pct":
                cell.number_format = PCT_FMT
            elif f == "date":
                cell.number_format = DATE_FMT
            elif f == "money":
                cell.number_format = "#,##0"
    last_data = first_data + len(rows) - 1

    if total_row and rows:
        tr = last_data + 1
        for j in range(ncol):
            cell = ws.cell(tr, c + j)
            cell.fill = solid(pal["light"])
            cell.border = box(pal["mid"])
            cell.font = Font(bold=True, color=SLATE, size=10)
            if j == 0:
                cell.value = "TOTAL"
                cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            else:
                if fmts[j] in ("int", "money"):
                    colvals = [row[j] for row in rows if isinstance(row[j], (int, float))]
                    cell.value = sum(colvals)
                    cell.number_format = INT_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif fmts[j] == "pct":
                    cell.value = 1.0
                    cell.number_format = PCT_FMT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        last_data = tr

    if databar_col is not None and rows:
        col_letter = get_column_letter(c + databar_col)
        rng = f"{col_letter}{first_data}:{col_letter}{first_data + len(rows) - 1}"
        ws.conditional_formatting.add(
            rng, DataBarRule(start_type="num", start_value=0, end_type="max",
                             color=pal["accent"], showValue=True))
    return last_data + 2


# --------------------------------------------------------------------------
# product-level aggregation
# --------------------------------------------------------------------------
def ordered_types(recs: list[dict]) -> list[str]:
    present = {r["atype"] for r in recs}
    extra = sorted(present - set(KNOWN_TYPES))
    return [t for t in KNOWN_TYPES if t in present] + extra


def meaningful_days(recs: list[dict]) -> list:
    """Distinct plan days that belong to the report window.

    Activities are planned as a single week (e.g. Mon 01-06 .. Sat 06-06).
    A few isolated advance/late entries can land on far-off dates; those form
    a separate cluster.  We split the distinct days into consecutive runs
    (a gap of more than 3 days starts a new run, so weekends stay joined) and
    keep the run holding the most activities — i.e. the real plan week."""
    dated = [r for r in recs if r["start"]]
    if not dated:
        return []
    ct = Counter(r["start"].date() for r in dated)
    days = sorted(ct)
    runs, cur = [], [days[0]]
    for prev, d in zip(days, days[1:]):
        if (d - prev).days <= 3:
            cur.append(d)
        else:
            runs.append(cur)
            cur = [d]
    runs.append(cur)
    best = max(runs, key=lambda run: sum(ct[d] for d in run))
    return best


def date_range(recs: list[dict]):
    """Report period = first to last *meaningful* plan day (Expected_Start_Date)."""
    days = meaningful_days(recs)
    if days:
        return days[0], days[-1]
    created = [r["created"] for r in recs if r["created"]]
    if not created:
        return None, None
    return min(created).date(), max(created).date()


# --------------------------------------------------------------------------
# SHEET 1  Summary
# --------------------------------------------------------------------------
def tl_coverage(recs: list[dict], roster: list[dict]):
    """Return (planned_count, not_planned_list) comparing roster vs Created_By."""
    planned = {norm_name(r["tl"]) for r in recs}
    not_planned = [t for t in roster if t["norm"] not in planned]
    not_planned.sort(key=lambda t: t["name"].lower())
    planned_in_roster = sum(1 for t in roster if t["norm"] in planned)
    return planned_in_roster, not_planned


def build_summary(ws, product: str, recs: list[dict], pal: dict,
                  roster: list[dict] | None = None):
    NCOLS = 12
    n = len(recs)
    tls = {r["tl"] for r in recs}
    agents = {(r["tl"], r["agent"]) for r in recs}
    uniq_agents = {r["agent"] for r in recs}
    plan_days = meaningful_days(recs)
    completed = sum(1 for r in recs if r["status"] == "COMPLETED")
    lo, hi = date_range(recs)

    row = write_title(ws, product, lo, hi, n, pal, NCOLS)

    # KPI cards (6 cards × 2 cols)
    cards = [
        (n, "TOTAL ROUTES", INT_FMT),
        (len(tls), "TEAM LEADERS", INT_FMT),
        (len(uniq_agents), "AGENTS", INT_FMT),
        (len(plan_days), "PLAN DAYS", INT_FMT),
        (round(n / len(uniq_agents), 1) if uniq_agents else 0, "AVG ROUTES / AGENT", "#,##0.0"),
        (completed / n if n else 0, "COMPLETED %", PCT_FMT),
    ]
    for i, (val, lab, fmt) in enumerate(cards):
        kpi_card(ws, row, 1 + i * 2, val, lab, pal, fmt)
    row += 3

    # ---- Routes by Activity Type (left)  |  Routes by Status (right) ----
    types = ordered_types(recs)
    type_ct = Counter(r["atype"] for r in recs)
    status_ct = Counter(r["status"] for r in recs)

    left_start = row
    rr = section(ws, row, 1, 5, "Routes by Activity Type", pal)
    trows = [[t, type_ct.get(t, 0), (type_ct.get(t, 0) / n if n else 0)] for t in types]
    end_left = write_table(ws, rr, 1, ["Activity Type", "Routes", "% of Total"],
                           trows, ["text", "int", "pct"], pal, databar_col=1, total_row=True)

    rr = section(ws, left_start, 7, 5, "Routes by Status", pal)
    status_order = ["COMPLETED", "INPROGRESS", "PENDING"]
    s_present = [s for s in status_order if s in status_ct] + \
                sorted(set(status_ct) - set(status_order))
    srows = [[s.title(), status_ct[s], (status_ct[s] / n if n else 0)] for s in s_present]
    end_right = write_table(ws, rr, 7, ["Status", "Routes", "% of Total"],
                            srows, ["text", "int", "pct"], pal, databar_col=1, total_row=True)
    row = max(end_left, end_right)

    # ---- Routes by Zone (full width-ish) ----
    zone_ct = Counter(r["zone"] for r in recs)
    zone_agents = defaultdict(set)
    for r in recs:
        zone_agents[r["zone"]].add(r["agent"])
    rr = section(ws, row, 1, 6, "Routes by Zone", pal)
    zrows = [[z, len(zone_agents[z]), c, (c / n if n else 0)]
             for z, c in sorted(zone_ct.items(), key=lambda x: -x[1])]
    row = write_table(ws, rr, 1, ["Zone", "Agents", "Routes", "% of Total"],
                      zrows, ["text", "int", "int", "pct"], pal, databar_col=2, total_row=True)

    # ---- Top 10 TLs (left)  |  Top 10 Agents (right) ----
    tl_ct = Counter(r["tl"] for r in recs)
    tl_agentn = defaultdict(set)
    for r in recs:
        tl_agentn[r["tl"]].add(r["agent"])
    top_start = row
    rr = section(ws, row, 1, 5, "Top 10 Team Leaders (by routes)", pal)
    tl_rows = [[tl, len(tl_agentn[tl]), c]
               for tl, c in sorted(tl_ct.items(), key=lambda x: -x[1])[:10]]
    end_left = write_table(ws, rr, 1, ["Team Leader", "Agents", "Routes"],
                           tl_rows, ["text", "int", "int"], pal, databar_col=2)

    ag_ct = Counter(r["agent"] for r in recs)
    rr = section(ws, top_start, 7, 5, "Top 10 Agents (by routes)", pal)
    ag_rows = [[a, c] for a, c in sorted(ag_ct.items(), key=lambda x: -x[1])[:10]]
    end_right = write_table(ws, rr, 7, ["Agent", "Routes"],
                            ag_rows, ["text", "int"], pal, databar_col=1)
    row = max(end_left, end_right)

    # ---- Team Leader coverage (vs user_data roster) ----
    if roster:
        expected = len(roster)
        planned_n, not_planned = tl_coverage(recs, roster)
        cov = planned_n / expected if expected else 0
        rr = section(ws, row, 1, 6,
                     "Team Leader Coverage  (Role = 'Team Leader' in user_data)", pal)
        cov_rows = [[expected, planned_n, len(not_planned), cov]]
        row = write_table(ws, rr, 1,
                          ["Expected TLs", "Created Plans", "No Plan Created", "Coverage %"],
                          cov_rows, ["int", "int", "int", "pct"], pal, total_row=False)

        if not_planned:
            rr = section(ws, row, 1, 6,
                         f"Team Leaders WITHOUT a Route Plan  ({len(not_planned)})", pal)
            np_rows = [[t["name"], t["tenant"], t["zone"]] for t in not_planned]
            row = write_table(ws, rr, 1, ["Team Leader", "Branch / Tenant", "Zone"],
                              np_rows, ["text", "text", "text"], pal,
                              row_fill=BAD_FILL, text_color=BAD_TXT)

    # methodology note
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    note = ws.cell(row, 1,
                   "Route = one planned activity.  Team Leader = Created_By.  Agent = Agent_Name.  "
                   "Plan Days = distinct Expected Start dates.  Zone & Product looked up from user_data (Tenant = Branch).")
    note.font = Font(italic=True, size=9, color=MUTED)
    note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    note.fill = solid(pal["tint"])

    # widths
    widths = [22, 12, 12, 12, 12, 4, 22, 12, 12, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"


# --------------------------------------------------------------------------
# SHEET 2  TL Summary
# --------------------------------------------------------------------------
def build_tl_summary(ws, recs: list[dict], pal: dict,
                     roster: list[dict] | None = None):
    types = ordered_types(recs)
    by_tl = defaultdict(list)
    for r in recs:
        by_tl[r["tl"]].append(r)

    headers = ["Team Leader", "Zone(s)", "# Agents", "Total Routes", "Plan Days"] \
        + types + ["Completed", "Pending", "Avg Routes / Agent"]
    fmts = ["text", "text", "int", "int", "int"] + ["int"] * len(types) + ["int", "int", "money"]

    rows = []
    for tl, rs in by_tl.items():
        agents = {r["agent"] for r in rs}
        zones = sorted({r["zone"] for r in rs})
        days = {r["start"].date() for r in rs if r["start"]}
        tcount = Counter(r["atype"] for r in rs)
        comp = sum(1 for r in rs if r["status"] == "COMPLETED")
        pend = sum(1 for r in rs if r["status"] == "PENDING")
        avg = round(len(rs) / len(agents), 1) if agents else 0
        zone_label = zones[0] if len(zones) == 1 else f"{zones[0]} (+{len(zones)-1})" if zones else ""
        rows.append([tl, zone_label, len(agents), len(rs), len(days)]
                    + [tcount.get(t, 0) for t in types] + [comp, pend, avg])
    rows.sort(key=lambda x: -x[3])

    ncols = len(headers)
    row = write_title(ws, "Team Leaders", *date_range(recs), len(recs), pal, ncols)
    section(ws, row, 1, ncols, "Team Leader Summary — routes, agents & activity mix", pal)
    row += 1
    routes_col = headers.index("Total Routes")
    end = write_table(ws, row, 1, headers, rows, fmts, pal, databar_col=routes_col, total_row=True)

    # ---- Team Leaders who created NO route plan (from user_data roster) ----
    if roster:
        _, not_planned = tl_coverage(recs, roster)
        if not_planned:
            srow = end + 1
            section(ws, srow, 1, ncols,
                    f"Team Leaders WITHOUT a Route Plan  ({len(not_planned)})  — "
                    f"Role = 'Team Leader' in user_data, no activity created", pal)
            srow += 1
            np_rows = [[t["name"], t["tenant"], t["zone"]] for t in not_planned]
            write_table(ws, srow, 1, ["Team Leader", "Branch / Tenant", "Zone"],
                        np_rows, ["text", "text", "text"], pal,
                        row_fill=BAD_FILL, text_color=BAD_TXT)

    widths = [26, 22, 9, 12, 10] + [12] * len(types) + [11, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row + 1, 1).coordinate
    ws.auto_filter.ref = f"A{row}:{get_column_letter(ncols)}{row + len(rows)}"


# --------------------------------------------------------------------------
# SHEET 3  Agent Summary  (per Team Leader, per Agent)
# --------------------------------------------------------------------------
def build_agent_summary(ws, recs: list[dict], pal: dict):
    types = ordered_types(recs)
    by_pair = defaultdict(list)
    for r in recs:
        by_pair[(r["tl"], r["agent"])].append(r)

    headers = ["Team Leader", "Agent", "Branch", "Zone", "Role", "Total Routes",
               "Plan Days", "Locations"] + types + ["Completed", "Pending",
               "First Plan", "Last Plan"]
    fmts = ["text", "text", "text", "text", "text", "int", "int", "int"] \
        + ["int"] * len(types) + ["int", "int", "date", "date"]

    rows = []
    for (tl, agent), rs in by_pair.items():
        branch = Counter(r["branch"] for r in rs if r["branch"]).most_common(1)
        zone = Counter(r["zone"] for r in rs).most_common(1)
        role = Counter(r["role"] for r in rs if r["role"]).most_common(1)
        days = {r["start"].date() for r in rs if r["start"]}
        locs = {r["location"] for r in rs if r["location"]}
        tcount = Counter(r["atype"] for r in rs)
        comp = sum(1 for r in rs if r["status"] == "COMPLETED")
        pend = sum(1 for r in rs if r["status"] == "PENDING")
        starts = [r["start"] for r in rs if r["start"]]
        rows.append([tl, agent,
                     branch[0][0] if branch else "",
                     zone[0][0] if zone else "",
                     role[0][0] if role else "",
                     len(rs), len(days), len(locs)]
                    + [tcount.get(t, 0) for t in types]
                    + [comp, pend,
                       min(starts) if starts else None,
                       max(starts) if starts else None])
    rows.sort(key=lambda x: (x[0].lower(), -x[5]))

    ncols = len(headers)
    row = write_title(ws, "Agents", *date_range(recs), len(recs), pal, ncols)
    section(ws, row, 1, ncols, "Agent Summary — one row per Team Leader × Agent", pal)
    row += 1
    routes_col = headers.index("Total Routes")
    write_table(ws, row, 1, headers, rows, fmts, pal, databar_col=routes_col, total_row=True)

    widths = [24, 24, 18, 20, 16, 12, 10, 10] + [12] * len(types) + [11, 10, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(row + 1, 3).coordinate
    ws.auto_filter.ref = f"A{row}:{get_column_letter(ncols)}{row + len(rows)}"


# --------------------------------------------------------------------------
# SHEET 4  Daily Plan  (agent × day matrix — who does what each day)
# --------------------------------------------------------------------------
def build_daily_plan(ws, product: str, recs: list[dict], pal: dict):
    """Matrix of activities per agent per calendar day.
    Colour: red = 0 (no plan that day), amber = 1 (below target), green = >= 2 (on target).
    Trivial outlier days (< 1% of routes) are dropped so the core plan week stays clean."""
    dated = [r for r in recs if r["start"]]
    days = meaningful_days(recs)

    # counts[(tl, agent)][day] = number of activities
    counts: dict = defaultdict(lambda: defaultdict(int))
    meta: dict = {}
    for r in dated:
        key = (r["tl"], r["agent"])
        counts[key][r["start"].date()] += 1
        if key not in meta:
            meta[key] = {"branch": r["branch"], "zone": r["zone"]}

    n_day_cols = len(days)
    TRAIL = 4   # Week Total, Active Days, Below Target, Status
    ncols = 4 + n_day_cols + TRAIL
    lo, hi = date_range(recs)
    row = write_title(ws, "Daily Plan", lo, hi, len(recs), pal, ncols)

    # legend
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    leg = ws.cell(row, 1, f"Daily activities per agent  —  target {DAILY_TARGET}+ per working day.   "
                          f"Green = on target (>={DAILY_TARGET})    Amber = below target (1)    "
                          f"Red = no activity that day.   'Below Target' counts working days under {DAILY_TARGET}.")
    leg.font = Font(italic=True, size=9, color=MUTED)
    leg.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    leg.fill = solid(pal["tint"])
    row += 1

    # header row
    headers = ["Team Leader", "Agent", "Branch", "Zone"] \
        + [f"{d:%a}\n{d:%d-%m}" for d in days] \
        + ["Week Total", "Active Days", "Below Target", "Status"]
    for j, h in enumerate(headers):
        cell = ws.cell(row, 1 + j, h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = solid(pal["dark"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box(pal["dark"])
    ws.row_dimensions[row].height = 30
    hdr_row = row
    row += 1

    def paint(cell, fill, txt, bold=True):
        cell.fill = solid(fill)
        cell.font = Font(color=txt, size=10, bold=bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = box()
        cell.number_format = INT_FMT

    # sort by TL then agent
    keys = sorted(counts.keys(), key=lambda k: (k[0].lower(), k[1].lower()))
    first_data = row
    for i, key in enumerate(keys):
        tl, agent = key
        dd = counts[key]
        base = solid(ALT_ROW) if i % 2 else solid(WHITE)
        for j, val in enumerate([tl, agent, meta[key]["branch"], meta[key]["zone"]]):
            c = ws.cell(row, 1 + j, val)
            c.fill = base
            c.border = box()
            c.font = Font(color=SLATE, size=10)
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        # day columns
        week_total = active = below = 0
        for k, d in enumerate(days):
            n = dd.get(d, 0)
            week_total += n
            c = ws.cell(row, 5 + k, n)
            if n == 0:
                paint(c, BAD_FILL, BAD_TXT)
            elif n < DAILY_TARGET:
                paint(c, WARN_FILL, WARN_TXT)
                active += 1
                below += 1
            else:
                paint(c, GOOD_FILL, GOOD_TXT)
                active += 1
        # trailing metrics
        c = ws.cell(row, 5 + n_day_cols, week_total)
        c.fill = base; c.border = box(); c.number_format = INT_FMT
        c.font = Font(color=SLATE, size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row, 6 + n_day_cols, active)
        c.fill = base; c.border = box(); c.number_format = INT_FMT
        c.font = Font(color=SLATE, size=10, bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")

        c = ws.cell(row, 7 + n_day_cols, below)
        paint(c, (GOOD_FILL if below == 0 else WARN_FILL),
              (GOOD_TXT if below == 0 else WARN_TXT))

        status = "NO PLAN" if week_total == 0 else ("ON TARGET" if below == 0 else "PARTIAL")
        c = ws.cell(row, 8 + n_day_cols, status)
        c.border = box()
        c.alignment = Alignment(horizontal="center", vertical="center")
        if status == "ON TARGET":
            c.fill = solid(GOOD_FILL); c.font = Font(color=GOOD_TXT, size=10, bold=True)
        elif status == "PARTIAL":
            c.fill = solid(WARN_FILL); c.font = Font(color=WARN_TXT, size=10, bold=True)
        else:
            c.fill = solid(BAD_FILL); c.font = Font(color=BAD_TXT, size=10, bold=True)
        row += 1

    # daily totals footer (activities planned that day)
    foot = row
    lab = ws.cell(foot, 1, "DAILY TOTALS (activities)")
    ws.merge_cells(start_row=foot, start_column=1, end_row=foot, end_column=4)
    lab.font = Font(bold=True, color=WHITE, size=10)
    lab.fill = solid(pal["mid"])
    lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for k, d in enumerate(days):
        c = ws.cell(foot, 5 + k, sum(counts[key].get(d, 0) for key in keys))
        c.fill = solid(pal["light"]); c.border = box(pal["mid"])
        c.font = Font(bold=True, color=SLATE, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = INT_FMT
    gt = ws.cell(foot, 5 + n_day_cols,
                 sum(counts[key].get(d, 0) for key in keys for d in days))
    gt.fill = solid(pal["light"]); gt.border = box(pal["mid"])
    gt.font = Font(bold=True, color=SLATE, size=10)
    gt.alignment = Alignment(horizontal="center", vertical="center")
    gt.number_format = INT_FMT
    for cc in (6 + n_day_cols, 7 + n_day_cols, 8 + n_day_cols):
        c = ws.cell(foot, cc); c.fill = solid(pal["light"]); c.border = box(pal["mid"])

    # widths
    for col, w in zip("ABCD", (24, 24, 20, 20)):
        ws.column_dimensions[col].width = w
    for k in range(n_day_cols):
        ws.column_dimensions[get_column_letter(5 + k)].width = 9
    ws.column_dimensions[get_column_letter(5 + n_day_cols)].width = 11
    ws.column_dimensions[get_column_letter(6 + n_day_cols)].width = 11
    ws.column_dimensions[get_column_letter(7 + n_day_cols)].width = 12
    ws.column_dimensions[get_column_letter(8 + n_day_cols)].width = 12
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = ws.cell(first_data, 5).coordinate
    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(ncols)}{first_data + len(keys) - 1}"


# --------------------------------------------------------------------------
# product stats (shared by the email)
# --------------------------------------------------------------------------
def product_stats(product: str, recs: list[dict], roster: list[dict]) -> dict:
    n = len(recs)
    days = meaningful_days(recs)
    uniq_agents = {r["agent"] for r in recs}
    completed = sum(1 for r in recs if r["status"] == "COMPLETED")
    planned_n, not_planned = tl_coverage(recs, roster) if roster else (0, [])
    expected = len(roster) if roster else 0
    lo, hi = date_range(recs)
    return {
        "product": product,
        "total_routes": n,
        "team_leaders": len({r["tl"] for r in recs}),
        "agents": len(uniq_agents),
        "plan_days": len(days),
        "avg_routes_agent": round(n / len(uniq_agents), 1) if uniq_agents else 0,
        "completed_pct": (completed / n) if n else 0,
        "expected_tls": expected,
        "created_tls": planned_n,
        "not_created": len(not_planned),
        "coverage_pct": (planned_n / expected) if expected else 0,
        "not_planned_list": not_planned,
        "lo": lo, "hi": hi,
    }


# --------------------------------------------------------------------------
# beautiful HTML email per product
# --------------------------------------------------------------------------
def build_email_html(stats: dict, pal: dict) -> str:
    product = stats["product"]
    full = PRODUCT_FULL.get(product, product)
    dark, mid, accent = "#" + pal["dark"], "#" + pal["mid"], "#" + pal["accent"]
    light, tint = "#" + pal["light"], "#" + pal["tint"]
    lo, hi = stats["lo"], stats["hi"]
    period = (f"{lo:%d-%m-%Y}  to  {hi:%d-%m-%Y}" if lo and hi else "—")
    gen = datetime.now().strftime("%d-%m-%Y %H:%M")

    def kpi(value, label):
        return f"""
          <td style="padding:6px;">
            <div style="background:{tint};border:1px solid {light};border-top:3px solid {accent};
                        border-radius:8px;padding:14px 10px;text-align:center;">
              <div style="font-size:22px;font-weight:700;color:{dark};line-height:1;">{value}</div>
              <div style="font-size:10px;font-weight:700;color:#64748B;margin-top:6px;
                          letter-spacing:.4px;text-transform:uppercase;">{label}</div>
            </div>
          </td>"""

    cards = (kpi(f"{stats['total_routes']:,}", "Total Activities")
             + kpi(f"{stats['expected_tls']:,}", "Expected TLs")
             + kpi(f"{stats['created_tls']:,}", "Created Plans")
             + kpi(f"{stats['not_created']:,}", "Not Created"))
    cards2 = (kpi(f"{stats['coverage_pct']*100:.1f}%", "Coverage")
              + kpi(f"{stats['agents']:,}", "Agents")
              + kpi(f"{stats['plan_days']:,}", "Plan Days")
              + kpi(f"{stats['avg_routes_agent']}", "Avg / Agent"))

    # summary table (mirrors the Excel cover stats)
    def trow(lbl, val, strong=False):
        vstyle = f"color:{dark};font-weight:700;" if strong else "color:#0F172A;font-weight:600;"
        return (f'<tr><td style="padding:9px 14px;border-bottom:1px solid #EEF2F7;'
                f'color:#475569;font-size:13px;">{lbl}</td>'
                f'<td style="padding:9px 14px;border-bottom:1px solid #EEF2F7;text-align:right;'
                f'font-size:13px;{vstyle}">{val}</td></tr>')

    summary_tbl = (
        trow("Total activities (routes) created", f"{stats['total_routes']:,}")
        + trow("Team Leaders expected (user_data)", f"{stats['expected_tls']:,}")
        + trow("Team Leaders who created a plan", f"{stats['created_tls']:,}", True)
        + trow("Team Leaders with NO plan", f"{stats['not_created']:,}", True)
        + trow("Coverage", f"{stats['coverage_pct']*100:.1f}%", True)
        + trow("Routes completed", f"{stats['completed_pct']*100:.1f}%")
    )

    # "Team Leaders WITHOUT a plan" — fixed scrollable box (≈3 rows tall)
    np_list = stats["not_planned_list"]
    if np_list:
        items = ""
        for i, t in enumerate(np_list):
            bg = "#FFFFFF" if i % 2 == 0 else "#FFF5F5"
            items += (
                f'<tr style="background:{bg};">'
                f'<td style="padding:8px 12px;border-bottom:1px solid #FEE2E2;color:#991B1B;'
                f'font-weight:600;font-size:13px;">{t["name"]}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #FEE2E2;color:#7F1D1D;'
                f'font-size:12px;">{t["tenant"] or "&mdash;"}</td>'
                f'<td style="padding:8px 12px;border-bottom:1px solid #FEE2E2;color:#7F1D1D;'
                f'font-size:12px;">{t["zone"] or "&mdash;"}</td></tr>')
        no_plan_block = f"""
        <div style="margin:6px 0 2px;font-size:11px;color:#991B1B;font-weight:700;">
          Showing {min(3, len(np_list))} of {len(np_list)} &mdash; scroll inside the box to view all.
        </div>
        <div style="max-height:150px;overflow-y:auto;border:1px solid #FECACA;border-radius:8px;
                    background:#FEF2F2;">
          <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
                 style="border-collapse:collapse;">
            <thead><tr style="position:sticky;top:0;background:#FEE2E2;">
              <th style="text-align:left;padding:8px 12px;color:#991B1B;font-size:11px;
                         text-transform:uppercase;letter-spacing:.3px;">Team Leader</th>
              <th style="text-align:left;padding:8px 12px;color:#991B1B;font-size:11px;
                         text-transform:uppercase;letter-spacing:.3px;">Branch / Tenant</th>
              <th style="text-align:left;padding:8px 12px;color:#991B1B;font-size:11px;
                         text-transform:uppercase;letter-spacing:.3px;">Zone</th>
            </tr></thead>
            <tbody>{items}</tbody>
          </table>
        </div>"""
    else:
        no_plan_block = ('<div style="padding:14px;border:1px solid #BBF7D0;border-radius:8px;'
                         'background:#F0FDF4;color:#166534;font-weight:700;font-size:13px;">'
                         '&#10004; Every expected Team Leader created a route plan.</div>')

    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#F1F5F9;
             font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;color:#0F172A;">
  <div style="max-width:680px;margin:0 auto;background:#FFFFFF;">
    <!-- header -->
    <div style="background:{dark};padding:22px 24px;">
      <div style="color:#FFFFFF;font-size:21px;font-weight:700;">Route Plan Report &mdash; {full}</div>
      <div style="color:#E2E8F0;font-size:13px;margin-top:6px;">
        Planned period: <b>{period}</b> &nbsp;&bull;&nbsp; Generated: {gen}
      </div>
    </div>
    <div style="height:4px;background:{accent};"></div>

    <!-- KPI cards -->
    <div style="padding:16px 16px 0;">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>{cards}</tr></table>
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0"><tr>{cards2}</tr></table>
    </div>

    <!-- summary table -->
    <div style="padding:18px 22px 4px;">
      <div style="font-size:15px;font-weight:700;color:{dark};margin-bottom:10px;
                  border-left:4px solid {accent};padding-left:10px;">Summary</div>
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #EEF2F7;border-radius:8px;border-collapse:separate;
                    overflow:hidden;">{summary_tbl}</table>
    </div>

    <!-- TLs without a plan -->
    <div style="padding:14px 22px 22px;">
      <div style="font-size:15px;font-weight:700;color:#991B1B;margin-bottom:8px;
                  border-left:4px solid #DC2626;padding-left:10px;">
        Team Leaders WITHOUT a Route Plan ({stats['not_created']})</div>
      {no_plan_block}
    </div>

    <!-- footer -->
    <div style="background:#0F172A;padding:16px 24px;color:#94A3B8;font-size:11px;text-align:center;">
      The full Excel workbook (Summary &bull; TL Summary &bull; Agent Summary &bull; Daily Plan)
      is attached.<br>Plan Days = number of plan days from the first to the last Expected Start Date in this report.
    </div>
  </div>
</body></html>"""


def send_product_email(stats: dict, pal: dict, attachment: Path,
                       sender: str, password: str) -> bool:
    product = stats["product"]
    recipients = recipients_for(product)
    lo, hi = stats["lo"], stats["hi"]
    span = (f"{lo:%d-%m-%Y} to {hi:%d-%m-%Y}" if lo and hi else "")
    subject = f"ROUTE PLAN REPORT - {PRODUCT_FULL.get(product, product)}  ({span})"
    html = build_email_html(stats, pal)

    try:
        try:
            ctx = ssl.create_default_context()
            server = smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=40, context=ctx)
        except Exception:
            server = smtplib.SMTP("smtp.gmail.com", 587, timeout=40)
            server.starttls()
        server.login(sender, password)

        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = ", ".join(recipients)
        msg["Subject"] = subject
        msg["Date"] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
        msg["Reply-To"] = sender
        msg.attach(MIMEText(html, "html"))

        if attachment and attachment.exists():
            with open(attachment, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition",
                            f'attachment; filename="{attachment.name}"')
            msg.attach(part)

        server.send_message(msg)
        server.quit()
        print(f"  [SENT] {product}: {len(recipients)} recipients -> {', '.join(recipients)}")
        return True
    except Exception as e:
        print(f"  [FAIL] {product}: {e}")
        return False


# --------------------------------------------------------------------------
# build one product workbook
# --------------------------------------------------------------------------
def build_report(product: str, recs: list[dict],
                 roster: list[dict] | None = None) -> Path | None:
    if not recs:
        print(f"  {product}: no records — skipped.")
        return None
    pal = PALETTES[product]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    s1 = wb.create_sheet("Summary")
    build_summary(s1, product, recs, pal, roster)
    s1.sheet_properties.tabColor = pal["dark"]

    s2 = wb.create_sheet("TL Summary")
    build_tl_summary(s2, recs, pal, roster)
    s2.sheet_properties.tabColor = pal["mid"]

    s3 = wb.create_sheet("Agent Summary")
    build_agent_summary(s3, recs, pal)
    s3.sheet_properties.tabColor = pal["accent"]

    s4 = wb.create_sheet("Daily Plan")
    build_daily_plan(s4, product, recs, pal)
    s4.sheet_properties.tabColor = pal["light"]

    lo, hi = date_range(recs)
    lo_s = f"{lo:%d-%m-%Y}" if lo else "NA"
    hi_s = f"{hi:%d-%m-%Y}" if hi else "NA"
    out = ROUTE_PLAN_DIR / product / f"ROUTE_PLAN_FROM_{lo_s}_{hi_s}.xlsx"
    wb.save(out)
    print(f"  {product}: {len(recs):,} routes -> {out.name}")
    return out


# --------------------------------------------------------------------------
def main():
    print("=" * 60)
    print("ROUTE PLAN REPORT BUILDER")
    print("=" * 60)
    setup_folders()
    act_path, cs_path = stage_files()
    lookup = load_lookup(cs_path)
    roster = load_tl_roster(cs_path)
    recs = load_activities(act_path, lookup)
    print(f"Loaded {len(recs):,} activity rows.")

    by_prod = defaultdict(list)
    skipped = 0
    for r in recs:
        if r["product"] in PRODUCT_FOLDERS:
            by_prod[r["product"]].append(r)
        else:
            skipped += 1
    if skipped:
        print(f"Note: {skipped} rows had no resolvable product (skipped).")

    print("\nBuilding reports:")
    built = {}   # product -> (path, stats)
    for product in PRODUCT_FOLDERS:
        prod_recs = by_prod.get(product, [])
        prod_roster = roster.get(product, [])
        out = build_report(product, prod_recs, prod_roster)
        if out:
            built[product] = (out, product_stats(product, prod_recs, prod_roster))

    print("\nDone building reports.")

    # ---- email step ----
    if not built:
        return
    try:
        ans = input("\nSend the reports by email now? [y/N]: ").strip().lower()
    except EOFError:
        ans = "n"
    if ans not in ("y", "yes"):
        print("Email step skipped.")
        return

    if load_dotenv:
        load_dotenv(ENV_PATH)
    sender = os.getenv("EMAIL_USERNAME")
    password = os.getenv("EMAIL_PASSWORD")
    if not sender or not password:
        print(f"EMAIL_USERNAME / EMAIL_PASSWORD not found (looked in {ENV_PATH}). Cannot send.")
        return

    print("\nSending emails:")
    for product, (out, stats) in built.items():
        send_product_email(stats, PALETTES[product], out, sender, password)
    print("\nEmail step complete.")


if __name__ == "__main__":
    main()
