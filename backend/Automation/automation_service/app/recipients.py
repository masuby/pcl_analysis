"""Read / write email recipients for each pipeline.

CRM   -> Master_crm_emails.xlsx (sheets Actual/Specific/Test, header row 4,
         department columns CS/LBF/SME) — the existing workbook.
Call Center -> master_email_call_center.xlsx (single sheet, header row 1,
               columns CS/LBF/RO) — created + seeded on first use.
MTD   -> master_emails_mtd.xlsx (single sheet, columns CS/LBF) — created here.

The scripts read these same files, so UI edits flow straight into the runs.
"""
from __future__ import annotations

from openpyxl import Workbook, load_workbook

from . import config as C

# --------------------------------------------------------------------------- #
# Store definitions
# --------------------------------------------------------------------------- #
CRM_MODE_SHEET = {"actual": "Actual", "specific": "Specific", "test": "Test"}

STORES = {
    "crm": {"file": C.CRM_EMAIL_MASTER, "kind": "crm",
            "departments": ["CS", "LBF", "SME"]},
    "call_center": {"file": C.CALL_CENTER_DIR / "master_email_call_center.xlsx",
                    "kind": "simple", "departments": ["CS", "LBF", "RO"]},
    "mtd": {"file": C.MTD_DIR / "master_emails_mtd.xlsx",
            "kind": "simple", "departments": ["CS", "LBF"]},
    "reps": {"file": C.REPS_DIR / "master_emails_reps.xlsx",
             "kind": "simple", "departments": ["Managers"]},
}

# Seed data used when a simple master doesn't exist yet.
_SEED = {
    "call_center": {
        "CS": ["raphael@platinumcredit.co.tz", "dorice@platinumcredit.co.tz",
               "sigfrid@platinumcredit.co.tz", "yusuph@platinumcredit.co.tz",
               "allan@platinumcredit.co.tz", "daniel@platinumcredit.co.tz"],
        "LBF": ["raphael@platinumcredit.co.tz", "allan@platinumcredit.co.tz",
                "dorice@platinumcredit.co.tz", "sigfrid@platinumcredit.co.tz",
                "thomas@platinumcredit.co.tz", "daniel@platinumcredit.co.tz"],
        "RO": ["raphael@platinumcredit.co.tz", "allan@platinumcredit.co.tz",
               "relationshipofficers@platinumcredit.co.tz", "daniel@platinumcredit.co.tz"],
    },
    "mtd": {
        "CS": ["daniel@platinumcredit.co.tz"],
        "LBF": ["daniel@platinumcredit.co.tz"],
    },
    "reps": {
        "Managers": [
            "daniel@platinumcredit.co.tz", "sigfrid@platinumcredit.co.tz",
            "augustine@platinumcredit.co.tz", "doris@platinumcredit.co.tz",
            "abdulhakim.khalfan@platinumcredit.co.tz", "fragrance@platinumcredit.co.tz",
            "allan@platinumcredit.co.tz", "raphael@platinumcredit.co.tz",
            "kelvin.mwasala@platinumcredit.co.tz", "irene.mmari@platinumcredit.co.tz",
        ],
    },
}


def _clean(emails: list[str]) -> list[str]:
    """Accept a list OR a pasted blob; tolerate commas / semicolons / spaces /
    trailing punctuation, extract valid addresses, dedupe (order kept)."""
    import re
    text = "\n".join(str(e) for e in (emails or []) if e is not None)
    found = re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", text)
    seen, out = set(), []
    for e in found:
        key = e.lower()
        if key not in seen:
            seen.add(key)
            out.append(e)
    return out


# --------------------------------------------------------------------------- #
# Simple single-sheet master (Call Center / MTD)
# --------------------------------------------------------------------------- #
_SIMPLE_SHEET = "Recipients"


def _ensure_simple(pid: str):
    store = STORES[pid]
    path = store["file"]
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = _SIMPLE_SHEET
    depts = store["departments"]
    for j, d in enumerate(depts, 1):
        ws.cell(row=1, column=j, value=d)
    seed = _SEED.get(pid, {})
    for j, d in enumerate(depts, 1):
        for i, email in enumerate(seed.get(d, []), 2):
            ws.cell(row=i, column=j, value=email)
    wb.save(path)


def _simple_col(ws, dept: str) -> int | None:
    for cell in ws[1]:
        if str(cell.value).strip().upper() == dept.upper():
            return cell.column
    return None


def _simple_get(pid: str, dept: str) -> list[str]:
    _ensure_simple(pid)
    wb = load_workbook(STORES[pid]["file"], read_only=True, data_only=True)
    ws = wb[_SIMPLE_SHEET] if _SIMPLE_SHEET in wb.sheetnames else wb.active
    col = _simple_col(ws, dept)
    out = []
    if col:
        for row in range(2, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v and "@" in str(v):
                out.append(str(v).strip())
    wb.close()
    return out


def _simple_set(pid: str, dept: str, emails: list[str]) -> int:
    _ensure_simple(pid)
    wb = load_workbook(STORES[pid]["file"])
    ws = wb[_SIMPLE_SHEET] if _SIMPLE_SHEET in wb.sheetnames else wb.active
    col = _simple_col(ws, dept)
    if not col:
        col = ws.max_column + 1
        ws.cell(row=1, column=col, value=dept)
    clean = _clean(emails)
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col).value = None
    for i, email in enumerate(clean, 2):
        ws.cell(row=i, column=col).value = email
    wb.save(STORES[pid]["file"])
    return len(clean)


# --------------------------------------------------------------------------- #
# CRM master (sheets by mode, header row 4)
# --------------------------------------------------------------------------- #
_CRM_HEADER_ROW = 4


def _crm_col(ws, dept: str) -> int | None:
    for cell in ws[_CRM_HEADER_ROW]:
        if str(cell.value).strip().upper() == dept.upper():
            return cell.column
    return None


def _crm_get(mode: str, dept: str) -> list[str]:
    path = C.CRM_EMAIL_MASTER
    if not path.exists():
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = CRM_MODE_SHEET.get(mode, "Actual")
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    col = _crm_col(ws, dept)
    out = []
    if col:
        for row in range(_CRM_HEADER_ROW + 1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v and "@" in str(v):
                out.append(str(v).strip())
    wb.close()
    return out


def _crm_set(mode: str, dept: str, emails: list[str]) -> int:
    wb = load_workbook(C.CRM_EMAIL_MASTER)
    sheet = CRM_MODE_SHEET.get(mode, "Actual")
    ws = wb[sheet]
    col = _crm_col(ws, dept)
    if not col:
        raise ValueError(f"column '{dept}' not found")
    clean = _clean(emails)
    for row in range(_CRM_HEADER_ROW + 1, ws.max_row + 1):
        ws.cell(row=row, column=col).value = None
    for i, email in enumerate(clean):
        ws.cell(row=_CRM_HEADER_ROW + 1 + i, column=col).value = email
    wb.save(C.CRM_EMAIL_MASTER)
    return len(clean)


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #
def departments(pid: str) -> list[str]:
    return STORES.get(pid, {}).get("departments", [])


def is_editable(pid: str) -> bool:
    return pid in STORES


def get(pid: str, dept: str, mode: str = "actual") -> list[str]:
    if pid == "crm":
        return _crm_get(mode, dept)
    return _simple_get(pid, dept)


def set_(pid: str, dept: str, emails: list[str], mode: str = "actual") -> int:
    if pid == "crm":
        return _crm_set(mode, dept, emails)
    return _simple_set(pid, dept, emails)


def all_(pid: str, mode: str = "actual") -> dict:
    return {d: get(pid, d, mode) for d in departments(pid)}
