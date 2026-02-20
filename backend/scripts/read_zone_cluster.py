#!/usr/bin/env python3
"""
Read Zone and cluster.xlsx - branch to Zone, Cluster, Product mapping.

Returns: dict branch_name -> {"zone": str, "cluster": str, "product": str}

Run: python backend/scripts/read_zone_cluster.py
"""

import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
ZONE_CLUSTER_FILE = BACKEND_DIR / "scripts" / "Zone and cluster.xlsx"


def load_zone_cluster_mapping(path=None):
    """Alias: returns lookup list for strict vlookup."""
    return load_zone_cluster_lookup(path)


def load_zone_cluster_lookup(path=None):
    """Load branch -> zone, cluster, product. First occurrence wins per branch."""
    from openpyxl import load_workbook

    fp = path or ZONE_CLUSTER_FILE
    if not fp.exists():
        return []

    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    zi = next((i for i, h in enumerate(headers) if h and "zone" in str(h).lower() and "cluster" not in str(h).lower()), 0)
    bi = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), 1)
    ci = next((i for i, h in enumerate(headers) if h and "cluster" in str(h).lower()), 2)
    pi = next((i for i, h in enumerate(headers) if h and "product" in str(h).lower()), 3)

    out = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(zi, bi, ci, pi):
            continue
        branch = str(row[bi] or "").strip()
        if not branch or branch in seen:
            continue
        seen.add(branch)
        out.append({
            "branch": branch,
            "zone": str(row[zi] or "").strip(),
            "cluster": str(row[ci] or "").strip(),
            "product": str(row[pi] or "").strip(),
        })
    wb.close()
    return out


def vlookup_branch(branch, lookup):
    """Strict vlookup: exact match only. Returns zone/cluster/product or ERR if no match."""
    x = str(branch or "").strip()
    matches = [r for r in lookup if r.get("branch") == x]
    if matches:
        return matches[0]
    return {"zone": "ERR", "cluster": "ERR", "product": "ERR"}


if __name__ == "__main__":
    m = load_zone_cluster_mapping()
    print(f"Loaded {len(m)} branch mappings from {ZONE_CLUSTER_FILE.name}\n")
    for r in sorted(m, key=lambda x: x["branch"])[:30]:
        print(f"  {r['branch']} -> Zone: {r['zone']}, Cluster: {r['cluster']}, Product: {r['product']}")
