#!/usr/bin/env python3
"""
Correct Management Reports - Using Clients file (true source of truth)

Uses openpyxl to ONLY modify cell values. Preserves colors, freeze panes, links.

Logic (same as correct_single_report_test.py):
- Leaf branches, Zone, Cluster, Product, Country: all from Clients file via branch summary
- Zone/Cluster/Product mapping from Zone and cluster.xlsx
- Product totals and Country: from product_sums (not summed from MR structure)

Output: Saves corrected files to ManagementCorrection/readable/ with readable names
        (Management_Report_YYYY-MM-DD.xlsx) and metadata.

Run from project root:
  python backend/scripts/correct_management_reports.py
"""

import json
import os
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
PROJECT_ROOT = BACKEND_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

import warnings
warnings.filterwarnings("ignore", message="Workbook contains no default style")

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install: pip install openpyxl")
    sys.exit(1)

MANAGEMENT_CORRECTION_DIR = BACKEND_DIR / "ManagementCorrection"
READABLE_OUTPUT_DIR = MANAGEMENT_CORRECTION_DIR / "readable"
CLIENTS_FILE = BACKEND_DIR / "Clients-platinumtanzania-dmasubi-2026-02-14T08_37_08.734_03_00.xlsx"
ZONE_CLUSTER_FILE = BACKEND_DIR / "scripts" / "Zone and cluster.xlsx"
UPLOADS_MANAGEMENT = BACKEND_DIR / "uploads" / "ALL" / "MANAGEMENT"


def load_env():
    """Load .env from project root or backend."""
    for p in [PROJECT_ROOT / ".env", BACKEND_DIR / ".env"]:
        if p.exists():
            with open(p) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        os.environ[k.strip()] = v.strip().strip("'\"")


def parse_created_date(val):
    """Parse Excel date or string to datetime."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return base + timedelta(days=int(float(val)))
        except Exception:
            return None
    if isinstance(val, str):
        val = val.strip()
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(val[:19].replace("Z", "").replace("T", " ")[:19], fmt)
            except Exception:
                pass
        try:
            return datetime.strptime(val[:10], "%Y-%m-%d")
        except Exception:
            pass
    return None


def load_zone_cluster_lookup(path=None):
    """Load Zone and cluster.xlsx as lookup table."""
    fp = path or ZONE_CLUSTER_FILE
    if not fp.exists():
        return []
    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    zi = next((i for i, h in enumerate(headers) if h and "zone" in str(h).lower() and "cluster" not in str(h).lower()), 0)
    bi = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), 1)
    ci = next((i for i, h in enumerate(headers) if h and "cluster" in str(h).lower()), 2)
    pi = next((i for i, h in enumerate(headers) if h and "product" in str(h).lower()), 3)
    out = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(zi, bi, ci, pi):
            continue
        branch = str(row[bi] or "").strip()
        if not branch or branch in seen:
            continue
        seen.add(branch)
        out.append({
            "branch": branch,
            "zone": str(row[zi] or "").strip(),
            "cluster": str(row[ci] or "").strip(),
            "product": str(row[pi] or "").strip(),
        })
    wb.close()
    return out


def vlookup_branch(branch, lookup):
    """Strict vlookup: exact match only."""
    x = str(branch or "").strip()
    matches = [r for r in lookup if r.get("branch") == x]
    if matches:
        return matches[0]
    return {"zone": "ERR", "cluster": "ERR", "product": "ERR"}


_clients_cache = {}


def load_clients_by_date(cutoff_date):
    """Load client counts per branch from Clients file. Only clients with Created <= cutoff_date."""
    if not CLIENTS_FILE.exists():
        raise FileNotFoundError(f"Clients file not found: {CLIENTS_FILE}")

    cache_key = cutoff_date.strftime("%Y-%m-%d") if cutoff_date else "none"
    if cache_key in _clients_cache:
        return _clients_cache[cache_key]

    wb = load_workbook(CLIENTS_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    branch_col = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), -1)
    state_col = next((i for i, h in enumerate(headers) if h and "client" in str(h).lower() and "state" in str(h).lower()), -1)
    created_col = next((i for i, h in enumerate(headers) if h and "created" in str(h).lower()), -1)

    if branch_col < 0 or state_col < 0 or created_col < 0:
        wb.close()
        raise ValueError("Clients file missing Branch, Client State, or Created column")

    cutoff = cutoff_date if cutoff_date else datetime(9999, 12, 31)
    by_branch = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) if row else []
        if len(row) <= max(branch_col, state_col, created_col):
            continue
        created = parse_created_date(row[created_col])
        # Exclude clients with unparseable Created - we need Created <= report date (plan §5 Step 2)
        if created is None:
            continue
        if created > cutoff:
            continue
        branch = str(row[branch_col] or "").strip()
        if not branch:
            continue
        state = str(row[state_col] or "").strip().lower()
        if state not in ("active", "inactive"):
            continue
        if branch not in by_branch:
            by_branch[branch] = {"active": 0, "inactive": 0}
        if state == "inactive":
            by_branch[branch]["inactive"] += 1
        else:
            by_branch[branch]["active"] += 1

    wb.close()
    _clients_cache[cache_key] = by_branch
    return by_branch


def build_branch_summary(by_branch, zone_cluster):
    """Build zone_sums, cluster_sums, product_sums from by_branch + Zone and cluster lookup."""
    zone_sums = {}
    cluster_sums = {}
    product_sums = {}
    for branch, counts in by_branch.items():
        lookup = vlookup_branch(branch, zone_cluster)
        z, c, p = lookup["zone"], lookup["cluster"], lookup["product"]
        if z == "ERR" or c == "ERR" or p == "ERR":
            continue
        for d, k in [(zone_sums, z), (cluster_sums, c), (product_sums, p)]:
            if k and k not in d:
                d[k] = {"active": 0, "inactive": 0}
            if k:
                d[k]["active"] += counts["active"]
                d[k]["inactive"] += counts["inactive"]
    return zone_sums, cluster_sums, product_sums


def match_branch(report_branch, client_branch):
    """Match report branch name to Clients file branch."""
    r = str(report_branch or "").strip().lower()
    c = str(client_branch or "").strip().lower()
    if r == c:
        return True
    if r in c:
        return True
    return False


def get_counts(report_branch, by_branch):
    """Get client counts for a report branch from Clients data."""
    for cb, counts in by_branch.items():
        if match_branch(report_branch, cb):
            return counts
    return None


def is_leaf_branch(val):
    """True if this is a leaf branch row (e.g. Bariadi, Chato)."""
    if not val or not isinstance(val, str):
        return False
    v = val.lower()
    return not re.search(
        r"cluster|zone|^cs$|^lbf$|zanzibar|call center|^sme$|maziwa|agrifinance|lbf zone|smes",
        v,
    )


def is_aggregate_row(val):
    """Returns 'CS', 'LBF', 'SME', 'Agrifinance' or None."""
    if not val or not isinstance(val, str):
        return None
    v = val.strip().lower()
    if v == "cs":
        return "CS"
    if v == "lbf":
        return "LBF"
    if v == "sme":
        return "SME"
    if "agrifinance" in v or v == "maziwa":
        return "Agrifinance"
    return None


def find_columns(ws):
    """Find column indices for Branch, Number of Clients, Active clients, Inactive clients."""
    headers = [c.value for c in ws[1]]
    idx = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        h = str(h).strip().lower()
        if "branch" in h:
            idx["branch"] = i
        elif "number" in h and "clients" in h:
            idx["num_clients"] = i
        elif "inactive" in h and "clients" in h:
            idx["inactive"] = i
        elif "active" in h and "clients" in h:
            idx["active"] = i
    if "num_clients" not in idx:
        idx["num_clients"] = next((i for i, h in enumerate(headers) if h and "clients" in str(h).lower()), -1)
    if "active" not in idx:
        idx["active"] = next((i for i, h in enumerate(headers) if h and "active" in str(h).lower() and "clients" in str(h).lower() and "inactive" not in str(h).lower()), -1)
    if "inactive" not in idx:
        idx["inactive"] = next((i for i, h in enumerate(headers) if h and "inactive" in str(h).lower() and "clients" in str(h).lower()), -1)
    return idx


def build_rows_info(ws, branch_col_m):
    """Build list of (row_idx, branch, is_leaf_row, agg) for Country sheet."""
    rows_info = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row) if row else []
        if len(row) <= branch_col_m:
            rows_info.append((row_idx, None, False, None))
            continue
        val = row[branch_col_m]
        if val is None or (isinstance(val, str) and not val.strip()):
            rows_info.append((row_idx, None, False, None))
            continue
        branch = str(val).strip()
        agg = is_aggregate_row(branch)
        leaf = is_leaf_branch(branch) if not agg else False
        rows_info.append((row_idx, branch, leaf, agg))

    sections = {"CS": [], "LBF": [], "SME": [], "Agrifinance": []}
    current = None
    for row_idx, branch, is_leaf_row, agg in rows_info:
        if not branch:
            continue
        b = branch.lower()
        if "cluster 1" in b or "cluster 2" in b or "cluster 3" in b:
            current = "CS"
        elif "lbf cluster" in b:
            current = "LBF"
        elif "sme" in b and "zone" not in b and "call" not in b:
            if "smes" in b or "branch" in b or is_leaf_row:
                current = "SME"
        elif "agrifinance" in b or "maziwa" in b:
            current = "Agrifinance"
        elif "zanzibar" in b and "zone" not in b:
            current = "ZANZIBAR"
        elif "call center" in b:
            current = None
        if is_leaf_row and current and current in sections:
            sections[current].append(row_idx)
    return rows_info


def correct_single_report(
    report_id,
    file_path,
    report_date,
    by_branch,
    zone_sums,
    cluster_sums,
    product_sums,
    zone_cluster,
):
    """
    Correct one management report. Returns (wb, values_by_row) or (None, None) on skip.
    Caller must close wb. Source file is read from ManagementCorrection or uploads.
    """
    basename = Path(file_path).name
    candidates = [
        MANAGEMENT_CORRECTION_DIR / f"{report_id}_{basename}",
        MANAGEMENT_CORRECTION_DIR / basename,
        UPLOADS_MANAGEMENT / basename,
    ]
    src_file = None
    for p in candidates:
        if p.exists():
            src_file = p
            break
    if not src_file or not src_file.exists():
        return None, None

    wb = load_workbook(src_file, data_only=False, rich_text=True, keep_links=True)
    country_sheet = next((s for s in wb.sheetnames if "country" in s.lower()), wb.sheetnames[0])
    ws = wb[country_sheet]

    idx = find_columns(ws)
    branch_col_m = idx.get("branch", -1)
    num_col1 = idx.get("num_clients", -1) + 1
    active_col1 = idx.get("active", -1) + 1
    inactive_col1 = idx.get("inactive", -1) + 1

    if branch_col_m < 0 or num_col1 <= 0:
        return None, None

    rows_info = build_rows_info(ws, branch_col_m)

    PRODUCT_ROW_NAMES = ("CS", "LBF", "SME", "Agrifinance", "Maziwa")
    values_by_row = {}

    # All values from Clients / branch summary (true source of truth)
    for row_idx, branch, is_leaf_row, agg in rows_info:
        if not branch:
            continue
        if branch == "Country":
            continue
        if branch in PRODUCT_ROW_NAMES:
            if branch in product_sums:
                v = product_sums[branch]
                values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}
            continue
        if is_leaf_row:
            counts = get_counts(branch, by_branch)
            if counts:
                total = counts["active"] + counts["inactive"]
                values_by_row[row_idx] = {"num": total, "active": counts["active"], "inactive": counts["inactive"]}
        elif branch in zone_sums:
            v = zone_sums[branch]
            values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}
        elif branch in cluster_sums:
            v = cluster_sums[branch]
            values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}

    # Country: sum of all product_sums
    for row_idx, branch, _, _ in rows_info:
        if branch != "Country":
            continue
        c_num = c_active = c_inactive = 0
        for v in product_sums.values():
            c_num += v["active"] + v["inactive"]
            c_active += v["active"]
            c_inactive += v["inactive"]
        values_by_row[row_idx] = {"num": c_num, "active": c_active, "inactive": c_inactive}

    for row_idx, vals in values_by_row.items():
        if num_col1 > 0:
            ws.cell(row=row_idx, column=num_col1).value = vals.get("num", 0)
        if active_col1 > 0:
            ws.cell(row=row_idx, column=active_col1).value = vals.get("active", 0)
        if inactive_col1 > 0:
            ws.cell(row=row_idx, column=inactive_col1).value = vals.get("inactive", 0)

    return wb, values_by_row


def main():
    load_env()

    metadata_path = MANAGEMENT_CORRECTION_DIR / "management_reports_metadata.json"
    if not metadata_path.exists():
        print(f"Metadata not found: {metadata_path}")
        print("Run backup first: npm run backup-management-reports")
        print("Then copy backup contents to ManagementCorrection/")
        sys.exit(1)

    with open(metadata_path) as f:
        reports = json.load(f)

    if not CLIENTS_FILE.exists():
        print(f"Clients file not found: {CLIENTS_FILE}")
        sys.exit(1)

    zone_cluster = load_zone_cluster_lookup()
    if not zone_cluster:
        print(f"Warning: Zone and cluster file empty or not found: {ZONE_CLUSTER_FILE}")
    else:
        print(f"Zone/cluster lookup: {len(zone_cluster)} branches")

    READABLE_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    used_names = {}
    output_metadata = []

    print("=" * 60)
    print("  Management Reports Correction (Clients = source of truth)")
    print("=" * 60)
    print(f"\nWorking dir: {MANAGEMENT_CORRECTION_DIR}")
    print(f"Output:     {READABLE_OUTPUT_DIR} (readable names + metadata)")
    print(f"Clients:    {CLIENTS_FILE.name}")
    print(f"Reports:    {len(reports)}\n")

    updated = 0
    for i, r in enumerate(reports):
        report_id = r["id"]
        file_path = r["file_path"]
        report_date = r.get("date") or "unknown"
        basename = Path(file_path).name

        try:
            cutoff = datetime.strptime(str(report_date)[:10], "%Y-%m-%d") if report_date and report_date != "unknown" else None
        except (ValueError, TypeError):
            cutoff = None
        if cutoff is None:
            print(f"  Skip [{i+1}/{len(reports)}] {basename} (invalid/missing date: {report_date})")
            continue
        by_branch = load_clients_by_date(cutoff)
        zone_sums, cluster_sums, product_sums = build_branch_summary(by_branch, zone_cluster)

        wb, values_by_row = correct_single_report(
            report_id, file_path, report_date,
            by_branch, zone_sums, cluster_sums, product_sums, zone_cluster,
        )

        if wb is None or not values_by_row:
            print(f"  Skip [{i+1}/{len(reports)}] {basename} (not found or missing columns)")
            continue

        # Readable filename: Management_Report_YYYY-MM-DD.xlsx
        safe_date = str(report_date).replace("/", "-").replace(" ", "_")[:10]
        base_out = f"Management_Report_{safe_date}.xlsx"
        if base_out in used_names:
            used_names[base_out] += 1
            stem = base_out.replace(".xlsx", "")
            out_name = f"{stem}_{used_names[base_out]}.xlsx"
        else:
            used_names[base_out] = 1
            out_name = base_out

        dst_file = READABLE_OUTPUT_DIR / out_name
        wb.save(dst_file)
        wb.close()
        updated += 1

        output_metadata.append({
            "report_id": report_id,
            "file_path": file_path,
            "date": report_date,
            "readable_filename": out_name,
            "rows_updated": len(values_by_row),
        })

        print(f"  OK [{updated}/{len(reports)}] {out_name} ({report_date}) - {len(values_by_row)} rows")

    # Save metadata in readable folder
    meta_out = READABLE_OUTPUT_DIR / "readable_metadata.json"
    with open(meta_out, "w") as f:
        json.dump({"generated": datetime.now().isoformat(), "reports": output_metadata}, f, indent=2)

    print("\n" + "=" * 60)
    print(f"Done: {updated} files corrected and saved to {READABLE_OUTPUT_DIR}")
    print(f"Metadata: {meta_out}")
    print("=" * 60)


if __name__ == "__main__":
    main()
