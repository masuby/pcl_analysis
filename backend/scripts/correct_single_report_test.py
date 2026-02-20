#!/usr/bin/env python3
"""
Single-file test: Correct ONE management report step by step.

Processes only: Management Report for date 2026-01-30
(id: 25b735f6-15fb-4a25-b38a-f45eba770032)

Steps (all outputs go to ManagementCorrection/Test/):
1. clients_active_inactive_only.xlsx  - Clients filtered to Active/Inactive only
2. clients_filtered_by_date.xlsx      - Clients with Created <= 2026-01-30
3. branch_summary_2026-01-30.xlsx     - Unique branches with Active & Inactive counts
4. Management_Report_2026-01-30.xlsx  - Corrected management report (inspect this)

Run from project root:
  python backend/scripts/correct_single_report_test.py
"""

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
BACKEND_DIR = SCRIPT_DIR.parent
PROJECT_ROOT = BACKEND_DIR.parent

import warnings
warnings.filterwarnings("ignore", message="Workbook contains no default style")

try:
    from openpyxl import load_workbook, Workbook
except ImportError:
    print("Install: pip install openpyxl")
    sys.exit(1)

# Paths
TEST_DIR = BACKEND_DIR / "ManagementCorrection" / "Test"
CLIENTS_FILE = BACKEND_DIR / "Clients-platinumtanzania-dmasubi-2026-02-14T08_37_08.734_03_00.xlsx"
MANAGEMENT_CORRECTION_DIR = BACKEND_DIR / "ManagementCorrection"
ZONE_CLUSTER_FILE = BACKEND_DIR / "scripts" / "Zone and cluster.xlsx"

# Single report to process
REPORT_ID = "25b735f6-15fb-4a25-b38a-f45eba770032"
REPORT_DATE = "2026-01-30"
REPORT_BASENAME = "1769947443099_Management_Report2026-01.xlsx"


def parse_created_date(val):
    """Parse Excel date or string to datetime."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, (int, float)):
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return base + timedelta(days=int(float(val)))
        except Exception:
            return None
    if isinstance(val, str):
        val = val.strip()
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(val[:19].replace("Z", "").replace("T", " ")[:19], fmt)
            except Exception:
                pass
        try:
            return datetime.strptime(val[:10], "%Y-%m-%d")
        except Exception:
            pass
    return None


def load_zone_cluster_lookup(path=None):
    """Load Zone and cluster.xlsx as lookup table (list of rows with branch, zone, cluster, product)."""
    fp = path or ZONE_CLUSTER_FILE
    if not fp.exists():
        return []
    wb = load_workbook(fp, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    zi = next((i for i, h in enumerate(headers) if h and "zone" in str(h).lower() and "cluster" not in str(h).lower()), 0)
    bi = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), 1)
    ci = next((i for i, h in enumerate(headers) if h and "cluster" in str(h).lower()), 2)
    pi = next((i for i, h in enumerate(headers) if h and "product" in str(h).lower()), 3)
    out = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= max(zi, bi, ci, pi):
            continue
        branch = str(row[bi] or "").strip()
        if not branch or branch in seen:
            continue
        seen.add(branch)
        out.append({
            "branch": branch,
            "zone": str(row[zi] or "").strip(),
            "cluster": str(row[ci] or "").strip(),
            "product": str(row[pi] or "").strip(),
        })
    wb.close()
    return out


def vlookup_branch(branch, lookup):
    """Strict vlookup: exact match only. lookup is list of dicts with 'branch' key."""
    x = str(branch or "").strip()
    matches = [r for r in lookup if r.get("branch") == x]
    if matches:
        return matches[0]
    return {"zone": "ERR", "cluster": "ERR", "product": "ERR"}


def main():
    TEST_DIR.mkdir(parents=True, exist_ok=True)

    if not CLIENTS_FILE.exists():
        # Try Test folder
        alt = TEST_DIR / CLIENTS_FILE.name
        if alt.exists():
            clients_path = alt
        else:
            print(f"Clients file not found: {CLIENTS_FILE}")
            sys.exit(1)
    else:
        clients_path = CLIENTS_FILE

    cutoff_date = datetime.strptime(REPORT_DATE, "%Y-%m-%d")

    print("=" * 60)
    print("  Single Report Test - Step by Step")
    print("=" * 60)
    print(f"\nReport date: {REPORT_DATE}")
    print(f"Output dir:  {TEST_DIR}\n")

    # --- Step 1: Filter to Active/Inactive only (preserve formatting) ---
    print("Step 1: Filter Clients to Active/Inactive only...")
    wb_src = load_workbook(clients_path, read_only=True, data_only=True)
    ws_src = wb_src.active
    headers = [c.value for c in ws_src[1]]
    branch_col = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), -1)
    state_col = next((i for i, h in enumerate(headers) if h and "client" in str(h).lower() and "state" in str(h).lower()), -1)
    created_col = next((i for i, h in enumerate(headers) if h and "created" in str(h).lower()), -1)

    if branch_col < 0 or state_col < 0 or created_col < 0:
        wb_src.close()
        print("Error: Clients file missing Branch, Client State, or Created column")
        sys.exit(1)

    step1_rows = [headers]
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        row = list(row) if row else []
        if len(row) <= state_col:
            continue
        state = str(row[state_col] or "").strip().lower()
        if state not in ("active", "inactive"):
            continue
        step1_rows.append(row)
    wb_src.close()

    # Save with formatting: copy column widths + freeze panes from original
    wb_orig = load_workbook(clients_path, read_only=False, data_only=False)
    ws_orig = wb_orig.active
    col_dims = {k: v.width for k, v in ws_orig.column_dimensions.items() if v.width}
    freeze = ws_orig.freeze_panes
    wb_orig.close()

    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Clients Active Inactive only"
    for ri, row in enumerate(step1_rows, 1):
        for ci, val in enumerate(row, 1):
            ws1.cell(row=ri, column=ci, value=val)
    for col, width in col_dims.items():
        ws1.column_dimensions[col].width = width
    if freeze:
        ws1.freeze_panes = "A2"
    out1 = TEST_DIR / "clients_active_inactive_only.xlsx"
    wb1.save(out1)
    wb1.close()
    print(f"  Saved: {out1.name} ({len(step1_rows)-1} rows)")

    # --- Step 2: Filter to Created <= date (preserve formatting) ---
    print("\nStep 2: Filter to Created <= " + REPORT_DATE + "...")
    step2_rows = [headers]
    for row in step1_rows[1:]:
        if len(row) <= created_col:
            continue
        created = parse_created_date(row[created_col])
        # Only include clients with Created <= report date (exclude unparseable)
        if created is not None and created <= cutoff_date:
            step2_rows.append(row)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Clients by " + REPORT_DATE
    for ri, row in enumerate(step2_rows, 1):
        for ci, val in enumerate(row, 1):
            ws2.cell(row=ri, column=ci, value=val)
    for col, width in col_dims.items():
        ws2.column_dimensions[col].width = width
    if freeze:
        ws2.freeze_panes = "A2"
    out2 = TEST_DIR / "clients_filtered_by_date.xlsx"
    wb2.save(out2)
    wb2.close()
    print(f"  Saved: {out2.name} ({len(step2_rows)-1} rows)")

    # --- Step 3: Branch summary with Zone, Cluster, Product ---
    print("\nStep 3: Create branch summary...")
    zone_cluster = load_zone_cluster_lookup()

    by_branch = {}
    for row in step2_rows[1:]:
        if len(row) <= max(branch_col, state_col):
            continue
        branch = str(row[branch_col] or "").strip()
        if not branch:
            continue
        state = str(row[state_col] or "").strip().lower()
        if state not in ("active", "inactive"):
            continue
        if branch not in by_branch:
            by_branch[branch] = {"active": 0, "inactive": 0}
        if state == "inactive":
            by_branch[branch]["inactive"] += 1
        else:
            by_branch[branch]["active"] += 1

    zone_sums = {}
    cluster_sums = {}
    product_sums = {}
    for branch, counts in by_branch.items():
        lookup = vlookup_branch(branch, zone_cluster)
        z, c, p = lookup["zone"], lookup["cluster"], lookup["product"]
        if z == "ERR" or c == "ERR" or p == "ERR":
            continue
        for d, k in [(zone_sums, z), (cluster_sums, c), (product_sums, p)]:
            if k and k not in d:
                d[k] = {"active": 0, "inactive": 0}
            if k:
                d[k]["active"] += counts["active"]
                d[k]["inactive"] += counts["inactive"]

    out3 = TEST_DIR / f"branch_summary_{REPORT_DATE}.xlsx"
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Branch Summary"
    ws3["A1"] = "Branch"
    ws3["B1"] = "Zone"
    ws3["C1"] = "Cluster"
    ws3["D1"] = "Product"
    ws3["E1"] = "Active"
    ws3["F1"] = "Inactive"
    ws3["G1"] = "Total"
    row_num = 2
    for branch, counts in sorted(by_branch.items()):
        lookup = vlookup_branch(branch, zone_cluster)
        z = lookup["zone"] if lookup.get("zone") != "ERR" else "ERR"
        c = lookup["cluster"] if lookup.get("cluster") != "ERR" else "ERR"
        p = lookup["product"] if lookup.get("product") != "ERR" else "ERR"
        ws3.cell(row=row_num, column=1, value=branch)
        ws3.cell(row=row_num, column=2, value=z)
        ws3.cell(row=row_num, column=3, value=c)
        ws3.cell(row=row_num, column=4, value=p)
        ws3.cell(row=row_num, column=5, value=counts["active"])
        ws3.cell(row=row_num, column=6, value=counts["inactive"])
        ws3.cell(row=row_num, column=7, value=counts["active"] + counts["inactive"])
        row_num += 1
    if zone_sums:
        ws3.cell(row=row_num, column=1, value="--- ZONE TOTALS ---")
        row_num += 1
        for z, v in sorted(zone_sums.items()):
            ws3.cell(row=row_num, column=1, value=z)
            ws3.cell(row=row_num, column=2, value=z)
            ws3.cell(row=row_num, column=5, value=v["active"])
            ws3.cell(row=row_num, column=6, value=v["inactive"])
            ws3.cell(row=row_num, column=7, value=v["active"] + v["inactive"])
            row_num += 1
    if cluster_sums:
        ws3.cell(row=row_num, column=1, value="--- CLUSTER TOTALS ---")
        row_num += 1
        for c, v in sorted(cluster_sums.items()):
            ws3.cell(row=row_num, column=1, value=c)
            ws3.cell(row=row_num, column=3, value=c)
            ws3.cell(row=row_num, column=5, value=v["active"])
            ws3.cell(row=row_num, column=6, value=v["inactive"])
            ws3.cell(row=row_num, column=7, value=v["active"] + v["inactive"])
            row_num += 1
    if product_sums:
        ws3.cell(row=row_num, column=1, value="--- PRODUCT TOTALS ---")
        row_num += 1
        for p, v in sorted(product_sums.items()):
            ws3.cell(row=row_num, column=1, value=p)
            ws3.cell(row=row_num, column=4, value=p)
            ws3.cell(row=row_num, column=5, value=v["active"])
            ws3.cell(row=row_num, column=6, value=v["inactive"])
            ws3.cell(row=row_num, column=7, value=v["active"] + v["inactive"])
            row_num += 1
    for col, w in [("A", 35), ("B", 22), ("C", 18), ("D", 12), ("E", 10), ("F", 10), ("G", 10)]:
        ws3.column_dimensions[col].width = w
    wb3.save(out3)
    wb3.close()
    print(f"  Saved: {out3.name} ({len(by_branch)} branches, {len(zone_sums)} zones, {len(cluster_sums)} clusters)")

    # --- Step 4: Update management report (preserve formatting) ---
    # Prefer ManagementCorrection - formatted files from API download
    print("\nStep 4: Update management report...")
    candidates = [
        MANAGEMENT_CORRECTION_DIR / f"{REPORT_ID}_{REPORT_BASENAME}",
        MANAGEMENT_CORRECTION_DIR / REPORT_BASENAME,
        TEST_DIR / f"{REPORT_ID}_{REPORT_BASENAME}",
        BACKEND_DIR / "uploads" / "ALL" / "MANAGEMENT" / REPORT_BASENAME,
    ]
    mgmt_src = None
    for p in candidates:
        if p.exists():
            mgmt_src = p
            break
    if not mgmt_src or not mgmt_src.exists():
        print("  Management report not found. Tried:")
        for p in candidates:
            print(f"    - {p}")
        sys.exit(1)
    print(f"  Using: {mgmt_src.name}")

    out4 = TEST_DIR / f"Management_Report_{REPORT_DATE}.xlsx"
    shutil.copy2(mgmt_src, out4)

    wb_mgmt = load_workbook(
        out4,
        data_only=False,
        rich_text=True,
        keep_links=True,
    )
    country_sheet = next((s for s in wb_mgmt.sheetnames if "country" in s.lower()), wb_mgmt.sheetnames[0])
    ws = wb_mgmt[country_sheet]

    # Find columns
    headers_m = [c.value for c in ws[1]]
    idx = {}
    for i, h in enumerate(headers_m):
        if not h:
            continue
        h = str(h).strip().lower()
        if "branch" in h:
            idx["branch"] = i
        elif "number" in h and "clients" in h:
            idx["num_clients"] = i
        elif "inactive" in h and "clients" in h:
            idx["inactive"] = i
        elif "active" in h and "clients" in h:
            idx["active"] = i
    if "num_clients" not in idx:
        idx["num_clients"] = next((i for i, h in enumerate(headers_m) if h and "clients" in str(h).lower()), -1)
    if "active" not in idx:
        idx["active"] = next((i for i, h in enumerate(headers_m) if h and "active" in str(h).lower() and "clients" in str(h).lower() and "inactive" not in str(h).lower()), -1)
    if "inactive" not in idx:
        idx["inactive"] = next((i for i, h in enumerate(headers_m) if h and "inactive" in str(h).lower() and "clients" in str(h).lower()), -1)

    branch_col_m = idx.get("branch", 0)
    num_col1 = idx.get("num_clients", -1) + 1
    active_col1 = idx.get("active", -1) + 1
    inactive_col1 = idx.get("inactive", -1) + 1

    def is_leaf(val):
        if not val or not isinstance(val, str):
            return False
        v = val.lower()
        return not re.search(r"cluster|zone|^cs$|^lbf$|zanzibar|call center|^sme$|maziwa|agrifinance|lbf zone|smes", v)

    def is_aggregate(val):
        if not val or not isinstance(val, str):
            return None
        v = val.strip().lower()
        if v == "cs":
            return "CS"
        if v == "lbf":
            return "LBF"
        if v == "sme":
            return "SME"
        if "agrifinance" in v or v == "maziwa":
            return "Agrifinance"
        return None

    def match_branch(report_br, client_br):
        r = str(report_br or "").strip().lower()
        c = str(client_br or "").strip().lower()
        if r == c:
            return True
        if r in c:
            return True
        return False

    def get_counts(report_br):
        for cb, counts in by_branch.items():
            if match_branch(report_br, cb):
                return counts
        return None

    def build_section_map():
        rows_info = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = list(row) if row else []
            if len(row) <= branch_col_m:
                rows_info.append((row_idx, None, False, None))
                continue
            val = row[branch_col_m]
            if val is None or (isinstance(val, str) and not val.strip()):
                rows_info.append((row_idx, None, False, None))
                continue
            branch = str(val).strip()
            agg = is_aggregate(branch)
            leaf = is_leaf(branch) if not agg else False
            rows_info.append((row_idx, branch, leaf, agg))

        sections = {"CS": [], "LBF": [], "SME": [], "Agrifinance": []}
        current = None
        for row_idx, branch, is_leaf_row, agg in rows_info:
            if not branch:
                continue
            b = branch.lower()
            if "cluster 1" in b or "cluster 2" in b or "cluster 3" in b:
                current = "CS"
            elif "lbf cluster" in b:
                current = "LBF"
            elif "sme" in b and "zone" not in b and "call" not in b:
                if "smes" in b or "branch" in b or is_leaf_row:
                    current = "SME"
            elif "agrifinance" in b or "maziwa" in b:
                current = "Agrifinance"
            elif "zanzibar" in b and "zone" not in b:
                current = "ZANZIBAR"
            elif "call center" in b:
                current = None
            if is_leaf_row and current and current in sections:
                sections[current].append(row_idx)
        return sections, rows_info

    values_by_row = {}
    _, rows_info = build_section_map()

    PRODUCT_ROW_NAMES = ("CS", "LBF", "SME", "Agrifinance", "Maziwa")

    # All values come from Clients file / branch summary (true source of truth):
    # - Leaf branches: by_branch
    # - Zone rows: zone_sums
    # - Cluster rows: cluster_sums
    # - Product rows: product_sums (CS, LBF, SME, Agrifinance, Maziwa)
    # - Country: sum of all product_sums

    for row_idx, branch, is_leaf_row, agg in rows_info:
        if not branch:
            continue
        if branch == "Country":
            continue
        if branch in PRODUCT_ROW_NAMES:
            # Product totals: use branch summary product_sums (true data from Clients)
            if branch in product_sums:
                v = product_sums[branch]
                values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}
            continue
        if is_leaf_row:
            counts = get_counts(branch)
            if counts:
                total = counts["active"] + counts["inactive"]
                values_by_row[row_idx] = {
                    "num": total,
                    "active": counts["active"],
                    "inactive": counts["inactive"],
                }
        elif branch in zone_sums:
            v = zone_sums[branch]
            values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}
        elif branch in cluster_sums:
            v = cluster_sums[branch]
            values_by_row[row_idx] = {"num": v["active"] + v["inactive"], "active": v["active"], "inactive": v["inactive"]}

    # Country: sum of ALL product_sums from branch summary (true total from Clients)
    for row_idx, branch, _, _ in rows_info:
        if branch != "Country":
            continue
        c_num = c_active = c_inactive = 0
        for v in product_sums.values():
            c_num += v["active"] + v["inactive"]
            c_active += v["active"]
            c_inactive += v["inactive"]
        values_by_row[row_idx] = {"num": c_num, "active": c_active, "inactive": c_inactive}

    for row_idx, vals in values_by_row.items():
        if num_col1 > 0:
            ws.cell(row=row_idx, column=num_col1).value = vals.get("num", 0)
        if active_col1 > 0:
            ws.cell(row=row_idx, column=active_col1).value = vals.get("active", 0)
        if inactive_col1 > 0:
            ws.cell(row=row_idx, column=inactive_col1).value = vals.get("inactive", 0)

    updated_count = len(values_by_row)

    wb_mgmt.save(out4)
    wb_mgmt.close()
    print(f"  Saved: {out4.name} ({updated_count} branch rows updated)")

    print("\n" + "=" * 60)
    print("Done. Inspect files in ManagementCorrection/Test/")
    print("=" * 60)


if __name__ == "__main__":
    main()
