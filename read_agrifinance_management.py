"""
Temporary script: read Management Report Excel, first sheet (Country),
and show how the Agrifinance row is structured (headers + row data).
Run: python read_agrifinance_management.py
"""
import openpyxl
from pathlib import Path

# Path to the Management Report (relative to project root)
EXCEL_PATH = Path(__file__).resolve().parent / "src" / "pages" / "Dashboard" / "components" / "DepartmentalDashboard" / "components" / "ScoreCardReports" / "Management Report2026-02.xlsx"


def main():
    if not EXCEL_PATH.exists():
        print(f"File not found: {EXCEL_PATH}")
        return
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False, data_only=True)
    sheet = wb.active
    name = sheet.title
    print("=" * 80)
    print(f"FIRST SHEET: {name!r}")
    print("=" * 80)
    print(f"Dimensions: max_row={sheet.max_row}, max_column={sheet.max_column}")
    print()

    # Get all rows as list of lists (1-based row numbers from openpyxl)
    rows = []
    for r in range(1, min(sheet.max_row + 1, 100)):
        row_vals = []
        for c in range(1, min(sheet.max_column + 1, 50)):
            cell = sheet.cell(row=r, column=c)
            row_vals.append(cell.value)
        rows.append(row_vals)

    # Find header row (row containing "Number of Clients" or "Active clients" or "Branch")
    header_row_idx = None
    for i, row in enumerate(rows):
        for j, val in enumerate(row):
            if val is None:
                continue
            s = str(val).strip().lower()
            if "number of clients" in s or "active clients" in s or "inactive clients" in s or (j == 0 and "branch" in s):
                header_row_idx = i
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        print("Could not find header row. Showing first 15 rows (first 25 cols):")
        for i, row in enumerate(rows[:15]):
            print(f"  Row {i+1}: {row[:25]}")
        wb.close()
        return

    headers = rows[header_row_idx]
    print(f"Header row (row {header_row_idx + 1}):")
    for col_idx, h in enumerate(headers):
        if h is not None and str(h).strip():
            print(f"  Col {col_idx}: {h!r}")
    print()

    # Find Agrifinance row (branch column usually 0 or 1)
    agri_row_idx = None
    for i in range(header_row_idx + 1, len(rows)):
        row = rows[i]
        for j in range(min(3, len(row))):
            val = row[j] if j < len(row) else None
            if val is None:
                continue
            s = str(val).strip().lower()
            if s == "agrifinance" or s == "agri finance" or s == "maziwa":
                agri_row_idx = i
                print(f"Found Agrifinance row at row {i + 1}, branch cell col {j}: {val!r}")
                break
        if agri_row_idx is not None:
            break

    if agri_row_idx is None:
        print("No Agrifinance row found. Rows after header (first 20 cols, branch in col 0):")
        for i in range(header_row_idx + 1, min(header_row_idx + 25, len(rows))):
            row = rows[i]
            branch = row[0] if row else None
            print(f"  Row {i+1}: branch={branch!r} | {row[:20]}")
        wb.close()
        return

    # Print Agrifinance row with header labels
    agri_row = rows[agri_row_idx]
    print()
    print("Agrifinance row (value per header):")
    print("-" * 60)
    for col_idx, h in enumerate(headers):
        if h is None or not str(h).strip():
            continue
        val = agri_row[col_idx] if col_idx < len(agri_row) else None
        print(f"  {h!r} => {val!r}")
    print("-" * 60)
    # Column indices (0-based) for client metrics on Country sheet
    col_number_of_clients = next((j for j, h in enumerate(headers) if h and "number of clients" == str(h).strip().lower()), None)
    col_active_clients = next((j for j, h in enumerate(headers) if h and "active clients" == str(h).strip().lower()), None)
    col_inactive_clients = next((j for j, h in enumerate(headers) if h and "inactive clients" == str(h).strip().lower()), None)
    print("Column indices (0-based) for client metrics:")
    print(f"  Number of Clients: col {col_number_of_clients} (value={agri_row[col_number_of_clients] if col_number_of_clients is not None else 'N/A'})")
    print(f"  Active clients:    col {col_active_clients} (value={agri_row[col_active_clients] if col_active_clients is not None else 'N/A'})")
    print(f"  Inactive clients:  col {col_inactive_clients} (value={agri_row[col_inactive_clients] if col_inactive_clients is not None else 'N/A'})")
    print()
    print("Full Agrifinance row (all columns):")
    print(agri_row[:42])
    wb.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
