#!/usr/bin/env python3
"""
Correct Management Reports - Preserves Excel Formatting
"""
import warnings
warnings.filterwarnings("ignore", message="Workbook contains no default style")

Uses openpyxl to ONLY modify cell values (preserves colors, freeze panes, links).

Flow:
1. Read from BACKUP folder (original files with formatting)
2. Correct Number of Clients, Active clients, Inactive clients
3. Update branch rows from Clients file + aggregate rows (CS, LBF, SME, Agrifinance)
4. Write to UPLOADS folder (replaces files)

First restore from backup if uploads are corrupted:
  python scripts/restore_from_backup.py

Then run correction:
  python scripts/correct_management_reports.py

Prerequisites:
  pip install -r requirements.txt
"""

import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from openpyxl import load_workbook
except ImportError:
    print("Install: pip install openpyxl")
    sys.exit(1)

try:
    import psycopg2
except ImportError:
    psycopg2 = None

BACKUP_DIR = PROJECT_ROOT / "backend" / "backup" / "management_reports"
UPLOADS_BASE = PROJECT_ROOT / "backend" / "uploads"
CLIENTS_FILE = PROJECT_ROOT / "backend" / "Clients-platinumtanzania-dmasubi-2026-02-14T08_37_08.734_03_00.xlsx"


def load_env():
    for p in [PROJECT_ROOT / ".env", PROJECT_ROOT / "backend" / ".env"]:
        if p.exists():
            with open(p) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        os.environ[k.strip()] = v.strip().strip("'\"")


def parse_created_date(val):
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return base + timedelta(days=int(float(val)))
        except Exception:
            return None
    if isinstance(val, str):
        val = val.strip()
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"]:
            try:
                return datetime.strptime(val[:19].replace("Z", "").replace("T", " ")[:19], fmt)
            except Exception:
                pass
        try:
            return datetime.strptime(val[:10], "%Y-%m-%d")
        except Exception:
            pass
    return None


_clients_cache = {}

def load_clients_by_date(cutoff_date):
    if not CLIENTS_FILE.exists():
        raise FileNotFoundError(f"Clients file not found: {CLIENTS_FILE}")

    cache_key = cutoff_date.strftime("%Y-%m-%d") if cutoff_date else "none"
    if cache_key in _clients_cache:
        return _clients_cache[cache_key]

    wb = load_workbook(CLIENTS_FILE, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    branch_col = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), -1)
    state_col = next((i for i, h in enumerate(headers) if h and "client" in str(h).lower() and "state" in str(h).lower()), -1)
    created_col = next((i for i, h in enumerate(headers) if h and "created" in str(h).lower()), -1)

    if branch_col < 0 or state_col < 0 or created_col < 0:
        wb.close()
        raise ValueError("Clients file missing Branch, Client State, or Created column")

    cutoff = cutoff_date if cutoff_date else datetime(9999, 12, 31)
    by_branch = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row) if row else []
        if len(row) <= max(branch_col, state_col, created_col):
            continue
        created = parse_created_date(row[created_col])
        if created and created > cutoff:
            continue
        branch = str(row[branch_col] or "").strip()
        if not branch:
            continue
        state = str(row[state_col] or "").lower()
        is_active = "active" in state

        if branch not in by_branch:
            by_branch[branch] = {"active": 0, "inactive": 0}
        if is_active:
            by_branch[branch]["active"] += 1
        else:
            by_branch[branch]["inactive"] += 1

    wb.close()
    _clients_cache[cache_key] = by_branch
    return by_branch


def norm(s):
    return str(s or "").strip().lower()


def match_branch(report_branch, client_branch):
    r = norm(report_branch)
    c = norm(client_branch)
    if r == c:
        return True
    if r in c:
        return True
    c_core = re.sub(r"\b(lbf|branch|cs|sme)\b", "", c, flags=re.I).strip()
    return r == c_core or r in c_core


def get_counts(report_branch, by_branch):
    for cb, counts in by_branch.items():
        if match_branch(report_branch, cb):
            return counts
    return None


def is_leaf_branch(val):
    if not val or not isinstance(val, str):
        return False
    v = val.lower()
    return not re.search(
        r"cluster|zone|^cs$|^lbf$|zanzibar|call center|^sme$|maziwa|agrifinance|lbf zone|smes",
        v,
    )


def is_aggregate_row(val):
    if not val or not isinstance(val, str):
        return None
    v = val.strip().lower()
    if v == "cs":
        return "CS"
    if v == "lbf":
        return "LBF"
    if v == "sme":
        return "SME"
    if "agrifinance" in v or v == "maziwa":
        return "Agrifinance"
    return None


def find_columns(ws):
    headers = [c.value for c in ws[1]]
    idx = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        h = str(h).strip().lower()
        if "branch" in h:
            idx["branch"] = i
        elif "number" in h and "clients" in h:
            idx["num_clients"] = i
        elif "active" in h and "clients" in h:
            idx["active"] = i
        elif "inactive" in h and "clients" in h:
            idx["inactive"] = i
    if "num_clients" not in idx:
        idx["num_clients"] = next((i for i, h in enumerate(headers) if h and "clients" in str(h).lower()), -1)
    if "active" not in idx:
        idx["active"] = next((i for i, h in enumerate(headers) if h and "active" in str(h).lower()), -1)
    if "inactive" not in idx:
        idx["inactive"] = next((i for i, h in enumerate(headers) if h and "inactive" in str(h).lower()), -1)
    return idx, headers


def build_section_map(rows_info):
    """
    Parse rows in order. Track current section based on headers.
    Returns: { "CS": [row_idx, ...], "LBF": [...], "SME": [...], "Agrifinance": [...] }
    """
    sections = {"CS": [], "LBF": [], "SME": [], "Agrifinance": []}
    current = None

    for row_idx, branch, is_leaf, agg in rows_info:
        if not branch:
            if current and current in sections:
                # Empty row might end section
                pass
            continue

        b = branch.lower()
        if "cluster 1" in b or "cluster 2" in b or "cluster 3" in b:
            current = "CS"
        elif "lbf cluster" in b:
            current = "LBF"
        elif "sme" in b and "zone" not in b and "call" not in b:
            if "smes" in b or "branch" in b or is_leaf:
                current = "SME"
        elif "agrifinance" in b or "maziwa" in b:
            current = "Agrifinance"
        elif "zanzibar" in b and "zone" not in b:
            current = "ZANZIBAR"  # not in our aggregate list, skip
        elif "call center" in b:
            current = None  # CS/LBF Call center - skip for product aggregates

        if is_leaf and current and current in sections:
            sections[current].append(row_idx)
    return sections


def main():
    load_env()

    metadata_path = BACKUP_DIR / "management_reports_metadata.json"
    if not metadata_path.exists():
        print("Run backup first: node scripts/management-reports-backup.mjs")
        sys.exit(1)

    with open(metadata_path) as f:
        reports = json.load(f)

    if not CLIENTS_FILE.exists():
        print(f"Clients file not found: {CLIENTS_FILE}")
        sys.exit(1)

    print("=" * 60)
    print("  Management Reports Correction (formatting preserved)")
    print("=" * 60)
    print(f"\nSource (backup): {BACKUP_DIR}")
    print(f"Output:          {UPLOADS_BASE}")
    print(f"Reports:         {len(reports)}\n")

    db_conn = None
    if psycopg2:
        try:
            db_conn = psycopg2.connect(
                host=os.getenv("DB_HOST", "localhost"),
                port=os.getenv("DB_PORT", "5432"),
                dbname=os.getenv("DB_NAME", "pcl_analysis"),
                user=os.getenv("DB_USER", "masubi"),
                password=os.getenv("DB_PASSWORD", "Masubi98%"),
            )
        except Exception as e:
            print(f"DB connection failed (skipping DB updates): {e}\n")

    updated = 0
    for r in reports:
        report_id = r["id"]
        file_path = r["file_path"]
        report_date = r.get("date")
        basename = Path(file_path).name

        backup_file = BACKUP_DIR / f"{report_id}_{basename}"
        if not backup_file.exists():
            print(f"  Skip (not in backup): {basename}")
            continue

        output_path = UPLOADS_BASE / file_path
        output_path.parent.mkdir(parents=True, exist_ok=True)

        cutoff = datetime.strptime(report_date, "%Y-%m-%d") if report_date else datetime(9999, 12, 31)
        by_branch = load_clients_by_date(cutoff)

        # openpyxl: load with data_only=False to preserve formatting
        wb = load_workbook(backup_file, data_only=False)
        country_sheet = next((s for s in wb.sheetnames if "country" in s.lower()), wb.sheetnames[0])
        ws = wb[country_sheet]

        cols, headers = find_columns(ws)
        if cols.get("branch", -1) < 0 or cols.get("num_clients", -1) < 0:
            print(f"  Skip (missing columns): {basename}")
            wb.close()
            continue

        branch_col = cols["branch"]
        num_col = cols.get("num_clients", -1)
        active_col = cols.get("active", -1)
        inactive_col = cols.get("inactive", -1)

        # 1-based column for openpyxl
        num_col1 = num_col + 1
        active_col1 = active_col + 1
        inactive_col1 = inactive_col + 1

        # Build row info: (row_idx, branch, is_leaf, agg_type)
        rows_info = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row = list(row) if row else []
            if len(row) <= branch_col:
                rows_info.append((row_idx, None, False, None))
                continue
            val = row[branch_col]
            if val is None or (isinstance(val, str) and not val.strip()):
                rows_info.append((row_idx, None, False, None))
                continue
            branch = str(val).strip()
            agg = is_aggregate_row(branch)
            leaf = is_leaf_branch(branch) if not agg else False
            rows_info.append((row_idx, branch, leaf, agg))

        values_by_row = {}
        section_map = build_section_map(rows_info)

        for row_idx, branch, is_leaf, agg in rows_info:
            if not branch:
                continue

            if is_leaf:
                counts = get_counts(branch, by_branch)
                if counts:
                    total = counts.get("active", 0) + counts.get("inactive", 0)
                    values_by_row[row_idx] = {
                        "num": total,
                        "active": counts.get("active", 0),
                        "inactive": counts.get("inactive", 0),
                    }
            elif agg and agg in section_map:
                c_num = c_active = c_inactive = 0
                for ri in section_map[agg]:
                    if ri in values_by_row:
                        v = values_by_row[ri]
                        c_num += v.get("num", 0)
                        c_active += v.get("active", 0)
                        c_inactive += v.get("inactive", 0)
                values_by_row[row_idx] = {"num": c_num, "active": c_active, "inactive": c_inactive}

        for row_idx, vals in values_by_row.items():
            if num_col1 > 0:
                ws.cell(row=row_idx, column=num_col1).value = vals.get("num", 0)
            if active_col1 > 0:
                ws.cell(row=row_idx, column=active_col1).value = vals.get("active", 0)
            if inactive_col1 > 0:
                ws.cell(row=row_idx, column=inactive_col1).value = vals.get("inactive", 0)

        wb.save(output_path)
        wb.close()
        updated += 1

        if db_conn:
            for row_idx, vals in values_by_row.items():
                branch_val = next((b for ri, b, _, _ in rows_info if ri == row_idx), None)
                if not branch_val:
                    continue
                metric_names = []
                if num_col >= 0 and num_col < len(headers) and headers[num_col]:
                    metric_names.append((str(headers[num_col]).strip(), "num"))
                if active_col >= 0 and active_col < len(headers) and headers[active_col]:
                    metric_names.append((str(headers[active_col]).strip(), "active"))
                if inactive_col >= 0 and inactive_col < len(headers) and headers[inactive_col]:
                    metric_names.append((str(headers[inactive_col]).strip(), "inactive"))
                for metric_name, key in metric_names:
                    if not metric_name or metric_name == "Branch":
                        continue
                    val = vals.get(key, 0)
                    cur = db_conn.cursor()
                    cur.execute(
                        """UPDATE report_data SET metric_value = %s
                        WHERE report_id = %s AND COALESCE(sheet_name, 'Country') = 'Country'
                        AND branch = %s AND metric_name = %s""",
                        (val, report_id, branch_val, metric_name),
                    )
                    cur.close()

        print(f"  OK [{updated}/{len(reports)}] {basename} ({report_date or 'no date'}) - {len(values_by_row)} rows")

    if db_conn:
        try:
            db_conn.commit()
            cur = db_conn.cursor()
            cur.execute("REFRESH MATERIALIZED VIEW dashboard_summary")
            cur.close()
            db_conn.commit()
        except Exception:
            try:
                cur = db_conn.cursor()
                cur.execute("REFRESH MATERIALIZED VIEW CONCURRENTLY dashboard_summary")
                cur.close()
                db_conn.commit()
            except Exception:
                pass
        db_conn.close()

    print("\n" + "=" * 60)
    print(f"Done: {updated} files corrected. Formatting preserved.")
    print("=" * 60)


if __name__ == "__main__":
    main()
