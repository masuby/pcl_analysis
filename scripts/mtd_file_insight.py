#!/usr/bin/env python3
"""
Read CS MTD and LBF MTD Excel files and print sheet names, headers, and sample rows.
Usage: python scripts/mtd_file_insight.py [path_to_cs] [path_to_lbf]
If no paths given, uses CS MTD as of 31st January 2026.xlsx and LBF MTD as of 31st January 2026.xlsx in project root.
"""

import sys
import os
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)


def find_header_row(ws, max_scan=12):
    """Find row that contains MONTH TARGET, VALUE, or NO. OF LOANS."""
    for row_idx in range(1, min(max_scan + 1, ws.max_row + 1)):
        for col_idx in range(1, min(25, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            s = str(val).upper().strip()
            if "MONTH TARGET" in s or "VALUE" == s or "NO. OF LOANS" in s or "NEW LOANS" in s:
                return row_idx
    return None


def inspect_file(file_path, label):
    path = Path(file_path)
    if not path.exists():
        print(f"\n[SKIP] {label}: File not found: {path}")
        return

    print(f"\n{'='*80}")
    print(f"{label}: {path.name}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\nSheets: {wb.sheetnames}")

    # First sheet = MTD data
    mtd_sheet = wb.worksheets[0]
    name = mtd_sheet.title
    print(f"\n--- First sheet (MTD): '{name}' ---")
    print(f"Max row: {mtd_sheet.max_row}, Max col: {mtd_sheet.max_column}")

    header_row = find_header_row(mtd_sheet)
    if header_row is None:
        print("Header row not found. First 8 rows (raw):")
        for r in range(1, min(9, mtd_sheet.max_row + 1)):
            row_vals = [mtd_sheet.cell(row=r, column=c).value for c in range(1, min(18, mtd_sheet.max_column + 1))]
            print(f"  Row {r}: {row_vals}")
        wb.close()
        return

    # Collect headers with exact strings and column index (1-based)
    headers = []
    for c in range(1, mtd_sheet.max_column + 1):
        val = mtd_sheet.cell(row=header_row, column=c).value
        if val is not None and str(val).strip():
            headers.append((c, str(val).strip()))
        else:
            headers.append((c, f"Col{c}"))

    print(f"\nHeader row index: {header_row}")
    print("Column index (1-based) -> Exact header string:")
    for col_idx, h in headers:
        print(f"  Col {col_idx}: '{h}'")

    # Sample data rows (skip subheaders if any)
    start_data = header_row + 1
    print(f"\nSample data rows (first 5 data rows, key columns):")
    key_headers = [h for _, h in headers if any(
        x in h.upper() for x in ["BRANCH", "TEAM", "NEW LOANS", "REFINANCE", "MONTH TARGET", "VALUE", "COMMENT", "ACTIVE"]
    )]
    for r in range(start_data, min(start_data + 5, mtd_sheet.max_row + 1)):
        row_dict = {}
        for col_idx, h in headers:
            val = mtd_sheet.cell(row=r, column=col_idx).value
            if val is not None and str(val).strip():
                row_dict[h] = val
        if row_dict:
            print(f"  Row {r}: {row_dict}")

    # Listing sheet if present
    for sheet in wb.worksheets[1:]:
        if "LIST" in sheet.title.upper():
            print(f"\n--- Listing sheet: '{sheet.title}' ---")
            for r in range(1, min(6, sheet.max_row + 1)):
                row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(12, sheet.max_column + 1))]
                print(f"  Row {r}: {row_vals}")
            break

    wb.close()


def main():
    base = Path(__file__).resolve().parent.parent

    cs_path = base / "CS MTD as of 31st January 2026.xlsx"
    lbf_path = base / "LBF MTD as of 31st January 2026.xlsx"

    if len(sys.argv) >= 2:
        cs_path = Path(sys.argv[1])
    if len(sys.argv) >= 3:
        lbf_path = Path(sys.argv[2])

    inspect_file(cs_path, "CS MTD")
    inspect_file(lbf_path, "LBF MTD")
    print("\nDone.")


if __name__ == "__main__":
    main()
