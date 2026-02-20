#!/usr/bin/env python3
"""
Export supervision To/CC email lists to an Excel file (.xlsx).
Run: python scripts/export_supervision_emails_to_xlsx.py
Output: lbf_supervision_emails.xlsx (in project root, or path set below)
Requires: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Replace with real addresses when sending to CALL-CENTER
TEST_EMAILS = ["test@platinumcredit.co.tz"]

# Supervision -> { "to": [...], "cc": [...] } — LBF email mapping
LBF_EMAIL_MAPPING = {
    "MIKOCHENI": {
        "to": [
            "fadhili.omary@platinumcredit.co.tz",
            "david.patrick.platinum@gmail.com",
            "digna.swai.platinum@gmail.com",
            "ester.nemes.platinum@gmail.com",
            "raphael.temu.platinum@gmail.com",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "NORTH-EAST": {
        "to": [
            "hadija.haji.platinum@gmail.com",
            "mrisho.katimle@platinumcredit.co.tz",
            "zulfa.jumanne@platinumcredit.co.tz",
            "chrizostom.thadeo@platinumcredit.co.tz",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "CITY-CENTRE": {
        "to": [
            "michael.manamba@platinumcredit.co.tz",
            "ruhindaedgar@gmail.com",
            "cletusgideon077@gmail.com",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "KIGAMBONI": {
        "to": [
            "veronica.mbasha.platinum@gmail.com",
            "david.kileo.platinum@gmail.com",
            "salim.ruwa@platinumcredit.co.tz",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "TEGETA": {
        "to": [
            "elvis.stephen.platinum@gmail.com",
        ],
        "cc": [
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "TABATA": {
        "to": [
            "allen.allan.platinum@gmail.com",
        ],
        "cc": [
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "CENTRAL": {
        "to": [
            "adam.tengeneza.platinum@gmail.com",
            "adam.tengeneza@platinumcredit.co.tz",
            "john.mdisa.platinum@gmail.com",
            "zulfa.jumanne@platinumcredit.co.tz",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "CITY-MALL": {
        "to": [
            "ashur.lusogo.platinum@gmail.com",
            "joseph.mambo.platinum@gmail.com",
            "sarah.galiatano.platinum@gmail.com",
            "nansi.luoga.platinum@gmail.com",
            "thobias.uchungu.platinum@gmail.com",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "MLIMANI": {
        "to": [
            "jamadin.mwahele.platinum@gmail.com",
            "john.kasanzu.platinum@gmail.com",
            "sarah.moshi@platinumcredit.co.tz",
            "abdallah.iddy@platinumcredit.co.tz",
            "abdallah.iddy.platinum@gmail.com",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "NYANZA": {
        "to": [
            "maganga.ally.platinum@gmail.com",
            "dickens.mathew.platinum@gmail.com",
            "ramadhani.luzila.platinum@gmail.com",
            "peter.amos.platinum@gmail.com",
            "mtatilo.joseph.platinum@gmail.com",
            "mashaka.simon.platinum@gmail.com",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "TANZARA": {
        "to": [
            "kenwood.madege.platinum@gmail.com",
            "kenwood.madege@platinumcredit.co.tz",
            "salehe.mgonja.platinum@gmail.com",
            "salehe.mgonja@platinumcredit.co.tz",
            "eutropia.pasian.platinum@gmail.com",
        ],
        "cc": [
            "irene.mmari@platinumcredit.co.tz",
            "daniel@platinumcredit.co.tz",
            "raphael@platinumcredit.co.tz",
        ],
    },
    "CALL-CENTER": {
        "to": TEST_EMAILS,
        "cc": [],
    },
}


def main():
    import os
    out_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "lbf_supervision_emails.xlsx")

    wb = Workbook()
    bold = Font(bold=True)
    separator_border = Border(bottom=Side(style="medium", color="2A5298"))

    # Sheet 1: To Emails
    ws_to = wb.active
    ws_to.title = "To Emails"
    ws_to.append(["Supervision", "Email"])
    for r in ws_to[1]:
        r.font = bold

    for sup, data in LBF_EMAIL_MAPPING.items():
        to_emails = [e.strip() for e in data["to"] if e and str(e).strip()]
        for email in to_emails:
            ws_to.append([sup, email])
        sep_row = ws_to.max_row + 1
        ws_to.append(["", ""])
        for col in range(1, 3):
            ws_to.cell(row=sep_row, column=col).border = separator_border

    ws_to.column_dimensions["A"].width = 22
    ws_to.column_dimensions["B"].width = 42

    # Sheet 2: CC Emails (one row per supervision–CC pair; empty CC list = no rows for that supervision)
    ws_cc = wb.create_sheet("CC Emails")
    ws_cc.append(["Supervision", "Email"])
    for r in ws_cc[1]:
        r.font = bold

    for sup, data in LBF_EMAIL_MAPPING.items():
        cc_emails = [e.strip() for e in data.get("cc", []) if e and str(e).strip()]
        for email in cc_emails:
            ws_cc.append([sup, email])
        if cc_emails:
            sep_row = ws_cc.max_row + 1
            ws_cc.append(["", ""])
            for col in range(1, 3):
                ws_cc.cell(row=sep_row, column=col).border = separator_border

    ws_cc.column_dimensions["A"].width = 22
    ws_cc.column_dimensions["B"].width = 42

    wb.save(out_path)
    print(f"Saved: {out_path}")
    print("  Sheets: 'To Emails' (Supervision, Email), 'CC Emails' (Supervision, Email). Line separator after each supervision.")


if __name__ == "__main__":
    main()
