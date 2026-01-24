#!/usr/bin/env python3
"""
Script to analyze MTD Excel file structure
Run: python scripts/analyze_mtd_excel.py <path_to_excel_file>
"""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl

def analyze_excel(file_path):
    print(f"\n{'='*60}")
    print(f"Analyzing: {file_path}")
    print(f"{'='*60}\n")
    
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print(f"Sheet names: {wb.sheetnames}")
    print(f"Total sheets: {len(wb.sheetnames)}\n")
    
    for idx, sheet_name in enumerate(wb.sheetnames):
        print(f"\n{'='*60}")
        print(f"Sheet {idx + 1}: '{sheet_name}'")
        print(f"{'='*60}")
        
        ws = wb[sheet_name]
        print(f"Dimensions: {ws.dimensions}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # Show first 10 rows
        print(f"\nFirst 10 rows:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    val_str = str(val)[:30]
                    row_values.append(f"{val_str}")
                else:
                    row_values.append("")
            print(f"  Row {row_idx}: {row_values}")
        
        # Check if this is the LISTING sheet
        if 'LISTING' in sheet_name.upper() or 'SALES' in sheet_name.upper():
            print(f"\n*** This looks like the Sales Listing sheet ***")
            
            # Find header row
            header_row = None
            for row_idx in range(1, min(15, ws.max_row + 1)):
                for col_idx in range(1, min(20, ws.max_column + 1)):
                    cell_val = ws.cell(row=row_idx, column=col_idx).value
                    if cell_val and ('SUPERVISION' in str(cell_val).upper() or 
                                     'SALES REP' in str(cell_val).upper() or
                                     'TEAM' in str(cell_val).upper()):
                        header_row = row_idx
                        break
                if header_row:
                    break
            
            if header_row:
                print(f"\nHeader row found at row {header_row}:")
                headers = []
                for col_idx in range(1, ws.max_column + 1):
                    val = ws.cell(row=header_row, column=col_idx).value
                    if val:
                        headers.append((col_idx, str(val)))
                
                print("  Column headers:")
                for col_idx, header in headers:
                    print(f"    Col {col_idx}: '{header}'")
                
                # Show a few data rows
                print(f"\nSample data rows (rows {header_row + 1} to {header_row + 5}):")
                for row_idx in range(header_row + 1, min(header_row + 6, ws.max_row + 1)):
                    row_data = {}
                    for col_idx, header in headers[:10]:  # First 10 columns
                        val = ws.cell(row=row_idx, column=col_idx).value
                        if val:
                            row_data[header] = str(val)[:25]
                    if row_data:
                        print(f"    Row {row_idx}: {row_data}")
    
    wb.close()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        # Try to find MTD files in uploads
        uploads_dir = "/home/masubi/Desktop/code/pcl_analysis/uploads"
        if os.path.exists(uploads_dir):
            print(f"Looking for MTD files in {uploads_dir}...")
            for root, dirs, files in os.walk(uploads_dir):
                for f in files:
                    if 'MTD' in f.upper() and f.endswith('.xlsx'):
                        file_path = os.path.join(root, f)
                        print(f"\nFound: {file_path}")
                        analyze_excel(file_path)
        else:
            print("Usage: python analyze_mtd_excel.py <path_to_excel_file>")
            print("Or place MTD files in uploads/ directory")
    else:
        analyze_excel(sys.argv[1])
