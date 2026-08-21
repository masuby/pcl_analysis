"""Build a downloadable Excel workbook of the acquired leads.

Same layout as the Google Sheet (see scraper/upload_to_sheet.py, which owns the
column definitions), so the file someone downloads from the dashboard and the
sheet the analysts work in cannot drift apart:

  LBF Leads    | SME Leads | Unique Leads | Summary

Every row keeps its Source Link, so a lead can always be traced to the advert.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import db

NAVY = "1F3864"
SCORE_FILL = {"Hot": "D8F3DC", "Warm": "FFF3C4", "Cold": "F1F3F5"}
SCORE_FONT = {"Hot": "0B5B29", "Warm": "92400E", "Cold": "4B5563"}
WIDE = {"Why", "What They Offer", "Source Link", "Business Name", "Offering"}


def _cols():
    """Column definitions live with the sheet publisher — import lazily to avoid
    a circular import at module load."""
    from scraper.upload_to_sheet import LBF_COLUMNS, SME_COLUMNS, UNIQUE_COLUMNS, _value
    return LBF_COLUMNS, SME_COLUMNS, UNIQUE_COLUMNS, _value


def _add_sheet(wb: Workbook, title: str, columns, leads: list, value_of) -> int:
    ws = wb.create_sheet(title)
    headers = [h for h, _ in columns]
    ws.append(headers)
    for lead in leads:
        ws.append([value_of(lead, f) for _, f in columns])

    head_fill = PatternFill("solid", fgColor=NAVY)
    for idx, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=idx)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = head_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = 46 if name in WIDE else 20

    # Colour the Score column so hot leads stand out in the downloaded file too.
    if "Score" in headers:
        s_idx = headers.index("Score") + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=s_idx)
            key = str(cell.value or "").strip()
            if key in SCORE_FILL:
                cell.fill = PatternFill("solid", fgColor=SCORE_FILL[key])
                cell.font = Font(bold=True, color=SCORE_FONT[key], size=10)

    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}"
    return max(0, ws.max_row - 1)


def _add_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary", 0)
    stats = db.stats()
    ws.append(["AI DIGITAL AGENT - ACQUIRED LEADS"])
    ws.append(["Generated", datetime.now().strftime("%d/%m/%Y %H:%M")])
    ws.append([])
    ws.append(["Listing links discovered", stats.get("links", 0)])
    ws.append(["Listings detail-scraped", stats.get("detailed", 0)])
    ws.append(["Leads structured by AI", stats.get("clean", 0)])
    ws.append(["Unique phone numbers (callable)", stats.get("unique", 0)])
    ws.append([])
    ws.append(["PRODUCT", "SOURCE", "LEADS", "WITH PHONE", "UNIQUE PHONES"])
    header_rows = {1, 9}
    for r in stats.get("by_product", []):
        ws.append([r["product"], r["source"], r["leads"], r["with_phone"], r["unique_phones"]])
    for row in header_rows:
        for c in range(1, 6):
            cell = ws.cell(row=row, column=c)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor=NAVY)
    for idx, w in enumerate([34, 44, 14, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_workbook(product: str = "") -> bytes:
    """Excel bytes for the requested product ('' = everything)."""
    lbf_cols, sme_cols, uniq_cols, value_of = _cols()
    product = (product or "").upper()

    wb = Workbook()
    wb.remove(wb.active)          # drop the default empty sheet

    if product in ("", "ALL", "LBF"):
        _add_sheet(wb, "LBF Leads", lbf_cols, db.all_clean_filtered(product="LBF"), value_of)
    if product in ("", "ALL", "SME"):
        _add_sheet(wb, "SME Leads", sme_cols, db.all_clean_filtered(product="SME"), value_of)
    if product in ("", "ALL"):
        _add_sheet(wb, "Unique Leads", uniq_cols, db.unique_clean(newest_first=False), value_of)
    _add_summary(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def filename(product: str = "") -> str:
    tag = (product or "all").upper()
    return f"digital_agent_leads_{tag}_{datetime.now():%Y-%m-%d}.xlsx"
