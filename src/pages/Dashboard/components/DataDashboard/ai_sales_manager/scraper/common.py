"""Shared helpers for the LBF scraping pipeline.

Pipeline (cartanzania.com, Phase 1):
  scrape_cartanzania.py  → data/lbf_ai_raw_data.xlsx   (raw_data, source_url)
  clean_with_ai.py       → data/lbf_ai_cleaned_data.xlsx (structured fields)
  upload_to_sheet.py     → appends cleaned rows to the Google Sheet
"""
from __future__ import annotations

import re
import time
from datetime import date
from pathlib import Path

import requests

# .../ai_sales_manager
PKG_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = PKG_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)

RAW_XLSX = DATA_DIR / "lbf_ai_raw_data.xlsx"
CLEANED_XLSX = DATA_DIR / "lbf_ai_cleaned_data.xlsx"

BASE = "https://www.cartanzania.com"

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}


def session() -> requests.Session:
    s = requests.Session()
    s.headers.update(_HEADERS)
    return s


def get(sess: requests.Session, url: str, *, tries: int = 3, pause: float = 1.0) -> str | None:
    """Polite GET with a small retry/back-off. Returns HTML text or None."""
    for attempt in range(1, tries + 1):
        try:
            r = sess.get(url, timeout=25)
            if r.status_code == 200 and r.text and r.text.strip() not in ("null", "[]"):
                return r.text
            if r.status_code in (429, 503):
                time.sleep(pause * attempt * 3)
                continue
            return None
        except requests.RequestException:
            time.sleep(pause * attempt)
    return None


# ── tiny xlsx helpers (openpyxl) ─────────────────────────────────────────────
def write_xlsx(path: Path, headers: list[str], rows: list[list], sheet_title: str = "Data") -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(["" if v is None else str(v) for v in row])
    # sensible column widths + wrap the long text column
    from openpyxl.styles import Alignment, Font

    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        wide = h.lower() in ("raw_data", "raw data", "seller comment", "reason")
        ws.column_dimensions[letter].width = 80 if wide else 22
        if wide:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=i).alignment = Alignment(wrap_text=True, vertical="top")
    path.parent.mkdir(exist_ok=True)
    wb.save(path)


def read_xlsx(path: Path) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(c) if c is not None else "" for c in next(it, [])]
    rows = [[("" if c is None else str(c)) for c in r] for r in it]
    return headers, rows


# ── dedup / dating helpers ───────────────────────────────────────────────────
def today_str() -> str:
    return date.today().isoformat()


def normalize_phone(raw: str) -> str:
    """Reduce a phone to comparable digits (last 9 — the TZ subscriber part),
    so +255 / 0 / spacing variants of the same number match."""
    digits = re.sub(r"\D", "", raw or "")
    return digits[-9:] if len(digits) >= 9 else digits


def load_existing(path: Path) -> tuple[list[str], list[list]]:
    """Return (headers, rows) for an accumulating store, or ([], []) if absent."""
    if not path.exists():
        return [], []
    try:
        return read_xlsx(path)
    except Exception:
        return [], []


def column_values(headers: list[str], rows: list[list], name: str) -> list[str]:
    if name not in headers:
        return []
    i = headers.index(name)
    return [r[i] if i < len(r) else "" for r in rows]


def dedup_by_phone(rows: list[list], phone_index: int) -> list[list]:
    """Keep one row per unique phone number (first occurrence). Rows without a
    usable phone are dropped, so the result is 'unique people we can call'."""
    seen: set[str] = set()
    out: list[list] = []
    for r in rows:
        phone = r[phone_index] if len(r) > phone_index else ""
        norm = normalize_phone(phone)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        out.append(r)
    return out
