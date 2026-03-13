"""
Temporary script to read Zone and cluster.xlsx: all column headers and sample data.
Zone column = Region (e.g. Nyasa region = Nyasa Zone in naming).

Run from repo root or from this folder:
  python src/pages/Dashboard/components/DepartmentalDashboard/components/KpiAnalysisReport/inspect_zone_cluster.py
  OR  cd .../KpiAnalysisReport && python inspect_zone_cluster.py
"""
from pathlib import Path
import openpyxl

# Prefer same folder as this script, then backend/scripts
SCRIPT_DIR = Path(__file__).resolve().parent
CANDIDATES = [
    SCRIPT_DIR / "Zone and cluster.xlsx",
    SCRIPT_DIR.parent.parent.parent.parent.parent / "backend" / "scripts" / "Zone and cluster.xlsx",
]
path = None
for p in CANDIDATES:
    if p.exists():
        path = p
        break
if not path:
    path = SCRIPT_DIR / "Zone and cluster.xlsx"
    print(f"File not found at {path}; trying backend...")
    path = Path(__file__).resolve().parents[5] / "backend" / "scripts" / "Zone and cluster.xlsx"

if not path.exists():
    print("Zone and cluster.xlsx not found in KpiAnalysisReport or backend/scripts.")
    exit(1)

print(f"Reading: {path}\n")
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
ws = wb.active
print(f"Active sheet: {ws.title}")
print(f"Max row: {ws.max_row}, max col: {ws.max_column}\n")

# Headers (row 1)
headers = []
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=c).value
    headers.append(v)
print("=== Column headers (row 1) ===")
for i, h in enumerate(headers, 1):
    print(f"  Col {i}: {repr(h)}")

print("\n=== Sample data (rows 2–20) ===")
for r in range(2, min(21, ws.max_row + 1)):
    row_vals = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        row_vals.append(str(v)[:30] if v is not None else "")
    print(f"  Row {r}: " + " | ".join(row_vals))

# Check for Zone / Region
zone_col = next((i for i, h in enumerate(headers) if h and "zone" in str(h).lower() and "cluster" not in str(h).lower()), None)
cluster_col = next((i for i, h in enumerate(headers) if h and "cluster" in str(h).lower()), None)
branch_col = next((i for i, h in enumerate(headers) if h and "branch" in str(h).lower()), None)
print(f"\n=== Detected: Zone col index (0-based)={zone_col}, Cluster={cluster_col}, Branch={branch_col} ===")
wb.close()
