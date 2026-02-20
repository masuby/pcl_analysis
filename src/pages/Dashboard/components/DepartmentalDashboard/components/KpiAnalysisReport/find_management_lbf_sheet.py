import os
import glob
import openpyxl
import re
import smtplib
import ssl
from datetime import datetime
from openpyxl.utils import get_column_letter
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from dotenv import load_dotenv

# Directory to search in
search_directory = r"C:\Users\Daniel\Desktop\Management"

# Colours for formatting
HEADER_FILL = openpyxl.styles.PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
FROZEN_COLUMNS_FILL = openpyxl.styles.PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Soft blue
BRANCH_SECTION_FILL = openpyxl.styles.PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold for branch headers
RED_FILL = openpyxl.styles.PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # 0%
ORANGE_FILL = openpyxl.styles.PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid")  # < 50%
GREEN_FILL = openpyxl.styles.PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")  # >= 50%

def get_percentage_fill(percentage):
    """Return fill colour for row based on % Achieved: 0%=red, <50%=orange, >=50%=green."""
    if percentage == 0:
        return RED_FILL
    elif percentage < 50:
        return ORANGE_FILL
    else:
        return GREEN_FILL

def format_branch_name(sheet_name):
    """
    Format branch name by adding spaces appropriately.
    Examples: 
    - LBFNJOMBE -> LBF NJOMBE
    - LBFOfficeDodoma -> LBF Office Dodoma
    - LBFIRINGABRANCH -> LBF IRINGA BRANCH
    - LBFCITYCENTER -> LBF CITY CENTER
    - LBFKIGAMBONIBRANCH -> LBF KIGAMBONI BRANCH
    """
    # Remove 'LBF' prefix
    if sheet_name.startswith('LBF'):
        rest = sheet_name[3:]
    else:
        rest = sheet_name
    
    # Known suffix words that should be separated
    known_suffixes = ['BRANCH', 'CENTER', 'OFFICE']
    
    # Check if the string ends with any known suffix
    suffix = None
    remaining = rest
    for known_suffix in known_suffixes:
        if rest.upper().endswith(known_suffix) and len(rest) > len(known_suffix):
            suffix = known_suffix
            remaining = rest[:-len(known_suffix)]
            break
    
    # Add space before capital letters (but not the first one)
    # This handles cases like "OfficeDodoma" -> "Office Dodoma" or "IRINGA" -> "IRINGA"
    formatted_remaining = re.sub(r'(?<!^)(?=[A-Z])', ' ', remaining)
    
    # Combine parts
    if suffix:
        formatted = f"{formatted_remaining} {suffix}"
    else:
        formatted = formatted_remaining
    
    return f"LBF {formatted}" if sheet_name.startswith('LBF') else formatted

def extract_team_leader_data(sheet, branch_name):
    """
    Extract data for each team leader in the sheet.
    Returns a list of dictionaries containing team leader information.
    """
    team_leaders = []
    current_team_leader_name = None
    current_team_leader_row = None
    
    # Column mappings based on actual structure
    # Row 1 contains headers, data starts from row 2
    # Column A = Team Leader/Sales Rep indicator
    # Column B = Branch Manager/Name
    # Column 3 = Target
    # Column 4 = New Business
    # Column 5 = Repeat Business
    # Column 9 = Disbursement this Month
    # Column 20 = Number of loans
    # Column 21 = Average loan size
    # Column 37 = Number of Clients
    # Column 38 = Active clients
    
    for row_idx in range(1, sheet.max_row + 1):
        col_a_value = sheet.cell(row_idx, 1).value
        col_b_value = sheet.cell(row_idx, 2).value
        
        if col_a_value and str(col_a_value).strip() == "Team Leader":
            # Save previous team leader data if exists
            if current_team_leader_name and current_team_leader_row:
                tl_data = extract_team_leader_metrics(
                    sheet, 
                    current_team_leader_row, 
                    row_idx - 1, 
                    current_team_leader_name,
                    branch_name
                )
                if tl_data:
                    team_leaders.append(tl_data)
            
            # Start new team leader
            current_team_leader_name = col_b_value
            current_team_leader_row = row_idx
    
    # Don't forget the last team leader
    if current_team_leader_name and current_team_leader_row:
        tl_data = extract_team_leader_metrics(
            sheet, 
            current_team_leader_row, 
            sheet.max_row, 
            current_team_leader_name,
            branch_name
        )
        if tl_data:
            team_leaders.append(tl_data)
    
    return team_leaders

def extract_team_leader_metrics(sheet, team_leader_row, end_row, team_leader_name, branch_name):
    """
    Extract metrics for a specific team leader section.
    """
    metrics = {
        "Branch": branch_name,
        "Team Leader": team_leader_name,
        "Target": 0,
        "New Business": 0,
        "Repeat Business": 0,
        "Disbursement this Month": 0,
        "Number of Loans": 0,
        "Average Loan Size": 0,
        "Number of client": 0,
        "Active Client": 0,
        "Active Agent": 0
    }
    
    # Extract metrics from the Team Leader row (same row as "Team Leader" in column A)
    # Column 3 = Target
    target_val = sheet.cell(team_leader_row, 3).value
    if target_val is not None:
        try:
            metrics["Target"] = float(target_val)
        except:
            metrics["Target"] = 0
    
    # Column 4 = New Business
    new_business_val = sheet.cell(team_leader_row, 4).value
    if new_business_val is not None:
        try:
            metrics["New Business"] = float(new_business_val)
        except:
            metrics["New Business"] = 0
    
    # Column 5 = Repeat Business
    repeat_business_val = sheet.cell(team_leader_row, 5).value
    if repeat_business_val is not None:
        try:
            metrics["Repeat Business"] = float(repeat_business_val)
        except:
            metrics["Repeat Business"] = 0
    
    # Column 9 = Disbursement this Month
    disbursement_val = sheet.cell(team_leader_row, 9).value
    if disbursement_val is not None:
        try:
            metrics["Disbursement this Month"] = float(disbursement_val)
        except:
            metrics["Disbursement this Month"] = 0
    
    # Column 20 = Number of loans
    num_loans_val = sheet.cell(team_leader_row, 20).value
    if num_loans_val is not None:
        try:
            metrics["Number of Loans"] = float(num_loans_val)
        except:
            metrics["Number of Loans"] = 0
    
    # Column 21 = Average loan size
    avg_loan_val = sheet.cell(team_leader_row, 21).value
    if avg_loan_val is not None:
        try:
            metrics["Average Loan Size"] = float(avg_loan_val)
        except:
            metrics["Average Loan Size"] = 0
    
    # Column 37 = Number of Clients
    num_client_val = sheet.cell(team_leader_row, 37).value
    if num_client_val is not None:
        try:
            metrics["Number of client"] = float(num_client_val)
        except:
            metrics["Number of client"] = 0
    
    # Column 38 = Active clients
    active_client_val = sheet.cell(team_leader_row, 38).value
    if active_client_val is not None:
        try:
            metrics["Active Client"] = float(active_client_val)
        except:
            metrics["Active Client"] = 0
    
    # Count Active Agents (sales reps with Number of Loans > 0)
    # Sales reps are rows after the Team Leader row until the next Team Leader
    active_agents = 0
    for row_idx in range(team_leader_row + 1, min(end_row + 1, sheet.max_row + 1)):
        col_a_value = sheet.cell(row_idx, 1).value
        
        # Check if this is a Sales Rep row
        if col_a_value and str(col_a_value).strip() == "Sales Rep":
            # Check if this sales rep has Number of Loans > 0 (Column 20)
            num_loans_val = sheet.cell(row_idx, 20).value
            if num_loans_val is not None:
                try:
                    num_loans = float(num_loans_val)
                    if num_loans > 0:
                        active_agents += 1
                except:
                    pass
        # Stop if we encounter another Team Leader
        elif col_a_value and str(col_a_value).strip() == "Team Leader":
            break
    
    metrics["Active Agent"] = active_agents
    
    # Calculate % Achieved
    if metrics["Target"] > 0:
        metrics["% Achieved"] = (metrics["Disbursement this Month"] / metrics["Target"]) * 100
    else:
        metrics["% Achieved"] = 0
    
    return metrics

def process_lbf_sheets():
    """
    Find Management Excel file, process all LBF sheets, and create summary Excel file.
    """
    print("="*80)
    print("Searching for Management Excel file")
    print("="*80)
    
    # Search for Excel files starting with "Management"
    pattern = os.path.join(search_directory, "Management*.xlsx")
    matching_files = glob.glob(pattern)
    
    # Also try .xlsm extension
    if not matching_files:
        pattern = os.path.join(search_directory, "Management*.xlsm")
        matching_files = glob.glob(pattern)
    
    if not matching_files:
        print(f"No Excel file starting with 'Management' found in {search_directory}")
        return None
    
    # Get the first matching file
    excel_file = matching_files[0]
    print(f"Found file: {excel_file}")
    print()
    
    try:
        # Open the workbook
        print("Opening workbook...")
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        
        # Find all LBF sheets
        lbf_sheets = [name for name in workbook.sheetnames if name.startswith('LBF')]
        print(f"Found {len(lbf_sheets)} LBF sheet(s):")
        for sheet_name in lbf_sheets:
            print(f"  - {sheet_name}")
        print()
        
        # Process each LBF sheet
        all_data = []
        
        for sheet_name in lbf_sheets:
            print(f"Processing sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # Format branch name
            branch_name = format_branch_name(sheet_name)
            
            # Extract team leader data
            team_leaders = extract_team_leader_data(sheet, branch_name)
            
            print(f"  Found {len(team_leaders)} team leader(s)")
            for tl in team_leaders:
                print(f"    - {tl['Team Leader']}: {tl.get('% Achieved', 0):.2f}%")
            
            all_data.extend(team_leaders)
        
        workbook.close()
        
        if not all_data:
            print("\n⚠️ No team leader data found!")
            return None
        
        # Sort by % Achieved (descending - top performing first)
        all_data.sort(key=lambda x: x.get("% Achieved", 0), reverse=True)
        
        # Create new Excel file
        output_file = os.path.join(search_directory, "LBF_Summary.xlsx")
        print()
        print(f"Creating summary file: {output_file}")
        
        wb_output = openpyxl.Workbook()
        ws_output = wb_output.active
        ws_output.title = "LBF Summary"
        
        # Write headers
        headers = ["Branch", "Branch Manager/TL", "Target", "New Business", "Repeat Business", 
                  "Disbursement this Month", "Number of Loans", "Average Loan Size", 
                  "Number of client", "Active Client", "Active Agent", "% Achieved"]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws_output.cell(1, col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
        
        # Write data
        for row_idx, data in enumerate(all_data, 2):
            pct = data.get("% Achieved", 0)
            row_fill = get_percentage_fill(pct)
            ws_output.cell(row_idx, 1).value = data.get("Branch", "")
            ws_output.cell(row_idx, 1).fill = FROZEN_COLUMNS_FILL
            ws_output.cell(row_idx, 2).value = data.get("Team Leader", "")
            ws_output.cell(row_idx, 2).fill = FROZEN_COLUMNS_FILL
            for col in range(3, 13):
                c = ws_output.cell(row_idx, col)
                if col == 3:
                    c.value = data.get("Target", 0)
                elif col == 4:
                    c.value = data.get("New Business", 0)
                elif col == 5:
                    c.value = data.get("Repeat Business", 0)
                elif col == 6:
                    c.value = data.get("Disbursement this Month", 0)
                elif col == 7:
                    c.value = data.get("Number of Loans", 0)
                elif col == 8:
                    c.value = data.get("Average Loan Size", 0)
                elif col == 9:
                    c.value = data.get("Number of client", 0)
                elif col == 10:
                    c.value = data.get("Active Client", 0)
                elif col == 11:
                    c.value = data.get("Active Agent", 0)
                elif col == 12:
                    percentage_value = pct / 100
                    c.value = percentage_value
                    c.number_format = '0.00%'
                c.fill = row_fill
        
        # Freeze panes: header row (1) and columns A & B; scrollable area starts at C2
        ws_output.freeze_panes = "C2"
        
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws_output[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_output.column_dimensions[column_letter].width = adjusted_width
        
        # Create second sheet grouped by Branch
        print("\nCreating Branch Grouped sheet...")
        ws_branch = wb_output.create_sheet("Grouped by Branch")
        
        # Write headers with dark blue background
        for col_idx, header in enumerate(headers, 1):
            cell = ws_branch.cell(1, col_idx)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
        
        # Group data by branch
        branches = {}
        for data in all_data:
            branch = data.get("Branch", "")
            if branch not in branches:
                branches[branch] = []
            branches[branch].append(data)
        
        # Sort branches alphabetically
        sorted_branches = sorted(branches.keys())
        
        # Write data grouped by branch
        current_row = 2
        for branch in sorted_branches:
            # Write branch name row (section header) - freeze columns A & B with gold, rest gold
            for col in range(1, 13):
                cell = ws_branch.cell(current_row, col)
                if col == 1:
                    cell.value = branch
                    cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = BRANCH_SECTION_FILL
            current_row += 1
            
            # Write team leaders for this branch - frozen cols A&B, data cols by percentage
            for tl_data in branches[branch]:
                pct = tl_data.get("% Achieved", 0)
                row_fill = get_percentage_fill(pct)
                ws_branch.cell(current_row, 1).value = tl_data.get("Branch", "")
                ws_branch.cell(current_row, 1).fill = FROZEN_COLUMNS_FILL
                ws_branch.cell(current_row, 2).value = tl_data.get("Team Leader", "")
                ws_branch.cell(current_row, 2).fill = FROZEN_COLUMNS_FILL
                ws_branch.cell(current_row, 3).value = tl_data.get("Target", 0)
                ws_branch.cell(current_row, 4).value = tl_data.get("New Business", 0)
                ws_branch.cell(current_row, 5).value = tl_data.get("Repeat Business", 0)
                ws_branch.cell(current_row, 6).value = tl_data.get("Disbursement this Month", 0)
                ws_branch.cell(current_row, 7).value = tl_data.get("Number of Loans", 0)
                ws_branch.cell(current_row, 8).value = tl_data.get("Average Loan Size", 0)
                ws_branch.cell(current_row, 9).value = tl_data.get("Number of client", 0)
                ws_branch.cell(current_row, 10).value = tl_data.get("Active Client", 0)
                ws_branch.cell(current_row, 11).value = tl_data.get("Active Agent", 0)
                percentage_value = pct / 100
                ws_branch.cell(current_row, 12).value = percentage_value
                ws_branch.cell(current_row, 12).number_format = '0.00%'
                for col in range(3, 13):
                    ws_branch.cell(current_row, col).fill = row_fill
                current_row += 1
        
        # Freeze panes: header row and columns A & B
        ws_branch.freeze_panes = "C2"
        
        # Auto-adjust column widths for branch sheet
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws_branch[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_branch.column_dimensions[column_letter].width = adjusted_width
        
        print(f"✅ Branch Grouped sheet created with {len(sorted_branches)} branches")
        
        wb_output.save(output_file)
        print(f"✅ Summary file created successfully with {len(all_data)} team leader entries")
        print(f"✅ File saved at: {output_file}")
        
        # Ask user if they want to send email
        print()
        print("="*80)
        user_input = input("Do you want to send the email? (yes/no): ").strip().lower()
        print("="*80)
        
        if user_input in ['yes', 'y']:
            print("Sending email with Excel file...")
            send_email_with_attachment(output_file, len(all_data))
        else:
            print("Email sending skipped by user.")
        
        return output_file
        
    except Exception as e:
        print(f"Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def create_email_html(num_entries):
    """Create HTML email content with professional styling"""
    date_str = datetime.now().strftime("%B %d, %Y")
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.6;
                color: #333333;
                max-width: 600px;
                margin: 0 auto;
                padding: 0;
                background-color: #ffffff;
            }}
            .header {{
                background-color: #0070C0;
                color: white;
                padding: 20px;
                text-align: center;
            }}
            .header h1 {{
                margin: 0;
                font-size: 24px;
                font-weight: 600;
            }}
            .content {{
                padding: 30px 20px;
            }}
            .content p {{
                margin: 15px 0;
                color: #333333;
                font-size: 15px;
            }}
            .content strong {{
                color: #0070C0;
            }}
            .footer {{
                margin-top: 30px;
                padding-top: 20px;
                border-top: 1px solid #e0e0e0;
                color: #888888;
                font-size: 12px;
            }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>LBF Summary Report</h1>
        </div>
        <div class="content">
            <p>Dear Augustine,</p>
            
            <p>Please find attached the LBF Summary Report containing performance data for all team leaders across all branches.</p>
            
            <p><strong>Report Details:</strong></p>
            <p>Total Team Leaders: <strong>{num_entries}</strong><br>
            Report Date: <strong>{date_str}</strong><br>
            Report Type: LBF Branch Performance Summary</p>
            
            <p>The report includes branch information, team leader performance metrics, targets, disbursements, and percentage achievements. All data is sorted by performance (highest to lowest).</p>
            
            <p>The Excel file contains two sheets: one sorted by performance and another grouped by branch for easier navigation.</p>
            
            <p>If you have any questions or need clarification on any metrics, please don't hesitate to reach out.</p>
            
            <p>Best regards,<br>
            Automated Reporting System</p>
        </div>
        <div class="footer">
            <p>This is an automated email. Please do not reply to this message.</p>
        </div>
    </body>
    </html>
    """
    return html_content

def send_email_with_attachment(excel_file_path, num_entries):
    """Send email with Excel file attachment to Augustine Mpollo and Denis Albert"""
    try:
        # Load environment variables
        load_dotenv()
        sender_email = os.getenv("EMAIL_USERNAME")
        sender_password = os.getenv("EMAIL_PASSWORD")
        receiver_emails = [
            "augustine@platinumcredit.co.tz",
            "denis.albert@platinumcredit.co.tz"
        ]
        
        if not sender_email or not sender_password:
            print("⚠️ Error: EMAIL_USERNAME or EMAIL_PASSWORD not found in .env file")
            return False
        
        print(f"Preparing email to: {', '.join(receiver_emails)}")
        
        # Create email content
        subject = f"LBF Summary Report - {datetime.now().strftime('%B %d, %Y')}"
        html_content = create_email_html(num_entries)
        
        # Create SMTP connection
        try:
            context = ssl.create_default_context()
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=30, context=context)
        except:
            server = smtplib.SMTP('smtp.gmail.com', 587, timeout=30)
            server.starttls()
        
        # Login
        server.login(sender_email, sender_password)
        print("✅ Connected to email server")
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = ', '.join(receiver_emails)
        msg['Subject'] = subject
        
        # Add date header
        msg['Date'] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
        
        # Add HTML content
        html_part = MIMEText(html_content, 'html')
        msg.attach(html_part)
        
        # Attach Excel file
        if os.path.exists(excel_file_path):
            with open(excel_file_path, 'rb') as f:
                excel_part = MIMEBase('application', 'octet-stream')
                excel_part.set_payload(f.read())
                encoders.encode_base64(excel_part)
                excel_part.add_header(
                    'Content-Disposition',
                    f'attachment; filename="{os.path.basename(excel_file_path)}"'
                )
                msg.attach(excel_part)
            print(f"✅ Excel file attached: {os.path.basename(excel_file_path)}")
        else:
            print(f"⚠️ Warning: Excel file not found at {excel_file_path}")
        
        # Send email
        server.send_message(msg)
        server.quit()
        
        print(f"✅ Email sent successfully to: {', '.join(receiver_emails)}")
        return True
        
    except Exception as e:
        print(f"❌ Error sending email: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    process_lbf_sheets()
