"""
Read the Country sheet from a Management Report Excel and print Branch column values.
Usage: python read_country_sheet.py [path_to_report.xlsx]
Default: Management Report2026-02.xlsx in this directory.
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

def main():
    script_dir = Path(__file__).resolve().parent
    default_path = script_dir / "Management Report2026-02.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Country" not in wb.sheetnames:
        print("Sheets:", wb.sheetnames)
        print("No 'Country' sheet.")
        wb.close()
        sys.exit(1)

    ws = wb["Country"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("Country sheet is empty.")
        return

    # Find Branch column (case-insensitive)
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    branch_col = None
    for i, h in enumerate(header):
        if h and "branch" in h.lower():
            branch_col = i
            break
    if branch_col is None:
        print("Header row:", header)
        print("No 'Branch' column found.")
        return

    # Also find Target and Disbursement for context
    target_col = None
    disb_col = None
    for i, h in enumerate(header):
        if h and "target" in h.lower() and "disbursement" not in h.lower():
            target_col = i
        if h and "disbursement" in h.lower():
            disb_col = i

    print("Country sheet - Branch column index:", branch_col)
    print("Header:", header[:max(branch_col + 1, (target_col or 0) + 1, (disb_col or 0) + 1)])
    print("-" * 60)
    print(f"{'Row':<5} {'Branch':<35} {'Target':<18} {'Disbursement':<18}")
    print("-" * 60)

    for r, row in enumerate(rows[1:], start=2):
        branch_val = row[branch_col] if branch_col < len(row) else ""
        branch_str = str(branch_val).strip() if branch_val is not None else ""
        target_val = row[target_col] if target_col is not None and target_col < len(row) else ""
        disb_val = row[disb_col] if disb_col is not None and disb_col < len(row) else ""
        if branch_str or target_val or disb_val:
            print(f"{r:<5} {branch_str:<35} {str(target_val):<18} {str(disb_val):<18}")

    # List unique branch names for cluster detection
    branches = []
    for row in rows[1:]:
        if branch_col < len(row) and row[branch_col] is not None:
            b = str(row[branch_col]).strip()
            if b and b not in branches:
                branches.append(b)
    print("-" * 60)
    print("Unique Branch values (in order):")
    for b in branches:
        print(" ", repr(b))

if __name__ == "__main__":
    main()
