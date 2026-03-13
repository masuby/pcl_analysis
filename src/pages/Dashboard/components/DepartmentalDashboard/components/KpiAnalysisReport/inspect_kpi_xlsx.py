"""
Temporary script to inspect CS_KPI_CLUSTER_TARGET_ROW_FILE.xlsx structure.
Run: python inspect_kpi_xlsx.py
"""
import openpyxl
from pathlib import Path

path = Path(__file__).parent / "CS_KPI_CLUSTER_TARGET_ROW_FILE.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

print("=== Sheet names ===")
print(wb.sheetnames)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== Sheet: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    # Print first 80 rows to see structure (tables, headers, etc.)
    for r in range(1, min(81, ws.max_row + 1)):
        row_vals = []
        for c in range(1, min(ws.max_column + 1, 25)):
            v = ws.cell(row=r, column=c).value
            row_vals.append(str(v)[:20] if v is not None else "")
        line = " | ".join(row_vals)
        if line.strip():
            print(f"  {r:3}: {line}")

wb.close()
