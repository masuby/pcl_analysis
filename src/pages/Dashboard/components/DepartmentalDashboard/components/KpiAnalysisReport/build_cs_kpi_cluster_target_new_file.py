"""
Build CS_KPI_CLUSTER_TARGET_NEW_FILE_2026.xlsx from CS_KPI_CLUSTER_TARGET_ROW_FILE.xlsx.

Each output sheet (Cluster 1, Cluster 2, Cluster 3, Zanzibar) has 4 columns:
  Month, New Business, Repeat Business, Total Target

Values are summed from three tables in the source sheet:
  1. Field Sales Team Leader: NEW BUSINESS, (REPEAT/REFINANCE + Reactivation), TOTAL SALES TARGET
  2. Branch Loan Officer: New Business, Repeat Business, Total
  3. Independent Team Leader: New Business, Refinance(Top up), Total

Repeat Business in output = (Field: Repeat/Refinance + Reactivation) + Branch: Repeat Business + Independent: Refinance(Top up)

Run: python build_cs_kpi_cluster_target_new_file.py
"""
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

SOURCE_FILE = "CS_KPI_CLUSTER_TARGET_ROW_FILE.xlsx"
OUTPUT_FILE = "CS_KPI_CLUSTER_TARGET_NEW_FILE_2026.xlsx"

# Sheets to process (exclude 'KPI'). Output: KPI + these 4 = 5 sheets.
KPI_SHEET_NAME = "KPI"
DATA_SHEET_NAMES = ["Cluster 1", "Cluster 2", "Cluster 3", "Zanzibar"]

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")
TOTAL_ROW_FILL = PatternFill(fill_type="solid", fgColor="D9E1F2")
TOTAL_ROW_FONT = Font(bold=True)


def _num(cell_value):
    """Return numeric value; 0 if empty or non-numeric."""
    if cell_value is None or cell_value == "":
        return 0
    try:
        return float(cell_value)
    except (TypeError, ValueError):
        return 0


def _month_str(cell_value):
    """Return YYYY-MM or short label from date or string."""
    if cell_value is None:
        return ""
    if hasattr(cell_value, "strftime"):
        return cell_value.strftime("%Y-%m")
    s = str(cell_value).strip()
    if not s:
        return ""
    # "2026-01-01 00:00:00" -> "2026-01"
    if "-" in s and len(s) >= 7:
        return s[:7]
    return s


def find_table_ranges(ws):
    """
    Find (start_row, header_row, data_end_row) for each of the three tables.
    Tables are identified by a row with 'MONTH' in col A and title in previous row (FIELD/BRANCH/INDEPENDENT).
    """
    tables = []  # list of (name, header_row_1based, last_data_row_1based)
    i = 1
    while i <= ws.max_row:
        a_val = ws.cell(row=i, column=1).value
        b_val = ws.cell(row=i, column=2).value
        a_str = (a_val or "").strip().upper() if a_val is not None else ""
        b_str = (b_val or "").strip().upper() if b_val is not None else ""
        if a_str == "MONTH" or (b_str == "MONTH" and a_str == ""):
            # This is a header row; table title is typically row above
            prev_b = (ws.cell(row=i - 1, column=2).value or "").strip().upper()
            if "FIELD" in prev_b and "TEAM" in prev_b:
                name = "field"
            elif "BRANCH" in prev_b and "LOAN" in prev_b:
                name = "branch"
            elif "INDEPENDENT" in prev_b and "TEAM" in prev_b:
                name = "independent"
            else:
                # Fallback: first table has NEW BUSINESS in col2, second has Repeat Business, third has Refinance
                col2_headers = []
                for c in range(2, 6):
                    v = ws.cell(row=i, column=c).value
                    if v is not None:
                        col2_headers.append(str(v).strip().upper())
                if "NEW BUSINESS" in str(col2_headers) and "REPEAT" not in str(col2_headers)[:30]:
                    name = "field"
                elif "REFINANCE" in str(col2_headers) or "TOP UP" in str(col2_headers):
                    name = "independent"
                else:
                    name = "branch"
            # Data rows: next 12 rows (months) or until blank in col A
            end_row = i
            for j in range(i + 1, min(i + 14, ws.max_row + 1)):
                month_cell = ws.cell(row=j, column=1).value
                if month_cell is None and ws.cell(row=j, column=2).value is None:
                    break
                end_row = j
            tables.append((name, i, end_row))
            i = end_row + 1
            continue
        i += 1
    return tables


def extract_table_data(ws, header_row, last_data_row, table_name):
    """
    Return list of dicts: { month, new_business, repeat_business, total }.
    Column layout from inspect:
    - Field: 1=Month, 2=New Business, 3=Repeat/Refinance, 4=Reactivation, 5=Total
    - Branch: 1=Month, 2=New Business, 3=Repeat Business, 4=Total
    - Independent: 1=Month, 2=New Business, 3=Refinance(Top up), 4=Total
    """
    rows = []
    for r in range(header_row + 1, last_data_row + 1):
        month_val = ws.cell(row=r, column=1).value
        month = _month_str(month_val)
        if not month:
            continue
        if table_name == "field":
            new_b = _num(ws.cell(row=r, column=2).value)
            repeat_refinance = _num(ws.cell(row=r, column=3).value)
            reactivation = _num(ws.cell(row=r, column=4).value)
            total = _num(ws.cell(row=r, column=5).value)
            repeat_business = repeat_refinance + reactivation
        elif table_name == "branch":
            new_b = _num(ws.cell(row=r, column=2).value)
            repeat_business = _num(ws.cell(row=r, column=3).value)
            total = _num(ws.cell(row=r, column=4).value)
        else:  # independent
            new_b = _num(ws.cell(row=r, column=2).value)
            repeat_business = _num(ws.cell(row=r, column=3).value)  # Refinance (Top up)
            total = _num(ws.cell(row=r, column=4).value)
        rows.append({"month": month, "new_business": new_b, "repeat_business": repeat_business, "total": total})
    return rows


def merge_by_month(table_rows_list):
    """Merge multiple table row lists by month; sum new_business, repeat_business, total."""
    by_month = {}
    for rows in table_rows_list:
        for row in rows:
            m = row["month"]
            if m not in by_month:
                by_month[m] = {"month": m, "new_business": 0, "repeat_business": 0, "total": 0}
            by_month[m]["new_business"] += row["new_business"]
            by_month[m]["repeat_business"] += row["repeat_business"]
            by_month[m]["total"] += row["total"]
    # Sort by month (YYYY-MM)
    months_sorted = sorted(by_month.keys())
    return [by_month[m] for m in months_sorted]


def process_sheet(ws):
    """Process one cluster sheet; return list of { month, new_business, repeat_business, total }."""
    tables = find_table_ranges(ws)
    field_data = []
    branch_data = []
    independent_data = []
    for name, hrow, end_row in tables:
        data = extract_table_data(ws, hrow, end_row, name)
        if name == "field":
            field_data = data
        elif name == "branch":
            branch_data = data
        elif name == "independent":
            independent_data = data
    return merge_by_month([field_data, branch_data, independent_data])


def main():
    base = Path(__file__).parent
    src_path = base / SOURCE_FILE
    out_path = base / OUTPUT_FILE

    if not src_path.exists():
        print(f"Source file not found: {src_path}")
        return

    wb_in = openpyxl.load_workbook(src_path, read_only=True, data_only=True)

    # Build output workbook
    wb_out = openpyxl.Workbook()
    # Remove default sheet if we create all from scratch
    if "Sheet" in wb_out.sheetnames:
        del wb_out["Sheet"]

    # Add KPI sheet first (yellow header row, total row styled, weights as percentage)
    if KPI_SHEET_NAME in wb_in.sheetnames:
        ws_kpi_in = wb_in[KPI_SHEET_NAME]
        ws_kpi_out = wb_out.create_sheet(KPI_SHEET_NAME, 0)
        for r in range(1, ws_kpi_in.max_row + 1):
            for c in range(1, ws_kpi_in.max_column + 1):
                val = ws_kpi_in.cell(row=r, column=c).value
                cell = ws_kpi_out.cell(row=r, column=c, value=val)
                if r == 1:
                    cell.fill = YELLOW_FILL
                if r == ws_kpi_in.max_row:
                    cell.fill = TOTAL_ROW_FILL
                    cell.font = TOTAL_ROW_FONT
                if c == 2 and val is not None:
                    cell.number_format = '0%'
        ws_kpi_out.column_dimensions["A"].width = 40
        ws_kpi_out.column_dimensions["B"].width = 12

    all_sheet_data = {}
    for sheet_name in DATA_SHEET_NAMES:
        if sheet_name not in wb_in.sheetnames:
            print(f"  Skip (not in source): {sheet_name}")
            continue
        ws_in = wb_in[sheet_name]
        rows = process_sheet(ws_in)
        all_sheet_data[sheet_name] = rows

        ws_out = wb_out.create_sheet(sheet_name)
        ws_out.append(["Month", "New Business", "Repeat Business", "Total Target"])
        for r in rows:
            ws_out.append([r["month"], r["new_business"], r["repeat_business"], r["total"]])

        # Accounting format for numeric columns (B, C, D) - 0 decimal places
        ACCOUNTING_FORMAT = "#,##0"
        for row_idx in range(2, ws_out.max_row + 1):
            for col_idx in range(2, 5):
                ws_out.cell(row=row_idx, column=col_idx).number_format = ACCOUNTING_FORMAT

        # Column widths
        ws_out.column_dimensions["A"].width = 12
        ws_out.column_dimensions["B"].width = 16
        ws_out.column_dimensions["C"].width = 18
        ws_out.column_dimensions["D"].width = 16

    wb_in.close()
    wb_out.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {wb_out.sheetnames}")


if __name__ == "__main__":
    main()
