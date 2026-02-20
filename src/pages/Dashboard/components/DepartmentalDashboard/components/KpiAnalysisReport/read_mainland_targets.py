"""
Read the MAINLAND sheet from CS KPI TARGET.xlsx to inspect MONTH column and target columns.
Usage: python read_mainland_targets.py [path_to_xlsx]
Default: CS KPI TARGET.xlsx in this directory.
"""
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

def excel_serial_to_date(serial):
    """Excel serial number to (year, month). Excel epoch 1899-12-30."""
    if serial is None or (isinstance(serial, float) and (serial != serial or serial < 0)):
        return None, None
    try:
        # Excel serial: days since 1899-12-30 (so 1 = 31 Dec 1899)
        from datetime import timedelta
        epoch = datetime(1899, 12, 30)
        d = epoch + timedelta(days=float(serial))
        return d.year, d.month
    except Exception:
        return None, None

def main():
    script_dir = Path(__file__).resolve().parent
    default_path = script_dir / "CS KPI TARGET.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "MAINLAND" not in wb.sheetnames:
        print("Sheets:", wb.sheetnames)
        wb.close()
        sys.exit(1)

    ws = wb["MAINLAND"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("MAINLAND sheet is empty.")
        return

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    print("Headers:", headers)
    month_col = next((i for i, h in enumerate(headers) if "MONTH" in h.upper()), None)
    if month_col is None:
        print("No MONTH column found.")
        return

    # Find target columns
    new_col = next((i for i, h in enumerate(headers) if "NEW BUSINESS" in h.upper()), None)
    total_col = next((i for i, h in enumerate(headers) if "TOTAL SALES" in h.upper()), None)

    print("\nRow | Raw MONTH value (type)          | YYYY-MM (from date) | New Biz  | Total")
    print("-" * 85)
    for r, row in enumerate(rows[1:], start=2):
        val = row[month_col] if month_col < len(row) else None
        raw_type = type(val).__name__
        if val is None:
            y, m = None, None
        elif isinstance(val, (int, float)):
            y, m = excel_serial_to_date(val)
        elif isinstance(val, datetime):
            y, m = val.year, val.month
        else:
            s = str(val).strip()
            y, m = None, None
            if s:
                try:
                    dt = datetime.strptime(s[:10], "%Y-%m-%d")
                    y, m = dt.year, dt.month
                except Exception:
                    try:
                        dt = datetime.strptime(s[:7], "%Y-%m")
                        y, m = dt.year, dt.month
                    except Exception:
                        pass
        key = f"{y}-{m:02d}" if y and m else "—"
        new_val = row[new_col] if new_col is not None and new_col < len(row) else ""
        total_val = row[total_col] if total_col is not None and total_col < len(row) else ""
        raw_str = str(val)[:28].ljust(28) if val is not None else "—"
        print(f"{r:3} | {raw_str} ({raw_type:6}) | {key:19} | {str(new_val):8} | {str(total_val)}")

if __name__ == "__main__":
    main()
