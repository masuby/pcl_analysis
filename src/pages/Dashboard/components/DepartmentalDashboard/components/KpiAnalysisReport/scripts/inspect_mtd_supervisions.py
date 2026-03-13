#!/usr/bin/env python3
"""
Inspect MTD CS Excel: list supervisions (first column) and key columns (MONTH TARGET, VALUE).
Run with path to an MTD CS file.

Usage: python scripts/inspect_mtd_supervisions.py [path/to/MTD_CS.xlsx]
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

def main():
    script_dir = Path(__file__).resolve().parent
    parent = script_dir.parent
    default_path = parent / "MTD_CS_sample.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    if not path.exists():
        print(f"File not found: {path}")
        print("Usage: python inspect_mtd_supervisions.py [path/to/MTD_CS.xlsx]")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    print("Sheet:", sheet.title)
    print("=== First 40 rows (all columns) ===")
    for r in range(1, min(41, sheet.max_row + 1)):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, min(sheet.max_column + 1, 20))]
        print(r, "|", row_vals)
    print("\nSupervisions = branch/supervision names in first column. Match with Zone and cluster.xlsx branches.")
    wb.close()

if __name__ == "__main__":
    main()
