#!/usr/bin/env python3
"""
Inspect Management report Excel: Country sheet structure (Branch, Target, Disbursements This Month).
Run with path to a Management report file, or place a sample in this folder as Management_Report_sample.xlsx.

Usage: python scripts/inspect_management_country.py [path/to/Management_Report.xlsx]
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

def main():
    script_dir = Path(__file__).resolve().parent
    parent = script_dir.parent
    default_path = parent / "Management_Report_sample.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    if not path.exists():
        print(f"File not found: {path}")
        print("Usage: python inspect_management_country.py [path/to/Management_Report.xlsx]")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Country" not in wb.sheetnames:
        print("No 'Country' sheet. Sheets:", wb.sheetnames)
        wb.close()
        return
    ws = wb["Country"]
    print("=== Country sheet (first 25 rows) ===")
    for r in range(1, min(26, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column + 1, 15))]
        print(r, "|", row_vals)
    print("\nLook for: Branch (Cluster 1, Cluster 2, Cluster 3, ZANZIBAR), Target, Disbursements This Month")
    wb.close()

if __name__ == "__main__":
    main()
