#!/usr/bin/env python3
"""
Inspect CS CRM Excel for Cluster KPIs 7 and 8.
- agent_activity: Product, Zone, Status, Target_Met (AT_LOCATION), Region
- Lead_Report: PRODUCT, Zone, Consent_Status (ACCEPTED)

Usage: python scripts/inspect_crm_cs.py [path/to/CS_CRM.xlsx]
"""
import sys
from pathlib import Path
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

def main():
    script_dir = Path(__file__).resolve().parent
    parent = script_dir.parent
    default_path = parent / "CS_CRM_11_03_2026.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    if not path.exists():
        print(f"File not found: {path}")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # agent_activity
    if "agent_activity" in wb.sheetnames:
        ws = wb["agent_activity"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        print("=== agent_activity ===")
        print("Headers:", headers)
        idx = {str(h).strip(): i for i, h in enumerate(headers) if h}
        product_idx = next((idx[k] for k in idx if "product" in k.lower()), None)
        zone_idx = next((idx[k] for k in idx if k.lower() == "zone"), None)
        status_idx = next((idx[k] for k in idx if k.lower() == "status"), None)
        target_met_idx = next((idx[k] for k in idx if "target_met" in k.lower() or "target met" in k.lower()), None)
        region_idx = next((idx[k] for k in idx if k.lower() == "region"), None)
        print("Column indices: Product=%s Zone=%s Status=%s Target_Met=%s Region=%s" % (product_idx, zone_idx, status_idx, target_met_idx, region_idx))

        by_zone = defaultdict(lambda: {"completed": 0, "at_location": 0, "total": 0})
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if len(row) <= max(i for i in [product_idx, zone_idx, status_idx, target_met_idx] if i is not None):
                continue
            product = str(row[product_idx] or "").strip().upper() if product_idx is not None else ""
            if product != "CS":
                continue
            zone = str(row[zone_idx] or "").strip() if zone_idx is not None else ""
            status = str(row[status_idx] or "").strip().upper() if status_idx is not None else ""
            target_met = str(row[target_met_idx] or "").strip().upper() if target_met_idx is not None else ""
            by_zone[zone]["total"] += 1
            if status == "COMPLETED":
                by_zone[zone]["completed"] += 1
                if target_met == "AT_LOCATION":
                    by_zone[zone]["at_location"] += 1
        print("By Zone (CS only):", dict(by_zone))
        print()

    # Lead_Report
    if "Lead_Report" in wb.sheetnames:
        ws = wb["Lead_Report"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        print("=== Lead_Report ===")
        print("Headers:", headers)
        idx = {str(h).strip(): i for i, h in enumerate(headers) if h}
        product_idx = next((idx[k] for k in idx if "product" in k.lower()), None)
        zone_idx = next((idx[k] for k in idx if k.lower() == "zone"), None)
        consent_idx = next((idx[k] for k in idx if "consent" in k.lower()), None)
        print("Column indices: PRODUCT=%s Zone=%s Consent_Status=%s" % (product_idx, zone_idx, consent_idx))

        by_zone = defaultdict(lambda: {"accepted": 0, "total": 0})
        for r in range(2, min(ws.max_row + 1, 500)):
            row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if product_idx is not None and len(row) <= product_idx:
                continue
            product = str(row[product_idx] or "").strip().upper() if product_idx is not None else ""
            if product != "CS":
                continue
            zone = str(row[zone_idx] or "").strip() if zone_idx is not None else ""
            consent = str(row[consent_idx] or "").strip().upper() if consent_idx is not None else ""
            by_zone[zone]["total"] += 1
            if consent == "ACCEPTED":
                by_zone[zone]["accepted"] += 1
        print("By Zone (CS only, first 500 rows):", dict(by_zone))
        print()

    wb.close()
    print("Use: Zone = region; filter Product=CS. agent_activity: Status=COMPLETED, Target_Met=AT_LOCATION. Lead_Report: Consent_Status=ACCEPTED.")

if __name__ == "__main__":
    main()
