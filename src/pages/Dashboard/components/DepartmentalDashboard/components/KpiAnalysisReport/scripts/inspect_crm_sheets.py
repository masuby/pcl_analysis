#!/usr/bin/env python3
"""
Inspect CRM CS Excel: list sheet names and columns for agent_activity and Lead_Report.
Used for Cluster KPI 7 (95% on location: Status=COMPLETED, Target_Met=AT_LOCATION) and KPI 8 (80% consent: Lead_Report Consent_Status=ACCEPTED).

Usage: python scripts/inspect_crm_sheets.py [path/to/CRM_CS.xlsx]
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)

def main():
    script_dir = Path(__file__).resolve().parent
    parent = script_dir.parent
    default_path = parent / "CRM_CS_sample.xlsx"
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_path
    if not path.exists():
        print(f"File not found: {path}")
        print("Usage: python inspect_crm_sheets.py [path/to/CRM_CS.xlsx]")
        return
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    print("Sheets:", wb.sheetnames)
    for name in ["agent_activity", "Agent_Activity", "Lead_Report", "Lead report"]:
        if name in wb.sheetnames:
            ws = wb[name]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            print(f"\n=== {name} (headers) ===")
            print(headers)
            print(f"Rows: {ws.max_row}")
            if ws.max_row > 1:
                print("Sample row 2:", [ws.cell(row=2, column=c).value for c in range(1, min(len(headers) + 1, 25))])
    wb.close()
    print("\nExpect: Product (CS), Zone, Status (COMPLETED), Target_Met (AT_LOCATION) in agent_activity; Lead_Report: Product, Zone, Consent_Status (ACCEPTED).")

if __name__ == "__main__":
    main()
