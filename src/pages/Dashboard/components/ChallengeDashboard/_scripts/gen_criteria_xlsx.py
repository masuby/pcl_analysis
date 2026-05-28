"""
Regenerate the four challenge-report criteria xlsx files so that each
filename + content match its PDF memo and folder.

Run from the repo root with: python <this script>.

Produces:
  TeamBuildingReport/TeamBuilding_Criteria_2026.xlsx
  LocalTripReport/LocalTrip_Criteria_2026.xlsx
  EATripReport/EATrip_Criteria_2026.xlsx
  EATeamBuildingReport/EATeamBuilding_Criteria_2026.xlsx

Also removes the stale `LocalTrip_Criteria_2026.xlsx` copies that were left
behind in the EA* + LocalTrip folders when the modules were scaffolded.
"""
from __future__ import annotations
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BASE = os.path.join(
    os.path.dirname(__file__), os.pardir,
)
BASE = os.path.abspath(BASE)

# ── styling helpers ──────────────────────────────────────────────────────────
NAVY    = "1F3864"
LIGHT   = "D6E4F0"
GOLD    = "C9A227"
RED     = "C0392B"
GREEN   = "1E8449"
GREY    = "7F8C8D"
WHITE   = "FFFFFF"
BAND    = "F4F7FB"

THIN = Side(style="thin", color="BFC9D9")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def title_font():  return Font(name="Calibri", size=14, bold=True, color=WHITE)
def hdr_font():    return Font(name="Calibri", size=11, bold=True, color=WHITE)
def cell_font(bold=False, color="111111", italic=False):
    return Font(name="Calibri", size=10, bold=bold, color=color, italic=italic)

def title_fill():  return PatternFill("solid", fgColor=NAVY)
def hdr_fill():    return PatternFill("solid", fgColor=NAVY)
def band_fill(i): return PatternFill("solid", fgColor=BAND if i % 2 else WHITE)
def accent_fill(c): return PatternFill("solid", fgColor=c)

def write_title(ws, text, span):
    ws.append([text] + [""] * (span - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1)
    cell.font = title_font()
    cell.fill = title_fill()
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28

def write_subtitle(ws, text, span):
    ws.append([text] + [""] * (span - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1)
    cell.font = cell_font(bold=True, color=NAVY, italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill("solid", fgColor=LIGHT)
    ws.row_dimensions[row].height = 20

def write_header(ws, headers):
    ws.append(headers)
    row = ws.max_row
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 26

def write_section(ws, label, span, color=GOLD):
    ws.append([label] + [""] * (span - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

def write_rows(ws, rows, start_band=0):
    for i, row in enumerate(rows):
        ws.append(row)
        rnum = ws.max_row
        for col in range(1, len(row) + 1):
            cell = ws.cell(row=rnum, column=col)
            cell.font = cell_font()
            cell.border = BOX
            cell.fill = band_fill(i + start_band)
            cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[rnum].height = 22

def _style_cell(cell, band_index, vbold=False, vcolor="111111"):
    cell.font = cell_font(bold=vbold, color=vcolor)
    cell.border = BOX
    cell.fill = band_fill(band_index)
    cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)

def write_grouped_rows(ws, groups, ncols, *, product_col=1, criteria_col=None, role_cols=None, start_band=0):
    """
    Render a grouped table where Product is merged down a group, and each
    block's `criteria` cell is merged across all rows in that block.

    groups format:
      [
        {
          'product': 'LBF',
          'blocks':  [
            {'role': 'Independent Team Leader', 'criteria': '…'},   # single-row block
            {'roles': ['A', 'B', 'C'], 'criteria': '…'},            # multi-row block w/ merged criteria
            {'rows': [['A', 'crit1'], ['B', 'crit2']]},             # raw rows, no merge
          ],
        }, ...
      ]

    role_cols: list of column indices that the role(s) span (default: just one column = product_col+1).
    criteria_col: column index for the criteria cell (default: last column).
    """
    role_cols = role_cols or [product_col + 1]
    role_col_start = role_cols[0]
    role_col_end   = role_cols[-1]
    if criteria_col is None:
        criteria_col = ncols
    band = start_band

    for grp in groups:
        product = grp['product']
        blocks  = grp['blocks']
        first_row_of_group = ws.max_row + 1
        last_row_of_group  = first_row_of_group - 1

        for block in blocks:
            # Resolve roles list (or raw rows)
            if 'rows' in block:
                # Each entry: [role, criteria]
                for entry in block['rows']:
                    band += 1
                    ws.append([""] * ncols)
                    rnum = ws.max_row
                    last_row_of_group = rnum
                    ws.cell(row=rnum, column=role_col_start, value=entry[0])
                    ws.cell(row=rnum, column=criteria_col,   value=entry[1])
                    for c in range(1, ncols + 1):
                        _style_cell(ws.cell(row=rnum, column=c), band)
                    ws.row_dimensions[rnum].height = 30
                    if role_col_end > role_col_start:
                        ws.merge_cells(start_row=rnum, start_column=role_col_start,
                                       end_row=rnum, end_column=role_col_end)
                continue

            roles    = block.get('roles') or [block['role']]
            criteria = block['criteria']
            block_first = ws.max_row + 1
            for role in roles:
                band += 1
                ws.append([""] * ncols)
                rnum = ws.max_row
                last_row_of_group = rnum
                ws.cell(row=rnum, column=role_col_start, value=role)
                # Criteria value only on the first row of the block; the merge takes care of display.
                if rnum == block_first:
                    ws.cell(row=rnum, column=criteria_col, value=criteria)
                for c in range(1, ncols + 1):
                    _style_cell(ws.cell(row=rnum, column=c), band)
                # Row-height tuned so wrapped criteria text is readable.
                lines = criteria.count("\n") + 1
                ws.row_dimensions[rnum].height = max(30, min(120, 18 + lines * 7))
                if role_col_end > role_col_start:
                    ws.merge_cells(start_row=rnum, start_column=role_col_start,
                                   end_row=rnum, end_column=role_col_end)
            # Merge the criteria column across the block
            if len(roles) > 1:
                ws.merge_cells(start_row=block_first, start_column=criteria_col,
                               end_row=block_first + len(roles) - 1, end_column=criteria_col)
                ws.cell(row=block_first, column=criteria_col).alignment = Alignment(
                    vertical="center", wrap_text=True, indent=1
                )

        # Set + merge product column down the entire group
        if last_row_of_group >= first_row_of_group:
            ws.cell(row=first_row_of_group, column=product_col, value=product)
            ws.cell(row=first_row_of_group, column=product_col).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ws.cell(row=first_row_of_group, column=product_col).font = cell_font(bold=True, color=NAVY)
            if last_row_of_group > first_row_of_group:
                ws.merge_cells(start_row=first_row_of_group, start_column=product_col,
                               end_row=last_row_of_group, end_column=product_col)

def write_note(ws, text, span):
    ws.append([text] + [""] * (span - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    cell = ws.cell(row=row, column=1)
    cell.font = cell_font(color=GREY, italic=True)
    cell.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[row].height = 24

def set_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─────────────────────────────────────────────────────────────────────────────
# TEAM BUILDING — currently uses the original code's pass/fail logic.
# (No fresh memo provided by the user; we keep the existing engine and
# document its criteria here so the xlsx matches what the engine enforces.)
# ─────────────────────────────────────────────────────────────────────────────
def build_team_building():
    wb = Workbook()
    ws = wb.active
    ws.title = "Team Building 2026"
    set_widths(ws, [14, 38, 56])
    write_title(ws, "TEAM BUILDING QUALIFICATION CRITERIA — 2026", 3)
    write_subtitle(ws, "Audience: ALL STAFF  •  Thresholds are CUMULATIVE (monthly value × months in window)", 3)

    leader_block = "≥ 100% cumulative target  AND  PAR > 30 ≤ 4%"

    write_section(ws, "AGENT THRESHOLDS", 3, color=GREEN)
    write_header(ws, ["PRODUCT", "AGENT CATEGORY", "CRITERIA"])
    write_grouped_rows(ws, [
        {
            'product': 'LBF',
            'blocks': [
                {'roles': ['Old Agent'],
                 'criteria': '• Minimum 4 loans / month\n• Disbursement ≥ TZS 20,000,000 / month'},
                {'roles': ['New Agent'],
                 'criteria': '• Minimum 3 loans / month\n• Disbursement ≥ TZS 15,000,000 / month'},
            ],
        },
        {
            'product': 'CS',
            'blocks': [
                {'roles': ['Old Agent — Tanzania Mainland'],
                 'criteria': '• Minimum 4 loans / month\n• Disbursement ≥ TZS 10,000,000 / month'},
                {'roles': ['Old Agent — Zanzibar'],
                 'criteria': '• Minimum 4 loans / month\n• Disbursement ≥ TZS 20,000,000 / month'},
                {'roles': ['New Agent — Tanzania Mainland'],
                 'criteria': '• Minimum 3 loans / month\n• Disbursement ≥ TZS 7,500,000 / month'},
                {'roles': ['New Agent — Zanzibar'],
                 'criteria': '• Minimum 3 loans / month\n• Disbursement ≥ TZS 15,000,000 / month'},
            ],
        },
        {
            'product': 'SME',
            'blocks': [
                {'roles': ['Old Agent'],
                 'criteria': '• Minimum 4 loans / month\n• Disbursement ≥ TZS 8,000,000 / month'},
                {'roles': ['New Agent'],
                 'criteria': '• Minimum 3 loans / month\n• Disbursement ≥ TZS 6,000,000 / month'},
            ],
        },
    ], ncols=3)

    write_section(ws, "LEADER THRESHOLDS", 3, color=GOLD)
    write_header(ws, ["PRODUCT", "LEVEL", "CRITERIA"])
    write_grouped_rows(ws, [
        {
            'product': 'LBF',
            'blocks': [{'roles': ['Team Leader', 'Region / BM'], 'criteria': leader_block}],
        },
        {
            'product': 'CS',
            'blocks': [{'roles': ['Team Leader', 'Region / BM'], 'criteria': leader_block}],
        },
        {
            'product': 'SME',
            'blocks': [{'roles': ['Team Leader', 'Region / BM'], 'criteria': leader_block}],
        },
    ], ncols=3)

    write_section(ws, "DEFINITIONS & RULES", 3, color=GREY)
    write_note(ws, "• Old agent  =  first Activities record  <  2026-01-01", 3)
    write_note(ws, "• New agent  =  first Activities record  in Jan – Apr 2026", 3)
    write_note(ws, "• Zanzibar   =  Supervision / Region label contains the word ZANZIBAR", 3)
    write_note(ws, "• IPF loans are excluded from the loan-count rollup.", 3)
    write_note(ws, "• PAR > 30 = (principal of loans with days-in-arrears > 30) ÷ (total principal).", 3)
    return wb

# ─────────────────────────────────────────────────────────────────────────────
# LOCAL TRIP — per the LOCAL TRIPS / SALES DEPARTMENTS pdf (2026)
# ─────────────────────────────────────────────────────────────────────────────
def build_local_trip():
    wb = Workbook()
    ws = wb.active
    ws.title = "Local Trip 2026"
    set_widths(ws, [14, 38, 56])
    write_title(ws, "LOCAL TRIP QUALIFICATION CRITERIA — 2026", 3)
    write_subtitle(ws, "Audience: SALES DEPARTMENTS  •  Source: LOCAL TRIPS memo (PCL Tanzania)", 3)
    write_header(ws, ["PRODUCT", "ROLE", "CRITERIA"])

    lbf_main = (
        "• Achieve 130% of your cumulative sales targets\n"
        "• Achieve 130% of your cumulative new-business targets\n"
        "• Achieve 130% of your loan counts\n"
        "• PAR 30 ≤ 4%"
    )
    cs_main = (
        "For Tanzania Mainland:\n"
        "• Achieve 150% of your cumulative Target YTD\n"
        "• Achieve 130% of your cumulative new-business targets\n"
        "• Achieve 130% of your loan counts\n"
        "• PAR 30 ≤ 4%"
    )
    cs_znz = (
        "For Zanzibar:\n"
        "• Achieve 130% of your cumulative Target YTD\n"
        "• Achieve 130% of your cumulative new-business targets\n"
        "• Achieve 130% of your loan counts\n"
        "• PAR 30 ≤ 4%"
    )
    cs_tele = (
        "• Achieve 130% of your cumulative new-business targets\n"
        "• Achieve 130% of your loan counts\n"
        "• PAR 30 ≤ 4%"
    )
    sme_block = (
        "• Achieve 120% of your cumulative sales target\n"
        "• Achieve 120% of your loan counts\n"
        "• PAR 30 ≤ 4%"
    )

    write_grouped_rows(ws, [
        {
            'product': 'LBF',
            'blocks': [
                {'roles': [
                    'Independent Team Leaders',
                    'Field Sales Team Leaders',
                    'Branch Managers & Call Center Supervisor (1 Trip)',
                    'Cluster Managers (1 Trip)',
                ], 'criteria': lbf_main},
                {'roles': [
                    'Telesales Team Leaders',
                    'Telesales Agents',
                ], 'criteria': lbf_main},
            ],
        },
        {
            'product': 'CS — Civil Servant',
            'blocks': [
                {'roles': [
                    'Independent Team Leaders',
                    'Field Sales Team Leaders',
                    'Branch Loan Officers',
                    'Regional Managers (1 Trip)',
                ], 'criteria': cs_main},
                {'roles': ['Cluster Managers (1 Trip)'], 'criteria': cs_znz},
                {'roles': [
                    'Telesales Team Leaders',
                    'Telesales Agents',
                ], 'criteria': cs_tele},
                {'roles': ['Sale Coordinator'],
                 'criteria': 'Ensure compliance with route plan creation at 95%'},
            ],
        },
        {
            'product': 'SME & AGRI',
            'blocks': [
                {'roles': [
                    'Senior Loan Officer',
                    'Regional Manager',
                    'Branch Manager',
                ], 'criteria': sme_block},
            ],
        },
    ], ncols=3)

    write_section(ws, "AGENT QUALIFICATION (only if their TL qualifies)", 3, color=GOLD)
    write_note(ws, "All staff & agents must have at least 3 months of work. Team Leaders who qualify select their agents from those who meet the criteria below.", 3)
    write_header(ws, ["PRODUCT", "CATEGORY", "AGENT CRITERIA"])

    write_grouped_rows(ws, [
        {
            'product': 'LBF',
            'blocks': [
                {'roles': ['Sales Agents (Old) — Joined before January 2026'],
                 'criteria':
                    '• Sale at least TZS 20,000,000 per month\n'
                    '• Minimum 4 loans per month (excluding IPF loans)\n'
                    '• Cumulative 48 loans (excluding IPF) until due date (6 × 8 months)'},
                {'roles': ['Sales Agents (New) — Joined from Jan 2026 – April 2026'],
                 'criteria':
                    '• Sale at least TZS 15,000,000 per month\n'
                    '• Minimum 3 loans per month (excluding IPF loans)\n'
                    '• Cumulative 20 loans (excluding IPF) until due date (5 × 4 months)'},
            ],
        },
        {
            'product': 'CS — Civil Servant',
            'blocks': [
                {'roles': ['Sales Agents (Old) — TZ Mainland'],
                 'criteria':
                    '• Sale at least TZS 10,000,000 per month\n'
                    '• Minimum 4 loans per month\n'
                    '• Cumulative 48 loans until due date (6 × 8 months)'},
                {'roles': ['Sales Agents (Old) — Zanzibar'],
                 'criteria':
                    '• Sale at least TZS 20,000,000 per month\n'
                    '• Minimum 4 loans per month\n'
                    '• Cumulative 48 loans until due date (6 × 8 months)'},
                {'roles': ['Sales Agents (New) — TZ Mainland'],
                 'criteria':
                    '• Sale at least TZS 7,500,000 per month\n'
                    '• Minimum 3 loans per month\n'
                    '• Cumulative 20 loans until due date (5 × 4 months)'},
                {'roles': ['Sales Agents (New) — Zanzibar'],
                 'criteria':
                    '• Sale at least TZS 15,000,000 per month\n'
                    '• Minimum 3 loans per month\n'
                    '• Cumulative 20 loans until due date (5 × 4 months)'},
            ],
        },
        {
            'product': 'SME & AGRI',
            'blocks': [
                {'roles': ['All sales agents'],
                 'criteria':
                    '• Sale at least TZS 8,000,000 per month\n'
                    '• Minimum 4 loans per month\n'
                    '• Cumulative 32 loans until due date (4 × 8 months)'},
            ],
        },
    ], ncols=3)

    write_section(ws, "NOTE", 3, color=GREY)
    write_note(ws, "• Qualification will also depend on the overall company performance.", 3)
    write_note(ws, "• The qualification selection is subject to review at the Management’s discretion.", 3)
    return wb

# ─────────────────────────────────────────────────────────────────────────────
# EA TRIP — per the EAST AFRICA TRIP / SALES MANAGERS pdf (2026)
# ─────────────────────────────────────────────────────────────────────────────
def build_ea_trip():
    wb = Workbook()
    ws = wb.active
    ws.title = "EA Trip 2026"
    set_widths(ws, [14, 42, 56])
    write_title(ws, "EAST AFRICA TRIP QUALIFICATION CRITERIA — 2026", 3)
    write_subtitle(ws, "Audience: SALES MANAGERS  •  Source: EAST AFRICA TRIP memo (PCL Tanzania)", 3)
    write_header(ws, ["PRODUCT", "ROLE", "CRITERIA"])

    write_grouped_rows(ws, [
        {
            'product': 'LBF',
            'blocks': [
                {'roles': ['Branch Managers & Call Center Supervisor'],
                 'criteria':
                    '• Achieve 130% of your cumulative sales targets YTD\n'
                    '• Achieve 130% of your loan counts\n'
                    '• PAR 30 ≤ 4%'},
            ],
        },
        {
            'product': 'CS',
            'blocks': [
                {'roles': ['Regional Managers (Tanzania Mainland)'],
                 'criteria':
                    'For Tanzania Mainland:\n'
                    '• Achieve 150% of your cumulative sales targets YTD\n'
                    '• Achieve 130% of your loan counts\n'
                    '• PAR 30 ≤ 4%'},
                {'roles': ['Regional Managers (Zanzibar)'],
                 'criteria':
                    'For Zanzibar:\n'
                    '• Achieve 130% of your cumulative sale target YTD\n'
                    '• Achieve 130% of your loan counts\n'
                    '• PAR 30 ≤ 4%'},
            ],
        },
        {
            'product': 'SME',
            'blocks': [
                {'roles': ['Regional Manager'],
                 'criteria':
                    '• Achieve 120% of your cumulative sales target YTD\n'
                    '• PAR 30 ≤ 4%'},
            ],
        },
    ], ncols=3)

    write_section(ws, "NOTE", 3, color=GREY)
    write_note(ws, "• All staff & agents must have at least 3 months of work.", 3)
    write_note(ws, "• Qualification will also depend on the overall company performance.", 3)
    write_note(ws, "• The qualification selection is subject to review at the Management’s discretion.", 3)
    return wb

# ─────────────────────────────────────────────────────────────────────────────
# EA TEAM BUILDING — per the EA TEAM BUILDING ALL STAFF pdf (2026) KE & UG
# ─────────────────────────────────────────────────────────────────────────────
def _write_eatb_table(ws, groups, ncols=4, start_band=0):
    """
    Render EA Team Building table with 4 columns (PRODUCT, ROLE, NO, CRITERIA).
    Merges PRODUCT down each group and CRITERIA across rows inside a block.

    groups: [
      {'product': 'LBF', 'blocks': [
          {'rows': [(role, no), (role, no), ...], 'criteria': '...'},
          ...
      ]}
    ]
    """
    band = start_band
    for grp in groups:
        product = grp['product']
        first   = ws.max_row + 1
        last    = first - 1
        for block in grp['blocks']:
            rows     = block['rows']
            criteria = block['criteria']
            blk_first = ws.max_row + 1
            for i, (role, no) in enumerate(rows):
                band += 1
                ws.append(["", role, no, criteria if i == 0 else ""])
                rnum = ws.max_row
                last = rnum
                for c in range(1, ncols + 1):
                    _style_cell(ws.cell(row=rnum, column=c), band)
                ws.cell(row=rnum, column=3).alignment = Alignment(horizontal="center", vertical="center")
                lines = criteria.count("\n") + 1
                ws.row_dimensions[rnum].height = max(26, min(110, 18 + lines * 7))
            if len(rows) > 1:
                ws.merge_cells(start_row=blk_first, start_column=4,
                               end_row=blk_first + len(rows) - 1, end_column=4)
        if last >= first:
            ws.cell(row=first, column=1, value=product)
            ws.cell(row=first, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=first, column=1).font = cell_font(bold=True, color=NAVY)
            if last > first:
                ws.merge_cells(start_row=first, start_column=1, end_row=last, end_column=1)

def build_ea_team_building():
    wb = Workbook()
    ws = wb.active
    ws.title = "EA Team Building 2026"
    set_widths(ws, [14, 30, 8, 52])
    write_title(ws, "EAST AFRICA TEAM BUILDING — 2026 (KE & UG)", 4)
    write_subtitle(ws, "Audience: ALL PCL STAFF  •  Source: EA TEAM BUILDING memo (PCL Tanzania)", 4)

    # ── KENYA ────────────────────────────────────────────────────────────────
    write_section(ws, "KENYA — 15 SLOTS", 4, color=RED)
    write_header(ws, ["PRODUCT", "ROLE", "NO", "CRITERIA"])
    _write_eatb_table(ws, [
        {'product': 'LBF', 'blocks': [
            {'rows': [('Sales Agents', 1)],     'criteria': 'Top performer (140% YTD)'},
            {'rows': [('Telesales',    1)],     'criteria': 'Agents must be active, month on month sales'},
            {'rows': [('Team Leaders', 1)],     'criteria': 'Top performer (130% YTD)'},
            {'rows': [('BM',           1)],     'criteria': 'Top performer (120% YTD)'},
        ]},
        {'product': 'CS', 'blocks': [
            {'rows': [('Sales Agents', 1)],     'criteria': 'Top performer (140% YTD)'},
            {'rows': [('FSTL',         1)],     'criteria': 'Top performer (150% YTD)'},
            {'rows': [('RSM', 1), ('Cluster Manager', 1)],
                                                'criteria': 'Top performer (120% YTD)'},
        ]},
        {'product': 'SME', 'blocks': [
            {'rows': [('Sales Agents', 1)],     'criteria': 'Top performer (130% YTD)'},
            {'rows': [('Team Leaders', 1)],     'criteria': 'Top performer (120% YTD)'},
            {'rows': [('RSM',          1)],     'criteria': 'Top performer (100% YTD)'},
        ]},
        {'product': 'AGRI', 'blocks': [
            {'rows': [('Sales Agents', 1)],     'criteria': 'Top performer (130% YTD)'},
        ]},
        {'product': 'BO', 'blocks': [
            {'rows': [('—',            2)],     'criteria': 'MGM Discretionary'},
        ]},
        {'product': 'RO', 'blocks': [
            {'rows': [('—',            1)],
             'criteria': 'Top performer\n• Collection Efficiency 90%\n• Retention 92%\n• PAR 30 < 1%'},
        ]},
    ])
    write_section(ws, "KENYA — TOTAL = 15", 4, color=NAVY)

    # ── UGANDA ───────────────────────────────────────────────────────────────
    write_section(ws, "UGANDA — 20 SLOTS", 4, color=RED)
    write_header(ws, ["PRODUCT", "ROLE", "NO", "CRITERIA"])
    _write_eatb_table(ws, [
        {'product': 'LBF', 'blocks': [
            {'rows': [('Sales Agents', 2)],     'criteria': 'Top performer (140% YTD)'},
            {'rows': [('Telesales',    1)],     'criteria': 'Agents must be active, month on month sales'},
            {'rows': [('Team Leaders', 1)],     'criteria': 'Top performer (130% YTD)'},
            {'rows': [('BM',           1)],     'criteria': 'Top performer (120% YTD)'},
            {'rows': [('Cluster Manager', 1)],  'criteria': '90% of TLs to be on target YTD'},
        ]},
        {'product': 'CS', 'blocks': [
            {'rows': [('Sales Agents Mainland', 2)], 'criteria': 'Top performer (140% YTD)'},
            {'rows': [('Sales Agents ZNZ', 1), ('BLO', 1)],
                                                'criteria': 'Top performer (150% YTD)'},
            {'rows': [('RSM',          1)],     'criteria': 'Top performer (120% YTD)'},
        ]},
        {'product': 'SME', 'blocks': [
            {'rows': [('Sales Agents', 2)],     'criteria': 'Top performer (130% YTD)'},
            {'rows': [('Team Leaders', 1)],     'criteria': 'Top performer (120% YTD)'},
        ]},
        {'product': 'AGRI', 'blocks': [
            {'rows': [('Sales Agents', 1)],     'criteria': 'Top performer (130% YTD)'},
            {'rows': [('Team Leaders', 1)],     'criteria': 'Top performer (120% YTD)'},
        ]},
        {'product': 'RO', 'blocks': [
            {'rows': [('—',            2)],
             'criteria': 'Top performer\n• Collection Efficiency 90%\n• Retention 92%\n• PAR 30 < 1%'},
        ]},
        {'product': 'BO', 'blocks': [
            {'rows': [('—',            2)],     'criteria': 'MGM Discretionary'},
        ]},
    ])
    write_section(ws, "UGANDA — TOTAL = 20", 4, color=NAVY)

    write_section(ws, "DISCLAIMER", 4, color=GREY)
    write_note(ws, "• If a staff member / agent qualifies for one trip, he/she won’t be eligible for the other trip.", 4)
    write_note(ws, "• The qualification selection is subject to review at the Management’s discretion.", 4)
    return wb

# ── output ────────────────────────────────────────────────────────────────────
def write(wb, folder, name):
    path = os.path.join(BASE, folder, name)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"  [ok] wrote {os.path.relpath(path, BASE)}")

def remove_if_exists(folder, name):
    path = os.path.join(BASE, folder, name)
    if os.path.exists(path):
        os.remove(path)
        print(f"  [rm] removed stale {os.path.relpath(path, BASE)}")

if __name__ == "__main__":
    print("Generating criteria files…")
    write(build_team_building(),      "TeamBuildingReport",     "TeamBuilding_Criteria_2026.xlsx")
    write(build_local_trip(),         "LocalTripReport",        "LocalTrip_Criteria_2026.xlsx")
    write(build_ea_trip(),            "EATripReport",           "EATrip_Criteria_2026.xlsx")
    write(build_ea_team_building(),   "EATeamBuildingReport",   "EATeamBuilding_Criteria_2026.xlsx")

    print("Cleaning up stale duplicates…")
    remove_if_exists("EATripReport",          "LocalTrip_Criteria_2026.xlsx")
    remove_if_exists("EATeamBuildingReport",  "LocalTrip_Criteria_2026.xlsx")
    # NB: LocalTripReport/LocalTrip_Criteria_2026.xlsx is the correct name for
    # that folder, so we leave it (we just overwrote it with fresh content).

    print("Done.")
