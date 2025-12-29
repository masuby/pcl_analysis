import pandas as pd
import numpy as np
import xlwings as xw
import matplotlib.pyplot as plt
import smtplib
import os
import glob
from datetime import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
import time
from dotenv import load_dotenv

# ============================================================================
# STEP 1: CONFIGURATION AND INITIALIZATION
# ============================================================================

def initialize_config():
    """Initialize configuration settings for SME script"""
    load_dotenv()
    config = {
        'base_dir': r"C:\Users\Daniel\Desktop\code\pcl\CRM\SME\NEW_EXCEL",  # Changed to SME directory
        'addin_path': r"C:\Users\Daniel\AppData\Roaming\Microsoft\AddIns\BranchFunctions.xlam",
        'screenshot_dir': 'screenshots',
        'sender_email': os.getenv("EMAIL_USERNAME"),
        'sender_password': os.getenv("EMAIL_PASSWORD"),
        'receiver_emails': [
            'raphael@platinumcredit.co.tz',
            'abdulhakim.khalfan@platinumcredit.co.tz',
            'abdulhakim.khalfan.platinum@gmail.com',
            'allan@platinumcredit.co.tz',
            'dorice@platinumcredit.co.tz',
            'sigfrid@platinumcredit.co.tz',
            'murigi@platinumcredit.co.ke',
            'wayne@platinumcredit.co.ke',
            'yusuph@platinumcredit.co.tz',
            'thomas@platinumcredit.co.tz',
            'wilhelm@platinumcredit.co.tz',
            'crmupdate.sme@platinumcredit.co.tz',
            'daniel@platinumcredit.co.tz'
        ]

    }
    return config

# ============================================================================
# STEP 2: EXCEL FILE MANAGEMENT
# ============================================================================

def get_latest_excel_file(base_dir):
    """
    Find the most recent SME Excel file in the specified directory
    Expected format: SME_CRM_DD_MM_YYYY.xlsx
    """
    try:
        # Pattern to match files like SME_CRM_28_11_2025.xlsx
        pattern = os.path.join(base_dir, "SME_CRM_*.xlsx")
        files = glob.glob(pattern)
        
        if not files:
            raise FileNotFoundError(f"No SME Excel files found matching pattern: {pattern}")
        
        # Sort files by embedded date in filename
        dated_files = []
        for file_path in files:
            filename = os.path.basename(file_path)
            # Extract date from filename
            parts = filename.replace('SME_CRM_', '').replace('.xlsx', '').split('_')
            if len(parts) == 3:
                day, month, year = parts
                try:
                    file_date = datetime(int(year), int(month), int(day))
                    dated_files.append((file_date, file_path))
                except ValueError:
                    dated_files.append((datetime.fromtimestamp(os.path.getmtime(file_path)), file_path))
        
        if dated_files:
            dated_files.sort(reverse=True)
            return dated_files[0][1]
        else:
            files.sort(key=os.path.getmtime, reverse=True)
            return files[0]
        
    except Exception as e:
        print(f"Error finding latest SME Excel file: {e}")
        return None

# ============================================================================
# STEP 3: EXCEL RECALCULATION WITH ADD-IN
# ============================================================================

def recalculate_excel_with_addin(excel_file, addin_path):
    """
    Recalculate Excel workbook with Add-in functions using xlwings
    """
    print(f"Recalculating SME Excel file: {excel_file}")
    
    app = None
    try:
        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False
        
        print(f"Loading Add-in: {addin_path}")
        if os.path.exists(addin_path):
            addin = app.books.open(addin_path)
            print("Add-in loaded successfully")
        
        wb = app.books.open(excel_file)
        print("Workbook opened")
        
        print("Performing full recalculation...")
        app.api.CalculateFull()
        
        time.sleep(2)
        
        wb.save()
        print("Recalculation completed and saved")
        
        wb.close()
        if 'addin' in locals():
            addin.close()
        
        return True
        
    except Exception as e:
        print(f"Error during Excel recalculation: {e}")
        return False
        
    finally:
        if app:
            app.quit()
            print("Excel application closed")

# ============================================================================
# STEP 4: DATA EXTRACTION AND CLEANING
# ============================================================================

def extract_data_from_excel(excel_file):
    """Extract data from SME Excel after recalculation"""
    print(f"Extracting data from SME file: {excel_file}")
    
    try:
        crm_sheets_data = pd.read_excel(
            excel_file,
            sheet_name=None,
            engine='openpyxl'
        )
        
        crm_email_data = crm_sheets_data.get('Email', pd.DataFrame())
        
        if crm_email_data.empty:
            print("Warning: 'Email' sheet is empty or not found")
            return pd.DataFrame()
        
        return crm_email_data
        
    except Exception as e:
        print(f"Error extracting data from Excel: {e}")
        return pd.DataFrame()

def clean_and_prepare_data(crm_email_data):
    """Clean the SME data - NO PERCENTAGE CONVERSION"""
    print("Cleaning and preparing SME data...")
    
    try:
        if 'Text' not in crm_email_data.columns:
            print("Error: 'Text' column not found in data")
            return crm_email_data
        
        crm_email_data = crm_email_data.dropna(subset=['Text']).copy()
        
        crm_email_data['Text'] = crm_email_data['Text'].astype(str).str.strip().str.lower()
        
        def clean_value(x):
            if pd.isna(x):
                return ''
            if isinstance(x, (int, float)):
                if isinstance(x, float):
                    return round(x, 2)
                return x
            return str(x).strip()
        
        crm_email_data['Value'] = crm_email_data['Value'].apply(clean_value)
        
        print(f"SME data cleaned successfully. {len(crm_email_data)} rows processed.")
        return crm_email_data
        
    except Exception as e:
        print(f"Error cleaning SME data: {e}")
        return crm_email_data

def format_cell_value(value):
    """Format cell values - simple version"""
    if pd.isna(value) or value == "" or str(value).strip() == "":
        return ""
    
    if isinstance(value, str):
        value = value.strip()
    
    if isinstance(value, str) and '%' in value:
        return value
    
    try:
        float_val = float(value)
        if float_val.is_integer():
            return str(int(float_val))
        else:
            return f"{float_val:.2f}"
    
    except (ValueError, TypeError):
        return str(value)

# ============================================================================
# STEP 5: TABLE IMAGE CREATION (UPDATED TO MATCH CS VERSION)
# ============================================================================

def calculate_dynamic_column_widths(data):
    """Calculate column widths based on actual content length"""
    if not data or len(data) == 0:
        return [0.1]  # Reduced from 0.15
    
    num_cols = len(data[0])
    max_content_lengths = [0] * num_cols
    
    for row in data:
        for col_idx, cell in enumerate(row):
            if col_idx < len(row):
                content_length = len(str(cell))
                max_content_lengths[col_idx] = max(max_content_lengths[col_idx], content_length)
    
    col_widths = []
    for max_len in max_content_lengths:
        base_width = 0.06  # REDUCED from 0.08
        length_factor = max_len * 0.02  # Reduced factor
        width = min(base_width + length_factor, 0.25)  # REDUCED max from 0.4 to 0.25
        col_widths.append(width)
    
    return col_widths

def format_cell_value_with_header(value, column_header=""):
    """
    Format cell values intelligently based on column header
    - Converts to percentage if column header contains '%', 'RATE', or 'PERCENTAGE'
    - Otherwise keeps numbers as they are
    """
    if pd.isna(value) or value == "" or str(value).strip() == "":
        return ""
    
    if isinstance(value, str):
        value = value.strip()
    
    # Check if it's already a formatted percentage
    if isinstance(value, str) and '%' in value:
        return value
    
    # Convert column header to lowercase for easier matching
    header_lower = str(column_header).lower()
    
    # Check if this column should contain percentages
    is_percentage_column = any(keyword in header_lower for keyword in 
                              ['%', 'rate', 'percentage', 'percent', 'ratio', 'share'])
    
    # Try to handle numbers
    try:
        # Try to convert to float
        float_val = float(value)
        
        if is_percentage_column:
            # Format as percentage with 2 decimal places
            return f"{float_val:.2%}"
        else:
            # Not a percentage column - format based on value
            if float_val.is_integer():
                return str(int(float_val))
            else:
                # Check if it's a decimal that looks like a percentage
                if 0 < float_val < 1 and 'rate' in header_lower:
                    # If it's a rate between 0-1, format as percentage
                    return f"{float_val:.2%}"
                else:
                    # Format as number with appropriate decimal places
                    # For whole numbers, no decimals
                    if abs(float_val) >= 100:
                        return f"{float_val:.0f}"
                    elif abs(float_val) >= 10:
                        return f"{float_val:.1f}"
                    else:
                        return f"{float_val:.2f}"
    
    except (ValueError, TypeError):
        # If conversion fails, return as string
        return str(value)

def extract_and_clean_data(workbook, sheet_name, min_col, max_col, min_row, max_row):
    """Extract data with cleaning and formatting"""
    try:
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found in workbook")
            return None
        
        sheet = workbook[sheet_name]
        data = []
        
        # First, extract column headers
        headers = []
        if min_row == 1:
            # Read header row
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else "")
        else:
            # If min_row > 1, we need to get headers from the first row
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=min_row, column=col).value
                headers.append(str(cell_value) if cell_value is not None else "")
        
        # Add headers to data
        if headers and any(h.strip() for h in headers):
            data.append(headers)
            start_data_row = min_row + 1 if min_row == 1 else min_row
        else:
            start_data_row = min_row
        
        # Extract data rows
        for row in range(start_data_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                # Pass column index to format_cell_value for context
                col_idx = col - min_col
                col_header = headers[col_idx] if col_idx < len(headers) else ""
                formatted_value = format_cell_value_with_header(cell_value, col_header)
                row_data.append(formatted_value)
            
            has_content = any(str(cell).strip() not in ['', 'None', 'NaN'] for cell in row_data)
            
            if has_content:
                data.append(row_data)
        
        if data:
            print(f"Extracted {len(data)} rows from {sheet_name}")
            # Print column headers for debugging
            if len(data) > 0:
                print(f"  Column headers: {data[0]}")
        else:
            print(f"No data extracted from {sheet_name}")
        
        return data if data else None
        
    except Exception as e:
        print(f"Extraction error from {sheet_name}: {e}")
        return None

def create_table_image(data, output_dir, filename, table_title=""):
    """Create table image with optimized column widths and layout - MATCHING CS VERSION"""
    try:
        if not data or len(data) == 0:
            print(f"No data to create image: {filename}")
            return None
        
        # REDUCED COLUMN WIDTHS
        col_widths = calculate_dynamic_column_widths(data)
        total_width = sum(col_widths) * 100  # Reduced multiplier
        num_rows = len(data)
        
        # Adjusted dimensions for better visibility
        fig_width = max(15, min(total_width, 30))  # Reduced for better fit
        fig_height = max(8, num_rows * 0.6)  # Increased height for larger fonts
        
        fig, ax = plt.subplots(figsize=(fig_width, fig_height))
        ax.axis('off')
        
        if table_title:
            plt.title(table_title, fontsize=20, fontweight='bold', pad=15, color='#2E75B6')
        
        # LEFT ALIGNMENT AND ENABLED TEXT WRAPPING
        table = ax.table(cellText=data, cellLoc='left', loc='center', colWidths=col_widths)
        table.auto_set_font_size(False)
        
        # Set font size for all cells - LARGER FONTS WITH BOLD TEXT
        for key, cell in table.get_celld().items():
            row, col = key
            if row < len(data) and col < len(data[0]):
                cell_text = str(data[row][col])
                # Set larger font size for body text with BOLD
                cell.set_fontsize(24)  # Body text font size
                cell.get_text().set_fontweight('bold')  # BOLD TEXT
                cell.get_text().set_wrap(True)  # Enable text wrapping
                # Set vertical alignment to top so text wraps to bottom
                cell.get_text().set_verticalalignment('top')
        
        # Format header row with LARGER FONT AND BOLD
        if len(data) > 0:
            for i in range(len(data[0])):
                table[(0, i)].set_facecolor('#2E75B6')
                # Set header font size to 24 with BOLD
                table[(0, i)].get_text().set_fontsize(24)
                table[(0, i)].get_text().set_fontweight('bold')
                table[(0, i)].get_text().set_color('white')
                table[(0, i)].get_text().set_wrap(True)  # Enable wrapping for headers too
                table[(0, i)].get_text().set_verticalalignment('top')
                table[(0, i)].set_height(0.15)  # Increased header height for larger font
                table[(0, i)].PAD = 0.03  # Slightly more padding for larger fonts
        
        # Alternate row colors
        for i in range(1, len(data)):
            color = '#F0F8FF' if i % 2 == 0 else '#FFFFFF'
            for j in range(len(data[0])):
                table[(i, j)].set_facecolor(color)
        
        # Set cell borders with adjusted heights for larger fonts
        for key, cell in table.get_celld().items():
            row, col = key
            cell.set_edgecolor('#666666')
            cell.set_linewidth(0.8)
            # Set larger cell height to accommodate wrapped text
            if row == 0:
                cell.set_height(0.15)  # Header row height
            else:
                cell.set_height(0.12)  # Body row height (increased for wrapping)
            cell.PAD = 0.03  # More padding for larger fonts
        
        # Increase padding for better layout with wrapped text
        plt.tight_layout(pad=3.0)
        image_path = os.path.join(output_dir, filename)
        
        # Save with optimized quality
        plt.savefig(image_path, dpi=200, bbox_inches='tight', pad_inches=0.5,
                   facecolor='white', edgecolor='none', transparent=False)
        plt.close()
        
        if os.path.exists(image_path) and os.path.getsize(image_path) > 1000:
            print(f"Table image created: {image_path} ({os.path.getsize(image_path)} bytes)")
            return image_path
        else:
            print(f"Failed to create table image: {image_path}")
            return None
            
    except Exception as e:
        print(f"Error creating table image {filename}: {e}")
        return None

def optimize_image_for_email(image_path, max_size_kb=500):
    """
    Optimize PNG image for email while maintaining good quality
    Uses different strategy than JPG compression
    """
    try:
        from PIL import Image
        
        if not os.path.exists(image_path):
            return image_path
        
        # Get current file size
        current_size_kb = os.path.getsize(image_path) / 1024
        
        if current_size_kb <= max_size_kb:
            print(f"  Image size OK: {current_size_kb:.1f}KB")
            return image_path
        
        print(f"  Optimizing image: {current_size_kb:.1f}KB -> target {max_size_kb}KB")
        
        with Image.open(image_path) as img:
            # For tables, PNG is better than JPG (no compression artifacts on text)
            # But we need to reduce file size
            
            # Strategy 1: Optimize PNG
            temp_path = image_path.replace('.png', '_opt.png')
            
            # Save with optimization
            img.save(temp_path, 'PNG', optimize=True)
            
            opt_size_kb = os.path.getsize(temp_path) / 1024
            print(f"  After PNG optimization: {opt_size_kb:.1f}KB")
            
            if opt_size_kb <= max_size_kb:
                os.replace(temp_path, image_path)
                return image_path
            
            # Strategy 2: Reduce color palette if needed (for simple tables)
            if len(img.getcolors(maxcolors=256)) is not None:
                # Convert to palette mode (reduces file size significantly)
                img_palette = img.convert('P', palette=Image.Palette.ADAPTIVE, colors=256)
                img_palette.save(temp_path, 'PNG', optimize=True)
                
                palette_size_kb = os.path.getsize(temp_path) / 1024
                print(f"  After palette conversion: {palette_size_kb:.1f}KB")
                
                if palette_size_kb <= max_size_kb:
                    os.replace(temp_path, image_path)
                    return image_path
            
            # Strategy 3: Slight reduction in dimensions if still too large
            if opt_size_kb > max_size_kb * 1.5:
                width, height = img.size
                if width > 2000 or height > 1500:
                    # Reduce by 20%
                    new_width = int(width * 0.8)
                    new_height = int(height * 0.8)
                    img_resized = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                    img_resized.save(temp_path, 'PNG', optimize=True)
                    
                    resize_size_kb = os.path.getsize(temp_path) / 1024
                    print(f"  After resizing: {resize_size_kb:.1f}KB")
                    
                    if os.path.exists(temp_path):
                        os.replace(temp_path, image_path)
            
            return image_path
            
    except ImportError:
        print("  PIL not available for optimization")
        return image_path
    except Exception as e:
        print(f"  Error optimizing image: {e}")
        return image_path

def create_jpg_alternative(png_path, jpg_path, quality=85):
    """Create a JPG alternative if PNG is still too large"""
    try:
        from PIL import Image
        
        if not os.path.exists(png_path):
            return None
        
        with Image.open(png_path) as img:
            # Convert to RGB for JPG
            rgb_img = img.convert('RGB')
            
            # Save as JPG with good quality
            rgb_img.save(jpg_path, 'JPEG', quality=quality, optimize=True)
            
            jpg_size_kb = os.path.getsize(jpg_path) / 1024
            print(f"  JPG alternative: {jpg_size_kb:.1f}KB (quality: {quality})")
            
            return jpg_path
            
    except Exception as e:
        print(f"  Error creating JPG: {e}")
        return None

# ============================================================================
# STEP 6: SCREENSHOT CREATION (MODIFIED FOR SME)
# ============================================================================

def create_screenshots(excel_file_path, output_dir=None):
    """Create screenshots with good quality for SME email"""
    print("Creating quality screenshots for SME...")
    
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(excel_file_path), "screenshots")
    
    os.makedirs(output_dir, exist_ok=True)
    print(f"Screenshot directory: {output_dir}")
    
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(excel_file_path, data_only=True)
        image_paths = {}
        
        # SME SPECIFIC CONFIGURATION - Different ranges than LBF
        screenshots_config = [
            {
                'name': 'leads_summary',
                'sheet': 'LEAD SUMMARY',
                'range': (2, 8, 2, 8),  # C2:H2 to C8:H8 (CHANGED for SME)
                'filename': 'leads_summary.png',
                'title': ''
            },
            {
                'name': 'agent_summary', 
                'sheet': 'summary',
                'range': (3, 19, 3, 9),  # C3:S3 to C9:S9 (CHANGED for SME)
                'filename': 'agent_summary.png',
                'title': ''
            },
            {
                'name': 'team_leader_summary',
                'sheet': 'summary', 
                'range': (23, 36, 3, 9),  # W3:AJ3 to W9:AJ9 (CHANGED for SME)
                'filename': 'team_leader_summary.png',
                'title': ''
            }
        ]
        
        for config in screenshots_config:
            print(f"\nProcessing SME {config['name']}...")
            min_col, max_col, min_row, max_row = config['range']
            data = extract_and_clean_data(workbook, config['sheet'], 
                                        min_col, max_col, min_row, max_row)
            
            if data and len(data) > 0:
                print(f"Creating image for {len(data)} rows, {len(data[0])} columns...")
                
                # Create PNG with updated styling (matching CS version)
                image_path = create_table_image(
                    data, output_dir, config['filename'], config.get('title', '')
                )
                
                if image_path and os.path.exists(image_path):
                    file_size_kb = os.path.getsize(image_path) / 1024
                    
                    # If PNG is still too large for email (>700KB), create JPG alternative
                    if file_size_kb > 700:
                        print(f"PNG too large ({file_size_kb:.1f}KB), creating JPG alternative...")
                        jpg_path = image_path.replace('.png', '.jpg')
                        jpg_created = create_jpg_alternative(image_path, jpg_path, quality=85)
                        
                        if jpg_created and os.path.exists(jpg_path):
                            jpg_size_kb = os.path.getsize(jpg_path) / 1024
                            
                            # If JPG is reasonable size, use it
                            if jpg_size_kb <= 500:
                                image_path = jpg_path
                                print(f"Using JPG: {jpg_size_kb:.1f}KB")
                            else:
                                # Try lower quality JPG
                                jpg_lower = create_jpg_alternative(image_path, 
                                                                  jpg_path.replace('.jpg', '_low.jpg'), 
                                                                  quality=75)
                                if jpg_lower and os.path.exists(jpg_lower):
                                    lower_size_kb = os.path.getsize(jpg_lower) / 1024
                                    if lower_size_kb <= 500:
                                        image_path = jpg_lower
                                        print(f"Using lower quality JPG: {lower_size_kb:.1f}KB")
                    
                    final_size_kb = os.path.getsize(image_path) / 1024
                    print(f"✓ SME {config['name']}: {final_size_kb:.1f}KB")
                    image_paths[config['name']] = image_path
                else:
                    print(f"✗ Failed to create SME {config['name']}")
            else:
                print(f"✗ No data found for SME {config['name']}")
        
        workbook.close()
        
        if image_paths:
            print(f"\n✓ Successfully created {len(image_paths)} SME screenshots")
            # Print final sizes and formats
            for name, path in image_paths.items():
                if os.path.exists(path):
                    size_kb = os.path.getsize(path) / 1024
                    format = os.path.splitext(path)[1].upper()
                    print(f"  - {name}: {size_kb:.1f}KB ({format})")
        else:
            print("✗ No SME screenshots were created")
        
        return image_paths
        
    except Exception as e:
        print(f"Error in SME screenshot process: {e}")
        import traceback
        traceback.print_exc()
        return {}

# ============================================================================
# STEP 7: HTML EMAIL GENERATION (UPDATED TO MATCH CS VERSION)
# ============================================================================

def generate_html_email(crm_email_data, image_paths, excel_file_path):
    """Generate HTML email content for SME - MATCHING CS VERSION STYLING"""
    print("Generating HTML email content for SME...")
    
    crm_email_data_clean = clean_and_prepare_data(crm_email_data)
    data_dict = dict(zip(crm_email_data_clean['Text'], crm_email_data_clean['Value']))
    
    current_date = datetime.now().strftime("%d-%m-%Y")
    
    filename = os.path.basename(excel_file_path)
    date_str = filename.replace('SME_CRM_', '').replace('.xlsx', '')  # Changed to SME
    try:
        day, month, year = date_str.split('_')
        report_date = f"{day} {datetime.strptime(month, '%m').strftime('%B')} {year}"
    except:
        report_date = datetime.now().strftime("%d %B %Y")
    
    def get_value(key, default="N/A"):
        """Get value from data dict with percentage formatting - MATCHING CS VERSION"""
        raw_value = data_dict.get(key.strip().lower(), default)
        
        # If value is N/A or empty, return as is
        if raw_value == "N/A" or raw_value == "" or raw_value is None:
            return raw_value
        
        # Check if key contains percentage indicators
        key_lower = key.lower()
        is_percentage_key = any(indicator in key_lower for indicator in 
                              ['percentage', 'percent', '%', 'rate', 'ratio'])
        
        try:
            # Try to convert to float
            if isinstance(raw_value, str):
                # Clean string value
                clean_str = raw_value.replace('%', '').strip()
                float_val = float(clean_str)
            else:
                float_val = float(raw_value)
            
            # Apply formatting - MATCHING CS VERSION LOGIC
            if is_percentage_key:
                # Convert to percentage with 2 decimal places
                if 0 <= float_val <= 1:
                    return f"{float_val:.2%}"  # 0.05 → 5.00%
                elif 0 <= float_val <= 100:
                    return f"{float_val:.2f}%"  # 5 → 5.00%
                else:
                    return f"{float_val:.2f}%"  # Fallback
            else:
                # Not a percentage - round to 0 decimal places
                if float_val.is_integer():
                    return str(int(float_val))
                else:
                    # For floats, round to 0 decimal places
                    return f"{float_val:.0f}"
        
        except (ValueError, TypeError):
            # If conversion fails, return as string
            return str(raw_value)
    
    # Calculate derived values for SME
    percentage_agent_location_reached = get_value('agent_location_reached', '0')
    try:
        if percentage_agent_location_reached != "N/A" and get_value('agent_location_planned', '0') != "N/A":
            reached = float(percentage_agent_location_reached)
            planned = float(get_value('agent_location_planned'))
            percentage_reached = f"{(reached/planned*100):.0f}%" if planned > 0 else "0%"
        else:
            percentage_reached = "0%"
    except:
        percentage_reached = "0%"
    
    # Calculate percentage_tl_location_reached
    tl_location_reached = get_value('tl_location_reached', '0')
    try:
        if tl_location_reached != "N/A" and get_value('tl_location_planned', '0') != "N/A":
            tl_reached = float(tl_location_reached)
            tl_planned = float(get_value('tl_location_planned'))
            tl_percentage_reached = f"{(tl_reached/tl_planned*100):.0f}%" if tl_planned > 0 else "0%"
        else:
            tl_percentage_reached = "0%"
    except:
        tl_percentage_reached = "0%"
    
    # HTML Template with OPTIMIZED CSS - MATCHING CS VERSION
    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>SME CRM Daily Report</title>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                line-height: 1.4;
                color: #333;
                margin: 0;
                padding: 0;
                width: 100%;
                background-color: #ffffff;
            }}
            .container {{
                width: 100%;
                margin: 0;
                padding: 0;
            }}
            .header {{
                background-color: #2E75B6;
                color: white;
                padding: 10px 15px;
                margin: 0;
                text-align: center;
                position: relative;
                overflow: hidden;
            }}
            .header::before {{
                content: '';
                position: absolute;
                bottom: 0;
                left: 10%;
                right: 10%;
                height: 3px;
                background: linear-gradient(90deg, rgba(255,255,255,0.3), rgba(255,255,255,0.8), rgba(255,255,255,0.3));
                border-radius: 2px;
            }}
            .header h1 {{
                margin: 0;
                font-size: 18px;
                font-weight: 600;
            }}
            .header p {{
                margin: 3px 0 0 0;
                font-size: 13px;
                opacity: 0.9;
            }}
            .section {{
                margin: 0;
                padding: 15px 15px;
                position: relative;
            }}
            .section::after {{
                content: '';
                position: absolute;
                bottom: 0;
                left: 10%;
                right: 10%;
                height: 2px;
                background: linear-gradient(90deg, transparent, #2E75B6 20%, #2E75B6 80%, transparent);
                border-radius: 1px;
                opacity: 0.7;
            }}
            .section:last-child::after {{
                display: none;
            }}
            .section-title {{
                color: #2E75B6;
                font-size: 16px;
                font-weight: 700;
                margin: 0 0 15px 0;
                padding: 0 0 12px 0;
                text-align: center;
                position: relative;
                letter-spacing: 0.5px;
            }}
            .section-title::after {{
                content: '';
                position: absolute;
                bottom: 0;
                left: 25%;
                right: 25%;
                height: 3px;
                background: linear-gradient(90deg, transparent, #2E75B6, transparent);
                border-radius: 2px;
            }}
            .subsection-title {{
                color: #2E75B6;
                font-size: 15px;
                font-weight: 600;
                margin: 15px 0 10px 0;
                padding: 0 0 10px 0;
                text-align: center;
                position: relative;
                letter-spacing: 0.3px;
            }}
            .subsection-title::after {{
                content: '';
                position: absolute;
                bottom: 0;
                left: 35%;
                right: 35%;
                height: 2px;
                background: linear-gradient(90deg, transparent, #2E75B6 40%, transparent);
                opacity: 0.8;
            }}
            .content {{
                font-size: 13px;
                line-height: 1.6;
                margin: 0;
                padding: 0;
            }}
            .content p {{
                margin: 8px 0;
                padding-left: 22px;
                position: relative;
                text-align: justify;
            }}
            .content p::before {{
                content: '●';
                position: absolute;
                left: 0;
                color: #000;
                font-size: 18px;
                line-height: 1;
                font-weight: bold;
                top: -1px;
            }}
            .highlight {{
                margin: 12px 0;
                padding: 0;
                position: relative;
            }}
            .highlight::before {{
                content: '';
                position: absolute;
                left: -5px;
                top: 0;
                bottom: 0;
                width: 3px;
                background: #2E75B6;
                border-radius: 2px;
                opacity: 0.6;
            }}
            .stat-box {{
                margin: 8px 0;
                padding: 0;
            }}
            .stat-value {{
                font-size: 14px;
                font-weight: bold;
                color: #2E75B6;
            }}
            .percentage-value {{
                font-size: 14px;
                font-weight: bold;
                color: #2E75B6;
            }}
            .number-value {{
                font-size: 14px;
                font-weight: bold;
                color: #2E75B6;
            }}
            .image-container {{
                margin: 15px 0;
                padding: 0;
                text-align: center;
                position: relative;
            }}
            .image-container::before {{
                content: '';
                position: absolute;
                top: 0;
                left: 10%;
                right: 10%;
                height: 1px;
                background: linear-gradient(90deg, transparent, #2E75B6 30%, transparent);
                opacity: 0.5;
            }}
            .image-container::after {{
                content: '';
                position: absolute;
                bottom: 0;
                left: 10%;
                right: 10%;
                height: 1px;
                background: linear-gradient(90deg, transparent, #2E75B6 30%, transparent);
                opacity: 0.5;
            }}
            .image-container img {{
                max-width: 100%;
                height: auto;
                display: block;
                margin: 0 auto;
            }}
            .image-title {{
                font-weight: 700;
                color: #2E75B6;
                margin: 0 0 8px 0;
                font-size: 14px;
                text-align: center;
                position: relative;
                padding-bottom: 8px;
            }}
            .image-title::after {{
                content: '';
                position: absolute;
                bottom: 0;
                left: 40%;
                right: 40%;
                height: 2px;
                background: #2E75B6;
                opacity: 0.6;
                border-radius: 1px;
            }}
            .image-wrapper {{
                overflow-x: auto;
                margin: 8px 0;
                padding: 0;
            }}
            .footer {{
                margin-top: 15px;
                padding: 15px 15px;
                color: #666;
                font-size: 12px;
                text-align: center;
                position: relative;
            }}
            .footer::before {{
                content: '';
                position: absolute;
                top: 0;
                left: 15%;
                right: 15%;
                height: 2px;
                background: linear-gradient(90deg, transparent, #2E75B6, transparent);
                opacity: 0.5;
            }}
            .signature {{
                margin-top: 10px;
                font-weight: 700;
                color: #333;
                text-align: left;
            }}
            .signature p {{
                margin: 5px 0;
                padding-left: 0;
            }}
            .signature p::before {{
                display: none;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>📊 SME CRM User, Activity and Leads Report</h1>
                <p>Report Date: {report_date} | Generated on: {current_date}</p>
            </div>
            
            <div class="section">
                <h2 class="section-title">📈 LEADS SUMMARY</h2>
                <div class="content">
                    <div class="stat-box">
                        <p><span class="number-value">{generated_lead}</span> leads were generated in the system across all SME branches.</p>
                    </div>
                    
                    <div class="stat-box">
                        <p><span class="percentage-value">{percentage_accepted_lead}</span> were accepted, 
                        <span class="percentage-value">{percentage_not_provided_lead}</span> (<span class="number-value">{number_not_provided_lead}</span>) were not provided with consent and 
                        <span class="percentage-value">{percentage_rejected_lead}</span> were rejected.</p>
                    </div>
                    
                    <div class="stat-box">
                        <p>Out of <span class="number-value">{generated_lead}</span> leads, <span class="number-value">{prospect_lead}</span> is a prospect.</p>
                    </div>
                    
                    <!-- Leads Summary Image -->
                    <div class="image-container">
                        <div class="image-title">📋 Detailed SME Leads Summary</div>
                        <div class="image-wrapper">
                            <img src="cid:leads_summary" alt="SME Leads Summary Table" 
                                 style="min-width: 1800px; height: auto; cursor: zoom-in;"
                                 onclick="this.style.transform = this.style.transform === 'scale(1.5)' ? 'scale(1)' : 'scale(1.5)'"
                                 title="Click to zoom in/out">
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <h2 class="section-title">🎯 MARKETING ACTIVITIES SUMMARY</h2>
                
                <div class="subsection-title">👥 Sales Agents</div>
                <div class="content">
                    <div class="stat-box">
                        <p>Total count of agents in CRM stood at <span class="number-value">{total_count_agent}</span>, 
                        and only <span class="number-value">{logged_in_agents}</span> agents logged in for the day.</p>
                    </div>
                    
                    <div class="stat-box">
                        <p>Out of <span class="number-value">{agent_assigned_activities}</span> agents assigned activities for the day, 
                        <span class="number-value">{agent_completed_at_location}</span> (<span class="percentage-value">{percentage_completed_at_location}</span>) 
                        agents completed at least one activity at the assigned location.</p>
                    </div>
                    
                    <div class="stat-box">
                        <p><span class="number-value">{agent_location_planned}</span> locations were planned for the day. 
                        Only <span class="number-value">{agent_location_reached}</span> (<span class="percentage-value">{percentage_agent_location_reached}</span>) 
                        locations were reached.</p>
                    </div>
                    
                    <div class="subsection-title">📅 For today {current_date}</div>
                    <div class="stat-box">
                        <p><span class="number-value">{today_location_planned}</span> locations have been planned.</p>
                        <p><span class="number-value">{today_agent_assigned_activities}</span> (<span class="percentage-value">{percentage_today_agent_assigned_activities}</span>) 
                        of total agents (<span class="number-value">{today_total_agents}</span>) have been assigned activities.</p>
                        <p>Average locations to be visited per agent is <span class="number-value">{average_location_to_be_visited}</span>.</p>
                    </div>
                    
                    <!-- Agent Summary Image -->
                    <div class="image-container">
                        <div class="image-title">👤 Detailed SME Agent Summary</div>
                        <div class="image-wrapper">
                            <img src="cid:agent_summary" alt="SME Agent Summary Table" 
                                 style="min-width: 2000px; height: auto; cursor: zoom-in;"
                                 onclick="this.style.transform = this.style.transform === 'scale(1.5)' ? 'scale(1)' : 'scale(1)'"
                                 title="Click to zoom in/out">
                        </div>
                    </div>
                </div>
                
                <div class="subsection-title">👨‍💼 Team Leaders</div>
                <div class="content">
                    <div class="stat-box">
                        <p>Total count of TLs in CRM stood at <span class="number-value">{total_count_team_leaders}</span>, 
                        only <span class="number-value">{logged_in_team_leaders}</span> logged in for the day.</p>
                    </div>
                    
                    <div class="stat-box">
                        <p>Out of <span class="number-value">{tl_assigned_activities}</span> TLs assigned activities, 
                        <span class="number-value">{tl_completed_activities_at_location}</span> (<span class="percentage-value">{percentage_tl_completed_activities_at_location}</span>) 
                        completed at least one activity at the assigned location.</p>
                    </div>
                    
                    <div class="stat-box">
                        <p><span class="number-value">{tl_location_planned}</span> locations were planned for the day, 
                        <span class="number-value">{tl_location_reached}</span> (<span class="percentage-value">{percentage_tl_location_reached}</span>) were reached.</p>
                    </div>
                    
                    <div class="highlight">
                        <p><span class="number-value">{tl_no_planned_visited_location}</span> had no planned location visited by a TL.</p>
                    </div>
                    
                    <div class="subsection-title">📅 For today {current_date}</div>
                    <div class="stat-box">
                        <p><span class="number-value">{today_tl_location_planned}</span> locations have been planned.</p>
                        <p><span class="number-value">{today_tl_assigned_activities}</span> (<span class="percentage-value">{percentage_tl_assigned_activities}</span>) of total TLs have been assigned activities.</p>
                        <p>Average locations to be visited per TL is <span class="number-value">{today_average_location_visited}</span>.</p>
                    </div>
                    
                    <!-- Team Leader Summary Image -->
                    <div class="image-container">
                        <div class="image-title">👨‍💼 Detailed SME Team Leader Summary</div>
                        <div class="image-wrapper">
                            <img src="cid:team_leader_summary" alt="SME Team Leader Summary Table" 
                                 style="min-width: 2000px; height: auto; cursor: zoom-in;"
                                 onclick="this.style.transform = this.style.transform === 'scale(1.5)' ? 'scale(1)' : 'scale(1)'"
                                 title="Click to zoom in/out">
                        </div>
                    </div>
                </div>
            </div>
            
            <div class="footer">
                <div class="signature">
                    <p>Best Regards,<br>
                    <strong>Daniel Masubi,</strong><br>
                    Senior Data Analyst and Sales Support</p>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Format the HTML with SME data
    html_content = html_template.format(
        report_date=report_date,
        current_date=current_date,
        
        # Leads Summary Data - SME SPECIFIC
        generated_lead=get_value('count_leads', '0'),
        percentage_accepted_lead=get_value('percentage_accepted_lead', '0%'),
        percentage_not_provided_lead=get_value('percentage_not_provided_lead', '0%'),
        number_not_provided_lead=get_value('not_provided_leads', '0'),
        percentage_rejected_lead=get_value('percentage_rejected_leads', '0%'),  # Note: different key name
        prospect_lead="none" if get_value('prospect_lead', '0') == "0" else get_value('prospect_lead'),
        
        # Agent Data - SME SPECIFIC
        total_count_agent=get_value('total_count_agent', '0'),
        logged_in_agents=get_value('logged_in_agents', '0'),
        agent_assigned_activities=get_value('agent_assigned_activities', '0'),
        agent_completed_at_location=get_value('agent_completed_at_location', '0'),
        percentage_completed_at_location=get_value('percentage_completed_at_location', '0%'),
        agent_location_planned=get_value('agent_location_planned', '0'),
        agent_location_reached=get_value('agent_location_reached', '0'),
        percentage_agent_location_reached=percentage_reached,  # Calculated value
        branch_agent_no_planned_location=get_value('branch_agent_no_planned_location', 'All branches'),
        branch_agent_no_planned_reached_location=get_value('branch_agent_no_planned_reached_location', 'All branches'),
        today_location_planned=get_value('today_location_planned', '0'),
        today_total_agents=get_value('today_total_agents', '0'),
        today_agent_assigned_activities=get_value('today_agent_assigned_activities', '0'),
        percentage_today_agent_assigned_activities=get_value('percentage_today_agent_assigned_activities', '0%'),
        average_location_to_be_visited=get_value('average_location_to_be_visited', '0'),
        
        # Team Leader Data - SME SPECIFIC
        total_count_team_leaders=get_value('total_count_team_leaders', '0'),
        logged_in_team_leaders=get_value('logged_in_team_leaders', '0'),
        tl_assigned_activities=get_value('tl_assigned_activities', '0'),
        tl_completed_activities_at_location=get_value('tl_completed_activities_at_location', '0'),
        percentage_tl_completed_activities_at_location=get_value('percentage_tl_completed_activities_at_location', '0%'),
        tl_location_planned=get_value('tl_location_planned', '0'),  # Note: tl_location_planned might need calculation
        tl_location_reached=tl_location_reached,
        percentage_tl_location_reached=tl_percentage_reached,  # Calculated value
        tl_no_planned_visited_location=get_value('tl_no_planned_visited_location', 'None'),
        today_tl_location_planned=get_value('today_tl_location_planned', '0'),
        today_tl_assigned_activities=get_value('today_tl_assigned_activities', '0'),
        percentage_tl_assigned_activities=get_value('percentage_tl_assigned_activities', '0%'),
        today_average_location_visited=get_value('today_average_location_visited', '0')
    )
    
    return html_content

# ============================================================================
# STEP 8: EMAIL SENDING (GROUP EMAIL WITH ALL IN TO FIELD)
# ============================================================================

def send_html_email_with_images(html_content, image_paths, config, excel_file_path, subject=None):
    """Send a single group email to all recipients - ALL in TO field for SME"""
    print("Sending SME CRM group email (all recipients in TO field)...")
    
    if subject is None:
        filename = os.path.basename(excel_file_path)
        try:
            date_str = filename.replace('SME_CRM_', '').replace('.xlsx', '')
            day, month, year = date_str.split('_')
            subject = f"SME CRM REPORT - {day}/{month}/{year}"
        except:
            subject = f"SME CRM REPORT - {datetime.now().strftime('%d %B %Y')}"
    
    try:
        if not config['receiver_emails']:
            print("⚠ No recipients configured")
            return False
            
        print(f"Preparing group email for {len(config['receiver_emails'])} recipients...")
        
        # Create fresh connection
        try:
            import ssl
            context = ssl.create_default_context()
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465, timeout=30, context=context)
        except:
            server = smtplib.SMTP('smtp.gmail.com', 587, timeout=30)
            server.starttls()
        
        # Login
        server.login(config['sender_email'], config['sender_password'])
        
        # Create message
        msg = MIMEMultipart()
        msg['From'] = config['sender_email']
        
        # ALL recipients in TO field (maximum visibility for group communication)
        msg['To'] = ', '.join(config['receiver_emails'])
        
        msg['Subject'] = subject
        
        # Add date header
        msg['Date'] = datetime.now().strftime("%a, %d %b %Y %H:%M:%S +0000")
        
        # Add reply-to header so replies go to the sender
        msg['Reply-To'] = config['sender_email']
        
        # Add headers for better email handling
        msg['X-Priority'] = '3'  # Normal priority
        msg['X-MSMail-Priority'] = 'Normal'
        msg['Importance'] = 'Normal'
        
        # HTML part
        html_part = MIMEText(html_content, 'html')
        msg.attach(html_part)
        
        # Attach images
        for image_name, image_path in image_paths.items():
            if os.path.exists(image_path):
                with open(image_path, 'rb') as f:
                    img = MIMEImage(f.read(), name=os.path.basename(image_path))
                    img.add_header('Content-ID', f'<{image_name}>')
                    msg.attach(img)
        
        # Attach Excel file
        if os.path.exists(excel_file_path):
            with open(excel_file_path, 'rb') as f:
                excel_part = MIMEBase('application', 'octet-stream')
                excel_part.set_payload(f.read())
                encoders.encode_base64(excel_part)
                excel_part.add_header(
                    'Content-Disposition',
                    f'attachment; filename="{os.path.basename(excel_file_path)}"'
                )
                msg.attach(excel_part)
        
        # Send the email to all recipients at once
        server.send_message(msg)
        
        print(f"\n✅ SME Group email sent successfully!")
        print(f"   Total recipients: {len(config['receiver_emails'])}")
        print(f"   All recipients in TO field: {msg['To']}")
        print(f"   Everyone can see all email addresses")
        print(f"   'Reply All' will include everyone in the conversation")
        
        # Close connection
        server.quit()
        
        return True
        
    except Exception as e:
        print(f"✗ Failed to send SME group email: {e}")
        import traceback
        traceback.print_exc()
        return False

# ============================================================================
# STEP 9: MAIN EXECUTION FUNCTION
# ============================================================================

def main():
    """Main execution function for SME"""
    print("=" * 60)
    print("SME CRM Email Automation System")
    print("=" * 60)
    
    print("\n[STEP 1] Initializing SME configuration...")
    config = initialize_config()
    
    print("\n[STEP 2] Finding latest SME Excel file...")
    excel_file = get_latest_excel_file(config['base_dir'])
    
    if not excel_file:
        print("❌ Error: No SME Excel file found. Exiting.")
        return
    
    print(f"✓ Found latest SME file: {excel_file}")
    
    print("\n[STEP 3] Recalculating Excel formulas with Add-in...")
    if not recalculate_excel_with_addin(excel_file, config['addin_path']):
        print("⚠ Warning: Excel recalculation had issues, but continuing...")
    
    print("\n[STEP 4] Extracting SME data from Excel...")
    crm_email_data = extract_data_from_excel(excel_file)
    
    if crm_email_data.empty:
        print("❌ Error: No SME data extracted from Excel. Exiting.")
        return
    
    print(f"✓ SME data extracted successfully: {len(crm_email_data)} rows")
    
    print("\n[STEP 5] Creating SME table images/screenshots...")
    image_paths = create_screenshots(excel_file)
    
    if not image_paths:
        print("❌ Error: Failed to create SME screenshots. Exiting.")
        return
    
    print(f"✓ Created {len(image_paths)} SME table images")
    
    print("\n[STEP 6] Generating SME HTML email content...")
    html_content = generate_html_email(crm_email_data, image_paths, excel_file)
    
    debug_html_path = os.path.join(config['base_dir'], f"sme_email_debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
    with open(debug_html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"✓ HTML saved for debugging: {debug_html_path}")
    
    print("\n[STEP 7] Sending SME HTML email with embedded images...")
    success = send_html_email_with_images(html_content, image_paths, config, excel_file)

    if success:
        print("\n" + "=" * 60)
        print("✅ SME PROCESS COMPLETED SUCCESSFULLY!")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("❌ SME PROCESS COMPLETED WITH ERRORS!")
        print("=" * 60)
    
    try:
        os.remove(debug_html_path)
        print(f"✓ Cleaned up debug file: {debug_html_path}")
    except:
        pass

# ============================================================================
# STEP 10: ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n⚠ SME process interrupted by user.")
    except Exception as e:
        print(f"\n\n❌ Unexpected error in SME process: {e}")
        import traceback
        traceback.print_exc()